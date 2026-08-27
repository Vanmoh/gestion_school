import hashlib
import io
import tempfile
import unicodedata
from urllib.parse import quote
from datetime import date
from pathlib import Path

from django.conf import settings
from django.core.paginator import Paginator
from django.db.models import Avg, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.permissions import AllowAny, IsAuthenticated

from apps.accounts.permissions import HasModuleAccess
from rest_framework.response import Response
from rest_framework.views import APIView
from apps.accounts.models import UserRole
from apps.school.models import (
    AcademicYear,
    ClassRoom,
    Etablissement,
    ExamPlanning,
    ExamResult,
    Expense,
    Grade,
    Payment,
    Student,
    StudentAcademicHistory,
    Subject,
    Teacher,
    TeacherAssignment,
)
from django.utils.html import escape

from apps.reports.card_verification import (
    signature_valide as signature_carte_valide,
)
from apps.reports.card_verification import signer as signer_carte
from apps.school.serializers import AcademicYearSerializer, PaymentSerializer, StudentSerializer
from apps.school.term_utils import normalize_term


def _pdf_text(value) -> str:
    return str(value or "").encode("latin-1", "replace").decode("latin-1")


def _ecourte(valeur: str, limite: int) -> str:
    """Coupe une valeur trop longue en le signalant.

    La coupe etait muette: « Mamadou Oualiyou Diallo » devenait « Mamadou
    Ouali », qui se lit comme un nom complet et faux. Les points de suspension
    disent au lecteur qu'il manque quelque chose.

    Trois points et non le caractere « … »: le PDF est encode en latin-1, qui
    ne le contient pas et l'aurait remplace par un point d'interrogation.
    """
    texte = str(valeur or "")
    if limite <= 0 or len(texte) <= limite:
        return texte
    return texte[: max(1, limite - 3)].rstrip() + "..."


def _school_logo_path() -> str | None:
    raw_path = str(getattr(settings, "SCHOOL_LOGO_PATH", "") or "").strip()
    if not raw_path:
        return None

    path = Path(raw_path)
    if not path.is_absolute():
        path = Path(settings.BASE_DIR) / path

    return str(path) if path.exists() else None


def _etablissement_logo_path(student: Student) -> str | None:
    etablissement = getattr(student, "etablissement", None)
    if etablissement is None and getattr(student, "classroom", None) is not None:
        etablissement = getattr(student.classroom, "etablissement", None)
    if etablissement is None:
        return None

    logo_field = getattr(etablissement, "logo", None)
    if not logo_field:
        return None

    try:
        direct_path = Path(getattr(logo_field, "path", "") or "")
    except Exception:
        direct_path = None

    if direct_path and direct_path.exists():
        return str(direct_path)

    logo_name = str(getattr(logo_field, "name", "") or "").strip()
    media_root = str(getattr(settings, "MEDIA_ROOT", "") or "").strip()
    if logo_name and media_root:
        candidate = Path(media_root) / logo_name
        if candidate.exists():
            return str(candidate)

    return None


def _student_etablissement(student: Student | None) -> Etablissement | None:
    if student is None:
        return None
    etablissement = getattr(student, "etablissement", None)
    if etablissement is None and getattr(student, "classroom", None) is not None:
        etablissement = getattr(student.classroom, "etablissement", None)
    return etablissement


def _payment_etablissement(payment: Payment | None) -> Etablissement | None:
    if payment is None:
        return None
    etablissement = getattr(payment, "etablissement", None)
    if etablissement is not None:
        return etablissement
    fee = getattr(payment, "fee", None)
    student = getattr(fee, "student", None) if fee is not None else None
    return _student_etablissement(student)


def _etablissement_media_field_path(etablissement: Etablissement | None, field_name: str) -> str | None:
    if etablissement is None:
        return None

    media_field = getattr(etablissement, field_name, None)
    if not media_field:
        return None

    try:
        direct_path = Path(getattr(media_field, "path", "") or "")
    except Exception:
        direct_path = None

    if direct_path and direct_path.exists():
        return str(direct_path)

    media_name = str(getattr(media_field, "name", "") or "").strip()
    media_root = str(getattr(settings, "MEDIA_ROOT", "") or "").strip()
    if media_name and media_root:
        candidate = Path(media_root) / media_name
        if candidate.exists():
            return str(candidate)

    return None


def _safe_scale_percent(value, default: int = 100) -> int:
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        parsed = default
    return max(40, min(200, parsed))


def _positioned_x(position: str, *, min_x: float, max_x: float, box_width: float, default_x: float) -> float:
    if max_x <= min_x:
        return min_x
    if position == "left":
        return min_x
    if position == "center":
        return min_x + max(0.0, (max_x - min_x - box_width) / 2.0)
    return min(max(default_x, min_x), max_x - box_width)


def _school_signature_asset_path() -> str | None:
    candidates = [
        Path(settings.BASE_DIR) / "assets" / "images" / "str_signature.png",
        Path(settings.BASE_DIR).parent / "frontend" / "gestion_school_app" / "assets" / "images" / "str_signature.png",
    ]
    for candidate in candidates:
        if candidate.exists():
            return str(candidate)
    return None


def _school_stamp_asset_path() -> str | None:
    candidates = [
        Path(settings.BASE_DIR) / "assets" / "images" / "str_cachet_signature.png",
        Path(settings.BASE_DIR).parent / "frontend" / "gestion_school_app" / "assets" / "images" / "str_cachet_signature.png",
    ]
    for candidate in candidates:
        if candidate.exists():
            return str(candidate)
    return None


def _pdf_compatible_image_path(source_path: str | None, *, cache_prefix: str) -> str | None:
    if not source_path:
        return None

    source = Path(source_path)
    if not source.exists():
        return None

    if source.suffix.lower() in {".jpg", ".jpeg"}:
        return str(source)

    try:
        cache_dir = Path(tempfile.gettempdir()) / "gestion_school_pdf_assets"
        cache_dir.mkdir(parents=True, exist_ok=True)

        stat = source.stat()
        cache_key = f"{source.resolve()}::{stat.st_mtime_ns}::{stat.st_size}"
        cache_hash = hashlib.sha1(cache_key.encode("utf-8")).hexdigest()[:12]
        cached_file = cache_dir / f"{cache_prefix}_{cache_hash}.jpg"

        if cached_file.exists():
            return str(cached_file)

        with Image.open(source) as image:
            if image.mode in {"RGBA", "LA"}:
                rgb = Image.new("RGB", image.size, (255, 255, 255))
                rgb.paste(image.convert("RGB"), mask=image.getchannel("A"))
            elif image.mode == "P":
                rgba = image.convert("RGBA")
                rgb = Image.new("RGB", rgba.size, (255, 255, 255))
                rgb.paste(rgba, mask=rgba.getchannel("A"))
            else:
                rgb = image.convert("RGB")

            rgb.save(cached_file, format="JPEG", quality=95, optimize=True)

        return str(cached_file)
    except Exception:
        # Last resort: return original path and let FPDF attempt loading it.
        return str(source)


def _photo_cadree(source_path: str | None, ratio: float) -> str | None:
    """Photo recadree au centre pour remplir un cadre de proportion donnee.

    `keep_aspect_ratio` seul empeche la deformation mais laisse deux bandes
    blanches quand la photo n'a pas les proportions du cadre. Un recadrage
    central remplit le cadre sans etirer personne, comme le ferait une photo
    d'identite passee sous massicot.
    """
    if not source_path or ratio <= 0:
        return None
    source = Path(source_path)
    if not source.exists():
        return None
    try:
        cache_dir = Path(tempfile.gettempdir()) / "gestion_school_pdf_assets"
        cache_dir.mkdir(parents=True, exist_ok=True)
        stat = source.stat()
        cle = f"{source.resolve()}::{stat.st_mtime_ns}::{stat.st_size}::{ratio:.4f}"
        cible = cache_dir / f"portrait_{hashlib.sha1(cle.encode()).hexdigest()[:12]}.jpg"
        if cible.exists():
            return str(cible)

        with Image.open(source) as image:
            image = image.convert("RGB")
            largeur, hauteur = image.size
            if largeur <= 0 or hauteur <= 0:
                return None
            if largeur / hauteur > ratio:
                neuve_l = int(hauteur * ratio)
                gauche = (largeur - neuve_l) // 2
                boite = (gauche, 0, gauche + neuve_l, hauteur)
            else:
                neuve_h = int(largeur / ratio)
                # Legerement au-dessus du centre: sur un portrait, le visage
                # se tient dans le tiers superieur, pas au milieu.
                haut = max(0, int((hauteur - neuve_h) * 0.35))
                boite = (0, haut, largeur, haut + neuve_h)
            image.crop(boite).save(cible, format="JPEG", quality=92, optimize=True)
        return str(cible)
    except Exception:
        return None


def _carte_verification_url(student, annee: str, base_url: str) -> str:
    """Adresse portee par le QR de la carte."""
    signature = signer_carte(getattr(student, "id", 0) or 0, annee)
    return f"{base_url.rstrip('/')}/api/reports/carte/{student.id}/{annee}/{signature}/"


def _carte_qr_image_path(url: str) -> str | None:
    """Image du QR, mise en cache par URL.

    Renvoie None si la generation echoue: une carte sans QR reste une carte,
    alors qu'une exception ici priverait l'ecole de toute sa planche.
    """
    if not url:
        return None
    try:
        import qrcode

        cache_dir = Path(tempfile.gettempdir()) / "gestion_school_pdf_assets"
        cache_dir.mkdir(parents=True, exist_ok=True)
        cache_hash = hashlib.sha1(url.encode("utf-8")).hexdigest()[:12]
        cached_file = cache_dir / f"qr_{cache_hash}.png"
        if cached_file.exists():
            return str(cached_file)

        code = qrcode.QRCode(
            version=None,
            # Correction haute: une carte se plie, se salit et se raye. Un QR
            # imprime a 15 mm ne se relit pas si le moindre module manque.
            error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=10,
            border=2,
        )
        code.add_data(url)
        code.make(fit=True)
        code.make_image(fill_color="black", back_color="white").save(cached_file)
        return str(cached_file)
    except Exception:
        return None


def pdf_output_response(pdf: FPDF, filename: str) -> HttpResponse:
    data = bytes(pdf.output())
    response = HttpResponse(data, content_type="application/pdf")
    # Un nom de classe accentue partait tel quel dans l'en-tete: « cartes_
    # 1ère_Année_EM2.pdf » y arrivait en octets non-ASCII, que la norme
    # n'admet pas. La RFC 6266 demande deux formes: un repli ASCII pour les
    # clients anciens, et `filename*` encode pour les autres.
    repli_ascii = (
        unicodedata.normalize("NFKD", filename)
        .encode("ascii", "ignore")
        .decode("ascii")
    ) or "document.pdf"
    response["Content-Disposition"] = (
        f'attachment; filename="{repli_ascii}"; '
        f"filename*=UTF-8''{quote(filename)}"
    )
    # Prevent stale browser/proxy cache for dynamically generated PDFs.
    response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    response["X-Card-Template-Version"] = "2026-03-14-2"
    return response


def _school_identity() -> dict[str, str]:
    return {
        "name": getattr(settings, "SCHOOL_NAME", "LYCEE TECHNIQUE OUMAR BAH"),
        "short": getattr(settings, "SCHOOL_SHORT", "LTOB"),
        "level": getattr(settings, "SCHOOL_LEVEL", "1er etage"),
        "phone": getattr(settings, "SCHOOL_PHONE", ""),
        "city": getattr(settings, "SCHOOL_CITY", "DAKAR"),
    }


_MOTS_LIAISON = {"de", "du", "des", "la", "le", "les", "et", "d", "l"}


def _school_acronym(name: str) -> str:
    """Sigle d'un etablissement, tire des initiales de ses mots.

    `name[:16]` coupait au milieu d'un mot: « Complexe Scolaire Oumar Bah »
    s'imprimait « COMPLEXE SCOLAIR ». Un nom deja court, ou deja sigle, est
    garde tel quel: abreger « LTOB » n'aurait aucun sens.
    """
    propre = " ".join(str(name or "").split())
    if not propre:
        return ""
    if len(propre) <= 12 or " " not in propre:
        return propre.upper()

    initiales = "".join(
        mot[0]
        for mot in propre.replace("-", " ").split()
        if mot and mot.lower().strip(".") not in _MOTS_LIAISON
    )
    # Deux lettres ne distinguent rien; on retombe alors sur le nom entier,
    # que la mise en page reduira si besoin.
    return initiales.upper() if len(initiales) >= 3 else propre.upper()


def _school_identity_for_student(student: Student) -> dict[str, str]:
    etablissement = getattr(student, "etablissement", None)
    if etablissement is None and getattr(student, "classroom", None) is not None:
        etablissement = getattr(student.classroom, "etablissement", None)

    if etablissement is None:
        return _school_identity()

    # Volontairement sans repli sur _school_identity(): ses valeurs decrivent
    # un etablissement precis (nom, telephone, etage du LTOB). Les heriter
    # faisait imprimer le telephone du LTOB sur les documents des trois autres
    # ecoles, dont les fiches n'ont ni telephone ni adresse. Mieux vaut une
    # ligne absente qu'une ligne qui renvoie vers la mauvaise ecole.
    nom = str(getattr(etablissement, "name", "") or "").strip()
    return {
        "name": nom,
        "short": _school_acronym(nom),
        "level": str(getattr(etablissement, "address", "") or "").strip(),
        "phone": str(getattr(etablissement, "phone", "") or "").strip(),
        "city": str(getattr(etablissement, "city", "") or "").strip(),
    }


def _active_academic_year(etablissement=None):
    """Annee scolaire active, ou la plus recente a defaut.

    La resolution vit desormais sur le modele: trois endroits la faisaient
    chacun a sa facon -- `.first()`, `-id`, `-start_date` -- et rien ne
    garantissait qu'ils designent la meme annee le meme jour.
    """
    courante = AcademicYear.courante(etablissement)
    if courante is not None:
        return courante

    repli = AcademicYear.objects.all()
    if etablissement is not None:
        repli = repli.filter(etablissement=etablissement)
    return repli.order_by("-start_date", "-id").first()


def _active_academic_year_label() -> str:
    year = _active_academic_year()

    if year is None:
        current_year = timezone.localdate().year
        return f"{current_year} - {current_year + 1}"

    year_name = str(getattr(year, "name", "") or "").strip()
    if year_name:
        return year_name

    if year.start_date and year.end_date:
        return f"{year.start_date.year} - {year.end_date.year}"

    return f"Annee {year.id}"


def _student_name_parts(student: Student) -> tuple[str, str, str]:
    student_user = student.user
    if not student_user:
        return "-", "-", "-"

    first_name = (student_user.first_name or "").strip()
    last_name = (student_user.last_name or "").strip()
    full_name = (student_user.get_full_name() or "").strip() or student_user.username

    if not first_name and full_name:
        first_name = full_name.split(" ", 1)[0]
    if not last_name and full_name and " " in full_name:
        last_name = full_name.split(" ", 1)[1]

    return first_name or "-", last_name or "-", full_name or "-"


def _student_photo_path(student: Student) -> str | None:
    photo_field = getattr(student, "photo", None)
    if not photo_field:
        return None

    try:
        direct_path = Path(getattr(photo_field, "path", "") or "")
    except Exception:
        direct_path = None

    if direct_path and direct_path.exists():
        return str(direct_path)

    photo_name = str(getattr(photo_field, "name", "") or "").strip()
    media_root = str(getattr(settings, "MEDIA_ROOT", "") or "").strip()
    if photo_name and media_root:
        candidate = Path(media_root) / photo_name
        if candidate.exists():
            return str(candidate)

    return None


def _format_fcfa(value) -> str:
    raw = str(value or "0").strip()
    if not raw:
        raw = "0"

    if "FCFA" in raw.upper():
        return raw

    normalized = raw.replace(" ", "").replace(",", ".")
    try:
        parsed = int(round(float(normalized)))
        return f"{parsed:,}".replace(",", " ") + " FCFA"
    except Exception:
        return f"{raw} FCFA"


def _draw_card_separator_line(pdf: FPDF, x1: float, y: float, x2: float) -> None:
    if x2 <= x1:
        return

    pdf.set_draw_color(165, 181, 205)
    try:
        pdf.dashed_line(x1, y, x2, y, dash_length=0.9, space_length=0.8)
    except Exception:
        pdf.line(x1, y, x2, y)


def _draw_student_card_template(
    pdf: FPDF,
    student: Student,
    *,
    school: dict[str, str],
    logo_path: str | None,
    x: float,
    y: float,
    width: float,
    height: float,
    verify_base_url: str = "",
) -> None:
    if width <= 0 or height <= 0:
        return

    compact = width < 72
    outer_line_w = max(0.12, min(0.44, width * 0.0034))
    inset = max(0.36, min(1.16, min(width, height) * 0.016))
    pad_x = max(0.55, min(1.9, width * 0.014))
    pad_y = max(0.45, min(1.45, height * 0.014))

    pdf.set_fill_color(246, 248, 252)
    pdf.set_draw_color(90, 99, 114)
    pdf.set_line_width(outer_line_w)
    pdf.rect(x, y, width, height, style="DF")

    pdf.set_draw_color(160, 169, 184)
    pdf.set_line_width(max(0.08, outer_line_w * 0.62))
    pdf.rect(x + inset, y + inset, width - (2 * inset), height - (2 * inset))

    content_x = x + inset + pad_x
    content_y = y + inset + pad_y
    content_w = width - (2 * (inset + pad_x))
    content_h = height - (2 * (inset + pad_y))
    if content_w <= 0 or content_h <= 0:
        return

    # Aucune valeur de repli: une carte qui affiche le telephone d'une autre
    # ecole invite a appeler le mauvais numero, ce qu'une ligne absente ne fait
    # pas. Trois des quatre etablissements n'ont ni telephone ni adresse.
    school_name_raw = (school.get("name") or "").strip().upper()
    school_short_raw = (school.get("short") or "").strip().upper()
    school_level_raw = (school.get("level") or "").strip().upper()
    school_phone_raw = (school.get("phone") or "").strip()

    school_name = _pdf_text(school_name_raw)
    # « LTOB » sous « LTOB » n'apprend rien: le sigle ne s'affiche que s'il
    # differe du nom deja ecrit au-dessus.
    sigle = school_short_raw if school_short_raw != school_name_raw else ""
    if sigle and school_level_raw:
        school_subtitle = _pdf_text(f"{sigle} ({school_level_raw})")
    else:
        school_subtitle = _pdf_text(sigle or school_level_raw)
    school_phone = _pdf_text(f"Tel : {school_phone_raw}") if school_phone_raw else ""

    # L'en-tete reservait toujours la place de trois lignes. Quand l'ecole n'a
    # ni sigle distinct ni telephone, cette reserve laissait une bande vide.
    lignes_entete = 1 + bool(school_subtitle) + bool(school_phone)
    header_h = max(8.7, min(16.2, content_h * 0.22)) * (0.52 + 0.16 * lignes_entete)
    header_name_font = 11.0 if width >= 120 else 8.8 if width >= 85 else 6.8 if width >= 72 else 5.8
    header_sub_font = header_name_font * 0.86
    header_phone_font = header_name_font * 0.66

    pdf.set_text_color(20, 70, 136)
    pdf.set_xy(content_x, content_y)
    pdf.set_font("Helvetica", "B", header_name_font)
    pdf.cell(content_w, max(2.7, header_h * 0.34), school_name, align="C")

    if school_subtitle:
        pdf.set_text_color(44, 45, 59)
        pdf.set_xy(content_x, content_y + max(2.5, header_h * 0.31))
        pdf.set_font("Helvetica", "B", header_sub_font)
        pdf.cell(content_w, max(2.3, header_h * 0.25), school_subtitle, align="C")

    if school_phone:
        pdf.set_text_color(177, 59, 67)
        pdf.set_xy(content_x, content_y + max(4.8, header_h * 0.56))
        pdf.set_font("Helvetica", "B", header_phone_font)
        pdf.cell(content_w, max(2.0, header_h * 0.18), school_phone, align="C")

    title_y = content_y + header_h + max(0.45, content_h * 0.008)
    title_h = max(3.1, min(5.9, content_h * 0.088))
    pdf.set_fill_color(27, 93, 168)
    pdf.rect(content_x, title_y, content_w, title_h, style="F")
    pdf.set_text_color(255, 255, 255)
    pdf.set_xy(content_x, title_y + 0.05)
    pdf.set_font(
        "Helvetica",
        "B",
        10.6 if width >= 120 else 8.6 if width >= 85 else 6.8 if width >= 72 else 5.5,
    )
    pdf.cell(content_w, max(2.0, title_h - 0.1), _pdf_text("CARTE SCOLAIRE"), align="C")

    footer_h = max(10.2, min(17.6, content_h * 0.24))
    body_top = title_y + title_h + max(0.62, content_h * 0.01)
    body_bottom = content_y + content_h - footer_h - max(0.32, content_h * 0.006)
    if body_bottom <= body_top:
        body_bottom = body_top + max(9.5, content_h * 0.28)

    photo_x = content_x
    photo_y = body_top
    photo_w = max(13.0, min(34.0, content_w * (0.30 if not compact else 0.34)))
    # Proportion d'une photo d'identite (35 x 45 mm). Le cadre occupait toute
    # la hauteur du corps, bien plus allongee, d'ou de larges bandes vides.
    photo_h = min(
        max(15.0, body_bottom - body_top),
        photo_w * (45.0 / 35.0),
    )

    pdf.set_fill_color(255, 255, 255)
    pdf.set_draw_color(50, 106, 176)
    pdf.set_line_width(max(0.09, outer_line_w * 0.57))
    pdf.rect(photo_x, photo_y, photo_w, photo_h, style="DF")

    photo_path = _photo_cadree(
        _student_photo_path(student),
        (photo_w - 1.0) / max(0.1, photo_h - 1.0),
    )
    if photo_path:
        try:
            pdf.image(
                photo_path,
                x=photo_x + 0.5,
                y=photo_y + 0.5,
                w=max(1.0, photo_w - 1.0),
                h=max(1.0, photo_h - 1.0),
            )
        except Exception:
            photo_path = None

    if not photo_path:
        pdf.set_fill_color(232, 238, 248)
        pdf.rect(photo_x + 0.5, photo_y + 0.5, photo_w - 1.0, photo_h - 1.0, style="F")
        pdf.set_text_color(95, 108, 128)
        pdf.set_xy(photo_x, photo_y + (photo_h / 2.0) - 1.4)
        pdf.set_font("Helvetica", "B", 6.1 if width >= 85 else 4.9)
        pdf.cell(photo_w, 2.6, _pdf_text("PHOTO"), align="C")

    first_name, last_name, _ = _student_name_parts(student)
    class_name = student.classroom.name if student.classroom else "Non attribuee"
    birth_date = student.birth_date.strftime("%d/%m/%Y") if student.birth_date else "-"
    year_label = _active_academic_year_label()

    # Une carte sans echeance reste valable indefiniment aux yeux de celui qui
    # la controle: celle de 2019 ressemble a celle de cette annee.
    annee = _active_academic_year()
    fin_annee = getattr(annee, "end_date", None)
    validity_label = (
        f"Valable jusqu'au {fin_annee.strftime('%d/%m/%Y')}" if fin_annee else ""
    )
    qr_url = (
        _carte_verification_url(student, year_label, verify_base_url)
        if verify_base_url and getattr(student, "id", 0)
        else ""
    )

    info_x = photo_x + photo_w + max(1.0, min(3.4, content_w * 0.024))
    info_w = max(8.0, (content_x + content_w) - info_x)
    label_w = max(6.5, min(info_w * 0.42, info_w - 4.0))
    # Six lignes suivies de leurs ecarts occupent 8,27 fois la hauteur d'une
    # ligne. A 0,135 elles debordaient de 12 % sous le pied de carte, ou la
    # derniere venait chevaucher la mention de validite.
    row_h = max(2.1, min(4.0, (body_bottom - body_top) * 0.118))
    row_gap = max(0.45, min(1.25, row_h * 0.42))
    label_font = 8.2 if width >= 120 else 6.7 if width >= 85 else 5.4 if width >= 72 else 4.7
    value_font = label_font * 1.02
    value_limit = 40 if width >= 120 else 31 if width >= 85 else 24 if width >= 72 else 17

    def _draw_info_row(row_y: float, label: str, value: str) -> None:
        pdf.set_xy(info_x, row_y)
        pdf.set_text_color(44, 48, 59)
        pdf.set_font("Helvetica", "B", label_font)
        pdf.cell(label_w, row_h, _pdf_text(f"{label} :"), align="L")

        pdf.set_xy(info_x + label_w, row_y)
        pdf.set_text_color(25, 72, 138)
        pdf.set_font("Helvetica", "B", value_font)
        pdf.cell(
            info_w - label_w,
            row_h,
            _pdf_text(_ecourte(value or "-", value_limit)),
            align="L",
        )

        _draw_card_separator_line(
            pdf,
            info_x,
            row_y + row_h + 0.05,
            info_x + info_w,
        )

    row_y = body_top + max(0.1, row_h * 0.05)
    for label, value in [
        ("Nom", last_name),
        ("Prénom", first_name),
        ("Classe", class_name),
        ("Année scolaire", year_label),
        ("Matricule", student.matricule or "-"),
    ]:
        _draw_info_row(row_y, label, value)
        row_y += row_h + row_gap

    row_y += max(0.25, row_gap * 0.4)
    _draw_info_row(row_y, "Né(e) le", birth_date)

    footer_y = content_y + content_h - footer_h

    # Le « No de Carte » affichait la cle primaire en base, sur cinq chiffres.
    # Elle ne signifie rien pour l'ecole, change si la base est restauree, et
    # faisait deux numeros concurrents avec le matricule. Sa place revient au
    # QR de verification et a la date de validite.
    number_x = content_x + max(0.15, content_w * 0.004)
    number_y = footer_y + max(0.46, footer_h * 0.11)
    number_label_font = 8.8 if width >= 120 else 6.9 if width >= 85 else 5.6 if width >= 72 else 4.8

    qr_d = max(7.0, min(20.0, footer_h * 0.92))
    qr_path = _carte_qr_image_path(qr_url) if qr_url else None

    etablissement = _student_etablissement(student)
    signature_source = _etablissement_media_field_path(etablissement, "principal_signature_image") or _school_signature_asset_path()
    stamp_source = _etablissement_media_field_path(etablissement, "stamp_image") or _school_stamp_asset_path()
    signature_asset_path = _pdf_compatible_image_path(signature_source, cache_prefix="signature")
    stamp_asset_path = _pdf_compatible_image_path(stamp_source, cache_prefix="stamp")
    signature_label = str(getattr(etablissement, "principal_signature_label", "") or "").strip() or "Le Principal"
    signature_position = str(getattr(etablissement, "principal_signature_position", "") or "right").strip().lower()
    stamp_position = str(getattr(etablissement, "stamp_position", "") or "right").strip().lower()
    signature_scale = _safe_scale_percent(getattr(etablissement, "principal_signature_scale", 100)) / 100.0
    stamp_scale = _safe_scale_percent(getattr(etablissement, "stamp_scale", 100)) / 100.0

    stamp_d = max(9.5, min(28.0, footer_h * 1.02 * stamp_scale))
    stamp_x = _positioned_x(
        stamp_position,
        min_x=content_x + max(0.15, content_w * 0.004),
        max_x=content_x + content_w - max(0.15, content_w * 0.004),
        box_width=stamp_d,
        default_x=content_x + content_w - stamp_d - max(0.25, content_w * 0.002),
    )
    stamp_y = footer_y + max(0.2, (footer_h - stamp_d) * 0.54)

    signature_w = max(13.0, min(40.0, content_w * 0.30 * signature_scale))
    signature_h = max(4.3, min(11.0, footer_h * 0.48 * signature_scale))
    signature_x = _positioned_x(
        signature_position,
        min_x=content_x + max(0.15, content_w * 0.004),
        max_x=content_x + content_w - max(0.15, content_w * 0.004),
        box_width=signature_w,
        default_x=stamp_x - signature_w - max(0.9, content_w * 0.015),
    )
    signature_y = footer_y + max(0.18, footer_h * 0.30)

    if signature_position == "right":
        number_max_x = max(number_x + 12.0, signature_x - max(0.8, content_w * 0.01))
    else:
        number_max_x = max(number_x + 12.0, content_x + (content_w * 0.58))
    number_line_w = max(8.0, number_max_x - number_x)

    texte_x = number_x
    if qr_path:
        try:
            pdf.image(
                qr_path,
                x=number_x,
                y=footer_y + max(0.1, (footer_h - qr_d) * 0.5),
                w=qr_d,
                h=qr_d,
            )
            texte_x = number_x + qr_d + max(0.8, content_w * 0.008)
        except Exception:
            texte_x = number_x

    if validity_label:
        pdf.set_xy(texte_x, number_y)
        pdf.set_text_color(44, 48, 59)
        pdf.set_font("Helvetica", "B", number_label_font * 0.82)
        pdf.cell(
            max(2.0, number_max_x - texte_x),
            max(2.1, footer_h * 0.26),
            _pdf_text(validity_label),
            align="L",
        )

    if signature_asset_path:
        try:
            pdf.image(
                signature_asset_path,
                x=signature_x,
                y=signature_y,
                w=signature_w,
                h=signature_h,
            )
        except Exception:
            signature_asset_path = None

    if not signature_asset_path:
        line_y = signature_y + (signature_h * 0.58)
        pdf.set_draw_color(55, 98, 156)
        pdf.set_line_width(max(0.08, outer_line_w * 0.5))
        pdf.line(signature_x + 0.4, line_y, signature_x + signature_w - 0.4, line_y)

    pdf.set_xy(signature_x, signature_y + signature_h + max(0.18, footer_h * 0.02))
    pdf.set_text_color(44, 48, 59)
    pdf.set_font("Helvetica", "B", 5.9 if width >= 85 else 4.8)
    pdf.cell(signature_w, max(1.8, footer_h * 0.2), _pdf_text(signature_label), align="C")

    if stamp_asset_path:
        try:
            pdf.image(stamp_asset_path, x=stamp_x, y=stamp_y, w=stamp_d, h=stamp_d)
        except Exception:
            stamp_asset_path = None

    if not stamp_asset_path:
        pdf.set_draw_color(31, 92, 158)
        pdf.set_line_width(max(0.08, outer_line_w * 0.52))
        try:
            pdf.ellipse(stamp_x, stamp_y, stamp_d, stamp_d)
            pdf.ellipse(
                stamp_x + (stamp_d * 0.17),
                stamp_y + (stamp_d * 0.17),
                stamp_d * 0.66,
                stamp_d * 0.66,
            )
        except Exception:
            pdf.rect(stamp_x, stamp_y, stamp_d, stamp_d)
            pdf.rect(
                stamp_x + (stamp_d * 0.17),
                stamp_y + (stamp_d * 0.17),
                stamp_d * 0.66,
                stamp_d * 0.66,
            )
        pdf.set_xy(stamp_x, stamp_y + (stamp_d * 0.48))
        pdf.set_text_color(31, 92, 158)
        pdf.set_font("Helvetica", "B", 4.2 if width >= 85 else 3.5)
        pdf.cell(stamp_d, stamp_d * 0.16, _pdf_text("Cachet"), align="C")


def _add_student_card_page(
    pdf: FPDF,
    student: Student,
    *,
    school: dict[str, str],
    logo_path: str | None,
    verify_base_url: str = "",
) -> None:
    pdf.add_page()
    pdf.set_auto_page_break(auto=False)

    page_w = pdf.w
    page_h = pdf.h
    _draw_student_card_template(
        pdf,
        student,
        school=school,
        logo_path=logo_path,
        x=4,
        y=4,
        width=page_w - 8,
        height=page_h - 8,
        verify_base_url=verify_base_url,
    )


def _draw_student_card_block(
    pdf: FPDF,
    student: Student,
    *,
    school: dict[str, str],
    logo_path: str | None,
    x: float,
    y: float,
    width: float,
    height: float,
    verify_base_url: str = "",
) -> None:
    _draw_student_card_template(
        pdf,
        student,
        school=school,
        logo_path=logo_path,
        x=x,
        y=y,
        width=width,
        height=height,
        verify_base_url=verify_base_url,
    )


ROSTER_COLUMNS = (
    ("N°", 12.0),
    ("Matricule", 34.0),
    ("Nom et prénoms", 68.0),
    ("Sexe", 14.0),
    ("Naissance", 26.0),
    ("Émargement", 36.0),
)


def _roster_gender_counts(students: list[Student]) -> tuple[int, int]:
    """Effectifs garcons/filles.

    Le genre peut manquer sur les fiches anciennes: ces eleves comptent dans
    le total sans etre attribues, plutot que d'etre ranges d'office dans une
    colonne et de fausser les deux chiffres.
    """
    garcons = sum(1 for student in students if (student.gender or "").upper() == "M")
    filles = sum(1 for student in students if (student.gender or "").upper() == "F")
    return garcons, filles


def _draw_roster_class_page(
    pdf: FPDF,
    *,
    classroom_name: str,
    students: list[Student],
    school: dict[str, str],
    logo_path: str | None,
    year_label: str,
) -> None:
    """Une page de liste d'appel pour une classe."""
    pdf.add_page()

    if logo_path:
        try:
            pdf.image(logo_path, x=12, y=10, w=18)
        except Exception:
            # Un logo illisible ne doit pas priver l'ecole de sa liste.
            pass

    pdf.set_xy(34, 11)
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 6, _pdf_text(school.get("name", "")), new_x="LMARGIN", new_y="NEXT")

    pdf.set_x(34)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 5, _pdf_text(f"Année scolaire {year_label}"), new_x="LMARGIN", new_y="NEXT")

    pdf.ln(6)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 7, _pdf_text(f"LISTE DES ÉLÈVES - Classe {classroom_name}"),
             align="C", new_x="LMARGIN", new_y="NEXT")

    garcons, filles = _roster_gender_counts(students)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(
        0, 5,
        _pdf_text(f"Effectif : {len(students)}  ({garcons} G / {filles} F)"),
        align="C", new_x="LMARGIN", new_y="NEXT",
    )
    pdf.ln(3)

    _draw_roster_table_header(pdf)

    pdf.set_font("Helvetica", "", 9)
    for index, student in enumerate(students, start=1):
        # Le tableau repart avec ses en-tetes en haut de chaque page: une
        # colonne sans titre ne se remplit pas correctement.
        if pdf.get_y() > pdf.h - 25:
            pdf.add_page()
            _draw_roster_table_header(pdf)
            pdf.set_font("Helvetica", "", 9)

        _, last_name, full_name = _student_name_parts(student)
        birth = student.birth_date.strftime("%d/%m/%Y") if student.birth_date else ""
        valeurs = (
            str(index),
            student.matricule or "",
            full_name if full_name != "-" else last_name,
            (student.gender or "").upper(),
            birth,
            "",
        )
        for (_, largeur), valeur in zip(ROSTER_COLUMNS, valeurs):
            texte = _ajuster_a_la_cellule(pdf, _pdf_text(valeur), largeur)
            pdf.cell(largeur, 7, texte, border=1)
        pdf.ln(7)

    if not students:
        pdf.set_font("Helvetica", "I", 9)
        largeur_totale = sum(largeur for _, largeur in ROSTER_COLUMNS)
        pdf.cell(largeur_totale, 8, _pdf_text("Aucun élève inscrit dans cette classe."),
                 border=1, align="C")
        pdf.ln(8)


def _draw_roster_table_header(pdf: FPDF) -> None:
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(232, 232, 240)
    for titre, largeur in ROSTER_COLUMNS:
        pdf.cell(largeur, 8, _pdf_text(titre), border=1, align="C", fill=True)
    pdf.ln(8)


def _draw_roster_summary_page(
    pdf: FPDF,
    *,
    par_classe: list[tuple[str, list[Student]]],
    school: dict[str, str],
    year_label: str,
) -> None:
    """Recapitulatif des effectifs, en fin de document multi-classes."""
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 7, _pdf_text(school.get("name", "")), align="C",
             new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, _pdf_text(f"RÉCAPITULATIF DES EFFECTIFS - {year_label}"),
             align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    colonnes = (("Classe", 80.0), ("Garçons", 30.0), ("Filles", 30.0), ("Total", 30.0))
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_fill_color(232, 232, 240)
    for titre, largeur in colonnes:
        pdf.cell(largeur, 8, _pdf_text(titre), border=1, align="C", fill=True)
    pdf.ln(8)

    total_garcons = total_filles = total_eleves = 0
    pdf.set_font("Helvetica", "", 10)
    for nom_classe, eleves in par_classe:
        garcons, filles = _roster_gender_counts(eleves)
        total_garcons += garcons
        total_filles += filles
        total_eleves += len(eleves)

        pdf.cell(colonnes[0][1], 7, _pdf_text(nom_classe), border=1)
        pdf.cell(colonnes[1][1], 7, str(garcons), border=1, align="C")
        pdf.cell(colonnes[2][1], 7, str(filles), border=1, align="C")
        pdf.cell(colonnes[3][1], 7, str(len(eleves)), border=1, align="C")
        pdf.ln(7)

    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(colonnes[0][1], 8, _pdf_text("TOTAL"), border=1, fill=True)
    pdf.cell(colonnes[1][1], 8, str(total_garcons), border=1, align="C", fill=True)
    pdf.cell(colonnes[2][1], 8, str(total_filles), border=1, align="C", fill=True)
    pdf.cell(colonnes[3][1], 8, str(total_eleves), border=1, align="C", fill=True)
    pdf.ln(8)


# Largeurs en mm; la somme fait la largeur utile d'une A4 portrait (190).
# « Matières » est la colonne qui deborde en pratique -- un professeur cumule
# volontiers trois intitules longs -- elle prend donc le plus de place apres
# le nom, l'emargement gardant de quoi signer.
STAFF_COLUMNS = (
    ("N°", 10.0),
    ("Code", 26.0),
    ("Nom et prénoms", 54.0),
    ("Matières", 62.0),
    ("Émargement", 38.0),
)


def _ajuster_a_la_cellule(pdf: FPDF, texte: str, largeur: float) -> str:
    """Tronque un texte trop long pour sa colonne.

    `FPDF.cell` dessine la bordure a la largeur demandee mais n'ecrete pas son
    contenu: un intitule trop long depasse simplement sur les colonnes
    suivantes, et vient se superposer a la colonne d'emargement. Rien ne le
    signale, sinon le document imprime.
    """
    if not texte:
        return texte

    disponible = largeur - 2  # marges internes de la cellule
    if pdf.get_string_width(texte) <= disponible:
        return texte

    # « ... » plutot que le caractere d'ellipse, absent du latin-1 du PDF.
    suffixe = "..."
    reste = disponible - pdf.get_string_width(suffixe)
    coupe = texte
    while coupe and pdf.get_string_width(coupe) > reste:
        coupe = coupe[:-1]
    return f"{coupe.rstrip()}{suffixe}" if coupe else suffixe


def _draw_staff_table_header(pdf: FPDF) -> None:
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(232, 232, 240)
    for titre, largeur in STAFF_COLUMNS:
        pdf.cell(largeur, 8, _pdf_text(titre), border=1, align="C", fill=True)
    pdf.ln(8)


def _teacher_subjects_label(teacher: Teacher) -> str:
    """Matieres enseignees, sans repetition.

    Un professeur de maths sur trois classes a trois affectations mais une
    seule matiere: les lister toutes remplirait la colonne sans rien apprendre.
    """
    matieres = []
    for affectation in teacher.assignments.all():
        nom = getattr(getattr(affectation, "subject", None), "name", "")
        if nom and nom not in matieres:
            matieres.append(nom)
    return ", ".join(matieres)


def _build_staff_roster_pdf(
    teachers: list[Teacher],
    *,
    school: dict[str, str],
    logo_path: str | None,
    year_label: str,
) -> FPDF:
    """Liste du personnel enseignant, avec colonne d'emargement."""
    pdf = FPDF(format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    if logo_path:
        try:
            pdf.image(logo_path, x=12, y=10, w=18)
        except Exception:
            pass

    pdf.set_xy(34, 11)
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 6, _pdf_text(school.get("name", "")), new_x="LMARGIN", new_y="NEXT")

    pdf.set_x(34)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 5, _pdf_text(f"Année scolaire {year_label}"), new_x="LMARGIN", new_y="NEXT")

    pdf.ln(6)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 7, _pdf_text("LISTE DU PERSONNEL ENSEIGNANT"),
             align="C", new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 5, _pdf_text(f"Effectif : {len(teachers)}"),
             align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    _draw_staff_table_header(pdf)

    pdf.set_font("Helvetica", "", 9)
    for index, teacher in enumerate(teachers, start=1):
        if pdf.get_y() > pdf.h - 25:
            pdf.add_page()
            _draw_staff_table_header(pdf)
            pdf.set_font("Helvetica", "", 9)

        _, last_name, full_name = _teacher_name_parts(teacher)
        valeurs = (
            str(index),
            teacher.employee_code or "",
            full_name if full_name != "-" else last_name,
            _teacher_subjects_label(teacher),
            "",
        )
        for (_, largeur), valeur in zip(STAFF_COLUMNS, valeurs):
            texte = _ajuster_a_la_cellule(pdf, _pdf_text(valeur), largeur)
            pdf.cell(largeur, 7, texte, border=1)
        pdf.ln(7)

    if not teachers:
        pdf.set_font("Helvetica", "I", 9)
        largeur_totale = sum(largeur for _, largeur in STAFF_COLUMNS)
        pdf.cell(largeur_totale, 8,
                 _pdf_text("Aucun enseignant enregistré."), border=1, align="C")
        pdf.ln(8)

    return pdf


def _teacher_name_parts(teacher: Teacher) -> tuple[str, str, str]:
    user = getattr(teacher, "user", None)
    if not user:
        return "-", "-", "-"

    first_name = (user.first_name or "").strip()
    last_name = (user.last_name or "").strip()
    full_name = (user.get_full_name() or "").strip() or user.username
    return first_name or "-", last_name or "-", full_name or "-"


def _build_class_roster_pdf(
    par_classe: list[tuple[str, list[Student]]],
    *,
    school: dict[str, str],
    logo_path: str | None,
    year_label: str,
    with_summary: bool,
) -> FPDF:
    """Liste d'appel: une classe par page, recapitulatif optionnel a la fin."""
    pdf = FPDF(format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)

    for nom_classe, eleves in par_classe:
        _draw_roster_class_page(
            pdf,
            classroom_name=nom_classe,
            students=eleves,
            school=school,
            logo_path=logo_path,
            year_label=year_label,
        )

    # Sur une classe unique, le recapitulatif ne ferait que repeter l'en-tete.
    if with_summary and len(par_classe) > 1:
        _draw_roster_summary_page(
            pdf, par_classe=par_classe, school=school, year_label=year_label
        )

    return pdf


# Dimensions en millimetres, paysage.
#   a6   : le format historique. Grand et lisible, mais ne rentre dans aucun
#          portefeuille, donc rarement porte par l'eleve.
#   cr80 : le format des cartes d'identite et bancaires.
CARD_FORMATS: dict[str, tuple[float, float]] = {
    "a6": (148.0, 105.0),
    "cr80": (85.6, 54.0),
}
CARD_FORMAT_DEFAUT = "a6"


def _draw_crop_marks(pdf: FPDF, x: float, y: float, w: float, h: float) -> None:
    """Traits de coupe aux quatre coins d'une carte.

    Les planches se decoupaient a vue, faute de reperes. Les traits sont
    places hors de la carte pour ne pas la barrer.
    """
    longueur = 3.0
    ecart = 0.8
    pdf.set_draw_color(150, 150, 150)
    pdf.set_line_width(0.1)
    for cx, sens_x in ((x, -1), (x + w, 1)):
        for cy, sens_y in ((y, -1), (y + h, 1)):
            pdf.line(cx + sens_x * ecart, cy, cx + sens_x * (ecart + longueur), cy)
            pdf.line(cx, cy + sens_y * ecart, cx, cy + sens_y * (ecart + longueur))


def _grille_a4(card_w: float, card_h: float) -> tuple[int, int, float, float]:
    """Nombre de cartes par planche A4, a leur taille reelle.

    L'ancienne grille imposait 3x3 et etirait chaque case aux dimensions
    obtenues: les cartes devenaient portrait alors que la maquette est
    paysage, d'ou un quart de vide par carte et un en-tete a 4 points.
    """
    marge = 10.0
    gap = 5.0
    utile_w = 210.0 - (2 * marge)
    utile_h = 297.0 - (2 * marge)
    cols = max(1, int((utile_w + gap) // (card_w + gap)))
    rows = max(1, int((utile_h + gap) // (card_h + gap)))
    return cols, rows, marge, gap


def _build_student_cards_pdf(
    students: list[Student],
    *,
    school: dict[str, str],
    logo_path: str | None,
    layout_mode: str,
    card_format: str = CARD_FORMAT_DEFAUT,
    verify_base_url: str = "",
) -> FPDF:
    card_w, card_h = CARD_FORMATS.get(card_format, CARD_FORMATS[CARD_FORMAT_DEFAUT])

    if layout_mode == "standard":
        pdf = FPDF(format=(card_w, card_h))
        for student in students:
            _add_student_card_page(
                pdf,
                student,
                school=school,
                logo_path=logo_path,
                verify_base_url=verify_base_url,
            )
        return pdf

    pdf = FPDF(format="A4")
    pdf.set_auto_page_break(auto=False)

    if layout_mode == "a4":
        cols, rows, marge, gap = _grille_a4(card_w, card_h)
    else:
        # a4_6up / a4_9up: grille imposee, mais la carte garde desormais ses
        # proportions et se centre dans sa case au lieu d'y etre etiree.
        cols, rows = (2, 3) if layout_mode == "a4_6up" else (3, 3)
        marge, gap = 10.0, 5.0
        case_w = (210.0 - (2 * marge) - ((cols - 1) * gap)) / cols
        case_h = (297.0 - (2 * marge) - ((rows - 1) * gap)) / rows
        echelle = min(case_w / card_w, case_h / card_h)
        card_w, card_h = card_w * echelle, card_h * echelle

    par_planche = cols * rows
    for index, student in enumerate(students):
        if index % par_planche == 0:
            pdf.add_page()

        slot = index % par_planche
        x = marge + (slot % cols) * (card_w + gap)
        y = marge + (slot // cols) * (card_h + gap)

        _draw_crop_marks(pdf, x, y, card_w, card_h)
        _draw_student_card_block(
            pdf,
            student,
            school=school,
            logo_path=logo_path,
            x=x,
            y=y,
            width=card_w,
            height=card_h,
            verify_base_url=verify_base_url,
        )

    return pdf


def _student_photo_url(student, request) -> str | None:
    """URL de la photo, ou None si l'eleve n'en a pas."""
    photo = getattr(student, "photo", None)
    if not photo:
        return None
    try:
        url = photo.url
    except Exception:
        return None
    return request.build_absolute_uri(url) if url.startswith("/") else url


def _requested_card_format(request) -> str | None:
    """Format demande, ou None si la valeur est inconnue.

    None plutot qu'un repli silencieux: une planche imprimee au mauvais format
    ne se rattrape qu'en la rejetant et en recommencant.
    """
    demande = str(request.query_params.get("card_format", CARD_FORMAT_DEFAUT)).strip().lower()
    return demande if demande in CARD_FORMATS else None


def _verify_base_url(request) -> str:
    """Racine publique a inscrire dans le QR.

    Deduite de la requete: le projet n'a pas de reglage d'URL publique, et
    l'ecrire en dur reproduirait exactement la faute que ce lot corrige.
    """
    try:
        return request.build_absolute_uri("/").rstrip("/")
    except Exception:
        return ""


def _requested_etablissement_id(request):
    raw_value = request.headers.get("X-Etablissement-Id") or request.query_params.get("etablissement")
    if raw_value in (None, ""):
        return None
    try:
        parsed = int(raw_value)
    except (TypeError, ValueError):
        return None
    return parsed if parsed > 0 else None


def _requested_etablissement_name(request):
    raw_name = request.headers.get("X-Etablissement-Name") or request.query_params.get("etablissement_name")
    if raw_name is None:
        return None
    cleaned = str(raw_name).strip()
    return cleaned or None


def _requested_etablissement(request):
    requested_id = _requested_etablissement_id(request)
    if requested_id:
        etablissement = Etablissement.objects.filter(id=requested_id).first()
        if etablissement:
            return etablissement

    requested_name = _requested_etablissement_name(request)
    if not requested_name:
        return None

    etablissement = Etablissement.objects.filter(name__iexact=requested_name).first()
    if etablissement:
        return etablissement

    return Etablissement.objects.filter(name__icontains=requested_name).order_by("name").first()


def _effective_etablissement_id(request):
    user = request.user
    role = getattr(user, "role", "")

    if role == UserRole.SUPER_ADMIN:
        requested = _requested_etablissement(request)
        return requested.id if requested else None

    return getattr(user, "etablissement_id", None)


def _ensure_reports_module_access(request) -> None:
    role = getattr(request.user, "role", "")
    if role not in {
        UserRole.SUPER_ADMIN,
        UserRole.DIRECTOR,
        UserRole.ACCOUNTANT,
        UserRole.PARENT,
        UserRole.STUDENT,
    }:
        raise PermissionDenied("Accès refusé au module rapports.")


def _ensure_sensitive_export_access(request) -> None:
    role = getattr(request.user, "role", "")
    if role not in {UserRole.SUPER_ADMIN, UserRole.DIRECTOR, UserRole.ACCOUNTANT}:
        raise PermissionDenied("Accès refusé: export sensible réservé à l'administration/finance.")


def _allowed_students_queryset(request):
    user = request.user
    queryset = Student.objects.select_related(
        "user",
        "classroom",
        "parent",
        "parent__user",
    ).all()
    role = getattr(user, "role", "")

    if role == UserRole.STUDENT:
        return queryset.filter(user_id=user.id)
    if role == UserRole.PARENT:
        return queryset.filter(parent__user_id=user.id)

    target_etablissement_id = _effective_etablissement_id(request)
    if target_etablissement_id:
        return queryset.filter(
            Q(etablissement_id=target_etablissement_id)
            | Q(etablissement__isnull=True, classroom__etablissement_id=target_etablissement_id)
        )
    return queryset.none()


def _allowed_payments_queryset(request):
    user = request.user
    queryset = Payment.objects.select_related(
        "fee",
        "fee__student",
        "fee__student__user",
        "fee__student__parent",
        "fee__student__parent__user",
        "fee__academic_year",
        "received_by",
    ).filter(is_cancelled=False)
    role = getattr(user, "role", "")

    if role == UserRole.STUDENT:
        return queryset.filter(fee__student__user_id=user.id)
    if role == UserRole.PARENT:
        return queryset.filter(fee__student__parent__user_id=user.id)

    if role == UserRole.SUPER_ADMIN:
        return queryset

    target_etablissement_id = _effective_etablissement_id(request)
    if target_etablissement_id:
        return queryset.filter(
            Q(fee__student__etablissement_id=target_etablissement_id)
            | Q(
                fee__student__etablissement__isnull=True,
                fee__student__classroom__etablissement_id=target_etablissement_id,
            )
        )
    return queryset.none()


def _allowed_expenses_queryset(request):
    user = request.user
    queryset = Expense.objects.select_related(
        "paid_by",
        "level_one_validated_by",
        "level_two_validated_by",
        "etablissement",
    ).all()
    role = getattr(user, "role", "")

    target_etablissement_id = _effective_etablissement_id(request)
    if target_etablissement_id:
        return queryset.filter(etablissement_id=target_etablissement_id)

    if role == UserRole.SUPER_ADMIN:
        return queryset.none()

    return queryset.filter(etablissement_id=getattr(user, "etablissement_id", None))


def _query_param_date(request, key: str) -> date | None:
    raw_value = str(request.query_params.get(key, "") or "").strip()
    if not raw_value:
        return None
    try:
        return date.fromisoformat(raw_value)
    except ValueError:
        return None


def _journal_period_bounds(request) -> tuple[date | None, date | None]:
    date_from = _query_param_date(request, "date_from")
    date_to = _query_param_date(request, "date_to")
    if date_from and date_to and date_to < date_from:
        date_from, date_to = date_to, date_from
    return date_from, date_to


def _apply_payment_journal_filters(queryset, request):
    search = str(request.query_params.get("search", "") or "").strip()
    method = str(request.query_params.get("method", "") or "").strip()
    date_from, date_to = _journal_period_bounds(request)

    if search:
        queryset = queryset.filter(
            Q(reference__icontains=search)
            | Q(fee__student__matricule__icontains=search)
            | Q(fee__student__user__first_name__icontains=search)
            | Q(fee__student__user__last_name__icontains=search)
            | Q(fee__fee_type__icontains=search)
        )
    if method:
        queryset = queryset.filter(method__iexact=method)
    if date_from:
        queryset = queryset.filter(created_at__date__gte=date_from)
    if date_to:
        queryset = queryset.filter(created_at__date__lte=date_to)

    return queryset


def _apply_expense_journal_filters(queryset, request):
    search = str(request.query_params.get("search", "") or "").strip()
    category = str(request.query_params.get("category", "") or "").strip()
    stage = str(request.query_params.get("stage", "") or "").strip().lower()
    date_from, date_to = _journal_period_bounds(request)

    if search:
        queryset = queryset.filter(Q(label__icontains=search) | Q(notes__icontains=search))
    if category:
        queryset = queryset.filter(category__iexact=category)
    if stage == "draft":
        queryset = queryset.filter(level_one_validated_at__isnull=True, level_two_validated_at__isnull=True)
    elif stage == "level_one":
        queryset = queryset.filter(level_one_validated_at__isnull=False, level_two_validated_at__isnull=True)
    elif stage == "level_two":
        queryset = queryset.filter(level_two_validated_at__isnull=False)
    if date_from:
        queryset = queryset.filter(date__gte=date_from)
    if date_to:
        queryset = queryset.filter(date__lte=date_to)

    return queryset


def _payment_journal_row(payment: Payment) -> dict:
    student = payment.fee.student if payment.fee else None
    student_user = student.user if student else None
    receiver = payment.received_by
    receiver_name = ""
    if receiver:
        receiver_name = receiver.get_full_name().strip() or receiver.username
    return {
        "id": payment.id,
        "created_at": timezone.localtime(payment.created_at).isoformat(),
        "student_full_name": (
            (student_user.get_full_name().strip() or student_user.username)
            if student_user
            else ""
        ),
        "student_matricule": student.matricule if student else "",
        "fee_type": payment.fee.get_fee_type_display() if payment.fee else "",
        "amount": float(payment.amount),
        "method": payment.method,
        "reference": payment.reference or "",
        "received_by": receiver_name,
    }


def _expense_journal_row(expense: Expense) -> dict:
    return {
        "id": expense.id,
        "date": expense.date.isoformat() if expense.date else "",
        "label": expense.label,
        "category": expense.category,
        "amount": float(expense.amount),
        "validation_stage": expense.validation_stage,
        "paid_on": expense.paid_on.isoformat() if expense.paid_on else "",
        "notes": expense.notes or "",
    }


def _parse_page_size(request, default: int = 100, max_size: int = 1000) -> int:
    try:
        parsed = int(request.query_params.get("page_size", default))
    except (TypeError, ValueError):
        parsed = default
    return max(1, min(max_size, parsed))


def _ensure_student_access(request, student: Student) -> None:
    user = request.user
    role = getattr(user, "role", "")
    if role == UserRole.STUDENT and student.user_id != user.id:
        raise PermissionDenied("Accès refusé à ce bulletin.")

    if role == UserRole.PARENT:
        parent_user_id = student.parent.user_id if student.parent else None
        if parent_user_id != user.id:
            raise PermissionDenied("Accès refusé à ce bulletin.")

    target_etablissement_id = _effective_etablissement_id(request)
    student_etablissement_id = getattr(student, "etablissement_id", None)
    if student_etablissement_id is None and getattr(student, "classroom", None) is not None:
        student_etablissement_id = getattr(student.classroom, "etablissement_id", None)

    if target_etablissement_id and student_etablissement_id and target_etablissement_id != student_etablissement_id:
        raise PermissionDenied("Accès refusé à ce bulletin.")

    if target_etablissement_id is None and role == UserRole.SUPER_ADMIN:
        raise PermissionDenied("Selectionnez un etablissement actif.")


def _ensure_payment_access(request, payment: Payment) -> None:
    user = request.user
    role = getattr(user, "role", "")
    student = payment.fee.student if payment.fee else None

    if role == UserRole.STUDENT:
        if not student or student.user_id != user.id:
            raise PermissionDenied("Accès refusé à ce reçu de paiement.")

    if role == UserRole.PARENT:
        parent_user_id = student.parent.user_id if student and student.parent else None
        if parent_user_id != user.id:
            raise PermissionDenied("Accès refusé à ce reçu de paiement.")

    if student is not None:
        student_etablissement_id = getattr(student, "etablissement_id", None)
        if student_etablissement_id is None and getattr(student, "classroom", None) is not None:
            student_etablissement_id = getattr(student.classroom, "etablissement_id", None)
        target_etablissement_id = _effective_etablissement_id(request)
        if target_etablissement_id and student_etablissement_id and target_etablissement_id != student_etablissement_id:
            raise PermissionDenied("Accès refusé à ce reçu de paiement.")
        if target_etablissement_id is None and role == UserRole.SUPER_ADMIN:
            raise PermissionDenied("Selectionnez un etablissement actif.")


def _term_variants(term: str) -> list[str]:
    raw = str(term or "").strip().upper()
    if not raw:
        return []

    variants = {raw}
    digits = "".join(ch for ch in raw if ch.isdigit())
    if raw.isdigit():
        digits = raw

    if digits:
        variants.update(
            {
                digits,
                f"T{digits}",
                f"TRIMESTRE{digits}",
                f"TRIMESTRE {digits}",
            }
        )

    return sorted(value for value in variants if value)


def _exam_term_title_tokens(term: str) -> list[str]:
    raw = str(term or "").strip().upper()
    if not raw:
        return []

    tokens = {raw}
    digits = "".join(ch for ch in raw if ch.isdigit())
    if digits:
        tokens.update(
            {
                digits,
                f"T{digits}",
                f"TRIMESTRE{digits}",
                f"TRIMESTRE {digits}",
                f"TERM{digits}",
                f"TERM {digits}",
            }
        )

    return sorted(token for token in tokens if token)


def _term_display_label(term: str) -> str:
    raw = str(term or "").strip().upper()
    if not raw:
        return "-"

    if raw.isdigit():
        return f"T{raw}"

    digits = "".join(ch for ch in raw if ch.isdigit())
    if raw.startswith("TRIMESTRE") and digits:
        return f"T{digits}"

    return raw


def _format_cell_value(value: float | None) -> str:
    if value is None:
        return "-"
    return f"{value:.2f}"


def _format_coef_value(value: float | None) -> str:
    if value is None:
        return "-"

    text = f"{value:.2f}"
    if text.endswith(".00"):
        return text[:-3]
    if text.endswith("0"):
        return text[:-1]
    return text


def _appreciation_from_score(value: float | None) -> str:
    if value is None:
        return "-"
    if value >= 16:
        return "Tres bien"
    if value >= 14:
        return "Bien"
    if value >= 12:
        return "Assez bien"
    if value >= 10:
        return "Passable"
    return "Insuffisant"


def _build_bulletin_rows(
    *,
    subjects,
    student_note_by_subject: dict[int, float],
    exam_note_by_subject: dict[int, float],
    class_average_by_subject: dict[int, float],
    conduite_note: float,
    conduite_coef: float = 2.0,
    conduite_moyenne_classe: float | None = None,
):
    weighted_sum = 0.0
    coef_sum = 0.0
    rows = [
        {
            "index": 1,
            "subject": "Conduite",
            "coef": conduite_coef,
            "note_classe": conduite_note,
            "note_examen": None,
            "note_finale": conduite_note,
            "appreciation": _appreciation_from_score(conduite_note),
            "moyenne_classe": conduite_moyenne_classe,
            "points": round(conduite_note * conduite_coef, 2),
        }
    ]

    weighted_sum += round(conduite_note * conduite_coef, 2)
    coef_sum += conduite_coef

    for index, subject in enumerate(subjects, start=2):
        coef = float(subject.coefficient)
        note_classe = student_note_by_subject.get(subject.id)
        note_examen = exam_note_by_subject.get(subject.id)

        if note_classe is not None and note_examen is not None:
            note_finale = round((note_classe + note_examen) / 2.0, 2)
            effective_coef = coef
        elif note_classe is not None:
            note_finale = round(note_classe, 2)
            effective_coef = coef
        elif note_examen is not None:
            note_finale = round(note_examen, 2)
            effective_coef = coef
        else:
            note_finale = None
            effective_coef = 0.0

        appreciation_score = note_finale

        note_moyenne_classe = class_average_by_subject.get(subject.id)
        points = round(note_finale * coef, 2) if note_finale is not None else None

        if points is not None and effective_coef > 0:
            weighted_sum += points
            coef_sum += effective_coef

        rows.append(
            {
                "index": index,
                "subject": subject.name,
                "coef": coef,
                "note_classe": note_classe,
                "note_examen": note_examen,
                "note_finale": note_finale,
                "appreciation": _appreciation_from_score(appreciation_score),
                "moyenne_classe": note_moyenne_classe,
                "points": points,
            }
        )

    average = round(weighted_sum / coef_sum, 2) if coef_sum else 0.0
    return rows, average, coef_sum


def _subject_name_key(name: str) -> str:
    return " ".join(str(name or "").strip().lower().split())


def _deduplicate_bulletin_subjects(
    *,
    subjects,
    student_note_by_subject: dict[int, float],
    exam_note_by_subject: dict[int, float],
    class_average_by_subject: dict[int, float],
):
    grouped: dict[str, dict] = {}

    for subject in subjects:
        key = _subject_name_key(getattr(subject, "name", ""))
        if not key:
            key = f"id:{subject.id}"

        entry = grouped.get(key)
        if entry is None:
            grouped[key] = {
                "subject": subject,
                "ids": [subject.id],
            }
            continue

        entry["ids"].append(subject.id)
        try:
            current_coef = float(entry["subject"].coefficient)
            new_coef = float(subject.coefficient)
            if new_coef > current_coef:
                entry["subject"].coefficient = subject.coefficient
        except Exception:
            pass

    deduped_subjects = []
    merged_student_note_by_subject: dict[int, float] = {}
    merged_exam_note_by_subject: dict[int, float] = {}
    merged_class_average_by_subject: dict[int, float] = {}

    for entry in grouped.values():
        subject = entry["subject"]
        ids = entry["ids"]
        rep_id = subject.id
        deduped_subjects.append(subject)

        class_notes = [student_note_by_subject[sid] for sid in ids if sid in student_note_by_subject]
        if class_notes:
            merged_student_note_by_subject[rep_id] = max(class_notes)

        exam_notes = [exam_note_by_subject[sid] for sid in ids if sid in exam_note_by_subject]
        if exam_notes:
            merged_exam_note_by_subject[rep_id] = max(exam_notes)

        class_averages = [class_average_by_subject[sid] for sid in ids if sid in class_average_by_subject]
        if class_averages:
            merged_class_average_by_subject[rep_id] = round(
                sum(class_averages) / len(class_averages),
                2,
            )

    deduped_subjects.sort(key=lambda subject: (str(subject.name or "").lower(), subject.id))
    return (
        deduped_subjects,
        merged_student_note_by_subject,
        merged_exam_note_by_subject,
        merged_class_average_by_subject,
    )


def _term_variants(term: str) -> list[str]:
    raw = str(term or "").strip().upper()
    if not raw:
        return []

    variants = {raw}
    digits = "".join(ch for ch in raw if ch.isdigit())
    if raw.isdigit():
        digits = raw

    if digits:
        variants.update(
            {
                digits,
                f"T{digits}",
                f"TRIMESTRE{digits}",
                f"TRIMESTRE {digits}",
            }
        )

    return sorted(value for value in variants if value)


def _exam_term_title_tokens(term: str) -> list[str]:
    raw = str(term or "").strip().upper()
    if not raw:
        return []

    tokens = {raw}
    digits = "".join(ch for ch in raw if ch.isdigit())
    if digits:
        tokens.update(
            {
                digits,
                f"T{digits}",
                f"TRIMESTRE{digits}",
                f"TRIMESTRE {digits}",
                f"TERM{digits}",
                f"TERM {digits}",
            }
        )

    return sorted(token for token in tokens if token)


def _term_display_label(term: str) -> str:
    raw = str(term or "").strip().upper()
    if not raw:
        return "-"

    if raw.isdigit():
        return f"T{raw}"

    digits = "".join(ch for ch in raw if ch.isdigit())
    if raw.startswith("TRIMESTRE") and digits:
        return f"T{digits}"

    return raw


def _format_cell_value(value: float | None) -> str:
    if value is None:
        return "-"
    return f"{value:.2f}"


def _format_coef_value(value: float | None) -> str:
    if value is None:
        return "-"

    text = f"{value:.2f}"
    if text.endswith(".00"):
        return text[:-3]
    if text.endswith("0"):
        return text[:-1]
    return text


class ReportsContextView(APIView):
    access_module = "reports"
    permission_classes = [IsAuthenticated, HasModuleAccess]

    def get(self, request):
        _ensure_reports_module_access(request)

        students = _allowed_students_queryset(request).order_by(
            "user__last_name",
            "user__first_name",
            "matricule",
        )
        payments = _allowed_payments_queryset(request).order_by("-created_at")
        years = AcademicYear.objects.all().order_by("-start_date", "-id")

        return Response(
            {
                "students": StudentSerializer(students, many=True).data,
                "academic_years": AcademicYearSerializer(years, many=True).data,
                "payments": PaymentSerializer(payments, many=True).data,
            }
        )


class BulletinPdfView(APIView):
    access_module = "reports"
    permission_classes = [IsAuthenticated, HasModuleAccess]

    def get(self, request, student_id: int, academic_year_id: int, term: str):
        normalized_term = normalize_term(term)
        if not normalized_term:
            return Response(
                {"detail": "Période invalide. Utilisez uniquement T1, T2 ou T3."},
                status=400,
            )

        student = get_object_or_404(
            Student.objects.select_related(
                "user",
                "classroom",
                "parent",
                "parent__user",
            ),
            id=student_id,
        )
        _ensure_student_access(request, student)

        payload = _build_bulletin_payload(
            student=student,
            academic_year_id=academic_year_id,
            normalized_term=normalized_term,
        )

        pdf = FPDF(orientation="L", format="A4")
        pdf.set_auto_page_break(auto=False)
        pdf.add_page()
        _render_bulletin_page(pdf, payload)

        safe_term = str(payload["period_label"] or term or "periode").replace("/", "-")
        return pdf_output_response(pdf, f"bulletin_{student.matricule}_{safe_term}.pdf")


def _build_bulletin_payload(*, student: Student, academic_year_id: int, normalized_term: str) -> dict:
    school = _school_identity_for_student(student)
    school_name = school["name"]
    school_short = school["short"]
    school_level = school["level"]
    school_phone = school["phone"]
    logo_path = _etablissement_logo_path(student) or _school_logo_path()

    student_name = student.user.get_full_name().strip() or student.user.username
    class_name = student.classroom.name if student.classroom else "N/A"
    period_label = normalized_term
    academic_year_name = (
        AcademicYear.objects.filter(id=academic_year_id)
        .values_list("name", flat=True)
        .first()
        or str(academic_year_id)
    )

    student_grades_qs = Grade.objects.filter(
        student_id=student.id,
        academic_year_id=academic_year_id,
        term=normalized_term,
    ).select_related("subject")

    student_note_by_subject: dict[int, float] = {}
    for grade in student_grades_qs.order_by("subject_id", "-created_at", "-id"):
        student_note_by_subject.setdefault(grade.subject_id, float(grade.value))

    classroom_id = student.classroom_id
    subject_ids: set[int] = set(student_note_by_subject.keys())
    class_average_by_subject: dict[int, float] = {}

    if classroom_id:
        class_grades_qs = Grade.objects.filter(
            classroom_id=classroom_id,
            academic_year_id=academic_year_id,
            term=normalized_term,
        )
        subject_ids.update(class_grades_qs.values_list("subject_id", flat=True))

        class_avg_rows = class_grades_qs.values("subject_id").annotate(avg_note=Avg("value"))
        class_average_by_subject = {
            int(row["subject_id"]): float(row["avg_note"])
            for row in class_avg_rows
            if row.get("avg_note") is not None
        }

        subject_ids.update(
            TeacherAssignment.objects.filter(classroom_id=classroom_id).values_list("subject_id", flat=True)
        )
        subject_ids.update(
            ExamPlanning.objects.filter(
                classroom_id=classroom_id,
                session__academic_year_id=academic_year_id,
            ).values_list("subject_id", flat=True)
        )

    student_exam_results_qs = ExamResult.objects.filter(
        student_id=student.id,
        session__academic_year_id=academic_year_id,
        session__term=normalized_term,
    )
    subject_ids.update(student_exam_results_qs.values_list("subject_id", flat=True))

    exam_note_by_subject: dict[int, float] = {}
    for exam_result in student_exam_results_qs.order_by(
        "subject_id",
        "-session__end_date",
        "-session__start_date",
        "-created_at",
        "-id",
    ):
        exam_note_by_subject.setdefault(exam_result.subject_id, float(exam_result.score))

    subjects = list(Subject.objects.filter(id__in=subject_ids).order_by("name", "id"))
    (
        subjects,
        student_note_by_subject,
        exam_note_by_subject,
        class_average_by_subject,
    ) = _deduplicate_bulletin_subjects(
        subjects=subjects,
        student_note_by_subject=student_note_by_subject,
        exam_note_by_subject=exam_note_by_subject,
        class_average_by_subject=class_average_by_subject,
    )

    conduite_note = float(student.conduite if student.conduite is not None else 18)
    conduite_coef = 2.0
    conduite_moyenne_classe = None
    if classroom_id:
        conduite_moyenne_classe = (
            Student.objects.filter(classroom_id=classroom_id)
            .aggregate(avg_conduite=Avg("conduite"))
            .get("avg_conduite")
        )
        if conduite_moyenne_classe is not None:
            conduite_moyenne_classe = float(conduite_moyenne_classe)

    rows, average, coef_sum = _build_bulletin_rows(
        subjects=subjects,
        student_note_by_subject=student_note_by_subject,
        exam_note_by_subject=exam_note_by_subject,
        class_average_by_subject=class_average_by_subject,
        conduite_note=conduite_note,
        conduite_coef=conduite_coef,
        conduite_moyenne_classe=conduite_moyenne_classe,
    )

    rank_value = None
    if classroom_id:
        history = (
            StudentAcademicHistory.objects.filter(
                student_id=student.id,
                academic_year_id=academic_year_id,
                classroom_id=classroom_id,
            )
            .order_by("-updated_at", "-id")
            .first()
        )
        if history and history.rank:
            rank_value = int(history.rank)

    if average >= 16:
        mention = "Tres bien"
    elif average >= 14:
        mention = "Bien"
    elif average >= 12:
        mention = "Assez bien"
    elif average >= 10:
        mention = "Passable"
    else:
        mention = "Insuffisant"

    etablissement = _student_etablissement(student)
    signature_source = _etablissement_media_field_path(etablissement, "principal_signature_image") or _school_signature_asset_path()
    stamp_source = _etablissement_media_field_path(etablissement, "stamp_image") or _school_stamp_asset_path()
    signature_asset_path = _pdf_compatible_image_path(signature_source, cache_prefix="bulletin_signature")
    stamp_asset_path = _pdf_compatible_image_path(stamp_source, cache_prefix="bulletin_stamp")
    signature_label = str(getattr(etablissement, "principal_signature_label", "") or "").strip() or "Direction"
    signature_position = str(getattr(etablissement, "principal_signature_position", "") or "right").strip().lower()
    stamp_position = str(getattr(etablissement, "stamp_position", "") or "right").strip().lower()
    signature_scale = _safe_scale_percent(getattr(etablissement, "principal_signature_scale", 100)) / 100.0
    stamp_scale = _safe_scale_percent(getattr(etablissement, "stamp_scale", 100)) / 100.0

    return {
        "logo_path": logo_path,
        "school_name": school_name,
        "school_short": school_short,
        "school_level": school_level,
        "school_phone": school_phone,
        "student_name": student_name,
        "student_matricule": student.matricule,
        "class_name": class_name,
        "academic_year_name": academic_year_name,
        "period_label": period_label,
        "rank": rank_value,
        "rank_display": str(rank_value) if rank_value else "-",
        "rows": rows,
        "average": average,
        "coef_sum": coef_sum,
        "mention": mention,
        "signature_asset_path": signature_asset_path,
        "stamp_asset_path": stamp_asset_path,
        "signature_label": signature_label,
        "signature_position": signature_position,
        "stamp_position": stamp_position,
        "signature_scale": signature_scale,
        "stamp_scale": stamp_scale,
    }


def _render_bulletin_page(pdf: FPDF, payload: dict) -> None:
    left_margin = 8
    right_margin = pdf.w - 8

    logo_path = payload["logo_path"]
    if logo_path:
        try:
            pdf.image(logo_path, x=left_margin, y=6, w=17)
        except Exception:
            pass

    title_x = 28 if logo_path else left_margin
    pdf.set_xy(title_x, 6)
    pdf.set_font("Helvetica", "B", 12.5)
    pdf.cell(0, 5.5, _pdf_text(payload["school_name"]), ln=True)

    pdf.set_x(title_x)
    pdf.set_font("Helvetica", size=8.8)
    header_line = (
        f"{payload['school_level']} | Tel: {payload['school_phone']}"
        if payload["school_phone"]
        else payload["school_level"]
    )
    pdf.cell(0, 4.4, _pdf_text(header_line), ln=True)

    pdf.set_x(title_x)
    pdf.set_font("Helvetica", "B", 8.5)
    pdf.cell(
        0,
        4.4,
        _pdf_text(f"Application: {payload['school_short']} - GESTION SCHOOL"),
        ln=True,
    )

    top_line_y = max(pdf.get_y() + 1.5, 20)
    pdf.set_draw_color(60, 60, 60)
    pdf.line(left_margin, top_line_y, right_margin, top_line_y)
    pdf.set_y(top_line_y + 1.8)

    pdf.set_font("Helvetica", "B", 13.2)
    pdf.cell(0, 6.0, _pdf_text("BULLETIN SCOLAIRE"), ln=True, align="C")

    info_label_w = 22
    info_value_w = 48
    info_h = 5.6
    info_rows = [
        ("Eleve", payload["student_name"]),
        ("Matricule", payload["student_matricule"]),
        ("Classe", payload["class_name"]),
        ("Rang", payload["rank_display"]),
        ("Etablissement", payload["school_name"]),
        ("Annee", payload["academic_year_name"]),
        ("Periode", payload["period_label"]),
    ]

    for index, (label, value) in enumerate(info_rows):
        if index % 2 == 0:
            pdf.set_x(left_margin)
        pdf.set_font("Helvetica", "B", 8.6)
        pdf.cell(info_label_w, info_h, _pdf_text(label), border=1)
        pdf.set_font("Helvetica", size=8.3)
        pdf.cell(info_value_w, info_h, _pdf_text(value)[:32], border=1)
        if index % 2 == 1:
            pdf.ln(info_h)

    # If the info table has an odd number of cells, force a line break
    # so the grades table always starts below it.
    if len(info_rows) % 2 == 1:
        pdf.ln(info_h)

    rows = payload["rows"]
    table_columns = [
        ("N", 10, "index"),
        ("Matiere", 86, "subject"),
        ("Coef", 14, "coef"),
        ("Classe /20", 22, "note_classe"),
        ("Examen /20", 22, "note_examen"),
        ("Moyenne /20", 22, "note_finale"),
        ("Points", 22, "points"),
        ("Appreciation", 34, "appreciation"),
    ]
    table_width = sum(column[1] for column in table_columns)
    table_x = max(left_margin, (pdf.w - table_width) / 2)

    table_y = pdf.get_y() + 2.4
    # Keep enough bottom space for director signature + stamp block.
    summary_start_y = 154
    header_h = 5.6
    available_for_rows = max(26.0, summary_start_y - table_y - header_h)
    row_count = max(len(rows), 1)
    row_h = max(2.5, min(5.4, available_for_rows / row_count))
    body_font_size = max(6.1, min(8.2, row_h + 2.0))
    subject_max_len = max(22, min(64, int(64 * (row_h / 5.4))))

    pdf.set_y(table_y)
    pdf.set_x(table_x)
    pdf.set_font("Helvetica", "B", 8.1)
    pdf.set_fill_color(228, 234, 244)
    for title, width, key in table_columns:
        align = "L" if key == "subject" else "C"
        pdf.cell(width, header_h, _pdf_text(title), border=1, fill=True, align=align)
    pdf.ln(header_h)

    pdf.set_font("Helvetica", size=body_font_size)
    if not rows:
        pdf.set_x(table_x)
        pdf.cell(
            table_width,
            row_h,
            _pdf_text("Aucune note disponible pour cette periode."),
            border=1,
            align="C",
        )
        pdf.ln(row_h)
    else:
        for row in rows:
            fill_row = row["index"] % 2 == 0
            if fill_row:
                pdf.set_fill_color(248, 250, 253)
            pdf.set_x(table_x)
            pdf.cell(10, row_h, _pdf_text(str(row["index"])), border=1, align="C", fill=fill_row)
            pdf.cell(86, row_h, _pdf_text(str(row["subject"])[:subject_max_len]), border=1, fill=fill_row)
            pdf.cell(14, row_h, _pdf_text(_format_coef_value(row["coef"])), border=1, align="C", fill=fill_row)
            pdf.cell(22, row_h, _pdf_text(_format_cell_value(row["note_classe"])), border=1, align="C", fill=fill_row)
            pdf.cell(22, row_h, _pdf_text(_format_cell_value(row["note_examen"])), border=1, align="C", fill=fill_row)
            pdf.cell(22, row_h, _pdf_text(_format_cell_value(row["note_finale"])), border=1, align="C", fill=fill_row)
            pdf.cell(22, row_h, _pdf_text(_format_cell_value(row.get("points"))), border=1, align="C", fill=fill_row)
            pdf.cell(34, row_h, _pdf_text(str(row.get("appreciation") or "-")[:20]), border=1, align="C", fill=fill_row)
            pdf.ln(row_h)

    # Keep at least 1 cm (10 mm) between the end of the table and the summary block.
    summary_y = max(summary_start_y, pdf.get_y() + 10.0)
    pdf.set_y(summary_y)
    pdf.set_font("Helvetica", "B", 9.2)
    pdf.cell(0, 4.3, _pdf_text(f"Moyenne generale ponderee: {payload['average']:.2f}/20"), ln=True)
    pdf.cell(0, 4.3, _pdf_text(f"Total coefficients utilises: {_format_coef_value(payload['coef_sum'])}"), ln=True)
    pdf.cell(0, 4.3, _pdf_text(f"Rang: {payload['rank_display']}"), ln=True)
    pdf.cell(0, 4.3, _pdf_text(f"Mention: {payload['mention']}"), ln=True)

    pdf.set_font("Helvetica", size=7.2)
    pdf.set_text_color(70, 70, 70)
    pdf.multi_cell(
        0,
        3.8,
        _pdf_text(
            "Formule: Moyenne matiere = (Note classe + Note examen) / 2 si les deux existent, "
            "sinon la note disponible. Points = Moyenne matiere x Coefficient."
        ),
    )
    pdf.set_text_color(0, 0, 0)

    signature_y = 198
    left_sig_x1 = left_margin + 12
    left_sig_x2 = left_sig_x1 + 62
    right_sig_x2 = right_margin - 12
    right_sig_x1 = right_sig_x2 - 62

    signature_asset_path = payload.get("signature_asset_path")
    stamp_asset_path = payload.get("stamp_asset_path")
    signature_label = str(payload.get("signature_label") or "Direction")
    signature_position = str(payload.get("signature_position") or "right").strip().lower()
    stamp_position = str(payload.get("stamp_position") or "right").strip().lower()
    signature_scale = float(payload.get("signature_scale") or 1.0)
    stamp_scale = float(payload.get("stamp_scale") or 1.0)

    right_line_w = right_sig_x2 - right_sig_x1
    signature_w = max(16.0, min(46.0, right_line_w * 0.90 * signature_scale))
    signature_h = max(4.2, min(12.0, 5.8 * signature_scale))
    stamp_size = max(11.0, min(24.0, 16.0 * stamp_scale))

    # Compute the lowest safe baseline so images stay fully visible.
    signature_bottom_padding = 5.2 + signature_h + (stamp_size * 0.42) + 2.0
    signature_max_y = pdf.h - signature_bottom_padding
    signature_pref_y = max(pdf.get_y() + 2.0, 176.0)
    signature_y = min(signature_pref_y, signature_max_y)
    if signature_y < 166.0:
        signature_y = 166.0

    pdf.line(left_sig_x1, signature_y, left_sig_x2, signature_y)
    pdf.line(right_sig_x1, signature_y, right_sig_x2, signature_y)
    signature_default_x = right_sig_x1 + max(0.0, (right_line_w - signature_w) / 2.0)
    signature_x = _positioned_x(
        signature_position,
        min_x=right_sig_x1,
        max_x=right_sig_x2,
        box_width=signature_w,
        default_x=signature_default_x,
    )
    # Keep signature and stamp under the "Le Directeur" label.
    signature_img_y = signature_y + 5.2

    if signature_asset_path:
        try:
            pdf.image(signature_asset_path, x=signature_x, y=signature_img_y, w=signature_w, h=signature_h)
        except Exception:
            signature_asset_path = None

    if not signature_asset_path:
        fallback_line_y = signature_img_y + (signature_h * 0.62)
        pdf.set_draw_color(85, 96, 112)
        pdf.set_line_width(0.16)
        pdf.line(signature_x + 0.3, fallback_line_y, signature_x + signature_w - 0.3, fallback_line_y)

    stamp_default_x = min(right_sig_x2 - stamp_size, signature_x + signature_w - (stamp_size * 0.5))
    stamp_x = _positioned_x(
        stamp_position,
        min_x=right_sig_x1,
        max_x=right_sig_x2,
        box_width=stamp_size,
        default_x=stamp_default_x,
    )
    stamp_y = signature_img_y + max(0.2, signature_h - (stamp_size * 0.58))
    if stamp_asset_path:
        try:
            pdf.image(stamp_asset_path, x=stamp_x, y=stamp_y, w=stamp_size, h=stamp_size)
        except Exception:
            stamp_asset_path = None

    if not stamp_asset_path:
        pdf.set_draw_color(31, 90, 161)
        pdf.set_line_width(0.18)
        try:
            pdf.ellipse(stamp_x, stamp_y, stamp_size, stamp_size)
            pdf.ellipse(stamp_x + 1.7, stamp_y + 1.7, stamp_size - 3.4, stamp_size - 3.4)
        except Exception:
            pdf.rect(stamp_x, stamp_y, stamp_size, stamp_size)
            pdf.rect(stamp_x + 1.7, stamp_y + 1.7, stamp_size - 3.4, stamp_size - 3.4)

        pdf.set_xy(stamp_x, stamp_y + (stamp_size * 0.45))
        pdf.set_font("Helvetica", "B", 5.8)
        pdf.set_text_color(31, 90, 161)
        pdf.cell(stamp_size, 2.4, _pdf_text("Cachet"), align="C")

    pdf.set_y(signature_y + 1.2)
    pdf.set_font("Helvetica", size=7.8)
    pdf.set_x(left_sig_x1)
    pdf.cell(left_sig_x2 - left_sig_x1, 3.8, _pdf_text("Titulaire / Enseignant"), align="C")
    pdf.set_x(right_sig_x1)
    pdf.cell(right_sig_x2 - right_sig_x1, 3.8, _pdf_text(signature_label), align="C")


class ClassBulletinsPdfView(APIView):
    access_module = "reports"
    permission_classes = [IsAuthenticated, HasModuleAccess]

    def get(self, request, classroom_id: int, academic_year_id: int, term: str):
        _ensure_sensitive_export_access(request)

        normalized_term = normalize_term(term)
        if not normalized_term:
            return Response(
                {"detail": "Période invalide. Utilisez uniquement T1, T2 ou T3."},
                status=400,
            )

        if getattr(request.user, "role", "") in {UserRole.PARENT, UserRole.STUDENT}:
            raise PermissionDenied("Accès refusé aux bulletins de classe.")

        classroom = get_object_or_404(ClassRoom, id=classroom_id)
        target_etablissement_id = _effective_etablissement_id(request)
        if getattr(request.user, "role", "") == UserRole.SUPER_ADMIN and target_etablissement_id is None:
            raise PermissionDenied("Selectionnez un etablissement actif.")
        if target_etablissement_id and classroom.etablissement_id != target_etablissement_id:
            raise PermissionDenied("Accès refusé aux bulletins de cette classe.")

        students = list(
            _allowed_students_queryset(request)
            .filter(classroom_id=classroom.id, is_archived=False)
        )

        if students:
            rank_by_student_id = {
                int(student_id): int(rank)
                for student_id, rank in StudentAcademicHistory.objects.filter(
                    classroom_id=classroom.id,
                    academic_year_id=academic_year_id,
                    student_id__in=[student.id for student in students],
                ).values_list("student_id", "rank")
                if rank is not None and int(rank) > 0
            }

            def _student_rank_key(student: Student):
                rank = rank_by_student_id.get(student.id)
                last_name = str(getattr(student.user, "last_name", "") or "").lower()
                first_name = str(getattr(student.user, "first_name", "") or "").lower()
                matricule = str(getattr(student, "matricule", "") or "").lower()
                if rank is None:
                    return (1, 10**9, last_name, first_name, matricule, student.id)
                return (0, rank, last_name, first_name, matricule, student.id)

            students.sort(key=_student_rank_key)

        if not students:
            return Response({"detail": "Aucun élève trouvé pour cette classe."}, status=404)

        pdf = FPDF(orientation="L", format="A4")
        pdf.set_auto_page_break(auto=False)

        for student in students:
            payload = _build_bulletin_payload(
                student=student,
                academic_year_id=academic_year_id,
                normalized_term=normalized_term,
            )
            pdf.add_page()
            _render_bulletin_page(pdf, payload)

        safe_term = str(normalized_term or term or "periode").replace("/", "-")
        class_slug = str(classroom.name or f"classe_{classroom.id}").strip().replace(" ", "_")
        return pdf_output_response(pdf, f"bulletins_{class_slug}_{safe_term}.pdf")


def _render_payment_receipt_page(pdf: FPDF, payment: Payment):
        student = payment.fee.student
        school = _school_identity_for_student(student) if student else _school_identity()
        logo_path = (_etablissement_logo_path(student) if student else None) or _school_logo_path()
        etablissement = _payment_etablissement(payment)
        cashier_signature_source = _etablissement_media_field_path(etablissement, "cashier_signature_image") or _school_signature_asset_path()
        stamp_source = _etablissement_media_field_path(etablissement, "stamp_image") or _school_stamp_asset_path()
        cashier_signature_path = _pdf_compatible_image_path(cashier_signature_source, cache_prefix="receipt_cashier_signature")
        stamp_asset_path = _pdf_compatible_image_path(stamp_source, cache_prefix="receipt_stamp")
        cashier_signature_label = str(getattr(etablissement, "cashier_signature_label", "") or "").strip() or "Signature caissier"
        parent_signature_label = str(getattr(etablissement, "parent_signature_label", "") or "").strip() or "Signature parent / eleve"
        stamp_position = str(getattr(etablissement, "stamp_position", "") or "right").strip().lower()
        stamp_scale = _safe_scale_percent(getattr(etablissement, "stamp_scale", 100)) / 100.0
        signature_scale = _safe_scale_percent(getattr(etablissement, "principal_signature_scale", 100)) / 100.0

        student_user = student.user if student else None
        student_name = student_user.get_full_name().strip() if student_user else ""
        student_name = student_name or (student_user.username if student_user else "")
        class_name = student.classroom.name if student and student.classroom else "N/A"
        academic_year = payment.fee.academic_year.name if payment.fee and payment.fee.academic_year else "N/A"

        receiver = payment.received_by
        receiver_name = ""
        if receiver:
            receiver_name = receiver.get_full_name().strip() or receiver.username

        payer_name = ""
        if student and student.parent and student.parent.user:
            parent_user = student.parent.user
            payer_name = parent_user.get_full_name().strip() or parent_user.username
        if not payer_name:
            payer_name = student_name or "Parent / Eleve"

        payment_amount_label = _format_fcfa(payment.amount)
        remaining_balance_label = _format_fcfa(payment.fee.balance if payment.fee else 0)
        receipt_no = f"RC-{timezone.localtime(payment.created_at).strftime('%Y%m%d')}-{payment.id:05d}"
        issue_date = timezone.localtime(payment.created_at).strftime("%d/%m/%Y %H:%M")
        fee_type = payment.fee.get_fee_type_display() if payment.fee else "N/A"
        method = payment.method or "N/A"
        reference = payment.reference or "-"

        page_x = 7
        page_y = 7
        page_w = pdf.w - 14
        page_h = pdf.h - 14

        pdf.set_fill_color(245, 247, 251)
        pdf.set_draw_color(71, 92, 124)
        pdf.set_line_width(0.45)
        pdf.rect(page_x, page_y, page_w, page_h, style="DF")

        pdf.set_draw_color(141, 156, 179)
        pdf.set_line_width(0.18)
        pdf.rect(page_x + 0.8, page_y + 0.8, page_w - 1.6, page_h - 1.6)

        content_x = page_x + 3.4
        content_y = page_y + 3.0
        content_w = page_w - 6.8

        if logo_path:
            try:
                pdf.image(logo_path, x=content_x, y=content_y + 0.2, w=12)
            except Exception:
                pass

        header_x = content_x + (14 if logo_path else 0)
        header_w = content_w - (14 if logo_path else 0)

        pdf.set_text_color(23, 69, 137)
        pdf.set_xy(header_x, content_y)
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(header_w, 5.8, _pdf_text(school["name"]).upper()[:62], align="C")

        pdf.set_text_color(33, 38, 46)
        pdf.set_xy(header_x, content_y + 5.0)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(
            header_w,
            4,
            _pdf_text(f"{school['short']} ({school['level']})").upper()[:60],
            align="C",
        )

        if school["phone"]:
            pdf.set_text_color(182, 53, 59)
            pdf.set_xy(header_x, content_y + 8.6)
            pdf.set_font("Helvetica", "B", 8)
            pdf.cell(header_w, 3.6, _pdf_text(f"Tel : {school['phone']}"), align="C")

        title_y = content_y + 13.2
        pdf.set_fill_color(24, 93, 168)
        pdf.rect(content_x, title_y, content_w, 7.1, style="F")
        pdf.set_text_color(255, 255, 255)
        pdf.set_xy(content_x, title_y + 0.4)
        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(content_w, 5.8, _pdf_text("RECU DE PAIEMENT"), align="C")

        meta_y = title_y + 8.8
        meta_h = 5.8
        left_meta_w = content_w * 0.52
        center_meta_w = content_w * 0.27
        right_meta_w = content_w - left_meta_w - center_meta_w

        pdf.set_fill_color(235, 241, 250)
        pdf.set_draw_color(157, 173, 197)
        pdf.set_line_width(0.16)
        pdf.rect(content_x, meta_y, left_meta_w, meta_h, style="DF")
        pdf.rect(content_x + left_meta_w, meta_y, center_meta_w, meta_h, style="DF")
        pdf.rect(content_x + left_meta_w + center_meta_w, meta_y, right_meta_w, meta_h, style="DF")

        pdf.set_text_color(45, 50, 60)
        pdf.set_xy(content_x + 1, meta_y + 1.3)
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(left_meta_w - 2, 3.1, _pdf_text(f"Recu N° : {receipt_no}")[:42])

        pdf.set_xy(content_x + left_meta_w + 1, meta_y + 1.3)
        pdf.cell(center_meta_w - 2, 3.1, _pdf_text(f"Date : {issue_date}")[:28])

        pdf.set_xy(content_x + left_meta_w + center_meta_w + 1, meta_y + 1.3)
        pdf.cell(right_meta_w - 2, 3.1, _pdf_text(f"Annee : {academic_year}")[:20])

        section_gap = 2.6
        student_box_y = meta_y + meta_h + section_gap
        student_box_h = 28
        payment_box_y = student_box_y + student_box_h + section_gap
        payment_box_h = 35

        pdf.set_fill_color(252, 253, 255)
        pdf.set_draw_color(168, 181, 202)
        pdf.rect(content_x, student_box_y, content_w, student_box_h, style="DF")

        pdf.set_fill_color(237, 243, 251)
        pdf.rect(content_x, student_box_y, content_w, 5.6, style="F")
        pdf.set_text_color(28, 72, 136)
        pdf.set_xy(content_x + 1.2, student_box_y + 1.2)
        pdf.set_font("Helvetica", "B", 8.5)
        pdf.cell(content_w - 2.4, 3.2, _pdf_text("INFORMATIONS ELEVE"))

        info_x = content_x + 1.4
        info_y = student_box_y + 7.4
        info_rows = [
            ("Nom complet", student_name or "N/A"),
            ("Matricule", student.matricule if student else "N/A"),
            ("Classe", class_name),
            ("Payeur", payer_name),
        ]

        for index, (label, value) in enumerate(info_rows):
            row_y = info_y + (index * 4.9)
            pdf.set_xy(info_x, row_y)
            pdf.set_text_color(48, 55, 65)
            pdf.set_font("Helvetica", "B", 8)
            pdf.cell(27, 3.4, _pdf_text(f"{label} :"))
            pdf.set_text_color(28, 72, 136)
            pdf.set_font("Helvetica", "B", 8)
            pdf.cell(content_w - 31, 3.4, _pdf_text(str(value))[:56])
            pdf.set_draw_color(189, 200, 217)
            pdf.line(info_x, row_y + 3.7, content_x + content_w - 1.2, row_y + 3.7)

        pdf.set_fill_color(252, 253, 255)
        pdf.set_draw_color(168, 181, 202)
        pdf.rect(content_x, payment_box_y, content_w, payment_box_h, style="DF")

        pdf.set_fill_color(237, 243, 251)
        pdf.rect(content_x, payment_box_y, content_w, 5.6, style="F")
        pdf.set_text_color(28, 72, 136)
        pdf.set_xy(content_x + 1.2, payment_box_y + 1.2)
        pdf.set_font("Helvetica", "B", 8.5)
        pdf.cell(content_w - 2.4, 3.2, _pdf_text("DETAIL DU PAIEMENT"))

        details_x = content_x + 1.4
        details_y = payment_box_y + 7.4
        details_rows = [
            ("Type de frais", fee_type),
            ("Methode", method),
            ("Reference", reference),
            ("Encaisse par", receiver_name or "N/A"),
            ("Solde restant", remaining_balance_label),
        ]

        for index, (label, value) in enumerate(details_rows):
            row_y = details_y + (index * 4.3)
            pdf.set_xy(details_x, row_y)
            pdf.set_text_color(48, 55, 65)
            pdf.set_font("Helvetica", "B", 7.8)
            pdf.cell(25, 3.1, _pdf_text(f"{label} :"))
            pdf.set_text_color(28, 72, 136)
            pdf.set_font("Helvetica", "B", 7.8)
            pdf.cell((content_w * 0.56) - 2, 3.1, _pdf_text(str(value))[:34])

        amount_box_w = content_w * 0.36
        amount_box_x = content_x + content_w - amount_box_w - 1.2
        amount_box_y = payment_box_y + 10.4
        amount_box_h = 18.6
        pdf.set_fill_color(24, 93, 168)
        pdf.set_draw_color(18, 72, 133)
        pdf.rect(amount_box_x, amount_box_y, amount_box_w, amount_box_h, style="DF")

        pdf.set_text_color(255, 255, 255)
        pdf.set_xy(amount_box_x, amount_box_y + 2.0)
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(amount_box_w, 3.2, _pdf_text("MONTANT VERSE"), align="C")

        pdf.set_xy(amount_box_x, amount_box_y + 7.6)
        pdf.set_font("Helvetica", "B", 11.2)
        pdf.cell(amount_box_w, 5, _pdf_text(payment_amount_label)[:24], align="C")

        pdf.set_text_color(66, 72, 84)
        pdf.set_xy(content_x, payment_box_y + payment_box_h + 2.6)
        pdf.set_font("Helvetica", size=7.6)
        pdf.multi_cell(
            content_w,
            3.4,
            _pdf_text(
                f"Merci pour votre paiement. Ce recu certifie l'encaissement effectif de {payment_amount_label} au profit de {school['short']}."
            ),
        )

        signature_y = page_y + page_h - 9.5
        left_sign_x1 = content_x + 2
        left_sign_x2 = left_sign_x1 + 42
        right_sign_x2 = content_x + content_w - 2
        right_sign_x1 = right_sign_x2 - 42

        pdf.set_draw_color(115, 130, 154)
        pdf.set_line_width(0.2)
        pdf.line(left_sign_x1, signature_y, left_sign_x2, signature_y)
        pdf.line(right_sign_x1, signature_y, right_sign_x2, signature_y)

        cashier_signature_w = max(14.0, min(34.0, (left_sign_x2 - left_sign_x1) * 0.88 * signature_scale))
        cashier_signature_h = max(3.6, min(9.2, 4.6 * signature_scale))
        cashier_signature_x = left_sign_x1 + max(0.0, ((left_sign_x2 - left_sign_x1) - cashier_signature_w) / 2.0)
        cashier_signature_y = signature_y - cashier_signature_h - 0.8
        if cashier_signature_path:
            try:
                pdf.image(
                    cashier_signature_path,
                    x=cashier_signature_x,
                    y=cashier_signature_y,
                    w=cashier_signature_w,
                    h=cashier_signature_h,
                )
            except Exception:
                cashier_signature_path = None

        pdf.set_xy(left_sign_x1, signature_y + 0.6)
        pdf.set_font("Helvetica", "B", 7.3)
        pdf.set_text_color(66, 72, 84)
        pdf.cell(left_sign_x2 - left_sign_x1, 3.4, _pdf_text(cashier_signature_label), align="C")

        pdf.set_xy(right_sign_x1, signature_y + 0.6)
        pdf.cell(right_sign_x2 - right_sign_x1, 3.4, _pdf_text(parent_signature_label), align="C")

        stamp_size = max(12.0, min(28.0, 18 * stamp_scale))
        stamp_default_x = content_x + content_w - stamp_size - 1.4
        stamp_x = _positioned_x(
            stamp_position,
            min_x=content_x + 1.2,
            max_x=content_x + content_w - 1.2,
            box_width=stamp_size,
            default_x=stamp_default_x,
        )
        stamp_y = page_y + page_h - stamp_size - 13.0
        if stamp_asset_path:
            try:
                pdf.image(stamp_asset_path, x=stamp_x, y=stamp_y, w=stamp_size, h=stamp_size)
            except Exception:
                stamp_asset_path = None

        if not stamp_asset_path:
            pdf.set_draw_color(31, 90, 161)
            pdf.set_line_width(0.24)
            try:
                pdf.ellipse(stamp_x, stamp_y, stamp_size, stamp_size)
                pdf.ellipse(stamp_x + 2.8, stamp_y + 2.8, stamp_size - 5.6, stamp_size - 5.6)
            except Exception:
                pdf.rect(stamp_x, stamp_y, stamp_size, stamp_size)
                pdf.rect(stamp_x + 2.8, stamp_y + 2.8, stamp_size - 5.6, stamp_size - 5.6)

            pdf.set_text_color(31, 90, 161)
            pdf.set_xy(stamp_x, stamp_y + 6.0)
            pdf.set_font("Helvetica", "B", 7.8)
            pdf.cell(stamp_size, 3.2, _pdf_text(school["short"])[:12], align="C")
            pdf.set_xy(stamp_x, stamp_y + 9.6)
            pdf.set_font("Helvetica", "B", 6.5)
            pdf.cell(stamp_size, 2.8, _pdf_text(school["city"])[:12], align="C")

def _build_receipts_pdf(payments: list[Payment]) -> FPDF:
    pdf = FPDF(format="A5")
    pdf.set_auto_page_break(auto=False)
    for payment in payments:
        pdf.add_page()
        _render_payment_receipt_page(pdf, payment)
    return pdf


class PaymentReceiptPdfView(APIView):
    access_module = "reports"
    permission_classes = [IsAuthenticated, HasModuleAccess]

    def get(self, request, payment_id: int):
        payment = Payment.objects.select_related(
            "fee__student__user",
            "fee__student__parent",
            "fee__student__parent__user",
            "fee__student__classroom",
            "fee__student__etablissement",
            "fee__student__classroom__etablissement",
            "fee__academic_year",
            "received_by",
            "etablissement",
        ).get(id=payment_id, is_cancelled=False)
        _ensure_payment_access(request, payment)
        pdf = _build_receipts_pdf([payment])
        return pdf_output_response(pdf, f"receipt_{payment.id}.pdf")


class BatchPaymentReceiptsPdfView(APIView):
    access_module = "reports"
    # Export en lot: POST par commodite (liste d'identifiants dans le corps),
    # mais l'operation ne fait que lire. Exiger le niveau ecriture priverait
    # le comptable de ses recus.
    access_read_only = True
    permission_classes = [IsAuthenticated, HasModuleAccess]

    def post(self, request):
        raw_ids = request.data.get("payment_ids") if isinstance(request.data, dict) else None
        if not isinstance(raw_ids, list):
            raise ValidationError({"payment_ids": "Fournissez une liste d'identifiants de paiements."})

        payment_ids: list[int] = []
        for raw in raw_ids:
            try:
                parsed = int(raw)
            except (TypeError, ValueError):
                continue
            if parsed > 0:
                payment_ids.append(parsed)

        payment_ids = list(dict.fromkeys(payment_ids))
        if not payment_ids:
            raise ValidationError({"payment_ids": "Aucun identifiant de paiement valide."})
        if len(payment_ids) > 80:
            raise ValidationError({"payment_ids": "Limite maximale: 80 reçus par impression groupée."})

        queryset = Payment.objects.select_related(
            "fee__student__user",
            "fee__student__parent",
            "fee__student__parent__user",
            "fee__student__classroom",
            "fee__student__etablissement",
            "fee__student__classroom__etablissement",
            "fee__academic_year",
            "received_by",
            "etablissement",
        ).filter(id__in=payment_ids, is_cancelled=False)

        payment_by_id = {payment.id: payment for payment in queryset}
        payments: list[Payment] = []
        for payment_id in payment_ids:
            payment = payment_by_id.get(payment_id)
            if payment is None:
                continue
            _ensure_payment_access(request, payment)
            payments.append(payment)

        if not payments:
            raise ValidationError({"payment_ids": "Aucun paiement accessible pour cette sélection."})

        pdf = _build_receipts_pdf(payments)
        timestamp = timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M")
        return pdf_output_response(pdf, f"receipts_batch_{timestamp}.pdf")


class PaymentExcelExportView(APIView):
    access_module = "reports"
    permission_classes = [IsAuthenticated, HasModuleAccess]

    def get(self, request):
        _ensure_sensitive_export_access(request)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Paiements"

        school_name = getattr(settings, "SCHOOL_NAME", "LYCEE TECHNIQUE OUMAR BAH")
        school_short = getattr(settings, "SCHOOL_SHORT", "LTOB")
        school_level = getattr(settings, "SCHOOL_LEVEL", "1er etage")
        school_phone = getattr(settings, "SCHOOL_PHONE", "")
        logo_path = _school_logo_path()

        columns = [
            "Recu N°",
            "Eleve",
            "Matricule",
            "Classe",
            "Type de frais",
            "Montant (FCFA)",
            "Methode",
            "Reference",
            "Date",
            "Encaisse par",
        ]
        last_col = len(columns)
        last_col_letter = get_column_letter(last_col)

        sheet.merge_cells(f"A1:{last_col_letter}1")
        sheet.merge_cells(f"A2:{last_col_letter}2")
        sheet.merge_cells(f"A3:{last_col_letter}3")

        sheet["A1"] = school_name
        sheet["A2"] = f"{school_level} | Tel: {school_phone}" if school_phone else school_level
        sheet["A3"] = f"ETAT DES PAIEMENTS - {school_short}"

        sheet["A1"].font = Font(bold=True, size=16, color="1F3B63")
        sheet["A2"].font = Font(size=11, color="3F3F3F")
        sheet["A3"].font = Font(bold=True, size=12, color="FFFFFF")
        sheet["A3"].fill = PatternFill(fill_type="solid", fgColor="1F4E78")

        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["A2"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["A3"].alignment = Alignment(horizontal="center", vertical="center")

        sheet.row_dimensions[1].height = 28
        sheet.row_dimensions[2].height = 20
        sheet.row_dimensions[3].height = 24

        if logo_path:
            try:
                from openpyxl.drawing.image import Image as XLImage

                logo = XLImage(logo_path)
                logo.width = 50
                logo.height = 50
                logo_anchor_col = get_column_letter(max(1, last_col - 1))
                sheet.add_image(logo, f"{logo_anchor_col}1")
            except Exception:
                pass

        thin_side = Side(style="thin", color="C8CDD3")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        header_row = 5
        for col_index, title in enumerate(columns, start=1):
            cell = sheet.cell(row=header_row, column=col_index, value=title)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="3A6EA5")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        payments = _allowed_payments_queryset(request).order_by("-created_at")

        row_index = header_row + 1
        total_amount = 0.0

        for payment in payments:
            student = payment.fee.student if payment.fee else None
            student_user = student.user if student else None
            student_name = ""
            if student_user:
                student_name = student_user.get_full_name().strip() or student_user.username

            receiver = payment.received_by
            receiver_name = ""
            if receiver:
                receiver_name = receiver.get_full_name().strip() or receiver.username

            amount_value = float(payment.amount)
            total_amount += amount_value

            row_values = [
                payment.id,
                student_name,
                student.matricule if student else "N/A",
                student.classroom.name if student and student.classroom else "N/A",
                payment.fee.get_fee_type_display() if payment.fee else "N/A",
                amount_value,
                payment.method,
                payment.reference or "-",
                payment.created_at.strftime("%d/%m/%Y %H:%M"),
                receiver_name or "N/A",
            ]

            for col_index, value in enumerate(row_values, start=1):
                cell = sheet.cell(row=row_index, column=col_index, value=value)
                cell.border = thin_border
                if col_index == 6:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_index in (1, 9):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            row_index += 1

        if row_index == header_row + 1:
            sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=last_col)
            empty_cell = sheet.cell(row=row_index, column=1, value="Aucun paiement disponible.")
            empty_cell.alignment = Alignment(horizontal="center", vertical="center")
            empty_cell.font = Font(italic=True, color="6B7280")
            empty_cell.border = thin_border
            row_index += 1

        summary_row = row_index + 1
        sheet.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=5)
        summary_label_cell = sheet.cell(row=summary_row, column=1, value="TOTAL ENCAISSE")
        summary_label_cell.font = Font(bold=True, color="1F3B63")
        summary_label_cell.fill = PatternFill(fill_type="solid", fgColor="E8EEF7")
        summary_label_cell.alignment = Alignment(horizontal="center", vertical="center")
        summary_label_cell.border = thin_border

        summary_value_cell = sheet.cell(row=summary_row, column=6, value=total_amount)
        summary_value_cell.number_format = '#,##0.00'
        summary_value_cell.font = Font(bold=True, color="1F3B63")
        summary_value_cell.fill = PatternFill(fill_type="solid", fgColor="E8EEF7")
        summary_value_cell.alignment = Alignment(horizontal="right", vertical="center")
        summary_value_cell.border = thin_border

        for col_index in range(7, last_col + 1):
            empty = sheet.cell(row=summary_row, column=col_index, value="")
            empty.fill = PatternFill(fill_type="solid", fgColor="E8EEF7")
            empty.border = thin_border

        generated_row = summary_row + 2
        sheet.merge_cells(start_row=generated_row, start_column=1, end_row=generated_row, end_column=last_col)
        generated_by = request.user.get_full_name().strip() or request.user.username
        generated_at = timezone.localtime().strftime("%d/%m/%Y %H:%M")
        generated_cell = sheet.cell(
            row=generated_row,
            column=1,
            value=f"Genere le {generated_at} par {generated_by}",
        )
        generated_cell.font = Font(italic=True, color="6B7280")
        generated_cell.alignment = Alignment(horizontal="left", vertical="center")

        column_widths = [10, 28, 16, 18, 18, 16, 16, 18, 20, 22]
        for col_index, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(col_index)].width = width

        sheet.freeze_panes = "A6"

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="payments_export.xlsx"'
        workbook.save(response)
        return response


class PaymentJournalPageView(APIView):
    access_module = "reports"
    permission_classes = [IsAuthenticated, HasModuleAccess]

    def get(self, request):
        _ensure_sensitive_export_access(request)
        queryset = _apply_payment_journal_filters(
            _allowed_payments_queryset(request).order_by("-created_at", "-id"),
            request,
        )

        page_size = _parse_page_size(request, default=100, max_size=1000)
        paginator = Paginator(queryset, page_size)
        try:
            page_number = int(request.query_params.get("page", 1))
        except (TypeError, ValueError):
            page_number = 1
        page_number = max(1, page_number)
        page_obj = paginator.get_page(page_number)

        base_url = request.build_absolute_uri(request.path)
        query_dict = request.query_params.copy()

        def _page_link(page_no: int | None) -> str | None:
            if page_no is None:
                return None
            query_copy = query_dict.copy()
            query_copy["page"] = str(page_no)
            return f"{base_url}?{query_copy.urlencode()}"

        return Response(
            {
                "count": paginator.count,
                "next": _page_link(page_obj.next_page_number() if page_obj.has_next() else None),
                "previous": _page_link(page_obj.previous_page_number() if page_obj.has_previous() else None),
                "results": [_payment_journal_row(payment) for payment in page_obj.object_list],
            }
        )


class ExpenseJournalPageView(APIView):
    access_module = "reports"
    permission_classes = [IsAuthenticated, HasModuleAccess]

    def get(self, request):
        _ensure_sensitive_export_access(request)
        queryset = _apply_expense_journal_filters(
            _allowed_expenses_queryset(request).order_by("-date", "-id"),
            request,
        )

        page_size = _parse_page_size(request, default=100, max_size=1000)
        paginator = Paginator(queryset, page_size)
        try:
            page_number = int(request.query_params.get("page", 1))
        except (TypeError, ValueError):
            page_number = 1
        page_number = max(1, page_number)
        page_obj = paginator.get_page(page_number)

        base_url = request.build_absolute_uri(request.path)
        query_dict = request.query_params.copy()

        def _page_link(page_no: int | None) -> str | None:
            if page_no is None:
                return None
            query_copy = query_dict.copy()
            query_copy["page"] = str(page_no)
            return f"{base_url}?{query_copy.urlencode()}"

        return Response(
            {
                "count": paginator.count,
                "next": _page_link(page_obj.next_page_number() if page_obj.has_next() else None),
                "previous": _page_link(page_obj.previous_page_number() if page_obj.has_previous() else None),
                "results": [_expense_journal_row(expense) for expense in page_obj.object_list],
            }
        )


def _csv_response(*, filename: str, header: list[str], rows: list[list]) -> HttpResponse:
    def _csv_cell(value) -> str:
        text = str(value or "")
        return f'"{text.replace("\"", "\"\"")}"'

    output = io.StringIO()
    output.write("\ufeff")
    for line in [header, *rows]:
        output.write(",".join([_csv_cell(value) for value in line]))
        output.write("\n")

    response = HttpResponse(output.getvalue(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _build_journal_pdf(*, title: str, subtitle: str, headers: list[str], rows: list[list[str]]) -> FPDF:
    pdf = FPDF(orientation="L", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 15)
    pdf.cell(0, 8, _pdf_text(title), ln=True)
    pdf.set_font("Helvetica", size=9)
    pdf.cell(0, 5, _pdf_text(subtitle), ln=True)
    pdf.cell(0, 5, _pdf_text(f"Genere le: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}"), ln=True)
    pdf.ln(3)

    page_width = pdf.w - 20
    col_width = page_width / max(1, len(headers))
    row_h = 6.2

    pdf.set_fill_color(57, 99, 151)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 8)
    for title_cell in headers:
        pdf.cell(col_width, row_h, _pdf_text(title_cell)[:30], border=1, align="C", fill=True)
    pdf.ln(row_h)

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", size=7.4)
    for row in rows:
        for value in row:
            pdf.cell(col_width, row_h, _pdf_text(str(value or ""))[:36], border=1)
        pdf.ln(row_h)

    return pdf


class PaymentJournalExportView(APIView):
    access_module = "reports"
    permission_classes = [IsAuthenticated, HasModuleAccess]

    def get(self, request):
        _ensure_sensitive_export_access(request)
        export_format = str(request.query_params.get("export_format", "csv") or "csv").strip().lower()
        queryset = _apply_payment_journal_filters(
            _allowed_payments_queryset(request).order_by("-created_at", "-id"),
            request,
        )
        rows = [_payment_journal_row(payment) for payment in queryset.iterator(chunk_size=1000)]

        stamp = timezone.localtime().strftime("%Y%m%d_%H%M")
        if export_format == "pdf":
            pdf = _build_journal_pdf(
                title="Journal des encaissements",
                subtitle=f"{len(rows)} ligne(s)",
                headers=["Date", "Eleve", "Matricule", "Type frais", "Montant", "Methode", "Reference", "Encaisse par"],
                rows=[
                    [
                        row.get("created_at", ""),
                        row.get("student_full_name", ""),
                        row.get("student_matricule", ""),
                        row.get("fee_type", ""),
                        f"{row.get('amount', 0):.0f}",
                        row.get("method", ""),
                        row.get("reference", ""),
                        row.get("received_by", ""),
                    ]
                    for row in rows
                ],
            )
            return pdf_output_response(pdf, f"journal_encaissements_{stamp}.pdf")

        return _csv_response(
            filename=f"journal_encaissements_{stamp}.csv",
            header=["id", "date", "eleve", "matricule", "type_frais", "montant", "methode", "reference", "encaisse_par"],
            rows=[
                [
                    row.get("id", ""),
                    row.get("created_at", ""),
                    row.get("student_full_name", ""),
                    row.get("student_matricule", ""),
                    row.get("fee_type", ""),
                    f"{row.get('amount', 0):.0f}",
                    row.get("method", ""),
                    row.get("reference", ""),
                    row.get("received_by", ""),
                ]
                for row in rows
            ],
        )


class ExpenseJournalExportView(APIView):
    access_module = "reports"
    permission_classes = [IsAuthenticated, HasModuleAccess]

    def get(self, request):
        _ensure_sensitive_export_access(request)
        export_format = str(request.query_params.get("export_format", "csv") or "csv").strip().lower()
        queryset = _apply_expense_journal_filters(
            _allowed_expenses_queryset(request).order_by("-date", "-id"),
            request,
        )
        rows = [_expense_journal_row(expense) for expense in queryset.iterator(chunk_size=1000)]

        stamp = timezone.localtime().strftime("%Y%m%d_%H%M")
        if export_format == "pdf":
            pdf = _build_journal_pdf(
                title="Journal des depenses",
                subtitle=f"{len(rows)} ligne(s)",
                headers=["Date", "Libelle", "Categorie", "Montant", "Validation", "Paye le", "Notes"],
                rows=[
                    [
                        row.get("date", ""),
                        row.get("label", ""),
                        row.get("category", ""),
                        f"{row.get('amount', 0):.0f}",
                        row.get("validation_stage", ""),
                        row.get("paid_on", ""),
                        row.get("notes", ""),
                    ]
                    for row in rows
                ],
            )
            return pdf_output_response(pdf, f"journal_depenses_{stamp}.pdf")

        return _csv_response(
            filename=f"journal_depenses_{stamp}.csv",
            header=["id", "date", "libelle", "categorie", "montant", "validation", "paye_le", "notes"],
            rows=[
                [
                    row.get("id", ""),
                    row.get("date", ""),
                    row.get("label", ""),
                    row.get("category", ""),
                    f"{row.get('amount', 0):.0f}",
                    row.get("validation_stage", ""),
                    row.get("paid_on", ""),
                    row.get("notes", ""),
                ]
                for row in rows
            ],
        )


class StudentCardPdfView(APIView):
    access_module = "reports"
    permission_classes = [IsAuthenticated, HasModuleAccess]

    def get(self, request, student_id: int):
        student = get_object_or_404(
            Student.objects.select_related("user", "classroom", "parent", "parent__user"),
            id=student_id,
        )
        _ensure_student_access(request, student)

        card_format = _requested_card_format(request)
        if card_format is None:
            return Response(
                {"detail": f"card_format invalide. Valeurs: {', '.join(CARD_FORMATS)}."},
                status=400,
            )

        school = _school_identity_for_student(student)
        logo_path = _etablissement_logo_path(student) or _school_logo_path()

        pdf = _build_student_cards_pdf(
            [student],
            school=school,
            logo_path=logo_path,
            layout_mode="standard",
            card_format=card_format,
            verify_base_url=_verify_base_url(request),
        )

        return pdf_output_response(pdf, f"carte_eleve_{student.matricule}.pdf")


class ClassStudentCardsPdfView(APIView):
    access_module = "reports"
    permission_classes = [IsAuthenticated, HasModuleAccess]

    def get(self, request, classroom_id: int):
        _ensure_sensitive_export_access(request)

        if getattr(request.user, "role", "") in {UserRole.PARENT, UserRole.STUDENT}:
            raise PermissionDenied("Accès refusé aux cartes de classe.")

        classroom = get_object_or_404(ClassRoom, id=classroom_id)
        target_etablissement_id = _effective_etablissement_id(request)
        if getattr(request.user, "role", "") == UserRole.SUPER_ADMIN and target_etablissement_id is None:
            raise PermissionDenied("Selectionnez un etablissement actif.")
        if target_etablissement_id and classroom.etablissement_id != target_etablissement_id:
            raise PermissionDenied("Accès refusé aux cartes de cette classe.")

        include_archived = (
            str(request.query_params.get("include_archived", "false")).strip().lower()
            in {"1", "true", "yes"}
        )
        layout_mode = str(request.query_params.get("layout_mode", "standard")).strip().lower()
        if layout_mode not in {"standard", "a4", "a4_6up", "a4_9up"}:
            return Response(
                {"detail": "layout_mode invalide. Valeurs: standard, a4, a4_6up, a4_9up."},
                status=400,
            )

        card_format = _requested_card_format(request)
        if card_format is None:
            return Response(
                {"detail": f"card_format invalide. Valeurs: {', '.join(CARD_FORMATS)}."},
                status=400,
            )

        queryset = Student.objects.select_related(
            "user", "classroom", "parent", "parent__user"
        ).filter(classroom_id=classroom.id)
        if not include_archived:
            queryset = queryset.filter(is_archived=False)

        students = list(queryset.order_by("user__last_name", "user__first_name", "matricule"))
        if not students:
            return Response(
                {"detail": "Aucun élève trouvé pour cette classe."},
                status=404,
            )

        school = _school_identity_for_student(students[0])
        logo_path = _etablissement_logo_path(students[0]) or _school_logo_path()
        pdf = _build_student_cards_pdf(
            students,
            school=school,
            logo_path=logo_path,
            layout_mode=layout_mode,
            card_format=card_format,
            verify_base_url=_verify_base_url(request),
        )

        class_slug = classroom.name.replace(" ", "_")
        suffix = {
            "a4": "_plancheA4",
            "a4_6up": "_6parA4",
            "a4_9up": "_9parA4",
        }.get(layout_mode, "")
        if card_format != CARD_FORMAT_DEFAUT:
            suffix += f"_{card_format}"
        return pdf_output_response(pdf, f"cartes_{class_slug}{suffix}.pdf")


class StudentCardVerifyView(APIView):
    """Page affichee quand on scanne le QR d'une carte scolaire.

    Publique par necessite: celui qui controle au portail n'a pas de compte.
    Elle ne revele donc aucune identite — ni nom, ni date de naissance, ni
    classe. Elle repond « valide » ou non, nomme l'ecole et l'annee, et montre
    la photo pour que le controleur compare un visage. Une carte perdue et
    scannee par un inconnu n'apprend rien sur l'eleve.

    L'acces suppose la signature imprimee sur la carte: les identifiants ne
    s'enumerent pas.
    """

    authentication_classes = ()
    permission_classes = [AllowAny]

    def get(self, request, student_id: int, annee: str, signature: str):
        if not signature_carte_valide(student_id, annee, signature):
            return self._page(
                titre="Carte non reconnue",
                detail="Cette carte n'a pas ete emise par l'etablissement.",
                valide=False,
            )

        student = Student.objects.select_related("etablissement").filter(id=student_id).first()
        if student is None:
            return self._page(
                titre="Carte non reconnue",
                detail="Cette carte ne correspond a aucun eleve.",
                valide=False,
            )

        annee_courante = _active_academic_year_label()
        ecole = str(getattr(getattr(student, "etablissement", None), "name", "") or "")

        if student.is_archived:
            return self._page(
                titre="Carte revoquee",
                detail=f"{ecole} — l'eleve n'est plus inscrit.",
                valide=False,
            )
        if annee != annee_courante:
            return self._page(
                titre="Carte expiree",
                detail=f"{ecole} — carte de l'annee {annee}, annee en cours : {annee_courante}.",
                valide=False,
            )

        return self._page(
            titre="Carte valide",
            detail=f"{ecole} — annee scolaire {annee}.",
            valide=True,
            photo_url=_student_photo_url(student, request),
        )

    def _page(self, *, titre, detail, valide, photo_url=None):
        couleur = "#1b7f3b" if valide else "#a3241f"
        photo = (
            f'<img src="{escape(photo_url)}" alt="Photo de l\'eleve">' if photo_url else ""
        )
        html = f"""<!doctype html>
<html lang="fr"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<meta name="robots" content="noindex, nofollow">
<title>{escape(titre)}</title>
<style>
 body{{font-family:system-ui,sans-serif;margin:0;padding:2rem 1rem;
      background:#f5f7fb;color:#1c2230;text-align:center}}
 .etat{{color:{couleur};font-size:1.6rem;font-weight:700;margin:0 0 .5rem}}
 .detail{{color:#4a5265;margin:0 0 1.5rem}}
 img{{max-width:220px;width:60%;border-radius:8px;border:3px solid {couleur}}}
 .note{{margin-top:2rem;font-size:.85rem;color:#6b7385}}
</style></head><body>
<p class="etat">{escape(titre)}</p>
<p class="detail">{escape(detail)}</p>
{photo}
<p class="note">Comparez la photo au porteur de la carte.</p>
</body></html>"""
        response = HttpResponse(html, content_type="text/html; charset=utf-8")
        response["Cache-Control"] = "no-store"
        response["X-Robots-Tag"] = "noindex, nofollow"
        return response


class StaffRosterPdfView(APIView):
    """Liste imprimable du personnel enseignant, avec colonne d'emargement.

    Memes regles que la liste d'appel des eleves: le module "reports" ouvre
    l'acces au personnel encadrant, les familles en sont exclues. Un document
    nommant tout le corps enseignant n'a rien a faire entre leurs mains.
    """

    access_module = "reports"
    permission_classes = [IsAuthenticated, HasModuleAccess]

    def get(self, request):
        if getattr(request.user, "role", "") in {UserRole.PARENT, UserRole.STUDENT}:
            raise PermissionDenied("Accès refusé à la liste du personnel.")

        target_etablissement_id = _effective_etablissement_id(request)
        # Sans etablissement determine, le filtre ci-dessous ne s'applique pas
        # et le document listerait tout le personnel de toutes les ecoles. Le
        # cas n'est pas theorique: un compte directeur mal rattache suffit.
        if target_etablissement_id is None:
            raise PermissionDenied("Selectionnez un etablissement actif.")

        queryset = Teacher.objects.select_related("user").prefetch_related(
            # Sans ce prefetch, _teacher_subjects_label ferait deux requetes
            # par enseignant: une pour ses affectations, une par matiere.
            "assignments__subject"
        )
        if target_etablissement_id:
            queryset = queryset.filter(etablissement_id=target_etablissement_id)

        teachers = list(queryset.order_by("user__last_name", "user__first_name", "id"))

        etablissement = (
            Etablissement.objects.filter(id=target_etablissement_id).first()
            if target_etablissement_id
            else None
        )
        school = _school_identity()
        if etablissement is not None:
            school = {**school, "name": etablissement.name}
        logo_path = (
            _etablissement_media_field_path(etablissement, "logo") or _school_logo_path()
        )

        pdf = _build_staff_roster_pdf(
            teachers,
            school=school,
            logo_path=logo_path,
            year_label=_active_academic_year_label(),
        )
        return pdf_output_response(pdf, "liste_enseignants.pdf")


class ClassRosterPdfView(APIView):
    """Liste d'appel imprimable, pour une classe ou pour tout l'etablissement.

    Volontairement sans `_ensure_sensitive_export_access`: ce garde-fou reserve
    les exports a l'administration et a la finance, alors qu'une liste d'appel
    sert d'abord en classe. Enseignants, censeurs et surveillants y ont donc
    acces via le module "reports"; parents et eleves en sont exclus, comme pour
    les cartes de classe.
    """

    access_module = "reports"
    permission_classes = [IsAuthenticated, HasModuleAccess]

    def get(self, request, classroom_id: int | None = None):
        if getattr(request.user, "role", "") in {UserRole.PARENT, UserRole.STUDENT}:
            raise PermissionDenied("Accès refusé aux listes de classe.")

        target_etablissement_id = _effective_etablissement_id(request)
        # Meme raison que pour la liste du personnel: sans etablissement, le
        # filtre ne s'applique pas et le document sortirait les classes de
        # toutes les ecoles. Refuser vaut mieux que fuiter.
        if target_etablissement_id is None:
            raise PermissionDenied("Selectionnez un etablissement actif.")

        # Trois etats, comme le filtre de l'ecran. Un simple booleen
        # "inclure les archives" ne saurait pas rendre "archives seulement",
        # et le document ne correspondrait plus a ce qui est affiche.
        statut = str(request.query_params.get("status", "active")).strip().lower()
        if statut not in {"active", "archived", "all"}:
            return Response(
                {"detail": "status invalide. Valeurs: active, archived, all."},
                status=400,
            )

        classrooms = ClassRoom.objects.all()
        if target_etablissement_id:
            classrooms = classrooms.filter(etablissement_id=target_etablissement_id)

        if classroom_id is not None:
            classroom = get_object_or_404(ClassRoom, id=classroom_id)
            if target_etablissement_id and classroom.etablissement_id != target_etablissement_id:
                raise PermissionDenied("Accès refusé à cette classe.")
            classrooms = classrooms.filter(id=classroom.id)

        classrooms = list(classrooms.order_by("name", "id"))
        if not classrooms:
            return Response({"detail": "Aucune classe trouvée."}, status=404)

        students = Student.objects.select_related("user", "classroom").filter(
            classroom_id__in=[item.id for item in classrooms]
        )
        if statut == "active":
            students = students.filter(is_archived=False)
        elif statut == "archived":
            students = students.filter(is_archived=True)
        students = students.order_by("user__last_name", "user__first_name", "matricule")

        # Un seul parcours, puis regroupement en memoire: interroger la base
        # par classe ferait une requete de plus a chaque classe.
        par_classe_map: dict[int, list[Student]] = {item.id: [] for item in classrooms}
        for student in students:
            par_classe_map[student.classroom_id].append(student)

        # Les classes vides restent dans le document: leur absence se lirait
        # comme un oubli, alors qu'un effectif nul est une information.
        par_classe = [(item.name, par_classe_map[item.id]) for item in classrooms]

        premier_eleve = next((eleve for _, eleves in par_classe for eleve in eleves), None)
        if premier_eleve is not None:
            school = _school_identity_for_student(premier_eleve)
            logo_path = _etablissement_logo_path(premier_eleve) or _school_logo_path()
        else:
            etablissement = (
                Etablissement.objects.filter(id=target_etablissement_id).first()
                if target_etablissement_id
                else None
            )
            school = _school_identity()
            if etablissement is not None:
                school = {**school, "name": etablissement.name}
            logo_path = _etablissement_media_field_path(etablissement, "logo") or _school_logo_path()

        pdf = _build_class_roster_pdf(
            par_classe,
            school=school,
            logo_path=logo_path,
            year_label=_active_academic_year_label(),
            with_summary=classroom_id is None,
        )

        if classroom_id is not None:
            nom_fichier = f"liste_{classrooms[0].name.replace(' ', '_')}.pdf"
        else:
            nom_fichier = "listes_toutes_classes.pdf"
        return pdf_output_response(pdf, nom_fichier)
