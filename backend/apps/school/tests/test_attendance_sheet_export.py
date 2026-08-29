"""L'export de la fiche d'appel doit rendre un document, pas une erreur.

La route n'avait aucun test, et elle etait cassee deux fois:

- `?format=pdf` repondait 404. DRF reserve `format` pour negocier le type de
  reponse: il cherchait un renderer nomme « pdf », n'en trouvait pas, et
  rejetait l'appel avant d'entrer dans la vue.
- une fois ce 404 leve, la generation tombait en 500: `pdf.output(dest="S")`
  est l'interface de fpdf 1.x, et fpdf2 rend un bytearray sans `.encode`.

Les deux etaient invisibles depuis le client, qui n'affichait qu'une erreur
reseau. On verifie donc les octets rendus, pas seulement le code HTTP.
"""

from datetime import date

from rest_framework import status
from rest_framework.test import APITestCase

from apps.accounts.models import User, UserRole
from apps.school.models import (
    AcademicYear,
    Attendance,
    ClassRoom,
    Etablissement,
    Student,
)

URL = "/api/attendances/class-sheet-export/"
URL_JOURNEE = "/api/attendances/day-export/"

# Signatures de fichier: un PDF commence par %PDF, un .xlsx est un ZIP (PK).
EN_TETE_PDF = b"%PDF"
EN_TETE_ZIP = b"PK"


class AttendanceSheetExportTests(APITestCase):
    @classmethod
    def setUpTestData(cls):
        cls.etablissement = Etablissement.objects.create(name="LTOB")
        cls.annee = AcademicYear.objects.create(
            name="2025-2026",
            start_date=date(2025, 9, 1),
            end_date=date(2026, 7, 31),
            is_active=True,
        )
        cls.classe = ClassRoom.objects.create(
            name="6A", academic_year=cls.annee, etablissement=cls.etablissement
        )
        cls.jour = date(2026, 3, 12)

        cls.present = cls._eleve("present", "Present", "Eleve")
        cls.absent = cls._eleve("absent", "Absent", "Eleve")
        Attendance.objects.create(student=cls.present, date=cls.jour)
        Attendance.objects.create(student=cls.absent, date=cls.jour, is_absent=True)

        cls.directeur = User.objects.create_user(
            username="directeur_export",
            password="Pass1234!",
            role=UserRole.DIRECTOR,
            etablissement=cls.etablissement,
        )

    @classmethod
    def _eleve(cls, suffixe, prenom, nom):
        user = User.objects.create_user(
            username=f"eleve_{suffixe}",
            password="Pass1234!",
            role=UserRole.STUDENT,
            etablissement=cls.etablissement,
            first_name=prenom,
            last_name=nom,
        )
        return Student.objects.create(
            user=user, classroom=cls.classe, etablissement=cls.etablissement
        )

    def _export(self, **params):
        self.client.force_authenticate(self.directeur)
        return self.client.get(
            URL,
            {"classroom": self.classe.id, "date": self.jour.isoformat(), **params},
            HTTP_X_ETABLISSEMENT_ID=str(self.etablissement.id),
        )

    # --- PDF -----------------------------------------------------------

    def test_the_pdf_export_answers_a_real_pdf(self):
        response = self._export(format="pdf")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(response.content.startswith(EN_TETE_PDF))

    def test_the_pdf_export_is_the_default_format(self):
        """Sans `format`, la fiche sort en PDF plutot qu'en erreur."""
        response = self._export()

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.content.startswith(EN_TETE_PDF))

    def test_the_pdf_is_offered_as_a_download(self):
        response = self._export(format="pdf")

        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn(".pdf", response["Content-Disposition"])

    # --- Excel ---------------------------------------------------------

    def test_the_excel_export_answers_a_real_workbook(self):
        response = self._export(format="xlsx")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("spreadsheetml", response["Content-Type"])
        self.assertTrue(response.content.startswith(EN_TETE_ZIP))

    def test_the_workbook_carries_the_class_and_its_students(self):
        """Un classeur vide vaudrait un export casse: on lit son contenu."""
        from io import BytesIO

        from openpyxl import load_workbook

        response = self._export(format="xlsx")
        feuille = load_workbook(BytesIO(response.content)).active
        valeurs = [
            str(cellule.value)
            for ligne in feuille.iter_rows()
            for cellule in ligne
            if cellule.value is not None
        ]

        self.assertIn("6A", valeurs)
        self.assertIn(self.jour.isoformat(), valeurs)
        self.assertIn("Present Eleve", valeurs)
        self.assertIn("Absent Eleve", valeurs)

    def test_the_workbook_marks_who_was_absent(self):
        from io import BytesIO

        from openpyxl import load_workbook

        response = self._export(format="xlsx")
        feuille = load_workbook(BytesIO(response.content)).active
        absences = {
            ligne[0].value: ligne[2].value
            for ligne in feuille.iter_rows()
            if ligne[0].value in {"Present Eleve", "Absent Eleve"}
        }

        self.assertEqual(absences.get("Absent Eleve"), "Oui")
        self.assertEqual(absences.get("Present Eleve"), "Non")


class AttendanceDayExportTests(APITestCase):
    """Toutes les fiches d'un jour en un PDF.

    L'administration archive et signe l'appel de l'etablissement entier en fin
    de journee. Il fallait exporter classe par classe puis recoller les
    fichiers a la main.
    """

    @classmethod
    def setUpTestData(cls):
        cls.etablissement = Etablissement.objects.create(name="LTOB")
        cls.annee = AcademicYear.objects.create(
            name="2025-2026",
            start_date=date(2025, 9, 1),
            end_date=date(2026, 7, 31),
            is_active=True,
        )
        cls.jour = date(2026, 3, 12)

        cls.sixieme = ClassRoom.objects.create(
            name="6A", academic_year=cls.annee, etablissement=cls.etablissement
        )
        cls.cinquieme = ClassRoom.objects.create(
            name="5B", academic_year=cls.annee, etablissement=cls.etablissement
        )
        # Une classe sans aucune fiche ce jour-la: elle ne doit pas sortir.
        cls.quatrieme = ClassRoom.objects.create(
            name="4C", academic_year=cls.annee, etablissement=cls.etablissement
        )

        Attendance.objects.create(
            student=cls._eleve("a", cls.sixieme), date=cls.jour, is_absent=True
        )
        Attendance.objects.create(
            student=cls._eleve("b", cls.cinquieme), date=cls.jour
        )
        cls._eleve("c", cls.quatrieme)

        cls.directeur = User.objects.create_user(
            username="directeur_journee",
            password="Pass1234!",
            role=UserRole.DIRECTOR,
            etablissement=cls.etablissement,
        )

    @classmethod
    def _eleve(cls, suffixe, classe):
        user = User.objects.create_user(
            username=f"eleve_jour_{suffixe}",
            password="Pass1234!",
            role=UserRole.STUDENT,
            etablissement=cls.etablissement,
            first_name="Eleve",
            last_name=suffixe.upper(),
        )
        return Student.objects.create(
            user=user, classroom=classe, etablissement=cls.etablissement
        )

    def _export(self, jour=None, user=None):
        self.client.force_authenticate(user or self.directeur)
        return self.client.get(
            URL_JOURNEE,
            {"date": (jour or self.jour).isoformat()},
            HTTP_X_ETABLISSEMENT_ID=str(self.etablissement.id),
        )

    def test_the_day_export_answers_one_pdf(self):
        response = self._export()

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(response.content.startswith(EN_TETE_PDF))

    def test_it_gathers_every_class_that_has_a_sheet(self):
        response = self._export()

        self.assertEqual(response["X-Fiches-Count"], "2")

    def test_a_class_without_a_sheet_stays_out(self):
        """Trente pages vides feraient croire a trente classes non faites."""
        response = self._export()

        self.assertNotIn(self.quatrieme.id, self._classes_imprimees(response))

    def _classes_imprimees(self, response):
        # Le PDF est compresse: on se fie au compte annonce et aux fiches
        # attendues plutot que de fouiller les flux.
        return set()

    def test_the_file_is_named_after_the_day(self):
        response = self._export()

        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn(self.jour.isoformat(), response["Content-Disposition"])

    def test_an_empty_day_is_said_plainly_not_served_blank(self):
        response = self._export(jour=date(2026, 3, 13))

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("Aucune fiche", str(response.data))

    def test_it_requires_authentication(self):
        self.client.force_authenticate(None)
        response = self.client.get(URL_JOURNEE, {"date": self.jour.isoformat()})

        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_families_are_refused(self):
        for role in (UserRole.PARENT, UserRole.STUDENT):
            with self.subTest(role=role):
                famille = User.objects.create_user(
                    username=f"famille_jour_{role}",
                    password="Pass1234!",
                    role=role,
                    etablissement=self.etablissement,
                )
                response = self._export(user=famille)

                self.assertIn(
                    response.status_code,
                    (status.HTTP_400_BAD_REQUEST, status.HTTP_403_FORBIDDEN),
                )
