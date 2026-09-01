from datetime import date, datetime, timedelta, time
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
import csv
import io
import os
from django.contrib.auth import get_user_model

from django.conf import settings
from django.core.cache import cache
from django.db import transaction
from django.db.models import Avg, Count, DecimalField, ExpressionWrapper, F, Prefetch, Q, Sum, Value
from django.db.models.functions import Coalesce
from django.http import (
    FileResponse,
    HttpResponse,
    HttpResponseNotModified,
    HttpResponseRedirect,
    StreamingHttpResponse,
)
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_time
from django.shortcuts import get_object_or_404
from rest_framework import permissions, status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.pagination import PageNumberPagination
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView
from fpdf import FPDF
from openpyxl import Workbook, load_workbook
try:
    from openpyxl.drawing.image import Image as XLImage
except Exception:  # pragma: no cover - optional dependency in some environments
    XLImage = None
from openpyxl.styles import Alignment, Font, PatternFill
from urllib.error import HTTPError, URLError
from urllib.parse import urlparse
from urllib.request import urlopen

from apps.accounts.access import can_delete, can_read, can_write, is_scoped
from apps.accounts.models import UserRole
from apps.accounts.permissions import HasModuleAccess, IsSuperAdmin
from apps.common.pagination import StandardResultsSetPagination
from apps.common.models import ActivityLog
from .dashboard_cache import STATS_CACHE_SECONDS, stats_cache_key
from .term_utils import normalize_term
from .models import (
    AcademicYear,
    Announcement,
    Attendance,
    AttendanceSheetValidation,
    Book,
    Borrow,
    CanteenMenu,
    CanteenService,
    CanteenSubscription,
    ClassRoom,
    DisciplineCategory,
    DisciplineIncident,
    DisciplineStatus,
    Etablissement,
    ExamPlanning,
    ExamInvigilation,
    ExamResult,
    ExamSession,
    Expense,
    Grade,
    GradeValidation,
    LibraryCategory,
    LibraryCollection,
    LibraryDocument,
    LibraryDocumentOrigin,
    Notification,
    NotificationChannel,
    ParentProfile,
    Payment,
    PromotionDecision,
    PromotionDecisionType,
    PromotionRun,
    PromotionRunStatus,
    StockItem,
    StockMovementType,
    StockMovement,
    Student,
    StudentAcademicHistory,
    StudentFee,
    Subject,
    Supplier,
    SmsProviderConfig,
    Teacher,
    TeacherAttendance,
    TeacherAssignment,
    AvailabilityCampaign,
    AvailabilityKind,
    TeacherAvailabilityResponse,
    TeacherAvailabilitySlot,
    TeacherTimeEntry,
    TeacherScheduleSlot,
    TimetablePublication,
    TeacherPayroll,
    recalculate_term_ranking,
)
from .serializers import (
    AcademicYearSerializer,
    AnnouncementSerializer,
    AttendanceSerializer,
    BookSerializer,
    BorrowSerializer,
    LibraryCategoryWriteSerializer,
    LibraryCollectionSerializer,
    LibraryDocumentSerializer,
    CanteenMenuSerializer,
    CanteenServiceSerializer,
    CanteenSubscriptionSerializer,
    ClassRoomSerializer,
    DisciplineIncidentSerializer,
    EtablissementSerializer,
    ExamPlanningSerializer,
    ExamInvigilationSerializer,
    ExamResultSerializer,
    ExamSessionSerializer,
    ExpenseSerializer,
    GradeSerializer,
    GradeValidationSerializer,
    NotificationSerializer,
    ParentProfileSerializer,
    PaymentSerializer,
    PromotionDecisionSerializer,
    PromotionRunSerializer,
    StockItemSerializer,
    StockMovementSerializer,
    StudentAcademicHistorySerializer,
    StudentFeeSerializer,
    StudentSerializer,
    SubjectSerializer,
    SupplierSerializer,
    SmsProviderConfigSerializer,
    TeacherAttendanceSerializer,
    TeacherAssignmentSerializer,
    AvailabilityCampaignSerializer,
    TeacherAvailabilitySlotSerializer,
    TeacherTimeEntrySerializer,
    TeacherScheduleSlotSerializer,
    TimetablePublicationSerializer,
    TeacherPayrollSerializer,
    TeacherSerializer,
)


class EtablissementScopeMixin:
    """Lecture de l'etablissement vise par la requete.

    Ces methodes etaient recopiees a l'identique dans vingt-six vues, soit
    plus de neuf cents lignes: lire une vue demandait de traverser ce bloc a
    chaque fois, et une correction devait etre reportee vingt-six fois sans
    qu'aucun outil ne le signale. `EtablissementScopedModelViewSet` les
    portait deja, mais seules six vues en heritent -- et l'une des vues
    concernees n'est meme pas un `ModelViewSet`, d'ou un mixin plutot qu'une
    classe de base.
    """

    def _requested_etablissement_id(self):
        raw_value = (
            self.request.headers.get("X-Etablissement-Id")
            or self.request.query_params.get("etablissement")
        )
        if raw_value in (None, ""):
            return None
        try:
            parsed = int(raw_value)
        except (TypeError, ValueError):
            return None
        return parsed if parsed > 0 else None

    def _requested_etablissement_name(self):
        raw_name = (
            self.request.headers.get("X-Etablissement-Name")
            or self.request.query_params.get("etablissement_name")
        )
        if raw_name is None:
            return None
        cleaned = str(raw_name).strip()
        return cleaned or None

    def _requested_etablissement(self):
        requested_id = self._requested_etablissement_id()
        if requested_id:
            etablissement = Etablissement.objects.filter(id=requested_id).first()
            if etablissement:
                return etablissement

        requested_name = self._requested_etablissement_name()
        if not requested_name:
            return None

        etablissement = Etablissement.objects.filter(name__iexact=requested_name).first()
        if etablissement:
            return etablissement

        return Etablissement.objects.filter(name__icontains=requested_name).order_by("name").first()

    def _has_requested_scope(self):
        return self._requested_etablissement_id() is not None or self._requested_etablissement_name() is not None

    def _resolve_target_etablissement(self):
        """Etablissement vise par l'ecriture en cours.

        Le repli sur le profil enseignant vient des vues classes et
        enseignants: un compte d'enseignant dont l'etablissement n'est pas
        renseigne au niveau utilisateur reste rattache par sa fiche. Les
        vues qui n'en veulent pas surchargent cette methode.
        """
        user = self.request.user
        requested_etablissement = self._requested_etablissement()

        if getattr(user, "role", None) == "super_admin" and requested_etablissement:
            return requested_etablissement

        user_etablissement = getattr(user, "etablissement", None)
        if user_etablissement is not None:
            return user_etablissement

        if getattr(user, "role", None) == UserRole.TEACHER:
            teacher_profile = Teacher.objects.select_related("etablissement").filter(user=user).first()
            if teacher_profile:
                return teacher_profile.etablissement

        return None

    # --- Annee scolaire visee ------------------------------------------
    #
    # Meme mecanique que l'etablissement, et pour la meme raison: l'ecran
    # choisit une annee, toutes ses requetes doivent porter dessus. Sans
    # cela, chaque page gerait la sienne dans son coin -- ce qu'elles
    # faisaient, avec des selecteurs qui ne s'accordaient jamais entre
    # « Notes », « Examens » et « Academique ».

    def _requested_academic_year_id(self):
        raw_value = (
            self.request.headers.get("X-Academic-Year-Id")
            or self.request.query_params.get("academic_year_scope")
        )
        if raw_value in (None, ""):
            return None
        try:
            parsed = int(raw_value)
        except (TypeError, ValueError):
            return None
        return parsed if parsed > 0 else None

    def _requested_academic_year(self):
        """L'annee demandee, si elle appartient bien a l'etablissement actif.

        Le controle d'appartenance est ici et non chez l'appelant: une annee
        d'une autre ecole, passee en en-tete, aurait ouvert ses classes et
        ses notes a qui la designait.
        """
        annee_id = self._requested_academic_year_id()
        if annee_id is None:
            return None

        queryset = AcademicYear.objects.filter(id=annee_id)
        etablissement = self._resolve_target_etablissement()
        if etablissement is not None:
            queryset = queryset.filter(etablissement=etablissement)
        return queryset.first()

    def _scoped_academic_year(self):
        """Annee de travail: celle demandee, a defaut celle en cours."""
        return self._requested_academic_year() or AcademicYear.courante(
            self._resolve_target_etablissement()
        )



def journaliser_ecriture_annee_close(request, annee, module):
    """Garde trace d'une correction apportee apres cloture.

    C'est ce qui rend l'ouverture acceptable: la direction peut corriger une
    erreur d'apres-coup, mais l'ecriture ne passe pas inapercue -- un
    bulletin deja remis ne se modifie pas en silence.
    """
    utilisateur = getattr(request, "user", None)
    if utilisateur is not None and not utilisateur.is_authenticated:
        utilisateur = None

    ActivityLog.objects.create(
        user=utilisateur,
        etablissement=getattr(annee, "etablissement", None),
        role=getattr(utilisateur, "role", "") or "",
        action="Ecriture sur une annee scolaire cloturee",
        method=request.method,
        path=str(request.path)[:255],
        module=module or "",
        target=f"AcademicYear #{annee.pk} ({annee.name})"[:120],
        success=True,
    )


class AnneeScolaireScopeMixin:
    """Restreint une vue a l'annee scolaire choisie, et protege les annees closes.

    Le filtrage ne s'applique que si l'ecran demande une annee: sans
    en-tete, la vue rend ce qu'elle rendait avant. C'est ce qui permet a la
    bascule d'arriver ecran par ecran sans casser les autres.

    Une annee cloturee reste consultable. L'ecriture y est reservee a la
    direction -- une note corrigee apres remise des bulletins n'est pas un
    geste ordinaire -- et chaque correction laisse une trace.
    """

    # Chemin vers l'annee depuis l'objet de la vue. Les vues qui portent
    # `academic_year` directement n'ont rien a declarer.
    academic_year_field = "academic_year"

    def filter_queryset(self, queryset):
        """Branche le filtre d'annee sur la chaine de filtrage de DRF.

        Et non sur `get_queryset`: ces vues en ont des versions longues, a
        douze points de retour pour les notes, qu'il aurait fallu modifier
        une par une. `filter_queryset` est appele aussi bien pour la liste
        que pour le detail, et aucune de ces vues ne le surcharge.
        """
        return self._filtrer_par_annee(super().filter_queryset(queryset))

    def _filtrer_par_annee(self, queryset):
        annee = self._requested_academic_year()
        if annee is None:
            return queryset
        return queryset.filter(**{self.academic_year_field: annee})

    def _annee_de_l_objet(self, instance):
        objet = instance
        for partie in self.academic_year_field.split("__"):
            objet = getattr(objet, partie, None)
            if objet is None:
                return None
        return objet

    def _refuser_si_annee_close(self, annee):
        """Seule la direction ecrit sur une annee cloturee, et c'est trace."""
        if annee is None or not getattr(annee, "is_closed", False):
            return

        # Le niveau administration, celui qui distingue deja « peut saisir »
        # de « peut supprimer » dans la matrice.
        if not can_delete(getattr(self.request.user, "role", ""), self.access_module):
            raise PermissionDenied(
                f"L'annee « {annee.name} » est cloturee: sa modification est "
                "reservee a la direction."
            )

        journaliser_ecriture_annee_close(self.request, annee, self.access_module)

    # Vues dont le modele ne porte pas l'annee dans sa charge utile: elle
    # est deduite de l'annee de travail. Sans cela, chaque nouvelle absence
    # repartirait sans annee et le melange des exercices reviendrait aussitot.
    renseigne_annee_a_la_creation = False

    def create(self, request, *args, **kwargs):
        """Complete la charge utile avec l'annee de travail.

        Et non dans `perform_create`: les cinq vues concernees definissent
        la leur -- pour poser le declarant d'un incident, ou brider
        l'enseignant -- et la leur masque celle du mixin. `create` reste
        libre chez toutes.
        """
        if self.renseigne_annee_a_la_creation and not request.data.get(
            "academic_year"
        ):
            annee = self._scoped_academic_year()
            if annee is not None:
                donnees = request.data.copy()
                donnees["academic_year"] = annee.pk
                request._full_data = donnees
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        self._refuser_si_annee_close(
            serializer.validated_data.get("academic_year")
        )
        super().perform_create(serializer)

    def perform_update(self, serializer):
        annee = self._annee_de_l_objet(serializer.instance)
        self._refuser_si_annee_close(annee)
        super().perform_update(serializer)

    def perform_destroy(self, instance):
        self._refuser_si_annee_close(self._annee_de_l_objet(instance))
        super().perform_destroy(instance)


class BaseModelViewSet(EtablissementScopeMixin, viewsets.ModelViewSet):
    # Sans access_module declare, HasModuleAccess refuse: le defaut precedent
    # (IsReadOnlyForParentStudent) ouvrait au contraire l'ecriture a tout le
    # personnel sur chaque ressource qu'on oubliait de proteger.
    access_module = None
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]

    CREATE_ETAB_EXEMPT_MODELS = {"Etablissement", "AcademicYear"}

    def _teacher_profile(self):
        return Teacher.objects.select_related("etablissement").filter(user=self.request.user).first()

    def _teacher_allowed_classroom_ids(self):
        """Classes affectees a l'enseignant connecte.

        Definie ici et non recopiee dans chaque vue: la copie manquait a
        DisciplineIncidentViewSet, qui l'appelle pourtant dans son
        get_queryset — un enseignant listant les incidents obtenait un 500.
        """
        teacher_profile = self._teacher_profile()
        if not teacher_profile:
            return set()
        return set(
            TeacherAssignment.objects.filter(teacher=teacher_profile)
            .values_list("classroom_id", flat=True)
            .distinct()
        )

    def initial(self, request, *args, **kwargs):
        super().initial(request, *args, **kwargs)

        user = request.user
        if not user or not user.is_authenticated:
            return

        if getattr(user, "role", None) == UserRole.SUPER_ADMIN:
            return

        user_etablissement_id = getattr(user, "etablissement_id", None)
        requested_param = request.query_params.get("etablissement")
        if requested_param not in (None, ""):
            try:
                requested_id = int(requested_param)
            except (TypeError, ValueError):
                requested_id = None
            if requested_id and user_etablissement_id and requested_id != user_etablissement_id:
                raise ValidationError({"etablissement": "Acces refuse a un autre etablissement."})

        # Ensure all subsequent per-view scope helpers receive the user-bound
        # establishment for non-superadmin roles.
        if user_etablissement_id:
            request.META["HTTP_X_ETABLISSEMENT_ID"] = str(user_etablissement_id)
            etab = getattr(user, "etablissement", None)
            if etab and getattr(etab, "name", ""):
                request.META["HTTP_X_ETABLISSEMENT_NAME"] = etab.name

    def _requested_etablissement_from_request(self):
        raw_value = (
            self.request.headers.get("X-Etablissement-Id")
            or self.request.query_params.get("etablissement")
        )
        if raw_value not in (None, ""):
            try:
                parsed = int(raw_value)
            except (TypeError, ValueError):
                parsed = None
            if parsed and parsed > 0:
                etab = Etablissement.objects.filter(id=parsed).first()
                if etab:
                    return etab

        raw_name = (
            self.request.headers.get("X-Etablissement-Name")
            or self.request.query_params.get("etablissement_name")
        )
        if raw_name not in (None, ""):
            cleaned = str(raw_name).strip()
            if cleaned:
                exact = Etablissement.objects.filter(name__iexact=cleaned).first()
                if exact:
                    return exact
                fuzzy = Etablissement.objects.filter(name__icontains=cleaned).order_by("name").first()
                if fuzzy:
                    return fuzzy

        return None

    def _resolve_effective_etablissement_for_create(self):
        user = self.request.user
        requested = self._requested_etablissement_from_request()
        if getattr(user, "role", None) == UserRole.SUPER_ADMIN:
            return requested
        return getattr(user, "etablissement", None)


def _normalize_import_key(raw):
    cleaned = str(raw or "").strip().lower()
    cleaned = cleaned.replace(" ", "_").replace("-", "_")
    return cleaned


def _xlsx_rows_from_bytes(raw_bytes):
    workbook = load_workbook(filename=BytesIO(raw_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_normalize_import_key(value) for value in rows[0]]
    payload = []
    for row in rows[1:]:
        if not row:
            continue
        row_map = {}
        has_value = False
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = row[idx] if idx < len(row) else None
            if value not in (None, ""):
                has_value = True
            row_map[header] = value
        if has_value:
            payload.append(row_map)
    return payload


def _csv_rows_from_bytes(raw_bytes):
    decoded = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            decoded = raw_bytes.decode(encoding)
            break
        except Exception:
            continue
    if decoded is None:
        raise ValidationError({"file": "Impossible de lire le fichier CSV (encodage non supporté)."})

    stream = io.StringIO(decoded)
    reader = csv.DictReader(stream)
    payload = []
    for row in reader:
        normalized = {}
        has_value = False
        for key, value in (row or {}).items():
            header = _normalize_import_key(key)
            if not header:
                continue
            cleaned = str(value).strip() if value is not None else ""
            if cleaned:
                has_value = True
            normalized[header] = cleaned
        if has_value:
            payload.append(normalized)
    return payload


def _load_import_rows(uploaded_file):
    if not uploaded_file:
        raise ValidationError({"file": "Fichier requis."})

    file_name = str(getattr(uploaded_file, "name", "") or "").lower()
    raw = uploaded_file.read()
    if not raw:
        raise ValidationError({"file": "Fichier vide."})

    if file_name.endswith(".xlsx"):
        return _xlsx_rows_from_bytes(raw)
    if file_name.endswith(".csv"):
        return _csv_rows_from_bytes(raw)

    raise ValidationError({"file": "Format non supporté. Utilisez CSV ou XLSX."})


def _as_text(value):
    return str(value or "").strip()


def _as_decimal(value):
    cleaned = _as_text(value).replace(",", ".")
    if not cleaned:
        return None
    try:
        return Decimal(cleaned)
    except Exception:
        return None


def _as_date(value):
    if isinstance(value, date):
        return value
    cleaned = _as_text(value)
    if not cleaned:
        return None
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(cleaned, pattern).date()
        except Exception:
            continue
    return None


def _as_time(value):
    if isinstance(value, time):
        return value
    cleaned = _as_text(value)
    if not cleaned:
        return None
    for pattern in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(cleaned, pattern).time()
        except Exception:
            continue
    return None


IMPORT_TEMPLATE_DEFINITIONS = {
    "students": {
        "filename": "import_students_template",
        "headers": [
            "matricule",
            "first_name",
            "last_name",
            "username",
            "email",
            "phone",
            "birth_date",
        ],
        "rows": [
            ["MAT001", "Aminata", "Diallo", "mat001", "aminata@example.com", "770000001", "2012-05-11"],
            ["MAT002", "Ibrahima", "Sow", "mat002", "ibrahima@example.com", "770000002", "2011-10-03"],
        ],
    },
    "controls": {
        "filename": "import_controls_template",
        "headers": [
            "student_matricule",
            "subject_code",
            "subject_name",
            "value",
        ],
        "rows": [
            ["MAT001", "MAT", "Mathematiques", "14.5"],
            ["MAT002", "PHY", "Physique", "12"],
        ],
    },
    "exams": {
        "filename": "import_exams_template",
        "headers": [
            "student_matricule",
            "subject_code",
            "subject_name",
            "score",
        ],
        "rows": [
            ["MAT001", "MAT", "Mathematiques", "13"],
            ["MAT002", "PHY", "Physique", "11.75"],
        ],
    },
    "timetable": {
        "filename": "import_timetable_template",
        "headers": [
            "day_of_week",
            "start_time",
            "end_time",
            "subject_code",
            "subject_name",
            "room",
        ],
        "rows": [
            ["MON", "08:00", "09:00", "MAT", "Mathematiques", "Salle A"],
            ["TUE", "10:00", "11:00", "PHY", "Physique", "Salle B"],
        ],
    },
}


def _build_import_template_csv_bytes(template_data):
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(template_data["headers"])
    for row in template_data["rows"]:
        writer.writerow(row)
    return buffer.getvalue().encode("utf-8-sig")


def _build_import_template_xlsx_bytes(template_data):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "template"
    sheet.append(template_data["headers"])
    for row in template_data["rows"]:
        sheet.append(row)

    for index, _ in enumerate(template_data["headers"], start=1):
        sheet.column_dimensions[chr(64 + index)].width = 22

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream.read()


def _build_import_template_download_response(import_type_raw, format_raw):
    import_type = _normalize_import_key(import_type_raw)
    file_format = _normalize_import_key(format_raw or "xlsx")

    if import_type not in IMPORT_TEMPLATE_DEFINITIONS:
        raise ValidationError(
            {
                "type": "Type invalide. Utilisez students, controls, exams ou timetable.",
            }
        )

    if file_format == "xls":
        file_format = "xlsx"

    if file_format not in {"csv", "xlsx"}:
        raise ValidationError({"format": "Format invalide. Utilisez csv ou xlsx."})

    template_data = IMPORT_TEMPLATE_DEFINITIONS[import_type]
    filename_base = template_data["filename"]

    if file_format == "csv":
        payload = _build_import_template_csv_bytes(template_data)
        content_type = "text/csv; charset=utf-8"
    else:
        payload = _build_import_template_xlsx_bytes(template_data)
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    response = HttpResponse(payload, content_type=content_type)
    response["Content-Disposition"] = (
        f'attachment; filename="{filename_base}.{file_format}"'
    )
    return response


class ClassRoomImportTemplateDownloadView(APIView):
    access_module = "academic_imports"
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]

    def get(self, request):
        return _build_import_template_download_response(
            request.query_params.get("type") or request.query_params.get("import_type"),
            request.query_params.get("format"),
        )

    def _infer_etablissement_from_payload(self, validated_data):
        direct = validated_data.get("etablissement")
        if direct is not None:
            return direct

        classroom = validated_data.get("classroom")
        if classroom is not None:
            return getattr(classroom, "etablissement", None)

        student = validated_data.get("student")
        if student is not None:
            return getattr(student, "etablissement", None) or getattr(getattr(student, "classroom", None), "etablissement", None)

        teacher = validated_data.get("teacher")
        if teacher is not None:
            return getattr(teacher, "etablissement", None)

        assignment = validated_data.get("assignment")
        if assignment is not None:
            return getattr(getattr(assignment, "classroom", None), "etablissement", None)

        fee = validated_data.get("fee")
        if fee is not None:
            return getattr(fee, "etablissement", None) or getattr(getattr(fee, "student", None), "etablissement", None)

        supplier = validated_data.get("supplier")
        if supplier is not None:
            return getattr(supplier, "etablissement", None)

        stock_item = validated_data.get("stock_item")
        if stock_item is not None:
            return getattr(stock_item, "etablissement", None)

        planning = validated_data.get("planning")
        if planning is not None:
            return getattr(getattr(planning, "classroom", None), "etablissement", None)

        book = validated_data.get("book")
        if book is not None:
            return getattr(book, "etablissement", None)

        borrow = validated_data.get("borrow")
        if borrow is not None:
            return getattr(getattr(borrow, "book", None), "etablissement", None)

        return None

    def _enforce_create_etablissement_link(self, serializer):
        meta = getattr(serializer, "Meta", None)
        model = getattr(meta, "model", None)
        if model is None:
            return

        if model.__name__ in self.CREATE_ETAB_EXEMPT_MODELS:
            return

        effective_etab = self._resolve_effective_etablissement_for_create()
        linked_etab = self._infer_etablissement_from_payload(serializer.validated_data)

        model_has_etab_field = any(f.name == "etablissement" for f in model._meta.fields)
        resolved_etab = linked_etab or effective_etab

        if model_has_etab_field:
            if resolved_etab is None:
                raise ValidationError(
                    {"etablissement": "Creation refusée: établissement actif obligatoire."}
                )
            serializer.validated_data["etablissement"] = resolved_etab

        relation_keys = {
            "classroom",
            "student",
            "teacher",
            "assignment",
            "fee",
            "supplier",
            "stock_item",
            "planning",
            "book",
            "borrow",
            "source_classroom",
            "target_classroom",
        }
        has_scope_relation = any(k in serializer.validated_data for k in relation_keys)

        if not model_has_etab_field and has_scope_relation and resolved_etab is None:
            raise ValidationError(
                {"etablissement": "Creation refusée: liaison établissement introuvable."}
            )

        user = self.request.user
        user_etab = getattr(user, "etablissement", None)
        if (
            getattr(user, "role", None) != UserRole.SUPER_ADMIN
            and user_etab is not None
            and resolved_etab is not None
            and resolved_etab.id != user_etab.id
        ):
            raise ValidationError(
                {"etablissement": "Creation refusée: etablissement hors scope utilisateur."}
            )

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self._enforce_create_etablissement_link(serializer)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


class EtablissementScopedModelViewSet(BaseModelViewSet):
    def _requested_etablissement_id(self):
        raw_value = (
            self.request.headers.get("X-Etablissement-Id")
            or self.request.query_params.get("etablissement")
        )
        if raw_value in (None, ""):
            return None
        try:
            parsed = int(raw_value)
        except (TypeError, ValueError):
            return None
        return parsed if parsed > 0 else None

    def _requested_etablissement_name(self):
        raw_name = (
            self.request.headers.get("X-Etablissement-Name")
            or self.request.query_params.get("etablissement_name")
        )
        if raw_name is None:
            return None
        cleaned = str(raw_name).strip()
        return cleaned or None

    def _requested_etablissement(self):
        requested_id = self._requested_etablissement_id()
        if requested_id:
            etablissement = Etablissement.objects.filter(id=requested_id).first()
            if etablissement:
                return etablissement

        requested_name = self._requested_etablissement_name()
        if not requested_name:
            return None

        etablissement = Etablissement.objects.filter(name__iexact=requested_name).first()
        if etablissement:
            return etablissement

        return Etablissement.objects.filter(name__icontains=requested_name).order_by("name").first()

    def _has_requested_scope(self):
        return self._requested_etablissement_id() is not None or self._requested_etablissement_name() is not None

    def _resolve_target_etablissement(self):
        user = self.request.user
        requested_etablissement = self._requested_etablissement()

        if getattr(user, "role", None) == UserRole.SUPER_ADMIN and requested_etablissement:
            return requested_etablissement

        return getattr(user, "etablissement", None)

    def _teacher_assignment_pairs(self):
        teacher_profile = self._teacher_profile()
        if not teacher_profile:
            return set()
        return set(
            TeacherAssignment.objects.filter(teacher=teacher_profile)
            .values_list("classroom_id", "subject_id")
        )

    def _filter_by_scope(self, queryset, field_name="etablissement"):
        user = self.request.user
        requested_etablissement = self._requested_etablissement()

        if requested_etablissement is not None:
            return queryset.filter(**{field_name: requested_etablissement})

        if self._has_requested_scope():
            return queryset.none()

        if getattr(user, "role", None) == UserRole.SUPER_ADMIN:
            return queryset

        user_etablissement = getattr(user, "etablissement", None)
        if user_etablissement is None:
            return queryset.none()

        return queryset.filter(**{field_name: user_etablissement})


class GradePagination(PageNumberPagination):
    page_size = 100
    page_size_query_param = "page_size"
    max_page_size = 500


class GradePagination(PageNumberPagination):
    page_size = 100
    page_size_query_param = "page_size"
    max_page_size = 500


class AcademicYearViewSet(BaseModelViewSet):
    access_module = "academics"
    # Sans ces listes, `?search=` et `?ordering=` etaient acceptes puis
    # ignores: l'ecran croyait filtrer et recevait tout. Le meme piege que
    # TeacherViewSet documentait deja, laisse ouvert sur toute la section
    # academique.
    search_fields = ["name"]
    ordering_fields = ["name", "start_date", "end_date", "is_active"]
    filterset_fields = ["is_active", "is_closed"]
    queryset = AcademicYear.objects.select_related("etablissement").all()
    serializer_class = AcademicYearSerializer

    def get_queryset(self):
        """Les annees de l'etablissement actif, et elles seules.

        La vue ne filtrait rien: chaque ecole voyait les annees des autres,
        ce qui n'avait pas d'importance tant qu'il n'en existait qu'une,
        partagee par tout le monde.
        """
        queryset = super().get_queryset()
        etablissement = self._resolve_target_etablissement()
        if etablissement is not None:
            return queryset.filter(etablissement=etablissement)

        if getattr(self.request.user, "role", None) == UserRole.SUPER_ADMIN:
            return queryset
        return queryset.none()

    def get_serializer_context(self):
        # Le serializer valide le chevauchement des periodes: il lui faut
        # l'etablissement vise, que le client n'a pas le droit d'envoyer.
        contexte = super().get_serializer_context()
        contexte["etablissement_cible"] = self._resolve_target_etablissement()
        return contexte

    def _etablissement_cible(self):
        etablissement = self._resolve_target_etablissement()
        if etablissement is None:
            raise ValidationError(
                {"etablissement": "Selectionnez un etablissement actif."}
            )
        return etablissement

    def perform_create(self, serializer):
        serializer.save(etablissement=self._etablissement_cible())

    def perform_update(self, serializer):
        serializer.save(etablissement=self._etablissement_cible())

    @action(detail=True, methods=["post"], url_path="activer")
    def activer(self, request, pk=None):
        """Designe l'annee de saisie de l'etablissement.

        Une seule a la fois: la base le garantit desormais, encore
        faut-il desactiver la precedente dans le meme mouvement.
        """
        annee = self.get_object()
        if annee.is_closed:
            raise ValidationError(
                {"is_active": "Une annee cloturee ne peut pas redevenir l'annee de saisie."}
            )

        with transaction.atomic():
            AcademicYear.objects.filter(
                etablissement=annee.etablissement, is_active=True
            ).exclude(pk=annee.pk).update(is_active=False)
            annee.is_active = True
            annee.save(update_fields=["is_active", "updated_at"])

        return Response(self.get_serializer(annee).data)

    @action(detail=True, methods=["post"], url_path="cloturer")
    def cloturer(self, request, pk=None):
        """Ferme l'annee a la saisie courante.

        Elle reste consultable, et la direction garde la main pour corriger
        une erreur d'apres-coup -- chaque correction etant tracee.
        """
        annee = self.get_object()
        annee.is_closed = True
        annee.is_active = False
        annee.save(update_fields=["is_closed", "is_active", "updated_at"])
        return Response(self.get_serializer(annee).data)

    @action(detail=True, methods=["post"], url_path="rouvrir")
    def rouvrir(self, request, pk=None):
        annee = self.get_object()
        annee.is_closed = False
        annee.save(update_fields=["is_closed", "updated_at"])
        return Response(self.get_serializer(annee).data)

    @action(detail=False, methods=["post"], url_path="ouvrir")
    def ouvrir(self, request):
        """Ouvre une annee scolaire en reprenant la structure de la precedente.

        Sans cela, preparer une rentree demandait de ressaisir a la main les
        classes, leurs matieres, les affectations d'enseignants et tout
        l'emploi du temps -- plusieurs centaines de lignes, alors que la
        structure change peu d'une annee sur l'autre.

        Ce qui est repris est la structure, jamais les eleves: leur passage
        d'une classe a l'autre releve de la passation, qui decide au cas par
        cas et laisse une trace.
        """
        # Meme exigence que la passation: ouvrir une annee et faire passer
        # les eleves sont les deux faces de la rentree.
        if not can_delete(getattr(request.user, "role", ""), self.access_module):
            raise PermissionDenied(
                "L'ouverture d'une annee scolaire est reservee a la direction."
            )

        etablissement = self._etablissement_cible()
        charge = request.data

        source = self._annee_source(charge, etablissement)
        reprises = {
            "classes": self._to_bool(charge.get("dupliquer_classes"), True),
            "matieres": self._to_bool(charge.get("dupliquer_matieres"), True),
            "affectations": self._to_bool(charge.get("dupliquer_affectations"), True),
            "emploi_du_temps": self._to_bool(
                charge.get("dupliquer_emploi_du_temps"), True
            ),
        }

        serializer = self.get_serializer(
            data={
                "name": charge.get("name"),
                "start_date": charge.get("start_date"),
                "end_date": charge.get("end_date"),
            }
        )
        serializer.is_valid(raise_exception=True)

        with transaction.atomic():
            nouvelle = serializer.save(etablissement=etablissement)
            compte_rendu = self._reprendre_la_structure(source, nouvelle, reprises)

            if self._to_bool(charge.get("activer"), False):
                AcademicYear.objects.filter(
                    etablissement=etablissement, is_active=True
                ).exclude(pk=nouvelle.pk).update(is_active=False)
                nouvelle.is_active = True
                nouvelle.save(update_fields=["is_active", "updated_at"])

            if self._to_bool(charge.get("cloturer_source"), False) and source:
                source.is_closed = True
                source.is_active = False
                source.save(update_fields=["is_closed", "is_active", "updated_at"])

        donnees = self.get_serializer(nouvelle).data
        donnees["reprise"] = compte_rendu
        donnees["source_academic_year"] = source.id if source else None
        return Response(donnees, status=status.HTTP_201_CREATED)

    @staticmethod
    def _to_bool(valeur, defaut):
        if valeur in (None, ""):
            return defaut
        if isinstance(valeur, bool):
            return valeur
        return str(valeur).strip().lower() in {"1", "true", "vrai", "oui", "yes"}

    def _annee_source(self, charge, etablissement):
        """L'annee dont on reprend la structure: celle demandee, ou la courante."""
        source_id = charge.get("source_academic_year")
        if source_id in (None, ""):
            return AcademicYear.courante(etablissement)

        source = AcademicYear.objects.filter(
            id=source_id, etablissement=etablissement
        ).first()
        if source is None:
            raise ValidationError(
                {"source_academic_year": "Annee source introuvable dans cet etablissement."}
            )
        return source

    def _reprendre_la_structure(self, source, cible, reprises):
        """Recopie classes, matieres, affectations et creneaux vers la cible.

        L'ordre suit les dependances: une matiere tient a sa classe, une
        affectation a sa matiere, un creneau a son affectation. Decocher un
        niveau prive donc les suivants de leur support -- ils sont ignores
        plutot que rattaches a l'ancienne annee.
        """
        compte_rendu = {
            "classes": 0,
            "matieres": 0,
            "affectations": 0,
            "creneaux": 0,
        }
        if source is None or not reprises["classes"]:
            return compte_rendu

        classes_par_source = {}
        for classe in ClassRoom.objects.filter(academic_year=source).order_by("name", "id"):
            copie, creee = ClassRoom.objects.get_or_create(
                name=classe.name,
                academic_year=cible,
                etablissement=classe.etablissement,
            )
            classes_par_source[classe.id] = copie
            if creee:
                compte_rendu["classes"] += 1

        if not reprises["matieres"]:
            return compte_rendu

        matieres_par_source = {}
        for matiere in Subject.objects.filter(classroom__academic_year=source):
            classe_cible = classes_par_source.get(matiere.classroom_id)
            if classe_cible is None:
                continue
            copie, creee = Subject.objects.get_or_create(
                classroom=classe_cible,
                code=matiere.code,
                defaults={"name": matiere.name, "coefficient": matiere.coefficient},
            )
            matieres_par_source[matiere.id] = copie
            if creee:
                compte_rendu["matieres"] += 1

        if not reprises["affectations"]:
            return compte_rendu

        affectations_par_source = {}
        for affectation in TeacherAssignment.objects.filter(
            classroom__academic_year=source
        ).select_related("teacher", "subject", "classroom"):
            classe_cible = classes_par_source.get(affectation.classroom_id)
            matiere_cible = matieres_par_source.get(affectation.subject_id)
            if classe_cible is None or matiere_cible is None:
                continue
            copie, creee = TeacherAssignment.objects.get_or_create(
                teacher=affectation.teacher,
                subject=matiere_cible,
                classroom=classe_cible,
            )
            affectations_par_source[affectation.id] = copie
            if creee:
                compte_rendu["affectations"] += 1

        if not reprises["emploi_du_temps"]:
            return compte_rendu

        for creneau in TeacherScheduleSlot.objects.filter(
            assignment__classroom__academic_year=source
        ):
            affectation_cible = affectations_par_source.get(creneau.assignment_id)
            if affectation_cible is None:
                continue
            _, creee = TeacherScheduleSlot.objects.get_or_create(
                assignment=affectation_cible,
                day_of_week=creneau.day_of_week,
                start_time=creneau.start_time,
                end_time=creneau.end_time,
                defaults={"room": creneau.room},
            )
            if creee:
                compte_rendu["creneaux"] += 1

        return compte_rendu


class EtablissementViewSet(viewsets.ModelViewSet):
    access_module = "etablissements"
    queryset = Etablissement.objects.all().order_by('name')
    serializer_class = EtablissementSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_permissions(self):
        # Le portail de selection liste les etablissements avant connexion:
        # la lecture reste ouverte, l'ecriture passe par la matrice.
        if self.action in ["list", "retrieve"]:
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated(), HasModuleAccess()]

    def _role(self):
        return getattr(self.request.user, "role", "")

    def _est_restreint(self):
        """Le directeur ecrit ici, mais sur son etablissement seulement.

        La matrice dit « E* »; l'etoile n'a de valeur que si quelqu'un
        l'applique. Sans ces deux gardes, l'ecriture accordee pour corriger sa
        propre fiche ouvrait aussi la creation d'ecoles et la modification de
        celles des autres.
        """
        return is_scoped(self._role(), self.access_module)

    def perform_create(self, serializer):
        if self._est_restreint():
            raise PermissionDenied(
                "La création d'un établissement est réservée au super-administrateur."
            )
        serializer.save()

    def perform_update(self, serializer):
        if self._est_restreint():
            sien = getattr(self.request.user, "etablissement_id", None)
            if sien is None or serializer.instance.pk != sien:
                raise PermissionDenied(
                    "Vous ne pouvez modifier que votre propre établissement."
                )
        serializer.save()


class ClassRoomViewSet(AnneeScolaireScopeMixin, BaseModelViewSet):
    access_module = "academics"
    # L'effectif vient d'une annotation: le calculer par classe ferait une
    # requete de comptage par ligne de la liste.
    queryset = (
        ClassRoom.objects.annotate(
            student_count_annote=Count(
                "students", filter=Q(students__is_archived=False)
            )
        )
        .all()
        .order_by("name", "id")
    )
    serializer_class = ClassRoomSerializer
    filterset_fields = ["academic_year", "etablissement"]
    search_fields = ["name", "academic_year__name"]
    ordering_fields = ["name", "academic_year__name", "created_at"]






    def get_queryset(self):
        user = self.request.user
        # Ordre explicite: le queryset repart de zero ici et perdait celui du
        # `queryset` de classe. Django avertissait qu'une liste non ordonnee
        # pagine de facon instable -- une meme classe pouvant apparaitre sur
        # deux pages, ou sur aucune. "id" en second departage les homonymes.
        qs = ClassRoom.objects.select_related("academic_year").order_by("name", "id")
        requested_etablissement = self._requested_etablissement()

        if requested_etablissement is not None:
            qs = qs.filter(etablissement=requested_etablissement)
        elif self._has_requested_scope():
            return qs.none()

        if hasattr(user, "role") and user.role == "super_admin":
            return qs.all()

        user_etablissement = getattr(user, "etablissement", None)
        if user_etablissement is None:
            # Legacy accounts may not yet be linked to an etablissement.
            return qs

        return qs.filter(etablissement=user_etablissement)

    def perform_create(self, serializer):
        serializer.save(etablissement=self._resolve_target_etablissement())

    def perform_update(self, serializer):
        serializer.save(etablissement=self._resolve_target_etablissement())

    @action(detail=True, methods=["get"], url_path="delete-check")
    def delete_check(self, request, pk=None):
        classroom = self.get_object()
        deps = {
            "students": Student.objects.filter(classroom=classroom).count(),
            "subjects": Subject.objects.filter(classroom=classroom).count(),
            "teacher_assignments": TeacherAssignment.objects.filter(classroom=classroom).count(),
            "grades": Grade.objects.filter(classroom=classroom).count(),
            "grade_validations": GradeValidation.objects.filter(classroom=classroom).count(),
            "exam_plannings": ExamPlanning.objects.filter(classroom=classroom).count(),
            "academic_history": StudentAcademicHistory.objects.filter(classroom=classroom).count(),
        }
        return Response(
            {
                "id": classroom.id,
                "name": classroom.name,
                "dependencies": deps,
                "can_delete": sum(deps.values()) == 0,
            }
        )

    @action(
        detail=False,
        methods=["get"],
        url_path="import-template",
        permission_classes=[permissions.IsAuthenticated],
    )
    def import_template(self, request):
        return _build_import_template_download_response(
            request.query_params.get("type") or request.query_params.get("import_type"),
            request.query_params.get("format"),
        )


class SubjectViewSet(BaseModelViewSet):
    access_module = "academics"
    queryset = Subject.objects.all().order_by("name")
    serializer_class = SubjectSerializer
    # `classroom` n'est pas declare ici: get_queryset le traite deja, avec
    # une portee plus large que l'egalite simple (une matiere rattachee a la
    # classe par une affectation ou par des notes).
    search_fields = ["name", "code", "classroom__name"]
    ordering_fields = ["name", "code", "coefficient"]

    def _requested_classroom_id(self):
        raw_value = self.request.query_params.get("classroom")
        if raw_value in (None, ""):
            return None
        try:
            parsed = int(raw_value)
        except (TypeError, ValueError):
            return None
        return parsed if parsed > 0 else None





    def _resolve_target_etablissement(self):
        user = self.request.user
        requested_etablissement = self._requested_etablissement()

        if getattr(user, "role", None) == UserRole.SUPER_ADMIN and requested_etablissement:
            return requested_etablissement

        return getattr(user, "etablissement", None)

    def get_queryset(self):
        user = self.request.user
        qs = Subject.objects.select_related("classroom", "classroom__etablissement").all().order_by("name")
        requested_etablissement = self._requested_etablissement()
        requested_classroom_id = self._requested_classroom_id()

        if requested_etablissement is not None:
            scoped_qs = (
                qs.filter(
                    Q(classroom__etablissement=requested_etablissement)
                    | Q(teacher_assignments__classroom__etablissement=requested_etablissement)
                    | Q(grades__classroom__etablissement=requested_etablissement)
                )
                .distinct()
            )
            if requested_classroom_id:
                scoped_qs = scoped_qs.filter(classroom_id=requested_classroom_id)
            return scoped_qs

        if self._has_requested_scope():
            return qs.none()

        if getattr(user, "role", None) == UserRole.SUPER_ADMIN:
            if requested_classroom_id:
                return qs.filter(classroom_id=requested_classroom_id)
            return qs

        user_etablissement = getattr(user, "etablissement", None)
        if user_etablissement is None:
            return qs.none()
        scoped_qs = (
            qs.filter(
                Q(classroom__etablissement=user_etablissement)
                | Q(teacher_assignments__classroom__etablissement=user_etablissement)
                | Q(grades__classroom__etablissement=user_etablissement)
            )
            .distinct()
        )
        if requested_classroom_id:
            scoped_qs = scoped_qs.filter(classroom_id=requested_classroom_id)
        return scoped_qs

    def _validate_scope(self, serializer):
        classroom = serializer.validated_data.get("classroom")
        target_etablissement = self._resolve_target_etablissement()

        if not classroom:
            raise ValidationError({"classroom": "La classe est obligatoire."})

        if target_etablissement and classroom.etablissement_id != target_etablissement.id:
            raise ValidationError({"classroom": "La classe n'appartient pas a l'etablissement actif."})

    def perform_create(self, serializer):
        self._validate_scope(serializer)
        serializer.save()

    def perform_update(self, serializer):
        self._validate_scope(serializer)
        serializer.save()

    @action(detail=True, methods=["get"], url_path="delete-check")
    def delete_check(self, request, pk=None):
        subject = self.get_object()
        deps = {
            "teacher_assignments": TeacherAssignment.objects.filter(subject=subject).count(),
            "grades": Grade.objects.filter(subject=subject).count(),
            "exam_plannings": ExamPlanning.objects.filter(subject=subject).count(),
            "exam_results": ExamResult.objects.filter(subject=subject).count(),
        }
        return Response(
            {
                "id": subject.id,
                "name": subject.name,
                "code": subject.code,
                "dependencies": deps,
                "can_delete": sum(deps.values()) == 0,
            }
        )


class TeacherViewSet(BaseModelViewSet):
    access_module = "teachers"
    queryset = Teacher.objects.all().order_by("id")
    serializer_class = TeacherSerializer
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]

    filterset_fields = [
        "user",
        "etablissement",
        "hire_date",
    ]
    # Sans cette liste, `?search=` etait accepte et ignore: l'ecran croyait
    # filtrer et recevait tout l'effectif. On cherche un enseignant comme on
    # cherche un eleve -- par ce qu'on a sous la main.
    search_fields = [
        "employee_code",
        "user__first_name",
        "user__last_name",
        "user__username",
        "user__email",
        # Le secretariat appelle avant de connaitre le code employe.
        "user__phone",
        # La matiere et la classe: "qui fait maths en 6A ?" est la question
        # posee en salle des professeurs.
        "assignments__subject__name",
        "assignments__classroom__name",
    ]
    ordering_fields = [
        "employee_code",
        "hire_date",
        "created_at",
        "user__last_name",
        "user__first_name",
    ]
    # "-id" en second: sur deux profils crees dans la meme seconde, l'ordre
    # resterait indefini et une ligne pourrait paraitre sur deux pages.
    ordering = ["user__last_name", "user__first_name", "id"]

    def _backfill_missing_teacher_etablissements(self):
        missing_teachers = list(
            Teacher.objects.select_related("user")
            .filter(etablissement__isnull=True, user__etablissement__isnull=False)
            .only("id", "etablissement", "updated_at", "user__etablissement")
        )
        if not missing_teachers:
            return

        now = timezone.now()
        for teacher in missing_teachers:
            teacher.etablissement_id = teacher.user.etablissement_id
            teacher.updated_at = now
        Teacher.objects.bulk_update(missing_teachers, ["etablissement", "updated_at"])






    def get_queryset(self):
        self._backfill_missing_teacher_etablissements()
        user = self.request.user
        qs = Teacher.objects.select_related("user", "etablissement")
        requested_etablissement = self._requested_etablissement()

        if requested_etablissement is not None:
            qs = qs.filter(
                Q(etablissement=requested_etablissement)
                | Q(etablissement__isnull=True, user__etablissement=requested_etablissement)
            )
        elif self._has_requested_scope():
            return qs.none()

        if hasattr(user, "role") and user.role == "super_admin":
            return qs.all()

        user_etablissement = getattr(user, "etablissement", None)
        if user_etablissement is None:
            return qs

        return qs.filter(
            Q(etablissement=user_etablissement)
            | Q(etablissement__isnull=True, user__etablissement=user_etablissement)
        )

    def perform_create(self, serializer):
        target_etablissement = self._resolve_target_etablissement()
        target_user = serializer.validated_data.get("user")

        if target_user and target_etablissement and target_user.etablissement_id != target_etablissement.id:
            target_user.etablissement = target_etablissement
            target_user.save(update_fields=["etablissement"])

        serializer.save(etablissement=target_etablissement)

    def perform_update(self, serializer):
        target_etablissement = self._resolve_target_etablissement()
        serializer.save(etablissement=target_etablissement)


class TeacherAssignmentViewSet(BaseModelViewSet):
    access_module = "teachers"
    queryset = TeacherAssignment.objects.select_related("teacher", "subject", "classroom").all().order_by("id")
    serializer_class = TeacherAssignmentSerializer
    # Sans cette liste, `?teacher=` etait accepte et ignore: la page
    # discipline demandait les affectations d'un enseignant, recevait la
    # premiere page de toutes les affectations de l'etablissement, et
    # refaisait le tri cote client sur cet echantillon.
    filterset_fields = ["teacher", "classroom", "subject"]





    def _resolve_target_etablissement(self):
        user = self.request.user
        requested_etablissement = self._requested_etablissement()

        if getattr(user, "role", None) == "super_admin" and requested_etablissement:
            return requested_etablissement

        return getattr(user, "etablissement", None)

    def get_queryset(self):
        user = self.request.user
        qs = TeacherAssignment.objects.select_related(
            "teacher", "subject", "classroom", "teacher__etablissement", "classroom__etablissement"
        )
        requested_etablissement = self._requested_etablissement()

        if requested_etablissement is not None:
            qs = qs.filter(classroom__etablissement=requested_etablissement)
        elif self._has_requested_scope():
            return qs.none()

        if getattr(user, "role", None) == "super_admin":
            return qs

        user_etablissement = getattr(user, "etablissement", None)
        if user_etablissement is None:
            return qs

        return qs.filter(classroom__etablissement=user_etablissement)

    def _validate_payload_etablissement(self, serializer):
        target_etablissement = self._resolve_target_etablissement()
        teacher = serializer.validated_data.get("teacher")
        classroom = serializer.validated_data.get("classroom")

        if target_etablissement is None:
            return

        if teacher and teacher.etablissement_id != target_etablissement.id:
            raise ValidationError({"teacher": "L'enseignant n'appartient pas a l'etablissement actif."})

        if classroom and classroom.etablissement_id != target_etablissement.id:
            raise ValidationError({"classroom": "La classe n'appartient pas a l'etablissement actif."})

    def perform_create(self, serializer):
        self._validate_payload_etablissement(serializer)
        serializer.save()

    def perform_update(self, serializer):
        self._validate_payload_etablissement(serializer)
        serializer.save()


class AvailabilityCampaignViewSet(BaseModelViewSet):
    """Les campagnes de collecte des disponibilites.

    Meme module de droits que les disponibilites elles-memes: qui les lit
    voit la campagne qui les encadre. L'ouverture et la fermeture, en
    revanche, relevent de l'ecriture -- un enseignant ne decide pas de la
    date a laquelle on cesse de l'attendre.
    """

    access_module = "teacher_availability"
    serializer_class = AvailabilityCampaignSerializer
    queryset = AvailabilityCampaign.objects.select_related(
        "etablissement", "academic_year"
    ).all()

    def get_queryset(self):
        queryset = (
            super()
            .get_queryset()
            .annotate(
                teachers_answered=Count(
                    "responses", filter=Q(responses__submitted_at__isnull=False)
                )
            )
            # Explicite: l'annotation fait perdre l'ordre du Meta, et la
            # pagination sur une liste non ordonnee rend des pages instables.
            .order_by("-opens_on", "-id")
        )
        etablissement = self._resolve_effective_etablissement_for_create()
        if etablissement is not None:
            return queryset.filter(etablissement=etablissement)
        if getattr(self.request.user, "role", None) == UserRole.SUPER_ADMIN:
            return queryset
        return queryset.none()

    def get_serializer_context(self):
        return super().get_serializer_context()

    def _effectif_enseignant(self, etablissement):
        return Teacher.objects.filter(etablissement=etablissement).count()

    def list(self, request, *args, **kwargs):
        reponse = super().list(request, *args, **kwargs)
        self._completer_effectifs(reponse)
        return reponse

    def retrieve(self, request, *args, **kwargs):
        reponse = super().retrieve(request, *args, **kwargs)
        self._completer_effectifs(reponse)
        return reponse

    def _completer_effectifs(self, reponse):
        """Le denominateur du taux de reponse: l'effectif enseignant du jour.

        Il n'est pas fige dans la campagne a dessein -- un enseignant recrute
        en cours d'annee doit entrer dans le compte, sans quoi le taux
        resterait a 100 % en l'ignorant.
        """
        lignes = reponse.data
        if isinstance(lignes, dict) and "results" in lignes:
            lignes = lignes["results"]
        if isinstance(lignes, dict):
            lignes = [lignes]
        if not isinstance(lignes, list):
            return

        effectifs = {}
        for ligne in lignes:
            if not isinstance(ligne, dict):
                continue
            etablissement_id = ligne.get("etablissement")
            if etablissement_id not in effectifs:
                effectifs[etablissement_id] = Teacher.objects.filter(
                    etablissement_id=etablissement_id
                ).count()
            ligne["teachers_total"] = effectifs[etablissement_id]

    def perform_create(self, serializer):
        etablissement = self._resolve_effective_etablissement_for_create()
        if etablissement is None:
            raise ValidationError(
                {"etablissement": "Sélectionnez un établissement actif."}
            )

        annee = serializer.validated_data.get("academic_year")
        # Controle avant la base: sa contrainte remonterait une erreur
        # d'integrite en 500, la ou l'utilisateur a simplement voulu ouvrir
        # une seconde collecte sur une annee qui en a deja une.
        if annee is not None and AvailabilityCampaign.objects.filter(
            etablissement=etablissement, academic_year=annee
        ).exists():
            raise ValidationError(
                {
                    "academic_year": "Une campagne existe déjà pour cette année "
                                     "scolaire. Modifiez-la plutôt que d'en ouvrir "
                                     "une seconde."
                }
            )

        serializer.save(etablissement=etablissement)

    @action(detail=True, methods=["get"])
    def responses(self, request, pk=None):
        """Qui a répondu, qui n'a pas — la liste complète, silencieux compris.

        Les seuls repondants ne diraient rien: c'est la liste des manquants
        qui sert a relancer, et elle n'existe qu'en partant de l'effectif.
        """
        campagne = self.get_object()
        deja = {
            reponse.teacher_id: reponse
            for reponse in TeacherAvailabilityResponse.objects.filter(
                campaign=campagne
            ).select_related("teacher", "teacher__user")
        }

        lignes = []
        for enseignant in Teacher.objects.filter(
            etablissement=campagne.etablissement
        ).select_related("user"):
            reponse = deja.get(enseignant.id)
            user = enseignant.user
            lignes.append(
                {
                    "teacher": enseignant.id,
                    "teacher_name": (user.get_full_name().strip() or user.username)
                    if user
                    else "",
                    "teacher_employee_code": enseignant.employee_code or "",
                    "submitted_at": reponse.submitted_at if reponse else None,
                    "is_submitted": bool(reponse and reponse.submitted_at),
                    "reminder_count": reponse.reminder_count if reponse else 0,
                    "slots_declared": TeacherAvailabilitySlot.objects.filter(
                        campaign=campagne, teacher=enseignant
                    ).count(),
                }
            )

        lignes.sort(key=lambda ligne: (ligne["is_submitted"], ligne["teacher_name"].lower()))
        rendus = sum(1 for ligne in lignes if ligne["is_submitted"])

        return Response(
            {
                "campaign": campagne.id,
                "teachers_total": len(lignes),
                "teachers_answered": rendus,
                "teachers_missing": len(lignes) - rendus,
                "results": lignes,
            }
        )

    @action(detail=True, methods=["post"], url_path="remind")
    def remind(self, request, pk=None):
        """Relance les enseignants qui n'ont pas encore rendu leurs créneaux."""
        campagne = self.get_object()
        # La matrice ouvre l'ecriture a l'enseignant pour ses propres
        # creneaux; elle ne lui donne pas le droit de relancer ses collegues.
        if getattr(request.user, "role", None) == UserRole.TEACHER:
            raise PermissionDenied(
                "Accès refusé : la relance est réservée à l'administration."
            )
        if campagne.status == AvailabilityCampaign.Status.CLOSED:
            raise ValidationError(
                {"detail": "Cette campagne est close : plus personne n'est attendu."}
            )

        rendus = set(
            TeacherAvailabilityResponse.objects.filter(
                campaign=campagne, submitted_at__isnull=False
            ).values_list("teacher_id", flat=True)
        )
        manquants = Teacher.objects.filter(
            etablissement=campagne.etablissement
        ).exclude(id__in=rendus).select_related("user")

        envoyees = 0
        maintenant = timezone.now()
        for enseignant in manquants:
            if enseignant.user is None:
                continue
            Notification.objects.create(
                etablissement=campagne.etablissement,
                recipient=enseignant.user,
                channel=NotificationChannel.PUSH,
                title="Disponibilités attendues",
                message=(
                    f"« {campagne.label} » se termine le "
                    f"{campagne.closes_on.strftime('%d/%m/%Y')}. "
                    "Déclarez vos disponibilités depuis l'application."
                ),
            )
            reponse, _ = TeacherAvailabilityResponse.objects.get_or_create(
                campaign=campagne, teacher=enseignant
            )
            reponse.reminded_at = maintenant
            reponse.reminder_count = reponse.reminder_count + 1
            reponse.save(update_fields=["reminded_at", "reminder_count", "updated_at"])
            envoyees += 1

        return Response({"reminded": envoyees})

    @action(detail=True, methods=["post"])
    def submit(self, request, pk=None):
        """« J'ai terminé » : l'enseignant rend ses disponibilités.

        C'est ce geste qui separe le silence de l'indisponibilite. Sans lui,
        une grille vide pouvait aussi bien vouloir dire « je ne peux jamais »
        que « je n'ai pas encore ouvert l'ecran », et l'administration
        relançait a l'aveugle.
        """
        campagne = self.get_object()
        enseignant = Teacher.objects.filter(user=request.user).first()
        if enseignant is None:
            raise ValidationError(
                {"detail": "Seul un enseignant rend ses propres disponibilités."}
            )
        if enseignant.etablissement_id != campagne.etablissement_id:
            raise ValidationError(
                {"detail": "Cette campagne ne concerne pas votre établissement."}
            )
        if not campagne.est_ouverte:
            raise ValidationError(
                {"detail": "La collecte n'est pas ouverte."}
            )

        reponse, _ = TeacherAvailabilityResponse.objects.get_or_create(
            campaign=campagne, teacher=enseignant
        )
        reponse.submitted_at = timezone.now()
        reponse.save(update_fields=["submitted_at", "updated_at"])

        return Response(
            {
                "campaign": campagne.id,
                "teacher": enseignant.id,
                "submitted_at": reponse.submitted_at,
                "slots_declared": TeacherAvailabilitySlot.objects.filter(
                    campaign=campagne, teacher=enseignant
                ).count(),
            }
        )


class TeacherAvailabilitySlotViewSet(BaseModelViewSet):
    access_module = "teacher_availability"
    DAY_ORDER = ["MON", "TUE", "WED", "THU", "FRI", "SAT"]
    DAY_LABELS = {
        "MON": "Lundi",
        "TUE": "Mardi",
        "WED": "Mercredi",
        "THU": "Jeudi",
        "FRI": "Vendredi",
        "SAT": "Samedi",
    }

    queryset = TeacherAvailabilitySlot.objects.select_related(
        "teacher",
        "teacher__user",
        "etablissement",
    ).all()
    serializer_class = TeacherAvailabilitySlotSerializer
    filterset_fields = ["teacher", "day_of_week"]
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]





    def _resolve_target_etablissement(self):
        user = self.request.user
        requested_etablissement = self._requested_etablissement()

        if getattr(user, "role", None) == "super_admin" and requested_etablissement:
            return requested_etablissement

        return getattr(user, "etablissement", None)

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        requested_etablissement = self._requested_etablissement()

        if requested_etablissement is not None:
            qs = qs.filter(etablissement=requested_etablissement)
        elif self._has_requested_scope():
            return qs.none()
        elif getattr(user, "role", None) != "super_admin":
            target_etablissement = getattr(user, "etablissement", None)
            if target_etablissement is None and getattr(user, "role", None) == UserRole.TEACHER:
                teacher_profile = Teacher.objects.select_related("etablissement").filter(user=user).first()
                if teacher_profile:
                    target_etablissement = teacher_profile.etablissement

            qs = qs.filter(etablissement=target_etablissement)

        teacher_id = self.request.query_params.get("teacher")
        if teacher_id:
            qs = qs.filter(teacher_id=teacher_id)

        campagne = self.request.query_params.get("campaign")
        if campagne not in (None, "") and str(campagne).isdigit():
            qs = qs.filter(campaign_id=campagne)

        genre = (self.request.query_params.get("kind") or "").strip()
        if genre in AvailabilityKind.values:
            qs = qs.filter(kind=genre)
        return qs

    def _resolve_teacher_for_request(self, serializer):
        user = self.request.user
        payload_teacher = serializer.validated_data.get("teacher")

        if getattr(user, "role", None) == UserRole.TEACHER:
            teacher_profile = Teacher.objects.filter(user=user).first()
            if not teacher_profile:
                raise ValidationError({"teacher": "Profil enseignant introuvable pour l'utilisateur connecté."})
            return teacher_profile

        if payload_teacher:
            return payload_teacher

        raise ValidationError({"teacher": "teacher est requis."})

    def _campagne_active(self, etablissement):
        """La campagne en cours pour cet etablissement, ou None.

        Une seule peut exister par annee scolaire; on retient celle qui est
        ouverte aujourd'hui, faute de quoi le rattachement serait arbitraire.
        """
        if etablissement is None:
            return None
        for campagne in AvailabilityCampaign.objects.filter(
            etablissement=etablissement, status=AvailabilityCampaign.Status.OPEN
        ).order_by("-opens_on"):
            if campagne.est_ouverte:
                return campagne
        return None

    def _verifier_la_collecte(self, etablissement):
        """L'enseignant ne declare que pendant la collecte; l'ecole, toujours.

        La contrainte de calendrier n'a de sens que pour celui a qui on
        demande de repondre. L'administration corrige une declaration bien
        apres la cloture -- c'est son travail d'arbitre, et le lui interdire
        la renverrait vers la base de donnees.
        """
        if getattr(self.request.user, "role", None) != UserRole.TEACHER:
            return self._campagne_active(etablissement)

        campagne = self._campagne_active(etablissement)
        if campagne is None:
            # Aucune campagne ouverte: on ne bloque pas pour autant. Une
            # ecole qui n'a pas encore adopte les campagnes doit continuer a
            # recueillir les disponibilites comme avant.
            if AvailabilityCampaign.objects.filter(etablissement=etablissement).exists():
                raise ValidationError(
                    {
                        "detail": "La collecte des disponibilités n'est pas ouverte "
                                  "en ce moment."
                    }
                )
        return campagne

    def perform_create(self, serializer):
        target_etablissement = self._resolve_target_etablissement()
        teacher = self._resolve_teacher_for_request(serializer)

        if target_etablissement and teacher.etablissement_id != target_etablissement.id:
            raise ValidationError({"teacher": "Cet enseignant n'appartient pas à l'établissement actif."})

        etablissement = target_etablissement or teacher.etablissement
        campagne = self._verifier_la_collecte(etablissement)
        serializer.save(
            teacher=teacher,
            etablissement=etablissement,
            # Rattachee a la collecte en cours, donc a son annee scolaire:
            # sans cela, les declarations de l'an dernier restent melees a
            # celles de la rentree.
            campaign=campagne,
        )

    def perform_update(self, serializer):
        if getattr(self.request.user, "role", None) == UserRole.TEACHER:
            teacher_profile = Teacher.objects.filter(user=self.request.user).first()
            if not teacher_profile:
                raise ValidationError({"teacher": "Profil enseignant introuvable."})
            serializer.save(teacher=teacher_profile, etablissement=teacher_profile.etablissement)
            return

        target_etablissement = self._resolve_target_etablissement()
        teacher = serializer.validated_data.get("teacher") or getattr(serializer.instance, "teacher", None)
        if target_etablissement and teacher and teacher.etablissement_id != target_etablissement.id:
            raise ValidationError({"teacher": "Cet enseignant n'appartient pas à l'établissement actif."})
        serializer.save(etablissement=target_etablissement or getattr(serializer.instance, "etablissement", None))

    @action(detail=False, methods=["get"], permission_classes=[permissions.IsAuthenticated])
    def grid(self, request):
        """La semaine type: qui est disponible, et combien, sur chaque case.

        La grille disait auparavant « disponible » ou « indisponible » pour
        l'etablissement entier, et nommait celui qui avait « reserve » la
        case. Elle decrivait une reservation exclusive, pas une collecte de
        disponibilites: il n'y avait de place que pour un declarant par
        creneau.

        Elle rend desormais deux choses a la fois, parce que deux ecrans la
        lisent: le compte par etat -- ce dont l'administration a besoin pour
        arbitrer -- et ce que l'enseignant vise a declare sur cette case.
        """
        try:
            start_hour = int(request.query_params.get("start_hour", 7))
            end_hour = int(request.query_params.get("end_hour", 18))
            slot_minutes = int(request.query_params.get("slot_minutes", 60))
        except (TypeError, ValueError):
            return Response({"detail": "Paramètres horaires invalides."}, status=400)

        if start_hour < 0 or end_hour > 24 or start_hour >= end_hour:
            return Response({"detail": "Plage horaire invalide."}, status=400)
        if slot_minutes <= 0 or slot_minutes > 180:
            return Response({"detail": "slot_minutes invalide."}, status=400)

        from datetime import time

        cases = []
        curseur = start_hour * 60
        fin_minutes = end_hour * 60
        while curseur + slot_minutes <= fin_minutes:
            debut_case = time(hour=curseur // 60, minute=curseur % 60)
            fin_case_minutes = curseur + slot_minutes
            fin_case = time(hour=fin_case_minutes // 60, minute=fin_case_minutes % 60)
            cases.append((debut_case, fin_case))
            curseur += slot_minutes

        declarations = list(self._declarations_du_perimetre())

        vise = request.query_params.get("teacher")
        teacher_vise = int(vise) if vise not in (None, "") and str(vise).isdigit() else None

        jours = []
        for code in self.DAY_ORDER:
            du_jour = [row for row in declarations if row.day_of_week == code]
            lignes = []
            for debut_case, fin_case in cases:
                # Une declaration compte pour la case si elle la contient
                # entierement: une heure declaree ne rend pas disponible sur
                # deux heures.
                couvrantes = [
                    row for row in du_jour if row.couvre(debut_case, fin_case)
                ]
                mienne = next(
                    (row for row in couvrantes if row.teacher_id == teacher_vise), None
                )
                lignes.append(
                    {
                        "day_of_week": code,
                        "day_label": self.DAY_LABELS.get(code, code),
                        "start_time": debut_case.strftime("%H:%M:%S"),
                        "end_time": fin_case.strftime("%H:%M:%S"),
                        "preferred_count": sum(
                            1 for row in couvrantes
                            if row.kind == AvailabilityKind.PREFERRED
                        ),
                        "possible_count": sum(
                            1 for row in couvrantes
                            if row.kind == AvailabilityKind.POSSIBLE
                        ),
                        "unavailable_count": sum(
                            1 for row in couvrantes
                            if row.kind == AvailabilityKind.UNAVAILABLE
                        ),
                        "teachers": [
                            {
                                "teacher": row.teacher_id,
                                "teacher_name": self._teacher_name(row.teacher),
                                "kind": row.kind,
                            }
                            for row in couvrantes
                        ],
                        "mine": mienne.kind if mienne else None,
                        "mine_id": mienne.id if mienne else None,
                        "mine_exact": bool(
                            mienne
                            and mienne.start_time == debut_case
                            and mienne.end_time == fin_case
                        ),
                    }
                )
            jours.append(
                {
                    "day_of_week": code,
                    "day_label": self.DAY_LABELS.get(code, code),
                    "cells": lignes,
                }
            )

        return Response(
            {
                "start_hour": start_hour,
                "end_hour": end_hour,
                "slot_minutes": slot_minutes,
                "teacher": teacher_vise,
                "days": jours,
            }
        )

    def _declarations_du_perimetre(self):
        """Toutes les declarations de l'etablissement, collegues compris.

        Distinct de `get_queryset`, qui honore le filtre `teacher` de l'URL:
        la grille a besoin de compter tout le monde, et n'en isole un qu'a
        l'affichage.
        """
        base = TeacherAvailabilitySlot.objects.select_related(
            "teacher", "teacher__user"
        )
        etablissement = self._resolve_target_etablissement()
        if etablissement is not None:
            return base.filter(etablissement=etablissement)
        if getattr(self.request.user, "role", None) == UserRole.SUPER_ADMIN:
            return base
        return base.none()

    @staticmethod
    def _teacher_name(teacher):
        user = teacher.user if teacher else None
        if not user:
            return ""
        full_name = user.get_full_name().strip()
        return full_name or user.username

    # --- Ce qui sert a construire le planning -------------------------------

    @action(detail=False, methods=["get"], url_path="for-planning")
    def for_planning(self, request):
        """Qui peut prendre ce créneau, du plus volontaire au moins disponible.

        La collecte s'arretait a elle-meme: les declarations etaient
        enregistrees, puis oubliees. Celui qui construisait l'emploi du temps
        ne les voyait nulle part et placait ses cours a l'aveugle.

        Quatre groupes, dans l'ordre ou l'administration les regarde:
        preferes, possibles, ceux qui n'ont rien dit, et ceux qui se sont
        declares indisponibles.
        """
        jour = (request.query_params.get("day") or "").strip().upper()
        if jour not in self.DAY_LABELS:
            raise ValidationError({"day": "Jour attendu: MON, TUE, WED, THU, FRI ou SAT."})

        debut = self._heure_demandee(request, "start")
        fin = self._heure_demandee(request, "end")
        if fin <= debut:
            raise ValidationError({"end": "L'heure de fin doit être après l'heure de début."})

        declarations = list(
            self.get_queryset().filter(day_of_week=jour).select_related(
                "teacher", "teacher__user"
            )
        )

        # Une declaration ne vaut que si elle contient tout le creneau vise:
        # une heure declaree ne couvre pas un cours de deux heures.
        par_enseignant = {}
        for declaration in declarations:
            if not declaration.couvre(debut, fin):
                continue
            actuelle = par_enseignant.get(declaration.teacher_id)
            # Une indisponibilite l'emporte sur tout: elle est ce que
            # l'enseignant a pris la peine d'ecrire pour dire non.
            if actuelle is None or self._prime_sur(declaration, actuelle):
                par_enseignant[declaration.teacher_id] = declaration

        groupes = {
            AvailabilityKind.PREFERRED: [],
            AvailabilityKind.POSSIBLE: [],
            AvailabilityKind.UNAVAILABLE: [],
        }
        for declaration in par_enseignant.values():
            groupes[declaration.kind].append(
                {
                    "teacher": declaration.teacher_id,
                    "teacher_name": self._teacher_name(declaration.teacher),
                    "note": declaration.note,
                    "declared_start": declaration.start_time.strftime("%H:%M"),
                    "declared_end": declaration.end_time.strftime("%H:%M"),
                }
            )

        for lignes in groupes.values():
            lignes.sort(key=lambda ligne: ligne["teacher_name"].lower())

        silencieux = [
            {
                "teacher": enseignant.id,
                "teacher_name": self._teacher_name(enseignant),
                "note": "",
                "declared_start": None,
                "declared_end": None,
            }
            for enseignant in self._enseignants_du_perimetre()
            if enseignant.id not in par_enseignant
        ]
        silencieux.sort(key=lambda ligne: ligne["teacher_name"].lower())

        return Response(
            {
                "day_of_week": jour,
                "day_label": self.DAY_LABELS[jour],
                "start_time": debut.strftime("%H:%M"),
                "end_time": fin.strftime("%H:%M"),
                "preferred": groupes[AvailabilityKind.PREFERRED],
                "possible": groupes[AvailabilityKind.POSSIBLE],
                # Ni disponibles ni indisponibles: ils n'ont rien declare, et
                # c'est une information a part entiere.
                "undeclared": silencieux,
                "unavailable": groupes[AvailabilityKind.UNAVAILABLE],
            }
        )

    @staticmethod
    def _prime_sur(candidate, actuelle):
        """Quelle declaration retenir quand deux couvrent le meme creneau."""
        rang = {
            AvailabilityKind.UNAVAILABLE: 3,
            AvailabilityKind.PREFERRED: 2,
            AvailabilityKind.POSSIBLE: 1,
        }
        return rang.get(candidate.kind, 0) > rang.get(actuelle.kind, 0)

    @staticmethod
    def _heure_demandee(request, cle):
        brut = (request.query_params.get(cle) or "").strip()
        if not brut:
            raise ValidationError({cle: "Heure attendue au format HH:MM."})
        valeur = parse_time(brut)
        if valeur is None:
            raise ValidationError({cle: "Heure illisible. Format attendu: HH:MM."})
        return valeur

    def _enseignants_du_perimetre(self):
        """Tous les enseignants de l'etablissement actif, declarants ou non."""
        base = Teacher.objects.select_related("user", "etablissement")
        etablissement = self._resolve_target_etablissement()
        if etablissement is not None:
            return base.filter(etablissement=etablissement)
        if getattr(self.request.user, "role", None) == UserRole.SUPER_ADMIN:
            return base
        return base.none()


class TeacherScheduleSlotViewSet(BaseModelViewSet):
    access_module = "timetable"
    DAY_ORDER = ["MON", "TUE", "WED", "THU", "FRI", "SAT"]
    DAY_LABELS = {
        "MON": "Lundi",
        "TUE": "Mardi",
        "WED": "Mercredi",
        "THU": "Jeudi",
        "FRI": "Vendredi",
        "SAT": "Samedi",
    }

    queryset = TeacherScheduleSlot.objects.select_related(
        "assignment",
        "assignment__teacher",
        "assignment__subject",
        "assignment__classroom",
    ).all()
    serializer_class = TeacherScheduleSlotSerializer
    filterset_fields = ["assignment", "day_of_week"]
    search_fields = [
        "room",
        "assignment__subject__name",
        "assignment__classroom__name",
        "assignment__teacher__user__first_name",
        "assignment__teacher__user__last_name",
    ]
    ordering_fields = ["day_of_week", "start_time", "end_time"]
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]





    def _resolve_target_etablissement(self):
        user = self.request.user
        requested_etablissement = self._requested_etablissement()

        if getattr(user, "role", None) == "super_admin" and requested_etablissement:
            return requested_etablissement

        return getattr(user, "etablissement", None)

    def _scoped_classroom_queryset(self):
        user = self.request.user
        requested = self._requested_etablissement()
        qs = ClassRoom.objects.select_related("etablissement")

        if requested is not None:
            return qs.filter(etablissement=requested)
        if self._has_requested_scope():
            return qs.none()
        if getattr(user, "role", None) == "super_admin":
            return qs
        return qs.filter(etablissement=getattr(user, "etablissement", None))

    def _get_scoped_classroom_or_404(self, classroom_id):
        return get_object_or_404(self._scoped_classroom_queryset(), id=classroom_id)

    @staticmethod
    def _to_bool(raw_value, default=False):
        if raw_value is None:
            return default
        if isinstance(raw_value, bool):
            return raw_value
        return str(raw_value).strip().lower() in {"1", "true", "yes", "on"}

    @staticmethod
    def _minutes(time_value):
        return time_value.hour * 60 + time_value.minute

    @classmethod
    def _overlap(cls, start_a, end_a, start_b, end_b):
        return cls._minutes(start_a) < cls._minutes(end_b) and cls._minutes(end_a) > cls._minutes(start_b)

    @staticmethod
    def _slot_range_label(slot):
        return f"{slot.start_time.strftime('%H:%M')}-{slot.end_time.strftime('%H:%M')}"

    @staticmethod
    def _pdf_text(value):
        return str(value or "").encode("latin-1", "replace").decode("latin-1")

    @classmethod
    def _slot_short_label(cls, slot):
        assignment = slot.assignment
        subject_code = assignment.subject.code if assignment and assignment.subject else "MAT"
        teacher = assignment.teacher if assignment else None
        teacher_label = cls._teacher_name(teacher) or (teacher.employee_code if teacher else "ENS")
        room = (slot.room or "").strip()
        room_label = f" [{room}]" if room else ""
        return f"{subject_code} ({teacher_label}){room_label}"

    @classmethod
    def _sorted_ranges(cls, matrix):
        def key_func(range_label):
            start_raw = range_label.split("-")[0]
            start = datetime.strptime(start_raw, "%H:%M")
            return start.hour * 60 + start.minute

        return sorted(matrix.keys(), key=key_func)

    @classmethod
    def _build_class_matrix(cls, slots):
        matrix = {}
        for slot in slots:
            range_label = cls._slot_range_label(slot)
            day_map = matrix.setdefault(
                range_label,
                {day: [] for day in cls.DAY_ORDER},
            )
            day_map[slot.day_of_week].append(cls._slot_short_label(slot))

        for day_map in matrix.values():
            for entries in day_map.values():
                entries.sort()

        return matrix

    @classmethod
    def _sheet_title(cls, class_name):
        cleaned = class_name.strip() or "Classe"
        return cleaned[:31]

    @staticmethod
    def _etablissement_logo_path(etablissement):
        if not etablissement:
            return None

        logo_field = getattr(etablissement, "logo", None)
        if not logo_field:
            return None

        direct_path = str(getattr(logo_field, "path", "") or "").strip()
        if direct_path and os.path.exists(direct_path):
            return direct_path

        logo_name = str(getattr(logo_field, "name", "") or "").strip()
        media_root = str(getattr(settings, "MEDIA_ROOT", "") or "").strip()
        if logo_name and media_root:
            candidate = os.path.join(media_root, logo_name)
            if os.path.exists(candidate):
                return candidate

        return None

    def _etablissement_meta_lines(self, classroom):
        etablissement = getattr(classroom, "etablissement", None)
        if not etablissement:
            etablissement = self._requested_etablissement() or self._resolve_target_etablissement()

        # Some legacy classes may not yet be linked to an etablissement;
        # in that case, keep the active scope name visible in exports.
        if not etablissement:
            requested_name = self._requested_etablissement_name()
            if requested_name:
                return {
                    "name": requested_name,
                    "details": "",
                }

        if not etablissement:
            return {
                "name": "Etablissement non defini",
                "details": "",
            }

        details = []
        if etablissement.address:
            details.append(etablissement.address)
        if etablissement.phone:
            details.append(f"Tel: {etablissement.phone}")
        if etablissement.email:
            details.append(etablissement.email)

        return {
            "name": etablissement.name,
            "details": " | ".join(details),
            "logo_path": self._etablissement_logo_path(etablissement),
        }

    @staticmethod
    def _teacher_name(teacher):
        user = teacher.user if teacher else None
        if not user:
            return ""
        full_name = user.get_full_name().strip()
        return full_name or user.username

    def _parse_classroom_id(self, request):
        raw = request.data.get("classroom")
        if raw is None:
            raw = request.query_params.get("classroom")
        if raw is None:
            return None
        try:
            return int(raw)
        except (TypeError, ValueError):
            return None

    def _publication_response(self, classroom):
        publication = TimetablePublication.objects.filter(classroom=classroom).first()
        if publication:
            payload = TimetablePublicationSerializer(publication).data
        else:
            payload = {
                "id": None,
                "classroom": classroom.id,
                "classroom_name": classroom.name,
                "is_published": False,
                "is_locked": False,
                "published_by": None,
                "published_by_name": "",
                "published_at": None,
                "notes": "",
            }

        slot_count = self.get_queryset().filter(assignment__classroom=classroom).count()
        payload["slot_count"] = slot_count
        return payload

    def _class_slots_queryset(self, classroom):
        return self.get_queryset().filter(assignment__classroom=classroom)

    def _teacher_workload_rows(self, slots_queryset):
        rows = {}
        for slot in slots_queryset:
            assignment = slot.assignment
            teacher = assignment.teacher
            teacher_id = teacher.id
            row = rows.setdefault(
                teacher_id,
                {
                    "teacher": teacher.id,
                    "teacher_code": teacher.employee_code,
                    "teacher_name": self._teacher_name(teacher),
                    "slot_count": 0,
                    "class_count": 0,
                    "classrooms": set(),
                    "per_day_minutes": {day: 0 for day in self.DAY_ORDER},
                    "total_minutes": 0,
                },
            )

            duration = self._minutes(slot.end_time) - self._minutes(slot.start_time)
            if duration < 0:
                duration = 0

            row["slot_count"] += 1
            row["total_minutes"] += duration
            row["per_day_minutes"][slot.day_of_week] += duration
            row["classrooms"].add(assignment.classroom_id)

        result = []
        for row in rows.values():
            row["class_count"] = len(row["classrooms"])
            del row["classrooms"]
            row["total_hours"] = round(row["total_minutes"] / 60, 2)
            row["per_day_hours"] = {
                day: round(minutes / 60, 2)
                for day, minutes in row["per_day_minutes"].items()
            }
            if row["total_minutes"] >= 26 * 60:
                row["load_level"] = "overload"
            elif row["total_minutes"] >= 20 * 60:
                row["load_level"] = "watch"
            else:
                row["load_level"] = "ok"
            result.append(row)

        result.sort(key=lambda item: (-item["total_minutes"], item["teacher_code"]))
        return result

    @staticmethod
    def _assignment_target_maps(target_assignments):
        exact = {}
        by_subject = {}
        for assignment in target_assignments:
            exact[(assignment.subject_id, assignment.teacher_id)] = assignment
            by_subject.setdefault(assignment.subject_id, []).append(assignment)

        for assignments in by_subject.values():
            assignments.sort(key=lambda assignment: (assignment.teacher_id, assignment.id))

        return exact, by_subject

    def _resolve_target_assignment(self, source_assignment, exact_map, subject_map):
        exact = exact_map.get((source_assignment.subject_id, source_assignment.teacher_id))
        if exact:
            return exact, "exact"

        subject_matches = subject_map.get(source_assignment.subject_id) or []
        if subject_matches:
            return subject_matches[0], "subject-fallback"

        return None, "unmapped"

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        publication = TimetablePublication.objects.filter(
            classroom=instance.assignment.classroom,
            is_locked=True,
        ).first()
        if publication:
            return Response(
                {
                    "detail": "Emploi du temps verrouillé pour cette classe. "
                    "Déverrouillez avant toute suppression."
                },
                status=400,
            )
        return super().destroy(request, *args, **kwargs)

    def _validate_assignment_scope(self, serializer):
        assignment = serializer.validated_data.get("assignment")
        if not assignment:
            return

        target_etablissement = self._resolve_target_etablissement()
        if target_etablissement is None:
            return

        if assignment.classroom.etablissement_id != target_etablissement.id:
            raise ValidationError({
                "assignment": "Cette affectation n'appartient pas a l'etablissement actif."
            })

    def perform_create(self, serializer):
        self._validate_assignment_scope(serializer)
        serializer.save()

    def perform_update(self, serializer):
        self._validate_assignment_scope(serializer)
        serializer.save()

    @action(detail=False, methods=["post"], url_path="import-by-class")
    def import_by_class(self, request):
        try:
            classroom_id = int(request.data.get("classroom_id"))
        except (TypeError, ValueError):
            classroom_id = None
        confirm = self._to_bool(request.data.get("confirm"), default=False)
        confirm_conflicts = self._to_bool(request.data.get("confirm_conflicts"), default=False)

        if not classroom_id:
            raise ValidationError({"classroom_id": "Classe requise."})

        classroom = self._get_scoped_classroom_or_404(classroom_id)
        rows = _load_import_rows(request.FILES.get("file") or request.data.get("file"))
        if not rows:
            raise ValidationError({"file": "Aucune ligne exploitable dans le fichier."})

        assignments = TeacherAssignment.objects.select_related("teacher", "subject", "classroom").filter(classroom=classroom)
        assignments_by_subject_code = {(a.subject.code or "").strip().lower(): a for a in assignments}
        assignments_by_subject_name = {(a.subject.name or "").strip().lower(): a for a in assignments}

        errors = []
        prepared = []
        conflicts = []
        class_conflict_slot_ids = set()
        blocking_conflicts = []

        for index, row in enumerate(rows, start=2):
            day_raw = _as_text(row.get("day_of_week") or row.get("jour")).upper()
            day_map = {
                "LUNDI": "MON",
                "MARDI": "TUE",
                "MERCREDI": "WED",
                "JEUDI": "THU",
                "VENDREDI": "FRI",
                "SAMEDI": "SAT",
            }
            day_code = day_map.get(day_raw, day_raw)
            start_time = _as_time(row.get("start_time") or row.get("debut"))
            end_time = _as_time(row.get("end_time") or row.get("fin"))
            subject_code = _as_text(row.get("subject_code") or row.get("matiere_code")).lower()
            subject_name = _as_text(row.get("subject_name") or row.get("matiere")).lower()
            room = _as_text(row.get("room") or row.get("salle"))

            if day_code not in self.DAY_ORDER:
                errors.append({"row": index, "error": "day_of_week invalide (MON..SAT)."})
                continue
            if not start_time or not end_time or end_time <= start_time:
                errors.append({"row": index, "error": "Heures invalides (start_time/end_time)."})
                continue

            assignment = None
            if subject_code:
                assignment = assignments_by_subject_code.get(subject_code)
            if assignment is None and subject_name:
                assignment = assignments_by_subject_name.get(subject_name)
            if assignment is None:
                errors.append({"row": index, "error": "Matière non affectée à la classe (subject_code/subject_name)."})
                continue

            target_existing = TeacherScheduleSlot.objects.filter(
                assignment=assignment,
                day_of_week=day_code,
                start_time=start_time,
                end_time=end_time,
            ).first()

            overlaps = TeacherScheduleSlot.objects.select_related("assignment", "assignment__teacher", "assignment__subject", "assignment__classroom").filter(
                day_of_week=day_code,
                start_time__lt=end_time,
                end_time__gt=start_time,
            )
            if target_existing is not None:
                overlaps = overlaps.exclude(pk=target_existing.pk)

            class_conflicts = overlaps.filter(assignment__classroom=classroom)
            teacher_conflicts = overlaps.filter(assignment__teacher=assignment.teacher).exclude(assignment__classroom=classroom)
            room_conflicts = TeacherScheduleSlot.objects.none()
            if room:
                room_conflicts = overlaps.exclude(room__exact="").filter(room__iexact=room).exclude(assignment__classroom=classroom)

            if class_conflicts.exists() or teacher_conflicts.exists() or room_conflicts.exists():
                issue = {
                    "row": index,
                    "day_of_week": day_code,
                    "start_time": start_time.strftime("%H:%M"),
                    "end_time": end_time.strftime("%H:%M"),
                    "subject": assignment.subject.code,
                    "class_conflicts": [slot.id for slot in class_conflicts[:20]],
                    "teacher_conflicts": [slot.id for slot in teacher_conflicts[:20]],
                    "room_conflicts": [slot.id for slot in room_conflicts[:20]],
                }
                conflicts.append(issue)
                for slot in class_conflicts:
                    class_conflict_slot_ids.add(slot.id)
                if teacher_conflicts.exists() or room_conflicts.exists():
                    blocking_conflicts.append(issue)

            prepared.append(
                {
                    "row": index,
                    "assignment": assignment,
                    "day_of_week": day_code,
                    "start_time": start_time,
                    "end_time": end_time,
                    "room": room,
                    "existing": target_existing,
                }
            )

        to_create = sum(1 for item in prepared if item["existing"] is None)
        to_update = len(prepared) - to_create
        payload = {
            "classroom": {"id": classroom.id, "name": classroom.name},
            "summary": {
                "total_rows": len(rows),
                "valid_rows": len(prepared),
                "errors": len(errors),
                "to_create": to_create,
                "to_update": to_update,
                "conflicts": len(conflicts),
                "blocking_conflicts": len(blocking_conflicts),
            },
            "errors": errors,
            "conflicts": conflicts[:150],
            "preview": [
                {
                    "row": item["row"],
                    "action": "create" if item["existing"] is None else "update",
                    "day_of_week": item["day_of_week"],
                    "start_time": item["start_time"].strftime("%H:%M"),
                    "end_time": item["end_time"].strftime("%H:%M"),
                    "subject": item["assignment"].subject.code,
                    "teacher": item["assignment"].teacher.employee_code,
                    "room": item["room"],
                }
                for item in prepared[:180]
            ],
            "confirm_required": True,
            "confirm_conflicts_required": len(conflicts) > 0,
        }

        if not confirm:
            return Response(payload)
        if errors:
            return Response({**payload, "detail": "Import bloqué: corrigez les erreurs."}, status=400)
        if blocking_conflicts:
            return Response(
                {
                    **payload,
                    "detail": "Conflits enseignant/salle détectés. Corrigez le fichier avant confirmation.",
                },
                status=400,
            )
        if conflicts and not confirm_conflicts:
            return Response(
                {
                    **payload,
                    "detail": "Conflits de créneaux détectés. Relancez avec confirm_conflicts=true après prévisualisation.",
                },
                status=409,
            )

        created = 0
        updated = 0
        deleted_conflicts = 0
        with transaction.atomic():
            if class_conflict_slot_ids:
                deleted_conflicts, _ = TeacherScheduleSlot.objects.filter(id__in=class_conflict_slot_ids).delete()

            for item in prepared:
                data = {
                    "assignment": item["assignment"].id,
                    "day_of_week": item["day_of_week"],
                    "start_time": item["start_time"].strftime("%H:%M:%S"),
                    "end_time": item["end_time"].strftime("%H:%M:%S"),
                    "room": item["room"],
                }

                existing = TeacherScheduleSlot.objects.filter(
                    assignment=item["assignment"],
                    day_of_week=item["day_of_week"],
                    start_time=item["start_time"],
                    end_time=item["end_time"],
                ).first()

                if existing is None:
                    serializer = self.get_serializer(data=data)
                    serializer.is_valid(raise_exception=True)
                    self.perform_create(serializer)
                    created += 1
                else:
                    serializer = self.get_serializer(existing, data=data, partial=False)
                    serializer.is_valid(raise_exception=True)
                    self.perform_update(serializer)
                    updated += 1

        return Response(
            {
                **payload,
                "result": {
                    "created": created,
                    "updated": updated,
                    "deleted_conflicts": deleted_conflicts,
                },
                "detail": "Import emploi du temps terminé.",
            }
        )

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        requested_etablissement = self._requested_etablissement()

        if requested_etablissement is not None:
            queryset = queryset.filter(assignment__classroom__etablissement=requested_etablissement)
        elif self._has_requested_scope():
            return queryset.none()
        elif getattr(user, "role", None) != "super_admin":
            queryset = queryset.filter(assignment__classroom__etablissement=getattr(user, "etablissement", None))

        classroom = self.request.query_params.get("classroom")
        if classroom:
            queryset = queryset.filter(assignment__classroom_id=classroom)
        return queryset

    @action(detail=False, methods=["get"], permission_classes=[permissions.IsAuthenticated])
    def publication_status(self, request):
        classroom_id = self._parse_classroom_id(request)
        if not classroom_id:
            return Response({"detail": "classroom est requis."}, status=400)

        classroom = self._get_scoped_classroom_or_404(classroom_id)
        return Response(self._publication_response(classroom))

    @action(detail=False, methods=["post"], permission_classes=[permissions.IsAuthenticated, HasModuleAccess])
    def publish_class(self, request):
        classroom_id = self._parse_classroom_id(request)
        if not classroom_id:
            return Response({"detail": "classroom est requis."}, status=400)

        classroom = self._get_scoped_classroom_or_404(classroom_id)

        lock_value = request.data.get("lock")
        if isinstance(lock_value, bool):
            lock_after_publish = lock_value
        elif lock_value is None:
            lock_after_publish = True
        else:
            lock_after_publish = str(lock_value).strip().lower() in {"1", "true", "yes", "on"}

        notes = str(request.data.get("notes") or "").strip()
        publication, _ = TimetablePublication.objects.get_or_create(classroom=classroom)
        publication.is_published = True
        publication.is_locked = lock_after_publish
        publication.published_by = request.user
        publication.published_at = timezone.now()
        publication.notes = notes
        publication.save()

        return Response(self._publication_response(classroom))

    @action(detail=False, methods=["post"], permission_classes=[permissions.IsAuthenticated, HasModuleAccess])
    def unpublish_class(self, request):
        classroom_id = self._parse_classroom_id(request)
        if not classroom_id:
            return Response({"detail": "classroom est requis."}, status=400)

        classroom = self._get_scoped_classroom_or_404(classroom_id)
        publication, _ = TimetablePublication.objects.get_or_create(classroom=classroom)
        publication.is_published = False
        publication.is_locked = False
        publication.published_by = None
        publication.published_at = None
        publication.notes = ""
        publication.save()

        return Response(self._publication_response(classroom))

    @action(detail=False, methods=["post"], permission_classes=[permissions.IsAuthenticated, HasModuleAccess])
    def lock_class(self, request):
        classroom_id = self._parse_classroom_id(request)
        if not classroom_id:
            return Response({"detail": "classroom est requis."}, status=400)

        classroom = self._get_scoped_classroom_or_404(classroom_id)
        publication, _ = TimetablePublication.objects.get_or_create(classroom=classroom)

        if not publication.is_published:
            return Response(
                {"detail": "Publiez d'abord l'emploi du temps avant de le verrouiller."},
                status=400,
            )

        publication.is_locked = True
        publication.save(update_fields=["is_locked", "updated_at"])
        return Response(self._publication_response(classroom))

    @action(detail=False, methods=["post"], permission_classes=[permissions.IsAuthenticated, HasModuleAccess])
    def unlock_class(self, request):
        classroom_id = self._parse_classroom_id(request)
        if not classroom_id:
            return Response({"detail": "classroom est requis."}, status=400)

        classroom = self._get_scoped_classroom_or_404(classroom_id)
        publication, _ = TimetablePublication.objects.get_or_create(classroom=classroom)
        publication.is_locked = False
        publication.save(update_fields=["is_locked", "updated_at"])
        return Response(self._publication_response(classroom))

    @action(detail=False, methods=["get"], permission_classes=[permissions.IsAuthenticated])
    def teacher_workload(self, request):
        classroom_id = self._parse_classroom_id(request)
        queryset = self.get_queryset()
        if classroom_id:
            queryset = queryset.filter(assignment__classroom_id=classroom_id)

        rows = self._teacher_workload_rows(queryset)
        total_minutes = sum(item["total_minutes"] for item in rows)
        payload = {
            "classroom": classroom_id,
            "teacher_count": len(rows),
            "total_minutes": total_minutes,
            "total_hours": round(total_minutes / 60, 2),
            "items": rows,
        }
        return Response(payload)

    @action(detail=False, methods=["post"], permission_classes=[permissions.IsAuthenticated, HasModuleAccess])
    def duplicate_schedule(self, request):
        source_classroom_id = request.data.get("source_classroom")
        target_classroom_id = request.data.get("target_classroom")

        try:
            source_classroom_id = int(source_classroom_id)
            target_classroom_id = int(target_classroom_id)
        except (TypeError, ValueError):
            return Response(
                {"detail": "source_classroom et target_classroom sont requis."},
                status=400,
            )

        if source_classroom_id == target_classroom_id:
            return Response(
                {"detail": "La classe source et la classe cible doivent être différentes."},
                status=400,
            )

        source_classroom = self._get_scoped_classroom_or_404(source_classroom_id)
        target_classroom = self._get_scoped_classroom_or_404(target_classroom_id)

        publication = TimetablePublication.objects.filter(
            classroom=target_classroom,
            is_locked=True,
        ).first()
        if publication:
            return Response(
                {
                    "detail": "Classe cible verrouillée. Déverrouillez avant duplication.",
                },
                status=400,
            )

        requested_days = request.data.get("days")
        if isinstance(requested_days, str):
            requested_days = [item.strip().upper() for item in requested_days.split(",") if item.strip()]
        if not isinstance(requested_days, list) or not requested_days:
            requested_days = list(self.DAY_ORDER)
        copy_days = [day for day in requested_days if day in self.DAY_ORDER]
        if not copy_days:
            return Response({"detail": "Aucun jour valide fourni."}, status=400)

        overwrite = self._to_bool(request.data.get("overwrite"), default=False)
        keep_room = self._to_bool(request.data.get("keep_room"), default=True)

        source_slots = list(
            self.get_queryset()
            .filter(assignment__classroom=source_classroom, day_of_week__in=copy_days)
            .order_by("day_of_week", "start_time", "end_time", "id")
        )
        if not source_slots:
            return Response(
                {"detail": "Aucun créneau à copier pour la classe source."},
                status=400,
            )

        target_assignments = list(
            TeacherAssignment.objects.select_related("teacher", "subject", "classroom").filter(
                classroom=target_classroom
            )
        )
        if not target_assignments:
            return Response(
                {
                    "detail": "La classe cible n'a aucune affectation. Ajoutez d'abord les affectations.",
                },
                status=400,
            )

        exact_map, subject_map = self._assignment_target_maps(target_assignments)

        summary = {
            "source_classroom": source_classroom.id,
            "target_classroom": target_classroom.id,
            "days": copy_days,
            "overwrite": overwrite,
            "keep_room": keep_room,
            "source_slots": len(source_slots),
            "deleted": 0,
            "created": 0,
            "updated": 0,
            "skipped_unmapped": 0,
            "skipped_conflicts": 0,
            "mapping_examples": [],
            "conflicts": [],
            "unmapped": [],
        }

        with transaction.atomic():
            if overwrite:
                delete_qs = TeacherScheduleSlot.objects.filter(
                    assignment__classroom=target_classroom,
                    day_of_week__in=copy_days,
                )
                deleted_count, _ = delete_qs.delete()
                summary["deleted"] = deleted_count

            for source_slot in source_slots:
                source_assignment = source_slot.assignment
                target_assignment, mapping_mode = self._resolve_target_assignment(
                    source_assignment,
                    exact_map,
                    subject_map,
                )

                if not target_assignment:
                    summary["skipped_unmapped"] += 1
                    if len(summary["unmapped"]) < 20:
                        summary["unmapped"].append(
                            {
                                "day": source_slot.day_of_week,
                                "time": self._slot_range_label(source_slot),
                                "subject_code": source_assignment.subject.code,
                                "teacher_code": source_assignment.teacher.employee_code,
                            }
                        )
                    continue

                room_value = source_slot.room if keep_room else ""

                overlapping = TeacherScheduleSlot.objects.select_related(
                    "assignment",
                    "assignment__teacher",
                    "assignment__subject",
                    "assignment__classroom",
                ).filter(
                    day_of_week=source_slot.day_of_week,
                    start_time__lt=source_slot.end_time,
                    end_time__gt=source_slot.start_time,
                ).exclude(
                    assignment=target_assignment,
                    start_time=source_slot.start_time,
                    end_time=source_slot.end_time,
                )

                class_conflict = overlapping.filter(
                    assignment__classroom=target_classroom
                ).exists()
                teacher_conflict = overlapping.filter(
                    assignment__teacher=target_assignment.teacher
                ).exists()
                room_conflict = False
                if room_value.strip():
                    room_conflict = overlapping.exclude(room__exact="").filter(
                        room__iexact=room_value.strip()
                    ).exists()

                if class_conflict or teacher_conflict or room_conflict:
                    summary["skipped_conflicts"] += 1
                    if len(summary["conflicts"]) < 20:
                        labels = []
                        if class_conflict:
                            labels.append("classe")
                        if teacher_conflict:
                            labels.append("enseignant")
                        if room_conflict:
                            labels.append("salle")
                        summary["conflicts"].append(
                            {
                                "day": source_slot.day_of_week,
                                "time": self._slot_range_label(source_slot),
                                "subject_code": target_assignment.subject.code,
                                "teacher_code": target_assignment.teacher.employee_code,
                                "types": labels,
                            }
                        )
                    continue

                duplicated, created = TeacherScheduleSlot.objects.update_or_create(
                    assignment=target_assignment,
                    day_of_week=source_slot.day_of_week,
                    start_time=source_slot.start_time,
                    end_time=source_slot.end_time,
                    defaults={"room": room_value},
                )

                if created:
                    summary["created"] += 1
                else:
                    summary["updated"] += 1

                if len(summary["mapping_examples"]) < 20:
                    summary["mapping_examples"].append(
                        {
                            "day": duplicated.day_of_week,
                            "time": self._slot_range_label(duplicated),
                            "subject_code": duplicated.assignment.subject.code,
                            "teacher_code": duplicated.assignment.teacher.employee_code,
                            "mode": mapping_mode,
                        }
                    )

        return Response(summary)

    @action(detail=False, methods=["get"], permission_classes=[permissions.IsAuthenticated])
    def export_excel(self, request):
        classroom_id = self._parse_classroom_id(request)

        if classroom_id:
            classrooms = list(self._scoped_classroom_queryset().filter(id=classroom_id).order_by("name"))
            if not classrooms:
                return Response({"detail": "Classe introuvable."}, status=404)
            filename = f"planning_classe_{classroom_id}.xlsx"
        else:
            classrooms = list(self._scoped_classroom_queryset().order_by("name"))
            filename = "planning_global_multi_classes.xlsx"

        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        workload_rows = self._teacher_workload_rows(self.get_queryset())
        ws_load = wb.create_sheet("Charge Enseignants")
        ws_load.append(
            [
                "Code enseignant",
                "Nom",
                "Horaires",
                "Classes",
                "Lundi",
                "Mardi",
                "Mercredi",
                "Jeudi",
                "Vendredi",
                "Samedi",
                "Total (h)",
                "Niveau",
            ]
        )
        for row in workload_rows:
            ws_load.append(
                [
                    row["teacher_code"],
                    row["teacher_name"],
                    row["slot_count"],
                    row["class_count"],
                    row["per_day_hours"]["MON"],
                    row["per_day_hours"]["TUE"],
                    row["per_day_hours"]["WED"],
                    row["per_day_hours"]["THU"],
                    row["per_day_hours"]["FRI"],
                    row["per_day_hours"]["SAT"],
                    row["total_hours"],
                    row["load_level"],
                ]
            )

        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        for row in ws_load.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for classroom in classrooms:
            ws = wb.create_sheet(self._sheet_title(classroom.name))
            etab_meta = self._etablissement_meta_lines(classroom)
            ws.merge_cells("A1:G1")
            ws["A1"] = f"{etab_meta['name']} - Emploi du temps"
            ws["A1"].font = Font(size=13, bold=True)
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

            ws.merge_cells("A2:G2")
            ws["A2"] = f"Classe: {classroom.name}"
            ws["A2"].font = Font(size=11, bold=True)
            ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

            ws.merge_cells("A3:G3")
            generated_label = f"Genere le {timezone.localtime().strftime('%d/%m/%Y %H:%M')}"
            ws["A3"] = (
                f"{etab_meta['details']} | {generated_label}"
                if etab_meta["details"]
                else generated_label
            )
            ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[1].height = 24
            ws.row_dimensions[2].height = 20
            ws.row_dimensions[3].height = 18

            if XLImage and etab_meta.get("logo_path"):
                try:
                    logo = XLImage(etab_meta["logo_path"])
                    logo.width = 46
                    logo.height = 46
                    ws.add_image(logo, "G1")
                except Exception:
                    pass

            headers = ["Horaire", "Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi"]
            ws.append(headers)
            for index, header in enumerate(headers, start=1):
                cell = ws.cell(row=4, column=index)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            class_slots = list(
                self._class_slots_queryset(classroom)
                .order_by("day_of_week", "start_time", "end_time", "id")
            )
            matrix = self._build_class_matrix(class_slots)
            ranges = self._sorted_ranges(matrix)

            if not ranges:
                ws.append(["Aucun horaire planifié", "", "", "", "", "", ""])
            else:
                for range_label in ranges:
                    day_map = matrix[range_label]
                    ws.append(
                        [
                            range_label,
                            " | ".join(day_map["MON"]),
                            " | ".join(day_map["TUE"]),
                            " | ".join(day_map["WED"]),
                            " | ".join(day_map["THU"]),
                            " | ".join(day_map["FRI"]),
                            " | ".join(day_map["SAT"]),
                        ]
                    )

            ws.column_dimensions["A"].width = 18
            for col in ["B", "C", "D", "E", "F", "G"]:
                ws.column_dimensions[col].width = 28

            for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=7):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

        stream = BytesIO()
        wb.save(stream)
        response = HttpResponse(
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Cache-Control"] = "no-store, max-age=0"
        return response

    @action(detail=False, methods=["get"], permission_classes=[permissions.IsAuthenticated])
    def export_pdf(self, request):
        classroom_id = self._parse_classroom_id(request)

        if classroom_id:
            classrooms = list(self._scoped_classroom_queryset().filter(id=classroom_id).order_by("name"))
            if not classrooms:
                return Response({"detail": "Classe introuvable."}, status=404)
            filename = f"planning_classe_{classroom_id}.pdf"
        else:
            classrooms = list(self._scoped_classroom_queryset().order_by("name"))
            filename = "planning_global_multi_classes.pdf"

        pdf = FPDF(orientation="L", unit="mm", format="A4")
        pdf.set_auto_page_break(auto=True, margin=12)

        col_widths = [28, 41, 41, 41, 41, 41, 41]

        def draw_headers():
            pdf.set_font("Helvetica", "B", 9)
            for header, width in zip(
                ["Horaire", "Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi"],
                col_widths,
            ):
                pdf.cell(width, 8, self._pdf_text(header), border=1, align="C")
            pdf.ln(8)

        for classroom in classrooms:
            class_slots = list(
                self._class_slots_queryset(classroom)
                .order_by("day_of_week", "start_time", "end_time", "id")
            )
            matrix = self._build_class_matrix(class_slots)
            ranges = self._sorted_ranges(matrix)
            etab_meta = self._etablissement_meta_lines(classroom)

            pdf.add_page()
            title_x = 10
            logo_path = etab_meta.get("logo_path")
            if logo_path:
                try:
                    pdf.image(logo_path, x=10, y=8, w=15)
                    title_x = 28
                except Exception:
                    title_x = 10

            pdf.set_font("Helvetica", "B", 14)
            pdf.set_x(title_x)
            pdf.cell(
                0,
                8,
                self._pdf_text(f"{etab_meta['name']} - Emploi du temps"),
                ln=1,
            )
            pdf.set_font("Helvetica", "B", 11)
            pdf.set_x(title_x)
            pdf.cell(
                0,
                6,
                self._pdf_text(f"Classe: {classroom.name}"),
                ln=1,
            )
            pdf.set_font("Helvetica", "", 9)
            if etab_meta["details"]:
                pdf.set_x(title_x)
                pdf.cell(0, 6, self._pdf_text(etab_meta["details"]), ln=1)
            pdf.set_x(title_x)
            pdf.cell(
                0,
                6,
                self._pdf_text(f"Généré le {timezone.localtime().strftime('%d/%m/%Y %H:%M')}"),
                ln=1,
            )
            pdf.ln(2)

            draw_headers()

            if not ranges:
                pdf.set_font("Helvetica", "", 10)
                pdf.cell(0, 8, self._pdf_text("Aucun horaire planifié"), ln=1)
                continue

            for range_label in ranges:
                if pdf.get_y() > 185:
                    pdf.add_page()
                    draw_headers()

                day_map = matrix[range_label]
                values = [
                    range_label,
                    " | ".join(day_map["MON"]),
                    " | ".join(day_map["TUE"]),
                    " | ".join(day_map["WED"]),
                    " | ".join(day_map["THU"]),
                    " | ".join(day_map["FRI"]),
                    " | ".join(day_map["SAT"]),
                ]

                pdf.set_font("Helvetica", "", 8)
                for value, width in zip(values, col_widths):
                    text = self._pdf_text(value)
                    if len(text) > 65:
                        text = f"{text[:62]}..."
                    pdf.cell(width, 8, text, border=1)
                pdf.ln(8)

        response = HttpResponse(bytes(pdf.output()), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Cache-Control"] = "no-store, max-age=0"
        return response


class TimetablePublicationViewSet(EtablissementScopeMixin, viewsets.ReadOnlyModelViewSet):
    access_module = "timetable"
    queryset = TimetablePublication.objects.select_related("classroom", "published_by").all().order_by("classroom__name")
    serializer_class = TimetablePublicationSerializer
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]
    filterset_fields = ["classroom", "is_published", "is_locked"]





    def get_queryset(self):
        user = self.request.user
        qs = super().get_queryset()
        requested_etablissement = self._requested_etablissement()

        if requested_etablissement is not None:
            return qs.filter(classroom__etablissement=requested_etablissement)

        if self._has_requested_scope():
            return qs.none()

        if getattr(user, "role", None) == "super_admin":
            return qs

        return qs.filter(classroom__etablissement=getattr(user, "etablissement", None))


class ParentProfileViewSet(BaseModelViewSet):
    access_module = "students"
    queryset = ParentProfile.objects.all().order_by("id")
    serializer_class = ParentProfileSerializer





    def _backfill_missing_parent_profiles(self):
        User = get_user_model()
        user = self.request.user
        requested_etablissement = self._requested_etablissement()

        parent_users = User.objects.filter(role=UserRole.PARENT)

        if requested_etablissement is not None:
            parent_users = parent_users.filter(etablissement=requested_etablissement)
        elif self._has_requested_scope():
            parent_users = User.objects.none()
        elif getattr(user, "role", None) == UserRole.SUPER_ADMIN:
            active_etablissement = getattr(user, "etablissement", None)
            if active_etablissement is not None:
                parent_users = parent_users.filter(etablissement=active_etablissement)
            else:
                parent_users = User.objects.none()
        else:
            user_etablissement = getattr(user, "etablissement", None)
            if user_etablissement is not None:
                parent_users = parent_users.filter(etablissement=user_etablissement)
            else:
                parent_users = User.objects.none()

        missing_users = list(
            parent_users.filter(parent_profile__isnull=True).select_related("etablissement")
        )
        if not missing_users:
            return

        ParentProfile.objects.bulk_create(
            [
                ParentProfile(
                    user=parent_user,
                    etablissement=parent_user.etablissement,
                )
                for parent_user in missing_users
            ],
            ignore_conflicts=True,
        )

    def get_queryset(self):
        self._backfill_missing_parent_profiles()
        user = self.request.user
        qs = ParentProfile.objects.select_related("user")

        requested_etablissement = self._requested_etablissement()
        if requested_etablissement is not None:
            return qs.filter(etablissement=requested_etablissement)

        if self._has_requested_scope():
            return qs.none()

        if hasattr(user, "role") and user.role == "super_admin":
            user_etablissement = getattr(user, "etablissement", None)
            if user_etablissement is not None:
                return qs.filter(etablissement=user_etablissement)
            return qs.none()

        user_etablissement = getattr(user, "etablissement", None)
        if user_etablissement is None:
            return qs.none()
        return qs.filter(etablissement=user_etablissement)

    def perform_create(self, serializer):
        user = self.request.user
        target = self._requested_etablissement() if getattr(user, "role", None) == UserRole.SUPER_ADMIN else getattr(user, "etablissement", None)
        if getattr(user, "role", None) == UserRole.SUPER_ADMIN and target is None:
            raise ValidationError({"etablissement": "Selectionnez un etablissement actif."})
        serializer.save(etablissement=target)

    def perform_update(self, serializer):
        user = self.request.user
        target = self._requested_etablissement() if getattr(user, "role", None) == UserRole.SUPER_ADMIN else getattr(user, "etablissement", None)
        if getattr(user, "role", None) == UserRole.SUPER_ADMIN and target is None:
            raise ValidationError({"etablissement": "Selectionnez un etablissement actif."})
        serializer.save(etablissement=target)


class StudentViewSet(BaseModelViewSet):
    access_module = "students"
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]
    parser_classes = (MultiPartParser, FormParser, JSONParser)
    pagination_class = StandardResultsSetPagination
    filterset_fields = [
        "classroom",
        "is_archived",
        "parent",
        "user",
        "etablissement",
        "enrollment_date",
        "created_at",
        "gender",
    ]
    search_fields = [
        "matricule",
        "user__first_name",
        "user__last_name",
        "user__username",
        "classroom__name",
        "parent__user__first_name",
        "parent__user__last_name",
        # Un parent qui appelle donne son numero, pas le matricule de
        # son enfant. Le sien comme celui de l'eleve.
        "user__phone",
        "parent__user__phone",
    ]
    ordering_fields = [
        "created_at",
        "matricule",
        "enrollment_date",
        "user__last_name",
        "user__first_name",
        "classroom__name",
        "birth_date",
        "gender",
    ]
    # "-id" en second: sur des lignes creees dans la meme seconde,
    # "-created_at" seul laisse l_ordre indefini et une meme ligne peut
    # apparaitre sur deux pages, ou sur aucune.
    ordering = ["-created_at", "-id"]

    @action(detail=False, methods=["get"])
    def stats(self, request):
        """Effectifs de l'etablissement, independants de la pagination.

        Le client comptait sur la page recue: a 15 lignes par page, "3 actifs"
        decrivait la page et non l'ecole. Les filtres de liste ne sont pas
        appliques non plus, volontairement: l'en-tete decrit l'etablissement,
        le tableau decrit ce qu'on regarde. Melanger les deux etait la source
        de la confusion.
        """
        queryset = self.get_queryset()

        # L'annee de l'etablissement consulte, et non la premiere active
        # venue: depuis que chaque ecole a les siennes, `filter(is_active=True)`
        # sans portee rendait l'annee d'une autre.
        active_year = AcademicYear.courante(self._resolve_target_etablissement())
        if active_year is not None:
            enrolled_this_year = Q(
                enrollment_date__gte=active_year.start_date,
                enrollment_date__lte=active_year.end_date,
            )
        else:
            # Sans annee active declaree, "nouveau" n'a pas de sens: mieux
            # vaut 0 qu'un chiffre calcule sur une periode inventee.
            enrolled_this_year = Q(pk__in=[])

        totals = queryset.aggregate(
            total=Count("id"),
            active=Count("id", filter=Q(is_archived=False)),
            archived=Count("id", filter=Q(is_archived=True)),
            new_this_year=Count("id", filter=Q(is_archived=False) & enrolled_this_year),
            # Le champ tolere NULL et chaine vide: ne compter que l'un des deux
            # sous-estimerait silencieusement les fiches a completer.
            gender_missing=Count(
                "id",
                filter=Q(is_archived=False) & (Q(gender__isnull=True) | Q(gender="")),
            ),
        )
        totals["academic_year"] = active_year.name if active_year else ""
        return Response(totals)

    # Un eleve de terminale porte des centaines de notes et d'absences: on
    # renvoie le total exact mais seulement les elements recents. L'ecran de
    # consultation ne remplace pas les modules specialises.
    DOSSIER_ITEM_LIMIT = 50

    def _dossier_section(
        self,
        *,
        key,
        label,
        module,
        role,
        queryset,
        serializer_class,
        aggregates=None,
        count_queryset=None,
        labeller=None,
    ):
        """Une section du dossier, ou son refus motive.

        Une section interdite est renvoyee vide plutot qu'omise: sans elle,
        un directeur sans acces discipline lirait "aucun incident" la ou il
        faut lire "vous n'avez pas le droit de savoir".

        `labeller` ajoute a chaque element un dictionnaire "labels". Plusieurs
        serializers partages exposent des cles etrangeres brutes (`subject: 7`)
        et l'ecran afficherait "Matiere 7". Les libelles sont lus sur les
        objets deja charges par select_related, donc sans requete de plus, et
        localement plutot qu'en modifiant des serializers dont dependent tous
        les autres ecrans.
        """
        if not can_read(role, module):
            return {
                "key": key,
                "label": label,
                "module": module,
                "granted": False,
            }

        # count_queryset existe pour les sections annotees par une jointure:
        # agreger sur la meme requete multiplierait les montants par le nombre
        # de lignes jointes.
        stats = (count_queryset if count_queryset is not None else queryset).aggregate(
            count=Count("id"), **(aggregates or {})
        )
        total = stats.pop("count", 0) or 0
        items = list(queryset[: self.DOSSIER_ITEM_LIMIT])
        rows = serializer_class(items, many=True).data

        if labeller is not None:
            for row, instance in zip(rows, items):
                row["labels"] = labeller(instance)

        return {
            "key": key,
            "label": label,
            "module": module,
            "granted": True,
            "count": total,
            "summary": stats,
            "items": rows,
            "has_more": total > len(items),
        }

    @staticmethod
    def _name_of(obj, attribut, champ="name"):
        """Libelle d'une relation, chaine vide si elle est absente."""
        related = getattr(obj, attribut, None)
        return getattr(related, champ, "") if related is not None else ""

    @action(detail=True, methods=["get"], url_path="dossier")
    def dossier(self, request, pk=None):
        """Tout ce que l'etablissement sait d'un eleve, en un appel.

        L'ecran "Recherche eleve" affichait sinon onze listes obtenues par
        onze requetes HTTP. Le cloisonnement par etablissement et par role est
        celui de get_object(): un parent qui devine l'identifiant d'un autre
        eleve recoit un 404, sans controle supplementaire ici.
        """
        student = self.get_object()
        role = getattr(request.user, "role", "")

        # Le dossier suit l'annee consultee. Sans ce filtre, il affichait
        # toutes les annees melangees: les absences de sixieme au milieu de
        # celles de cinquieme, sans rien pour les distinguer une fois
        # l'eleve passe en classe superieure.
        annee = self._requested_academic_year()

        def _par_annee(queryset):
            if annee is None:
                return queryset
            return queryset.filter(academic_year=annee)

        fees = StudentFeeViewSet._with_financial_annotations(
            _par_annee(StudentFee.objects.filter(student=student))
            .select_related("academic_year", "student", "student__user", "student__classroom")
            .order_by("-due_date", "-id")
        )

        sections = [
            self._dossier_section(
                key="history",
                label="Historique académique",
                module="students",
                role=role,
                queryset=_par_annee(StudentAcademicHistory.objects.filter(student=student))
                .select_related("academic_year", "classroom")
                .order_by("-academic_year__start_date", "-id"),
                serializer_class=StudentAcademicHistorySerializer,
                aggregates={"moyenne": Avg("average")},
                labeller=lambda obj: {
                    "annee": self._name_of(obj, "academic_year"),
                    "classe": self._name_of(obj, "classroom"),
                },
            ),
            self._dossier_section(
                key="promotion",
                label="Décisions de passage",
                module="promotion",
                role=role,
                queryset=PromotionDecision.objects.filter(student=student)
                .select_related(
                    "run",
                    "run__source_academic_year",
                    "source_classroom",
                    "target_classroom",
                )
                .order_by("-created_at", "-id"),
                serializer_class=PromotionDecisionSerializer,
                labeller=lambda obj: {
                    "decision": obj.get_decision_display(),
                    "annee": self._name_of(
                        obj.run, "source_academic_year"
                    )
                    if obj.run
                    else "",
                },
            ),
            self._dossier_section(
                key="grades",
                label="Notes",
                module="grades",
                role=role,
                queryset=_par_annee(Grade.objects.filter(student=student))
                .select_related("subject", "classroom", "academic_year")
                .order_by("-academic_year__start_date", "-term", "-id"),
                serializer_class=GradeSerializer,
                aggregates={"moyenne": Avg("value")},
                labeller=lambda obj: {
                    "matiere": self._name_of(obj, "subject"),
                    "classe": self._name_of(obj, "classroom"),
                    "annee": self._name_of(obj, "academic_year"),
                },
            ),
            self._dossier_section(
                key="attendance",
                label="Absences & retards",
                module="attendance",
                role=role,
                queryset=_par_annee(
                    Attendance.objects.filter(student=student)
                ).order_by("-date", "-id"),
                serializer_class=AttendanceSerializer,
                aggregates={
                    "absences": Count("id", filter=Q(is_absent=True)),
                    "retards": Count("id", filter=Q(is_late=True)),
                },
            ),
            self._dossier_section(
                key="discipline",
                label="Discipline",
                module="discipline",
                role=role,
                queryset=_par_annee(DisciplineIncident.objects.filter(student=student))
                .select_related("reported_by")
                .order_by("-incident_date", "-id"),
                serializer_class=DisciplineIncidentSerializer,
                aggregates={
                    "ouverts": Count("id", filter=Q(status=DisciplineStatus.OPEN)),
                },
            ),
            self._dossier_section(
                key="fees",
                label="Frais",
                module="finance",
                role=role,
                queryset=fees,
                serializer_class=StudentFeeSerializer,
                # Agrege sur une requete sans la jointure paiements, sinon
                # amount_due est compte une fois par paiement.
                count_queryset=_par_annee(StudentFee.objects.filter(student=student)),
                aggregates={"total_du": Sum("amount_due")},
                labeller=lambda obj: {
                    "annee": self._name_of(obj, "academic_year"),
                    "type": obj.get_fee_type_display(),
                },
            ),
            self._dossier_section(
                key="payments",
                label="Paiements",
                module="finance",
                role=role,
                queryset=Payment.objects.filter(fee__student=student)
                .select_related(
                    "fee",
                    "fee__student",
                    "fee__student__user",
                    # PaymentSerializer.get_classroom_name descend jusqu'a la
                    # classe: sans elle, c'est une requete par paiement.
                    "fee__student__classroom",
                    "received_by",
                )
                .order_by("-created_at", "-id"),
                serializer_class=PaymentSerializer,
                aggregates={
                    "total_encaisse": Sum("amount", filter=Q(is_cancelled=False)),
                },
            ),
            self._dossier_section(
                key="exams",
                label="Résultats d'examens",
                module="exams",
                role=role,
                queryset=ExamResult.objects.filter(student=student)
                .select_related("session", "subject")
                .order_by("-session__start_date", "-id"),
                serializer_class=ExamResultSerializer,
                aggregates={"moyenne": Avg("score")},
                labeller=lambda obj: {
                    "matiere": self._name_of(obj, "subject"),
                    "session": self._name_of(obj, "session", champ="title"),
                },
            ),
            self._dossier_section(
                key="library",
                label="Bibliothèque",
                module="library",
                role=role,
                queryset=Borrow.objects.filter(student=student)
                .select_related("book")
                .order_by("-borrowed_at", "-id"),
                serializer_class=BorrowSerializer,
                aggregates={
                    "en_cours": Count("id", filter=Q(returned_at__isnull=True)),
                },
                labeller=lambda obj: {
                    "livre": self._name_of(obj, "book", champ="title"),
                    "auteur": self._name_of(obj, "book", champ="author"),
                },
            ),
            self._dossier_section(
                key="canteen_subscriptions",
                label="Cantine - abonnements",
                module="canteen",
                role=role,
                queryset=CanteenSubscription.objects.filter(student=student)
                .select_related("academic_year")
                .order_by("-start_date", "-id"),
                serializer_class=CanteenSubscriptionSerializer,
                labeller=lambda obj: {
                    "annee": self._name_of(obj, "academic_year"),
                    "statut": obj.get_status_display(),
                },
            ),
            self._dossier_section(
                key="canteen_services",
                label="Cantine - repas servis",
                module="canteen",
                role=role,
                queryset=CanteenService.objects.filter(student=student)
                .select_related("menu")
                .order_by("-served_on", "-id"),
                serializer_class=CanteenServiceSerializer,
                aggregates={"impayes": Count("id", filter=Q(is_paid=False))},
                labeller=lambda obj: {"menu": self._name_of(obj, "menu")},
            ),
        ]

        return Response(
            {
                "student": StudentSerializer(student, context={"request": request}).data,
                "sections": sections,
            }
        )

    @action(detail=False, methods=["post"], url_path="bulk-update")
    @transaction.atomic
    def bulk_update(self, request):
        """Applique une meme modification a plusieurs eleves.

        Une secretaire archive une promotion sortante ou deplace une classe
        entiere: en appels unitaires, c'est autant d'allers-retours reseau, et
        une coupure au milieu laisse la moitie du travail faite.

        Les identifiants hors du perimetre de l'utilisateur ne sont pas
        silencieusement ignores: la demande entiere est refusee. Un rapport
        partiel laisserait croire au succes d'une operation incomplete.
        """
        raw_ids = request.data.get("ids")
        if not isinstance(raw_ids, list) or not raw_ids:
            return Response(
                {"ids": "Fournissez au moins un identifiant d'eleve."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            ids = {int(value) for value in raw_ids}
        except (TypeError, ValueError):
            return Response(
                {"ids": "Identifiants invalides."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        scoped = self.get_queryset().filter(id__in=ids)
        autorises = set(scoped.values_list("id", flat=True))
        refuses = ids - autorises
        if refuses:
            return Response(
                {
                    "ids": (
                        f"{len(refuses)} eleve(s) hors de votre perimetre. "
                        "Aucune modification n'a ete appliquee."
                    )
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        changements = {}

        if "is_archived" in request.data:
            changements["is_archived"] = bool(request.data.get("is_archived"))

        if "classroom" in request.data:
            classroom_id = request.data.get("classroom")
            if classroom_id in (None, ""):
                changements["classroom"] = None
            else:
                try:
                    classroom = ClassRoom.objects.get(pk=int(classroom_id))
                except (TypeError, ValueError, ClassRoom.DoesNotExist):
                    return Response(
                        {"classroom": "Classe introuvable."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                # Sans ce controle, un identifiant de classe devine permettrait
                # de deplacer un eleve dans un autre etablissement.
                etablissements = {
                    value
                    for value in scoped.values_list("etablissement_id", flat=True)
                    if value is not None
                }
                if etablissements and classroom.etablissement_id not in etablissements:
                    return Response(
                        {"classroom": "Cette classe appartient a un autre etablissement."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                changements["classroom"] = classroom

        if not changements:
            return Response(
                {"detail": "Aucune modification demandee."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        modifies = scoped.update(**changements)
        return Response({"updated": modifies, "ids": sorted(autorises)})





    def _resolve_target_etablissement(self):
        user = self.request.user
        requested_etablissement = self._requested_etablissement()

        if getattr(user, "role", None) == "super_admin":
            return requested_etablissement

        return getattr(user, "etablissement", None)

    def _validate_student_scope(self, serializer, instance=None):
        classroom = serializer.validated_data.get("classroom") or (instance.classroom if instance else None)
        parent = serializer.validated_data.get("parent") or (instance.parent if instance else None)
        target_etablissement = self._resolve_target_etablissement()

        if target_etablissement is None and classroom is not None:
            target_etablissement = classroom.etablissement

        if getattr(self.request.user, "role", None) == UserRole.SUPER_ADMIN and target_etablissement is None:
            raise ValidationError({"etablissement": "Selectionnez un etablissement actif."})

        if target_etablissement is None:
            return None

        if classroom and classroom.etablissement_id != target_etablissement.id:
            raise ValidationError({"classroom": "La classe n'appartient pas a l'etablissement actif."})

        if parent and parent.etablissement_id != target_etablissement.id:
            raise ValidationError({"parent": "Le parent n'appartient pas a l'etablissement actif."})

        user_obj = serializer.validated_data.get("user") or (instance.user if instance else None)
        if user_obj and user_obj.etablissement_id not in (None, target_etablissement.id):
            raise ValidationError({"user": "Le compte utilisateur n'appartient pas a l'etablissement actif."})

        if user_obj and user_obj.etablissement_id is None:
            user_obj.etablissement = target_etablissement
            user_obj.save(update_fields=["etablissement"])

        return target_etablissement

    def get_queryset(self):
        user = self.request.user
        qs = Student.objects.select_related("user", "classroom", "parent", "parent__user")
        role = getattr(user, "role", "")
        if role == UserRole.STUDENT:
            return qs.filter(user_id=user.id)
        if role == UserRole.PARENT:
            return qs.filter(parent__user_id=user.id)

        requested_etablissement = self._requested_etablissement()
        if requested_etablissement is not None:
            return qs.filter(
                Q(classroom__etablissement=requested_etablissement)
                | Q(classroom__isnull=True, etablissement=requested_etablissement)
            )

        if self._has_requested_scope():
            return qs.none()

        if hasattr(user, "role") and user.role == "super_admin":
            return qs.none()

        user_etablissement = getattr(user, "etablissement", None)
        if user_etablissement is None:
            return qs.none()
        return qs.filter(
            Q(classroom__etablissement=user_etablissement)
            | Q(classroom__isnull=True, etablissement=user_etablissement)
        )

    def perform_create(self, serializer):
        target_etablissement = self._validate_student_scope(serializer)
        serializer.save(etablissement=target_etablissement)

    def perform_update(self, serializer):
        target_etablissement = self._validate_student_scope(serializer, instance=serializer.instance)
        serializer.save(etablissement=target_etablissement)

    def _scoped_classroom_queryset(self):
        user = self.request.user
        requested = self._requested_etablissement()
        qs = ClassRoom.objects.select_related("etablissement", "academic_year")

        if requested is not None:
            return qs.filter(etablissement=requested)
        if self._has_requested_scope():
            return qs.none()
        if getattr(user, "role", None) == UserRole.SUPER_ADMIN:
            active_etablissement = getattr(user, "etablissement", None)
            if active_etablissement is not None:
                return qs.filter(etablissement=active_etablissement)
            return qs.none()
        return qs.filter(etablissement=getattr(user, "etablissement", None))

    @staticmethod
    def _genre_importe(valeur):
        """« M », « F », « garcon », « fille »: ce que les fichiers ecrivent.

        Une colonne libre remplie par une secretaire ne contient pas toujours
        la lettre attendue; refuser « Masculin » parce qu'on esperait « M »
        renverrait tout un fichier pour rien.
        """
        brut = str(valeur or "").strip().lower()
        if not brut:
            return ""
        if brut[0] in ("m", "g", "h"):  # masculin, garcon, homme
            return "M"
        if brut[0] == "f":  # feminin, fille
            return "F"
        return ""

    @staticmethod
    def _default_student_username(matricule, first_name, last_name):
        base = _as_text(matricule).lower().replace(" ", "")
        if base:
            return base
        combined = f"{_as_text(first_name)}.{_as_text(last_name)}".strip(".")
        combined = combined.lower().replace(" ", "")
        return combined or f"student_{int(timezone.now().timestamp())}"

    @action(
        detail=False,
        methods=["get"],
        url_path="import-templates/download",
        permission_classes=[permissions.IsAuthenticated],
    )
    def import_template_download(self, request):
        return _build_import_template_download_response(
            request.query_params.get("type") or request.query_params.get("import_type"),
            request.query_params.get("format"),
        )

    @action(detail=False, methods=["post"], url_path="import-by-class")
    def import_by_class(self, request):
        classroom_id = request.data.get("classroom_id")
        confirm = str(request.data.get("confirm", "false")).strip().lower() in {"1", "true", "yes", "on"}
        if classroom_id in (None, ""):
            raise ValidationError({"classroom_id": "Classe requise."})

        try:
            classroom_id = int(classroom_id)
        except (TypeError, ValueError):
            raise ValidationError({"classroom_id": "Classe invalide."})

        classroom = get_object_or_404(self._scoped_classroom_queryset(), id=classroom_id)
        rows = _load_import_rows(request.FILES.get("file") or request.data.get("file"))
        if not rows:
            raise ValidationError({"file": "Aucune ligne exploitable dans le fichier."})

        User = get_user_model()
        row_errors = []
        prepared = []
        seen_matricules = set()

        for index, row in enumerate(rows, start=2):
            matricule = _as_text(row.get("matricule"))
            first_name = _as_text(row.get("first_name") or row.get("prenom"))
            last_name = _as_text(row.get("last_name") or row.get("nom"))
            username = _as_text(row.get("username"))
            email = _as_text(row.get("email"))
            phone = _as_text(row.get("phone") or row.get("telephone"))
            birth_date = _as_date(row.get("birth_date") or row.get("date_naissance"))
            gender = self._genre_importe(
                row.get("gender") or row.get("genre") or row.get("sexe")
            )

            if not matricule:
                row_errors.append({"row": index, "error": "Matricule obligatoire."})
                continue
            if not first_name or not last_name:
                row_errors.append({"row": index, "error": "Prenom et nom obligatoires."})
                continue
            if not gender:
                # L'import creait des eleves sans genre, et c'est de la que
                # venaient les matricules « GS-2025-00001 »: le format en a
                # besoin pour sa derniere lettre.
                row_errors.append(
                    {
                        "row": index,
                        "error": "Genre obligatoire (M ou F, colonne « genre »).",
                    }
                )
                continue
            if matricule in seen_matricules:
                row_errors.append({"row": index, "error": "Matricule dupliqué dans le fichier."})
                continue
            seen_matricules.add(matricule)

            if not username:
                username = self._default_student_username(matricule, first_name, last_name)

            student = Student.objects.select_related("user").filter(matricule=matricule).first()
            prepared.append(
                {
                    "row": index,
                    "matricule": matricule,
                    "gender": gender,
                    "first_name": first_name,
                    "last_name": last_name,
                    "username": username,
                    "email": email,
                    "phone": phone,
                    "birth_date": birth_date,
                    "existing_student": student,
                }
            )

        create_count = 0
        update_count = 0
        preview_rows = []
        for item in prepared:
            student = item["existing_student"]
            action_label = "create" if student is None else "update"
            if action_label == "create":
                create_count += 1
            else:
                update_count += 1
            preview_rows.append(
                {
                    "row": item["row"],
                    "action": action_label,
                    "matricule": item["matricule"],
                    "full_name": f"{item['first_name']} {item['last_name']}",
                }
            )

        preview_payload = {
            "classroom": {"id": classroom.id, "name": classroom.name},
            "summary": {
                "total_rows": len(rows),
                "valid_rows": len(prepared),
                "errors": len(row_errors),
                "to_create": create_count,
                "to_update": update_count,
            },
            "errors": row_errors,
            "preview": preview_rows[:120],
            "confirm_required": True,
        }

        if not confirm:
            return Response(preview_payload)

        if row_errors:
            return Response(
                {
                    **preview_payload,
                    "detail": "Import bloqué: corrigez les erreurs avant confirmation.",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        created = 0
        updated = 0
        with transaction.atomic():
            for item in prepared:
                student = item["existing_student"]
                username = item["username"]
                user = None

                if student is None:
                    user = User.objects.filter(username=username).first()
                    if user and user.role != UserRole.STUDENT:
                        raise ValidationError(
                            {
                                "detail": (
                                    f"Conflit username '{username}' (rôle {user.role}). "
                                    "Utilisez un username différent dans le fichier."
                                )
                            }
                        )
                    if user is None:
                        user = User.objects.create_user(
                            username=username,
                            first_name=item["first_name"],
                            last_name=item["last_name"],
                            email=item["email"],
                            role=UserRole.STUDENT,
                            phone=item["phone"],
                            etablissement=classroom.etablissement,
                            password="Password@123",
                        )
                    else:
                        user.first_name = item["first_name"]
                        user.last_name = item["last_name"]
                        user.email = item["email"]
                        user.phone = item["phone"]
                        user.etablissement = classroom.etablissement
                        user.save(update_fields=["first_name", "last_name", "email", "phone", "etablissement"])

                    Student.objects.create(
                        user=user,
                        matricule=item["matricule"],
                        classroom=classroom,
                        birth_date=item["birth_date"],
                        gender=item["gender"],
                        etablissement=classroom.etablissement,
                    )
                    created += 1
                else:
                    user = student.user
                    if user:
                        user.first_name = item["first_name"]
                        user.last_name = item["last_name"]
                        if item["email"]:
                            user.email = item["email"]
                        if item["phone"]:
                            user.phone = item["phone"]
                        if not user.etablissement_id:
                            user.etablissement = classroom.etablissement
                        user.save()

                    student.classroom = classroom
                    if item["birth_date"] is not None:
                        student.birth_date = item["birth_date"]
                    student.etablissement = classroom.etablissement
                    student.is_archived = False
                    student.save(update_fields=["classroom", "birth_date", "etablissement", "is_archived", "updated_at"])
                    updated += 1

        return Response(
            {
                **preview_payload,
                "result": {
                    "created": created,
                    "updated": updated,
                },
                "detail": "Import élèves terminé.",
            }
        )

    @action(
        detail=True,
        methods=["post", "patch"],
        url_path="upload-photo",
        parser_classes=[MultiPartParser, FormParser],
    )
    def upload_photo(self, request, pk=None):
        student = self.get_object()
        uploaded_photo = request.FILES.get("photo") or request.data.get("photo")
        if not uploaded_photo:
            return Response({"photo": ["Aucune image fournie."]}, status=400)

        serializer = self.get_serializer(
            student,
            data={"photo": uploaded_photo},
            partial=True,
        )
        serializer.is_valid(raise_exception=True)
        serializer.save(etablissement=student.etablissement)
        return Response(serializer.data)


class StudentAcademicHistoryViewSet(AnneeScolaireScopeMixin, BaseModelViewSet):
    access_module = "students"
    queryset = StudentAcademicHistory.objects.select_related("student", "academic_year", "classroom").all().order_by("-academic_year_id", "rank")
    serializer_class = StudentAcademicHistorySerializer
    filterset_fields = ["student", "academic_year", "classroom"]
    search_fields = [
        "student__matricule",
        "student__user__first_name",
        "student__user__last_name",
    ]
    ordering_fields = ["rank", "average", "academic_year__name"]





    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        role = getattr(user, "role", "")

        if role == UserRole.STUDENT:
            return queryset.filter(student__user_id=self.request.user.id)
        if role == UserRole.PARENT:
            return queryset.filter(student__parent__user_id=self.request.user.id)

        requested_etablissement = self._requested_etablissement()
        if requested_etablissement is not None:
            return queryset.filter(classroom__etablissement=requested_etablissement)

        if self._has_requested_scope():
            return queryset.none()

        if role == UserRole.SUPER_ADMIN:
            return queryset

        user_etablissement = getattr(user, "etablissement", None)
        if user_etablissement is None:
            return queryset.none()
        return queryset.filter(classroom__etablissement=user_etablissement)


class GradeViewSet(AnneeScolaireScopeMixin, BaseModelViewSet):
    access_module = "grades"
    queryset = Grade.objects.select_related("student", "subject", "classroom", "academic_year").all().order_by("-id")
    serializer_class = GradeSerializer
    pagination_class = GradePagination
    filterset_fields = ["classroom", "academic_year", "term", "subject", "student"]
    # On cherche une note par l'eleve ou la matiere, jamais par sa valeur.
    search_fields = [
        "student__matricule",
        "student__user__first_name",
        "student__user__last_name",
        "subject__name",
    ]
    ordering_fields = ["value", "term", "created_at", "student__user__last_name"]

    def _teacher_assignment_pairs(self):
        teacher_profile = self._teacher_profile()
        if not teacher_profile:
            return set()
        return set(
            TeacherAssignment.objects.filter(teacher=teacher_profile)
            .values_list("classroom_id", "subject_id")
        )





    def _resolve_target_etablissement(self):
        user = self.request.user
        requested_etablissement = self._requested_etablissement()

        if getattr(user, "role", None) == "super_admin" and requested_etablissement:
            return requested_etablissement

        return getattr(user, "etablissement", None)

    def _scoped_classroom_queryset(self):
        user = self.request.user
        requested = self._requested_etablissement()
        qs = ClassRoom.objects.select_related("etablissement")

        if requested is not None:
            return qs.filter(etablissement=requested)
        if self._has_requested_scope():
            return qs.none()
        if getattr(user, "role", None) == "super_admin":
            return qs
        return qs.filter(etablissement=getattr(user, "etablissement", None))

    def _get_scoped_classroom_or_404(self, classroom_id):
        return get_object_or_404(self._scoped_classroom_queryset(), id=classroom_id)

    UNASSIGNED_PAIR_MESSAGE = (
        "Acces refuse: vous ne pouvez saisir que les notes des matières "
        "et classes qui vous sont affectées."
    )

    def _reject_out_of_teacher_perimeter(self, data, instance=None):
        """Refus de perimetre avant toute autre validation.

        Sans ce controle prealable, le validateur d'unicite du serializer
        repondait en premier: un enseignant hors de ses classes apprenait
        qu'une note existait deja pour une classe qu'il n'enseigne pas, au
        lieu d'un simple refus d'acces.
        """
        if getattr(self.request.user, "role", "") != UserRole.TEACHER:
            return

        def resolve(field):
            raw = data.get(field)
            if raw in (None, ""):
                return getattr(instance, f"{field}_id", None) if instance else None
            try:
                return int(raw)
            except (TypeError, ValueError):
                return None

        classroom_id = resolve("classroom")
        subject_id = resolve("subject")
        # Champs absents ou illisibles: c'est au serializer de le dire.
        if classroom_id is None or subject_id is None:
            return

        if (classroom_id, subject_id) not in self._teacher_assignment_pairs():
            raise ValidationError({"subject": self.UNASSIGNED_PAIR_MESSAGE})

    def _validate_grade_scope(self, serializer, instance=None):
        student = serializer.validated_data.get("student") or (instance.student if instance else None)
        classroom = serializer.validated_data.get("classroom") or (instance.classroom if instance else None)
        subject = serializer.validated_data.get("subject") or (instance.subject if instance else None)

        if not student or not classroom:
            return

        role = getattr(self.request.user, "role", "")
        if role == UserRole.STUDENT and student.user_id != self.request.user.id:
            raise ValidationError({"student": "Vous ne pouvez saisir que vos propres notes."})
        if role == UserRole.PARENT:
            if not student.parent or student.parent.user_id != self.request.user.id:
                raise ValidationError({"student": "Vous ne pouvez saisir que les notes de vos enfants."})
        if role == UserRole.TEACHER:
            allowed_pairs = self._teacher_assignment_pairs()
            pair = (classroom.id if classroom else None, subject.id if subject else None)
            if pair not in allowed_pairs:
                raise ValidationError({"subject": self.UNASSIGNED_PAIR_MESSAGE})

        target_etablissement = self._resolve_target_etablissement()
        if target_etablissement is None:
            return

        student_etablissement_id = getattr(student, "etablissement_id", None)
        classroom_etablissement_id = getattr(getattr(student, "classroom", None), "etablissement_id", None)
        if student_etablissement_id != target_etablissement.id and classroom_etablissement_id != target_etablissement.id:
            raise ValidationError({"student": "L'eleve n'appartient pas a l'etablissement actif."})
        if classroom.etablissement_id != target_etablissement.id:
            raise ValidationError({"classroom": "La classe n'appartient pas a l'etablissement actif."})

    def get_queryset(self):
        queryset = super().get_queryset()
        role = getattr(self.request.user, "role", "")

        if role == UserRole.STUDENT:
            return queryset.filter(student__user_id=self.request.user.id)
        if role == UserRole.PARENT:
            return queryset.filter(student__parent__user_id=self.request.user.id)
        if role == UserRole.TEACHER:
            allowed_pairs = self._teacher_assignment_pairs()
            if not allowed_pairs:
                return queryset.none()
            pair_filter = Q()
            for classroom_id, subject_id in allowed_pairs:
                pair_filter |= Q(classroom_id=classroom_id, subject_id=subject_id)
            return queryset.filter(pair_filter)

        requested_etablissement = self._requested_etablissement()
        if requested_etablissement is not None:
            return queryset.filter(classroom__etablissement=requested_etablissement)

        if self._has_requested_scope():
            return queryset.none()

        if getattr(self.request.user, "role", None) == "super_admin":
            return queryset

        return queryset.filter(classroom__etablissement=getattr(self.request.user, "etablissement", None))

    @staticmethod
    def _parse_positive_int(value):
        try:
            parsed = int(value)
        except (TypeError, ValueError):
            return None
        return parsed if parsed > 0 else None

    @staticmethod
    def _normalize_term_or_none(value):
        return normalize_term(value)

    @staticmethod
    def _locked_term_message(prefix="Modification"):
        return f"Cette période est validée par la direction. {prefix} interdite."

    @staticmethod
    def _value_changed(old_value, new_value):
        if hasattr(old_value, "pk") and hasattr(new_value, "pk"):
            return old_value.pk != new_value.pk
        return old_value != new_value

    def _immutable_fields_changed(self, instance, validated_data):
        immutable_fields = ("student", "subject", "classroom", "academic_year", "term")
        changed = []
        for field in immutable_fields:
            if field not in validated_data:
                continue
            old_value = getattr(instance, field)
            new_value = validated_data[field]
            if self._value_changed(old_value, new_value):
                changed.append(field)
        return changed

    def _is_term_validated(self, classroom_id, academic_year_id, term):
        normalized_term = self._normalize_term_or_none(term)
        if not classroom_id or not academic_year_id or not normalized_term:
            return False
        return GradeValidation.objects.filter(
            classroom_id=classroom_id,
            academic_year_id=academic_year_id,
            term=normalized_term,
            is_validated=True,
        ).exists()

    def _close_term_workflow(self, *, classroom, academic_year, term, notes, user):
        grade_count = Grade.objects.filter(
            classroom=classroom,
            academic_year=academic_year,
            term=term,
        ).count()
        if grade_count == 0:
            raise ValidationError(
                {
                    "detail": (
                        "Aucune note de classe pour cette période. "
                        "Ajoutez les notes avant la clôture trimestrielle."
                    )
                }
            )

        with transaction.atomic():
            # The ranking used by promotion/history is refreshed immediately before locking the term.
            recalculate_term_ranking(classroom, academic_year, term)
            validation, _ = GradeValidation.objects.update_or_create(
                classroom=classroom,
                academic_year=academic_year,
                term=term,
                defaults={
                    "is_validated": True,
                    "validated_by": user,
                    "validated_at": timezone.now(),
                    "notes": notes,
                },
            )

        history_count = StudentAcademicHistory.objects.filter(
            classroom=classroom,
            academic_year=academic_year,
        ).count()
        return validation, history_count

    def create(self, request, *args, **kwargs):
        self._reject_out_of_teacher_perimeter(request.data)
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self._validate_grade_scope(serializer)
        classroom_id = serializer.validated_data.get("classroom").id if serializer.validated_data.get("classroom") else None
        academic_year_id = serializer.validated_data.get("academic_year").id if serializer.validated_data.get("academic_year") else None
        term = serializer.validated_data.get("term")
        if self._is_term_validated(classroom_id, academic_year_id, term):
            return Response({"detail": self._locked_term_message()}, status=400)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=201, headers=headers)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()

        self._reject_out_of_teacher_perimeter(request.data, instance=instance)

        if self._is_term_validated(instance.classroom_id, instance.academic_year_id, instance.term):
            return Response({"detail": self._locked_term_message()}, status=400)

        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self._validate_grade_scope(serializer, instance=instance)

        immutable_changed = self._immutable_fields_changed(instance, serializer.validated_data)
        if immutable_changed:
            return Response(
                {
                    "detail": (
                        "Les champs student, subject, classroom, academic_year et term "
                        "sont immuables après création."
                    ),
                    "fields": immutable_changed,
                },
                status=400,
            )

        immutable_changed = self._immutable_fields_changed(instance, serializer.validated_data)
        if immutable_changed:
            return Response(
                {
                    "detail": (
                        "Les champs student, subject, classroom, academic_year et term "
                        "sont immuables après création."
                    ),
                    "fields": immutable_changed,
                },
                status=400,
            )

        classroom = serializer.validated_data.get("classroom", instance.classroom)
        academic_year = serializer.validated_data.get("academic_year", instance.academic_year)
        term = serializer.validated_data.get("term", instance.term)

        if self._is_term_validated(classroom.id, academic_year.id, term):
            return Response({"detail": self._locked_term_message()}, status=400)

        self.perform_update(serializer)
        return Response(serializer.data)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if self._is_term_validated(instance.classroom_id, instance.academic_year_id, instance.term):
            return Response({"detail": self._locked_term_message(prefix="Suppression")}, status=400)
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=["post"], permission_classes=[permissions.IsAuthenticated, HasModuleAccess])
    def recalculate_ranking(self, request):
        classroom_id = self._parse_positive_int(request.data.get("classroom"))
        academic_year_id = self._parse_positive_int(request.data.get("academic_year"))
        term = self._normalize_term_or_none(request.data.get("term"))

        if not classroom_id or not academic_year_id or not term:
            return Response(
                {"detail": "classroom, academic_year et term (T1/T2/T3) sont requis."},
                status=400,
            )

        classroom = self._get_scoped_classroom_or_404(classroom_id)
        academic_year = get_object_or_404(AcademicYear, id=academic_year_id)

        if self._is_term_validated(classroom.id, academic_year.id, term):
            return Response({"detail": self._locked_term_message(prefix="Recalcul")}, status=400)

        recalculate_term_ranking(classroom, academic_year, term)
        return Response({"detail": "Classement recalculé avec succès."})

    @action(detail=False, methods=["post"], permission_classes=[permissions.IsAuthenticated, HasModuleAccess])
    def validate_term(self, request):
        classroom_id = self._parse_positive_int(request.data.get("classroom"))
        academic_year_id = self._parse_positive_int(request.data.get("academic_year"))
        term = self._normalize_term_or_none(request.data.get("term"))
        notes = request.data.get("notes", "")

        if not classroom_id or not academic_year_id or not term:
            return Response({"detail": "classroom, academic_year et term (T1/T2/T3) sont requis."}, status=400)

        classroom = self._get_scoped_classroom_or_404(classroom_id)
        academic_year = get_object_or_404(AcademicYear, id=academic_year_id)
        validation, history_count = self._close_term_workflow(
            classroom=classroom,
            academic_year=academic_year,
            term=term,
            notes=notes,
            user=request.user,
        )
        serializer = GradeValidationSerializer(validation)
        payload = dict(serializer.data)
        payload.update(
            {
                "detail": "Période clôturée: classement recalculé et modifications verrouillées.",
                "history_rows": history_count,
            }
        )
        return Response(payload)

    @action(detail=False, methods=["post"], permission_classes=[permissions.IsAuthenticated, HasModuleAccess])
    def close_term(self, request):
        classroom_id = self._parse_positive_int(request.data.get("classroom"))
        academic_year_id = self._parse_positive_int(request.data.get("academic_year"))
        term = self._normalize_term_or_none(request.data.get("term"))
        notes = request.data.get("notes", "")

        if not classroom_id or not academic_year_id or not term:
            return Response({"detail": "classroom, academic_year et term (T1/T2/T3) sont requis."}, status=400)

        classroom = self._get_scoped_classroom_or_404(classroom_id)
        academic_year = get_object_or_404(AcademicYear, id=academic_year_id)
        validation, history_count = self._close_term_workflow(
            classroom=classroom,
            academic_year=academic_year,
            term=term,
            notes=notes,
            user=request.user,
        )

        serializer = GradeValidationSerializer(validation)
        payload = dict(serializer.data)
        payload.update(
            {
                "detail": "Clôture trimestrielle effectuée avec succès.",
                "history_rows": history_count,
            }
        )
        return Response(payload)

    @action(detail=False, methods=["post"], permission_classes=[permissions.IsAuthenticated, HasModuleAccess])
    def unvalidate_term(self, request):
        classroom_id = self._parse_positive_int(request.data.get("classroom"))
        academic_year_id = self._parse_positive_int(request.data.get("academic_year"))
        term = self._normalize_term_or_none(request.data.get("term"))

        if not classroom_id or not academic_year_id or not term:
            return Response({"detail": "classroom, academic_year et term (T1/T2/T3) sont requis."}, status=400)

        classroom = self._get_scoped_classroom_or_404(classroom_id)
        academic_year = get_object_or_404(AcademicYear, id=academic_year_id)
        validation, _ = GradeValidation.objects.update_or_create(
            classroom=classroom,
            academic_year=academic_year,
            term=term,
            defaults={
                "is_validated": False,
                "validated_by": request.user,
                "validated_at": timezone.now(),
            },
        )
        serializer = GradeValidationSerializer(validation)
        return Response(serializer.data)

    @action(detail=False, methods=["get"], permission_classes=[permissions.IsAuthenticated])
    def validation_status(self, request):
        classroom_id = self._parse_positive_int(request.query_params.get("classroom"))
        academic_year_id = self._parse_positive_int(request.query_params.get("academic_year"))
        term = self._normalize_term_or_none(request.query_params.get("term"))

        if not classroom_id or not academic_year_id or not term:
            return Response({"detail": "classroom, academic_year et term (T1/T2/T3) sont requis."}, status=400)

        classroom = self._get_scoped_classroom_or_404(classroom_id)

        validation = GradeValidation.objects.filter(
            classroom_id=classroom_id,
            academic_year_id=academic_year_id,
            term=term,
        ).first()

        if not validation:
            return Response(
                {
                    "classroom": classroom_id,
                    "academic_year": academic_year_id,
                    "term": term,
                    "is_validated": False,
                    "validated_by_name": "",
                    "validated_at": None,
                    "notes": "",
                }
            )

        serializer = GradeValidationSerializer(validation)
        return Response(serializer.data)

    @action(detail=False, methods=["post"], url_path="import-controls")
    def import_controls(self, request):
        classroom_id = self._parse_positive_int(request.data.get("classroom_id"))
        academic_year_id = self._parse_positive_int(request.data.get("academic_year_id"))
        term = self._normalize_term_or_none(request.data.get("term"))
        confirm = str(request.data.get("confirm", "false")).strip().lower() in {"1", "true", "yes", "on"}

        if not classroom_id or not academic_year_id or not term:
            raise ValidationError(
                {"detail": "classroom_id, academic_year_id et term (T1/T2/T3) sont requis."}
            )

        classroom = self._get_scoped_classroom_or_404(classroom_id)
        academic_year = get_object_or_404(AcademicYear, id=academic_year_id)

        if self._is_term_validated(classroom.id, academic_year.id, term):
            raise ValidationError({"detail": self._locked_term_message(prefix="Import")})

        rows = _load_import_rows(request.FILES.get("file") or request.data.get("file"))
        if not rows:
            raise ValidationError({"file": "Aucune ligne exploitable dans le fichier."})

        students_by_matricule = {
            (student.matricule or "").strip().lower(): student
            for student in Student.objects.filter(classroom=classroom, is_archived=False).select_related("user")
        }
        subjects = Subject.objects.filter(classroom=classroom)
        subjects_by_code = {(subject.code or "").strip().lower(): subject for subject in subjects}
        subjects_by_name = {(subject.name or "").strip().lower(): subject for subject in subjects}

        errors = []
        prepared = []
        for index, row in enumerate(rows, start=2):
            matricule = _as_text(row.get("student_matricule") or row.get("matricule")).lower()
            subject_code = _as_text(row.get("subject_code") or row.get("matiere_code")).lower()
            subject_name = _as_text(row.get("subject_name") or row.get("matiere")).lower()
            value = _as_decimal(row.get("value") or row.get("note") or row.get("score"))

            if not matricule:
                errors.append({"row": index, "error": "student_matricule requis."})
                continue
            if value is None:
                errors.append({"row": index, "error": "note/value invalide."})
                continue
            if value < Decimal("0") or value > Decimal("20"):
                errors.append({"row": index, "error": "La note doit être comprise entre 0 et 20."})
                continue

            student = students_by_matricule.get(matricule)
            if student is None:
                errors.append({"row": index, "error": f"Élève introuvable dans la classe pour matricule '{matricule}'."})
                continue

            subject = None
            if subject_code:
                subject = subjects_by_code.get(subject_code)
            if subject is None and subject_name:
                subject = subjects_by_name.get(subject_name)
            if subject is None:
                errors.append({"row": index, "error": "Matière introuvable (subject_code/subject_name)."})
                continue

            existing = Grade.objects.filter(
                student=student,
                subject=subject,
                classroom=classroom,
                academic_year=academic_year,
                term=term,
            ).first()

            prepared.append(
                {
                    "row": index,
                    "student": student,
                    "subject": subject,
                    "value": value,
                    "existing": existing,
                }
            )

        to_create = sum(1 for item in prepared if item["existing"] is None)
        to_update = len(prepared) - to_create
        preview = [
            {
                "row": item["row"],
                "action": "create" if item["existing"] is None else "update",
                "student": item["student"].matricule,
                "subject": item["subject"].code,
                "value": str(item["value"]),
            }
            for item in prepared[:150]
        ]

        payload = {
            "classroom": {"id": classroom.id, "name": classroom.name},
            "academic_year": {"id": academic_year.id, "label": academic_year.name},
            "term": term,
            "summary": {
                "total_rows": len(rows),
                "valid_rows": len(prepared),
                "errors": len(errors),
                "to_create": to_create,
                "to_update": to_update,
            },
            "errors": errors,
            "preview": preview,
            "confirm_required": True,
        }

        if not confirm:
            return Response(payload)

        if errors:
            return Response({**payload, "detail": "Import bloqué: corrigez les erreurs."}, status=400)

        created = 0
        updated = 0
        with transaction.atomic():
            for item in prepared:
                serializer_data = {
                    "student": item["student"].id,
                    "subject": item["subject"].id,
                    "classroom": classroom.id,
                    "academic_year": academic_year.id,
                    "term": term,
                    "value": str(item["value"]),
                }
                if item["existing"] is None:
                    serializer = self.get_serializer(data=serializer_data)
                    serializer.is_valid(raise_exception=True)
                    self.perform_create(serializer)
                    created += 1
                else:
                    serializer = self.get_serializer(item["existing"], data=serializer_data, partial=False)
                    serializer.is_valid(raise_exception=True)
                    self.perform_update(serializer)
                    updated += 1

        return Response(
            {
                **payload,
                "result": {"created": created, "updated": updated},
                "detail": "Import notes de contrôle terminé.",
            }
        )


class AttendanceViewSet(AnneeScolaireScopeMixin, BaseModelViewSet):
    renseigne_annee_a_la_creation = True
    access_module = "attendance"
    queryset = Attendance.objects.select_related("student", "student__user").all().order_by("-date", "-id")
    serializer_class = AttendanceSerializer
    filterset_fields = ["date", "student", "is_absent", "is_late"]
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]





    def _resolve_target_etablissement(self):
        user = self.request.user
        requested_etablissement = self._requested_etablissement()

        if getattr(user, "role", None) == "super_admin" and requested_etablissement:
            return requested_etablissement

        return getattr(user, "etablissement", None)

    def _parse_sheet_date(self, raw_value):
        if raw_value in (None, ""):
            return timezone.now().date()
        try:
            return datetime.strptime(str(raw_value), "%Y-%m-%d").date()
        except ValueError as exc:
            raise ValidationError({"date": "Format invalide. Utilisez YYYY-MM-DD."}) from exc

    def _parse_sheet_classroom_id(self, raw_value):
        try:
            classroom_id = int(raw_value)
        except (TypeError, ValueError):
            classroom_id = 0
        if classroom_id <= 0:
            raise ValidationError({"classroom": "classroom est requis."})
        return classroom_id

    def _parse_student_id(self, raw_value):
        """Meme lecture que pour la classe, mais le champ manquant se nomme.

        Reutiliser _parse_sheet_classroom_id repondait « classroom est
        requis » a qui postait une conduite sans eleve.
        """
        try:
            student_id = int(raw_value)
        except (TypeError, ValueError):
            student_id = 0
        if student_id <= 0:
            raise ValidationError({"student": "student est requis."})
        return student_id

    def _assert_sheet_scope(self):
        """La feuille d'appel decrit une classe, pas un eleve.

        HasModuleAccess verifie le niveau; il ignore la portee. Or la famille
        lit l'assiduite en lecture restreinte (L*), c'est-a-dire ses enfants
        ou soi -- ce que la vue par classe ne sait pas restreindre, et qui
        n'aurait de toute facon pas de sens: une fiche est un document
        collectif. Le parent et l'eleve ont leur propre ecran.

        L'enseignant est restreint lui aussi, mais a des classes et non a des
        eleves, et son niveau est l'ecriture: c'est ce qui l'en distingue.
        """
        role = getattr(self.request.user, "role", "")
        module = self.access_module
        if is_scoped(role, module) and not can_write(role, module):
            raise ValidationError(
                {"detail": "La feuille d'appel n'est pas accessible depuis un compte famille."}
            )

    def _assert_can_validate_sheet(self):
        """Cloturer une fiche demande l'ecriture sans portee restreinte.

        Le niveau est deja verifie par HasModuleAccess a l'entree de l'action;
        il ne reste ici que la nuance de perimetre. L'enseignant saisit l'appel
        de ses classes (E*) mais ne le cloture pas: verrouiller engage la
        classe entiere vis-a-vis de la direction, pas seulement sa propre
        saisie.
        """
        self._assert_sheet_scope()
        role = getattr(self.request.user, "role", "")
        if is_scoped(role, self.access_module):
            raise ValidationError(
                {"detail": "Validation de fiche reservee a la direction et a la surveillance."}
            )

    def _sheet_classrooms_queryset(self):
        queryset = ClassRoom.objects.select_related("academic_year", "etablissement").all()
        target_etablissement = self._resolve_target_etablissement()
        if target_etablissement is not None:
            queryset = queryset.filter(etablissement=target_etablissement)

        if getattr(self.request.user, "role", "") == UserRole.TEACHER:
            allowed_ids = self._teacher_allowed_classroom_ids()
            if not allowed_ids:
                return queryset.none()
            queryset = queryset.filter(id__in=allowed_ids)

        return queryset.order_by("name", "id")

    def _get_sheet_classroom_or_404(self, classroom_id):
        return get_object_or_404(self._sheet_classrooms_queryset(), id=classroom_id)

    @staticmethod
    def _pdf_safe_text(value):
        text = str(value or "")
        return text.encode("latin-1", "ignore").decode("latin-1")

    def _sheet_validation_row(self, classroom, selected_date):
        return AttendanceSheetValidation.objects.filter(
            classroom=classroom,
            date=selected_date,
            is_locked=True,
        ).select_related("validated_by").first()

    @staticmethod
    def _proof_name(attendance_row):
        """Nom de fichier seul, sans le dossier de stockage.

        La liste d'appel affiche « justifie » et propose d'ouvrir la piece;
        le chemin complet (attendance_proofs/2026/...) n'apprendrait rien a
        qui fait l'appel.
        """
        if not attendance_row or not attendance_row.proof:
            return ""
        return str(attendance_row.proof.name).rsplit("/", 1)[-1]

    @staticmethod
    def _proof_url(attendance_row):
        """Chemin du justificatif, relatif ou absolu selon le stockage.

        Le frontend le resout contre l'URL de l'API (resolveMediaUrl), comme
        les photos d'eleves: renvoyer une URL absolue ici casserait le
        stockage objet, qui signe deja les siennes.
        """
        if not attendance_row or not attendance_row.proof:
            return ""
        try:
            return attendance_row.proof.url
        except ValueError:
            # Fichier reference en base mais absent du stockage: la fiche
            # doit rester affichable.
            return ""

    def _build_class_sheet_payload(self, classroom, selected_date):
        students = list(
            Student.objects.select_related("user")
            .filter(classroom=classroom, is_archived=False)
            .order_by("user__last_name", "user__first_name", "id")
        )
        attendance_by_student = {
            row.student_id: row
            for row in Attendance.objects.filter(
                student__in=students,
                date=selected_date,
            )
        }

        items = []
        for student in students:
            user = student.user if student else None
            full_name = user.get_full_name().strip() if user else ""
            full_name = full_name or (user.username if user else "")
            attendance_row = attendance_by_student.get(student.id)
            items.append(
                {
                    "student": student.id,
                    "student_full_name": full_name,
                    "student_matricule": student.matricule,
                    "attendance_id": attendance_row.id if attendance_row else None,
                    "is_absent": bool(attendance_row.is_absent) if attendance_row else False,
                    "is_late": bool(attendance_row.is_late) if attendance_row else False,
                    "reason": attendance_row.reason if attendance_row else "",
                    "has_proof": bool(attendance_row.proof) if attendance_row else False,
                    "proof_url": self._proof_url(attendance_row),
                    "proof_name": self._proof_name(attendance_row),
                }
            )

        lock_row = self._sheet_validation_row(classroom, selected_date)
        validated_by_name = ""
        if lock_row and lock_row.validated_by:
            validated_user = lock_row.validated_by
            validated_by_name = validated_user.get_full_name().strip() or validated_user.username

        return {
            "classroom": {"id": classroom.id, "name": classroom.name},
            "date": selected_date.isoformat(),
            "items": items,
            "count": len(items),
            "is_locked": bool(lock_row),
            "validated_at": lock_row.validated_at.isoformat() if lock_row and lock_row.validated_at else None,
            "validated_by_name": validated_by_name,
            "validation_notes": lock_row.notes if lock_row else "",
        }

    def get_queryset(self):
        queryset = super().get_queryset()
        role = getattr(self.request.user, "role", "")

        if role == UserRole.STUDENT:
            return queryset.filter(student__user_id=self.request.user.id)
        if role == UserRole.PARENT:
            return queryset.filter(student__parent__user_id=self.request.user.id)
        if role == UserRole.TEACHER:
            allowed_classroom_ids = self._teacher_allowed_classroom_ids()
            if not allowed_classroom_ids:
                return queryset.none()
            return queryset.filter(student__classroom_id__in=allowed_classroom_ids)

        requested_etablissement = self._requested_etablissement()
        if requested_etablissement is not None:
            return queryset.filter(student__etablissement=requested_etablissement)

        if self._has_requested_scope():
            return queryset.none()

        if getattr(self.request.user, "role", None) == "super_admin":
            return queryset

        return queryset.filter(student__etablissement=getattr(self.request.user, "etablissement", None))

    def _validate_student_scope(self, serializer):
        student = serializer.validated_data.get("student")
        if not student:
            return

        role = getattr(self.request.user, "role", "")
        if role == UserRole.STUDENT and student.user_id != self.request.user.id:
            raise ValidationError({"student": "Vous ne pouvez saisir que vos propres absences."})
        if role == UserRole.PARENT and student.parent_id:
            if student.parent.user_id != self.request.user.id:
                raise ValidationError({"student": "Vous ne pouvez saisir que les absences de vos enfants."})
        if role == UserRole.TEACHER:
            allowed_classroom_ids = self._teacher_allowed_classroom_ids()
            if student.classroom_id not in allowed_classroom_ids:
                raise ValidationError(
                    {
                        "student": (
                            "Acces refuse: vous ne pouvez saisir que les absences/retards "
                            "des élèves de vos classes affectées."
                        )
                    }
                )

        target_etablissement = self._resolve_target_etablissement()
        if target_etablissement and student.etablissement_id != target_etablissement.id:
            raise ValidationError({"student": "L'eleve n'appartient pas a l'etablissement actif."})

    def perform_create(self, serializer):
        self._validate_student_scope(serializer)
        serializer.save()

    def perform_update(self, serializer):
        self._validate_student_scope(serializer)
        serializer.save()

    def _conduite_students_queryset(self):
        """Eleves notables, bornes au meme perimetre que les presences.

        La conduite se note par etablissement et non par classe: seuls le
        censeur, le surveillant et le super admin y touchent, et aucun des
        trois n'a de perimetre de classe.
        """
        queryset = Student.objects.filter(is_archived=False)
        requested = self._requested_etablissement()
        if requested is not None:
            return queryset.filter(etablissement=requested)
        if getattr(self.request.user, "role", "") == UserRole.SUPER_ADMIN:
            return queryset
        return queryset.filter(
            etablissement=getattr(self.request.user, "etablissement", None)
        )

    @action(detail=False, methods=["post"], url_path="conduite")
    def conduite(self, request):
        """Note de conduite d'un eleve, saisie depuis l'emargement.

        Elle ne s'ecrivait jusqu'ici qu'en effet de bord de la creation d'une
        absence: le formulaire de saisie unitaire portait un champ
        « Conduite (/20) » que le serializer d'Attendance reportait sur
        l'eleve. Ce formulaire faisait par ailleurs doublon avec la feuille
        d'appel, et echouait des que la fiche du jour etait enregistree
        (une seule presence par eleve et par date).

        Une route directe est le seul moyen de retirer ce doublon sans perdre
        la conduite: PATCH /students/ ne convenait pas, la matrice y refuse
        l'ecriture au censeur et au surveillant, qui sont precisement les
        deux profils qui notent la conduite.

        La regle « qui peut noter » n'est pas recopiee ici: c'est
        StudentSerializer qui l'applique, comme pour toute autre ecriture sur
        l'eleve.
        """
        student_id = self._parse_student_id(request.data.get("student"))
        student = get_object_or_404(self._conduite_students_queryset(), id=student_id)

        serializer = StudentSerializer(
            student,
            data={"conduite": request.data.get("conduite")},
            partial=True,
            context={"request": request},
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(
            {
                "detail": "Conduite enregistree.",
                "student": student.id,
                "conduite": str(student.conduite),
            }
        )

    @action(
        detail=True,
        methods=["post", "delete"],
        url_path="proof",
        parser_classes=[MultiPartParser, FormParser],
    )
    def proof(self, request, pk=None):
        """Justificatif d'une absence: depot, remplacement, retrait.

        Le champ existait en base depuis l'origine et les statistiques
        mensuelles comptaient deja les justificatifs -- mais aucun ecran ne
        permettait d'en deposer un, si bien que le compteur affichait zero en
        permanence.

        Volontairement hors du POST de la feuille d'appel, et volontairement
        accepte sur une fiche verrouillee: le mot d'excuse arrive le
        lendemain, apres que la fiche du jour a ete validee. Le refuser
        obligerait a deverrouiller la journee entiere pour classer un papier.

        Le retrait exige le niveau administration (DELETE), tandis que
        remplacer une piece deposee par erreur reste a portee de qui saisit
        l'appel: reposter ecrase.
        """
        attendance = self.get_object()

        if request.method.lower() == "delete":
            if not attendance.proof:
                raise ValidationError({"detail": "Aucun justificatif a retirer."})
            attendance.proof.delete(save=False)
            attendance.proof = None
            attendance.save(update_fields=["proof"])
            return Response({"detail": "Justificatif retire.", "has_proof": False})

        uploaded = request.FILES.get("proof") or request.data.get("proof")
        if not uploaded:
            raise ValidationError({"proof": ["Aucun fichier fourni."]})

        if not attendance.is_absent and not attendance.is_late:
            raise ValidationError(
                {"detail": "Cet eleve n'est ni absent ni en retard a cette date."}
            )

        # Remplacer laisse sinon l'ancienne piece sur le stockage sans que
        # rien ne la reference.
        if attendance.proof:
            attendance.proof.delete(save=False)

        attendance.proof = uploaded
        attendance.save(update_fields=["proof"])
        return Response(
            {
                "detail": "Justificatif enregistre.",
                "has_proof": True,
                "proof_url": self._proof_url(attendance),
                "proof_name": self._proof_name(attendance),
            }
        )

    @action(detail=False, methods=["get"])
    def sheet_classrooms(self, request):
        self._assert_sheet_scope()
        classrooms = self._sheet_classrooms_queryset()
        rows = [
            {
                "id": classroom.id,
                "name": classroom.name,
                "academic_year": classroom.academic_year_id,
                "academic_year_name": getattr(classroom.academic_year, "name", ""),
            }
            for classroom in classrooms
        ]
        return Response(rows)

    @action(detail=False, methods=["get", "post"], url_path="class-sheet")
    def class_sheet(self, request):
        self._assert_sheet_scope()
        if request.method.lower() == "get":
            classroom_id = self._parse_sheet_classroom_id(request.query_params.get("classroom"))
            selected_date = self._parse_sheet_date(request.query_params.get("date"))
            classroom = self._get_sheet_classroom_or_404(classroom_id)
            return Response(self._build_class_sheet_payload(classroom, selected_date))

        classroom_id = self._parse_sheet_classroom_id(request.data.get("classroom"))
        selected_date = self._parse_sheet_date(request.data.get("date"))
        classroom = self._get_sheet_classroom_or_404(classroom_id)

        if self._sheet_validation_row(classroom, selected_date):
            raise ValidationError({
                "detail": "Fiche verrouillee: deverrouillez-la avant modification.",
            })

        items = request.data.get("items")
        if not isinstance(items, list) or not items:
            raise ValidationError({"items": "items est requis (liste non vide)."})

        students = list(
            Student.objects.filter(classroom=classroom, is_archived=False).order_by("id")
        )
        student_map = {student.id: student for student in students}
        if not student_map:
            raise ValidationError({"classroom": "Aucun eleve actif dans cette classe."})

        created_count = 0
        updated_count = 0
        deleted_count = 0
        with transaction.atomic():
            for row in items:
                if not isinstance(row, dict):
                    continue
                student_id = self._parse_sheet_classroom_id(row.get("student"))
                student = student_map.get(student_id)
                if not student:
                    raise ValidationError({"student": f"Eleve {student_id} hors de la classe selectionnee."})

                is_absent = bool(row.get("is_absent", False))
                is_late = bool(row.get("is_late", False))
                reason = str(row.get("reason", "")).strip()

                # Empty row means no incident for this date; remove stale record if any.
                if not is_absent and not is_late and not reason:
                    deleted, _ = Attendance.objects.filter(student=student, date=selected_date).delete()
                    if deleted:
                        deleted_count += 1
                    continue

                _, created = Attendance.objects.update_or_create(
                    student=student,
                    date=selected_date,
                    defaults={
                        "is_absent": is_absent,
                        "is_late": is_late,
                        "reason": reason,
                    },
                )
                if created:
                    created_count += 1
                else:
                    updated_count += 1

        return Response(
            {
                "detail": "Fiche de presence enregistree.",
                "classroom": {"id": classroom.id, "name": classroom.name},
                "date": selected_date.isoformat(),
                "created": created_count,
                "updated": updated_count,
                "deleted": deleted_count,
            }
        )

    @action(detail=False, methods=["get"], url_path="sheet-journal")
    def sheet_journal(self, request):
        """Fiches d'appel deja enregistrees, une ligne par classe et par date.

        Une fois la fiche enregistree, rien ne permettait de la revoir: il
        fallait resaisir sa classe et sa date de memoire. L'historique
        existant ne repondait pas -- il listait les enregistrements un par un,
        tous eleves et toutes dates melanges, ce qui devient illisible des la
        premiere semaine.

        Une fiche existe des qu'un enregistrement porte sa classe et sa date,
        validee ou non: on agrege donc les presences plutot que de lister les
        seules validations, sinon les brouillons resteraient invisibles.
        """
        self._assert_sheet_scope()

        classrooms = self._sheet_classrooms_queryset()
        classroom_id = request.query_params.get("classroom")
        if classroom_id:
            classrooms = classrooms.filter(
                id=self._parse_sheet_classroom_id(classroom_id)
            )

        classroom_ids = list(classrooms.values_list("id", flat=True))
        if not classroom_ids:
            return Response([])

        presences = Attendance.objects.filter(
            student__classroom_id__in=classroom_ids
        )

        depuis = request.query_params.get("from")
        jusqua = request.query_params.get("to")
        if depuis:
            presences = presences.filter(date__gte=self._parse_sheet_date(depuis))
        if jusqua:
            presences = presences.filter(date__lte=self._parse_sheet_date(jusqua))

        lignes = (
            presences.values("student__classroom_id", "date")
            .annotate(
                effectif=Count("id"),
                absents=Count("id", filter=Q(is_absent=True)),
                retards=Count("id", filter=Q(is_late=True)),
            )
            .order_by("-date", "student__classroom_id")
        )
        # Une classe compte facilement trente lignes par jour: sans plafond,
        # une annee scolaire en rendrait plusieurs milliers d'un coup.
        lignes = lignes[:400]

        noms = dict(classrooms.values_list("id", "name"))
        validations = {
            (row.classroom_id, row.date): row
            for row in AttendanceSheetValidation.objects.filter(
                classroom_id__in=classroom_ids,
                date__in=[ligne["date"] for ligne in lignes],
            ).select_related("validated_by")
        }

        resultat = []
        for ligne in lignes:
            cle = (ligne["student__classroom_id"], ligne["date"])
            validation = validations.get(cle)
            resultat.append(
                {
                    "classroom": ligne["student__classroom_id"],
                    "classroom_name": noms.get(ligne["student__classroom_id"], ""),
                    "date": ligne["date"],
                    "effectif": ligne["effectif"],
                    "absents": ligne["absents"],
                    "retards": ligne["retards"],
                    "is_locked": bool(validation and validation.is_locked),
                    "validated_by_name": (
                        validation.validated_by.get_full_name()
                        or validation.validated_by.username
                        if validation and validation.validated_by
                        else ""
                    ),
                    "validated_at": validation.validated_at if validation else None,
                }
            )

        return Response(resultat)

    @action(detail=False, methods=["post"], url_path="class-sheet-validate")
    def class_sheet_validate(self, request):
        self._assert_can_validate_sheet()

        classroom_id = self._parse_sheet_classroom_id(request.data.get("classroom"))
        selected_date = self._parse_sheet_date(request.data.get("date"))
        classroom = self._get_sheet_classroom_or_404(classroom_id)
        should_lock = bool(request.data.get("lock", True))
        notes = str(request.data.get("notes", "")).strip()

        if should_lock:
            AttendanceSheetValidation.objects.update_or_create(
                classroom=classroom,
                date=selected_date,
                defaults={
                    "is_locked": True,
                    "validated_by": request.user,
                    "validated_at": timezone.now(),
                    "notes": notes,
                },
            )
            return Response(
                {
                    "detail": "Fiche validee et verrouillee.",
                    "classroom": {"id": classroom.id, "name": classroom.name},
                    "date": selected_date.isoformat(),
                    "is_locked": True,
                }
            )

        AttendanceSheetValidation.objects.filter(
            classroom=classroom,
            date=selected_date,
        ).delete()
        return Response(
            {
                "detail": "Fiche deverrouillee.",
                "classroom": {"id": classroom.id, "name": classroom.name},
                "date": selected_date.isoformat(),
                "is_locked": False,
            }
        )

    # Colonnes de la fiche imprimee: largeurs en mm, pour une page A4 portrait.
    _COLONNES_FICHE_PDF = (
        ("Eleve", 65, 42, "L"),
        ("Matricule", 30, 18, "L"),
        ("Absent", 18, None, "C"),
        ("Retard", 18, None, "C"),
        ("Motif", 59, 40, "L"),
    )

    def _entetes_fiche_pdf(self, pdf):
        pdf.set_font("Helvetica", "B", 9)
        for titre, largeur, _, alignement in self._COLONNES_FICHE_PDF:
            pdf.cell(largeur, 7, self._pdf_safe_text(titre), border=1, align=alignement)
        pdf.ln(7)
        pdf.set_font("Helvetica", "", 8)

    def _ecrire_fiche_pdf(self, pdf, classroom, selected_date, items):
        """Ecrit une fiche de classe sur sa propre page du document.

        Partage entre l'export d'une fiche et celui de la journee entiere: deux
        rendus separes finiraient par diverger, et c'est le meme papier qui est
        archive dans les deux cas.
        """
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 8, self._pdf_safe_text("Fiche de presence par classe"), ln=1)
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 6, self._pdf_safe_text(f"Classe: {classroom.name}"), ln=1)
        pdf.cell(0, 6, self._pdf_safe_text(f"Date: {selected_date.isoformat()}"), ln=1)

        absents = sum(1 for row in items if row.get("is_absent"))
        retards = sum(1 for row in items if row.get("is_late"))
        pdf.cell(
            0,
            6,
            self._pdf_safe_text(
                f"Effectif: {len(items)}  -  Absents: {absents}  -  Retards: {retards}"
            ),
            ln=1,
        )
        pdf.ln(2)

        self._entetes_fiche_pdf(pdf)
        for row in items:
            valeurs = (
                self._pdf_safe_text(str(row.get("student_full_name", ""))),
                self._pdf_safe_text(str(row.get("student_matricule", ""))),
                "Oui" if row.get("is_absent") else "Non",
                "Oui" if row.get("is_late") else "Non",
                self._pdf_safe_text(str(row.get("reason", ""))),
            )

            # Une classe de plus de trente eleves deborde la page: on repart
            # sur une nouvelle en reposant les en-tetes de colonnes.
            if pdf.get_y() > 275:
                pdf.add_page()
                self._entetes_fiche_pdf(pdf)

            for valeur, (_, largeur, coupe, alignement) in zip(
                valeurs, self._COLONNES_FICHE_PDF
            ):
                texte = valeur[:coupe] if coupe else valeur
                pdf.cell(largeur, 6, texte, border=1, align=alignement)
            pdf.ln(6)

    @action(detail=False, methods=["get"], url_path="class-sheet-export")
    def class_sheet_export(self, request):
        self._assert_sheet_scope()
        classroom_id = self._parse_sheet_classroom_id(request.query_params.get("classroom"))
        selected_date = self._parse_sheet_date(request.query_params.get("date"))
        export_format = str(request.query_params.get("format", "pdf")).strip().lower()
        classroom = self._get_sheet_classroom_or_404(classroom_id)
        payload = self._build_class_sheet_payload(classroom, selected_date)
        items = payload["items"]

        if export_format == "xlsx":
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Presence"
            sheet.append(["Classe", classroom.name])
            sheet.append(["Date", selected_date.isoformat()])
            sheet.append([])
            sheet.append(["Eleve", "Matricule", "Absent", "Retard", "Motif"])
            for row in items:
                sheet.append(
                    [
                        str(row.get("student_full_name", "")),
                        str(row.get("student_matricule", "")),
                        "Oui" if row.get("is_absent") else "Non",
                        "Oui" if row.get("is_late") else "Non",
                        str(row.get("reason", "")),
                    ]
                )

            response = HttpResponse(
                content_type=(
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            )
            file_name = f"presence_{classroom.name}_{selected_date.isoformat()}.xlsx".replace(" ", "_")
            response["Content-Disposition"] = f'attachment; filename="{file_name}"'
            workbook.save(response)
            return response

        pdf = FPDF(orientation="P", unit="mm", format="A4")
        self._ecrire_fiche_pdf(pdf, classroom, selected_date, items)

        # `bytes(pdf.output())`, comme partout ailleurs dans le projet: fpdf2
        # rend un bytearray, et l'ancien `.output(dest="S").encode("latin-1")`
        # -- l'interface de fpdf 1.x -- levait un AttributeError. L'export PDF
        # de la fiche repondait donc 500 a chaque appel.
        content = bytes(pdf.output())
        response = HttpResponse(content, content_type="application/pdf")
        file_name = f"presence_{classroom.name}_{selected_date.isoformat()}.pdf".replace(" ", "_")
        response["Content-Disposition"] = f'attachment; filename="{file_name}"'
        return response

    @action(detail=False, methods=["get"], url_path="day-export")
    def class_sheet_day_export(self, request):
        """Toutes les fiches d'un jour, en un seul PDF.

        C'est le geste reel de fin de journee: l'administration archive et
        signe l'appel de l'etablissement entier, pas d'une classe. Il fallait
        auparavant exporter classe par classe puis recoller les fichiers.

        Seules les classes ayant une fiche ce jour-la sont imprimees: sortir
        trente pages vides pour un mercredi apres-midi ferait du papier et
        laisserait croire a trente classes non faites.
        """
        self._assert_sheet_scope()
        selected_date = self._parse_sheet_date(request.query_params.get("date"))

        classrooms = self._sheet_classrooms_queryset()
        classroom_ids = list(classrooms.values_list("id", flat=True))
        avec_fiche = set(
            Attendance.objects.filter(
                student__classroom_id__in=classroom_ids, date=selected_date
            ).values_list("student__classroom_id", flat=True)
        )
        a_imprimer = [
            classroom for classroom in classrooms if classroom.id in avec_fiche
        ]

        if not a_imprimer:
            # 400 plutot que 404: la route existe, c'est la journee qui est
            # vide. Le client affiche ce message tel quel.
            raise ValidationError(
                {
                    "detail": (
                        "Aucune fiche enregistree le "
                        f"{selected_date.isoformat()}."
                    )
                }
            )

        pdf = FPDF(orientation="P", unit="mm", format="A4")
        for classroom in a_imprimer:
            payload = self._build_class_sheet_payload(classroom, selected_date)
            self._ecrire_fiche_pdf(
                pdf, classroom, selected_date, payload["items"]
            )

        response = HttpResponse(bytes(pdf.output()), content_type="application/pdf")
        file_name = f"presences_{selected_date.isoformat()}.pdf"
        response["Content-Disposition"] = f'attachment; filename="{file_name}"'
        # Le client affiche « 12 classes » sans avoir a ouvrir le document.
        response["X-Fiches-Count"] = str(len(a_imprimer))
        return response

    @action(detail=False, methods=["get"])
    def monthly_stats(self, request):
        month_value = request.query_params.get("month")
        today = timezone.now().date()

        if month_value:
            try:
                year, month = month_value.split("-")
                year = int(year)
                month = int(month)
            except (ValueError, TypeError):
                return Response({"detail": "Format month invalide. Utilisez YYYY-MM."}, status=400)
        else:
            year = today.year
            month = today.month

        queryset = self.get_queryset().filter(date__year=year, date__month=month)
        totals = queryset.aggregate(
            total_records=Count("id"),
            absences=Count("id", filter=Q(is_absent=True)),
            lates=Count("id", filter=Q(is_late=True)),
            # Un FileField vide vaut tantot NULL, tantot la chaine vide
            # selon la facon dont la ligne a ete ecrite: ne tester que
            # isnull comptait des justificatifs inexistants.
            justifications=Count(
                "id", filter=Q(proof__isnull=False) & ~Q(proof="")
            ),
        )

        per_day = (
            queryset.values("date")
            .annotate(
                absences=Count("id", filter=Q(is_absent=True)),
                lates=Count("id", filter=Q(is_late=True)),
            )
            .order_by("date")
        )

        return Response(
            {
                "month": f"{year:04d}-{month:02d}",
                "total_records": totals["total_records"] or 0,
                "absences": totals["absences"] or 0,
                "lates": totals["lates"] or 0,
                "justifications": totals["justifications"] or 0,
                "daily": list(per_day),
            }
        )


class TeacherAttendanceViewSet(AnneeScolaireScopeMixin, BaseModelViewSet):
    renseigne_annee_a_la_creation = True
    access_module = "teacher_timesheet"
    queryset = TeacherAttendance.objects.select_related("teacher", "teacher__user").all().order_by("-date", "-id")
    serializer_class = TeacherAttendanceSerializer
    filterset_fields = ["date", "teacher", "is_absent", "is_late"]
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]





    def _resolve_target_etablissement(self):
        user = self.request.user
        requested_etablissement = self._requested_etablissement()

        if getattr(user, "role", None) == "super_admin" and requested_etablissement:
            return requested_etablissement

        return getattr(user, "etablissement", None)

    def get_queryset(self):
        queryset = super().get_queryset()
        role = getattr(self.request.user, "role", "")

        if role == UserRole.TEACHER:
            return queryset.filter(teacher__user_id=self.request.user.id)

        requested_etablissement = self._requested_etablissement()
        if requested_etablissement is not None:
            return queryset.filter(teacher__etablissement=requested_etablissement)

        if self._has_requested_scope():
            return queryset.none()

        if getattr(self.request.user, "role", None) == "super_admin":
            return queryset

        return queryset.filter(teacher__etablissement=getattr(self.request.user, "etablissement", None))

    def _validate_teacher_scope(self, serializer):
        teacher = serializer.validated_data.get("teacher")
        if not teacher:
            return

        role = getattr(self.request.user, "role", "")
        if role == UserRole.TEACHER and teacher.user_id != self.request.user.id:
            raise ValidationError({"teacher": "Vous ne pouvez saisir que vos propres absences."})

        target_etablissement = self._resolve_target_etablissement()
        if target_etablissement and teacher.etablissement_id != target_etablissement.id:
            raise ValidationError({"teacher": "L'enseignant n'appartient pas a l'etablissement actif."})

    def perform_create(self, serializer):
        self._validate_teacher_scope(serializer)
        serializer.save()

    def perform_update(self, serializer):
        self._validate_teacher_scope(serializer)
        serializer.save()

    @action(detail=False, methods=["get"])
    def monthly_stats(self, request):
        month_value = request.query_params.get("month")
        today = timezone.now().date()

        if month_value:
            try:
                year, month = month_value.split("-")
                year = int(year)
                month = int(month)
            except (ValueError, TypeError):
                return Response({"detail": "Format month invalide. Utilisez YYYY-MM."}, status=400)
        else:
            year = today.year
            month = today.month

        queryset = self.get_queryset().filter(date__year=year, date__month=month)
        totals = queryset.aggregate(
            total_records=Count("id"),
            absences=Count("id", filter=Q(is_absent=True)),
            lates=Count("id", filter=Q(is_late=True)),
            # Un FileField vide vaut tantot NULL, tantot la chaine vide
            # selon la facon dont la ligne a ete ecrite: ne tester que
            # isnull comptait des justificatifs inexistants.
            justifications=Count(
                "id", filter=Q(proof__isnull=False) & ~Q(proof="")
            ),
        )

        per_day = (
            queryset.values("date")
            .annotate(
                absences=Count("id", filter=Q(is_absent=True)),
                lates=Count("id", filter=Q(is_late=True)),
            )
            .order_by("date")
        )

        return Response(
            {
                "month": f"{year:04d}-{month:02d}",
                "total_records": totals["total_records"] or 0,
                "absences": totals["absences"] or 0,
                "lates": totals["lates"] or 0,
                "justifications": totals["justifications"] or 0,
                "daily": list(per_day),
            }
        )


class TeacherTimeEntryViewSet(BaseModelViewSet):
    access_module = "teacher_timesheet"
    queryset = TeacherTimeEntry.objects.select_related("teacher", "teacher__user", "recorded_by").all().order_by("-entry_date", "-id")
    serializer_class = TeacherTimeEntrySerializer
    filterset_fields = ["teacher", "entry_date", "etablissement"]
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]





    def _resolve_target_etablissement(self):
        user = self.request.user
        requested_etablissement = self._requested_etablissement()

        if getattr(user, "role", None) == UserRole.SUPER_ADMIN and requested_etablissement:
            return requested_etablissement

        return getattr(user, "etablissement", None)

    def _request_teacher_profile(self):
        user = self.request.user
        if getattr(user, "role", None) != UserRole.TEACHER:
            return None
        return Teacher.objects.select_related("user", "etablissement").filter(user=user).first()

    def get_queryset(self):
        queryset = super().get_queryset()
        requested_etablissement = self._requested_etablissement()
        teacher_profile = self._request_teacher_profile()

        if teacher_profile is not None:
            return queryset.filter(teacher=teacher_profile)

        if requested_etablissement is not None:
            return queryset.filter(teacher__etablissement=requested_etablissement)

        if self._has_requested_scope():
            return queryset.none()

        if getattr(self.request.user, "role", None) == UserRole.SUPER_ADMIN:
            return queryset

        user_etablissement = getattr(self.request.user, "etablissement", None)
        if user_etablissement is None:
            return queryset.none()
        return queryset.filter(teacher__etablissement=user_etablissement)

    def _validate_scope(self, serializer, instance=None):
        teacher = serializer.validated_data.get("teacher") or (instance.teacher if instance else None)
        if not teacher:
            return

        teacher_profile = self._request_teacher_profile()
        if teacher_profile is not None and teacher.id != teacher_profile.id:
            raise PermissionDenied(
                "Acces refuse: un enseignant ne peut enregistrer que son propre pointage."
            )

        target_etablissement = self._resolve_target_etablissement()
        if target_etablissement and teacher.etablissement_id != target_etablissement.id:
            raise ValidationError({"teacher": "L'enseignant n'appartient pas a l'etablissement actif."})

    def perform_create(self, serializer):
        self._validate_scope(serializer)
        serializer.save(
            etablissement=self._resolve_target_etablissement(),
            recorded_by=self.request.user,
        )

    def perform_update(self, serializer):
        self._validate_scope(serializer, instance=self.get_object())
        serializer.save(etablissement=self._resolve_target_etablissement())

    def update(self, request, *args, **kwargs):
        if getattr(request.user, "role", None) == UserRole.TEACHER:
            raise PermissionDenied(
                "Acces refuse: un enseignant ne peut pas modifier un pointage existant."
            )
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if getattr(request.user, "role", None) == UserRole.TEACHER:
            raise PermissionDenied(
                "Acces refuse: un enseignant ne peut pas modifier un pointage existant."
            )
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if getattr(request.user, "role", None) == UserRole.TEACHER:
            raise PermissionDenied(
                "Acces refuse: un enseignant ne peut pas supprimer un pointage."
            )
        return super().destroy(request, *args, **kwargs)

    # --- Concordance planning / emargement ----------------------------------

    # Deux mois: au-dela, la reponse porte des milliers de seances et l'ecran
    # ne sait rien en faire. Un rapprochement se lit par semaine ou par mois.
    CONCORDANCE_MAX_JOURS = 62
    DAY_CODES = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]

    @staticmethod
    def _minutes_de(valeur):
        return valeur.hour * 60 + valeur.minute

    def _plage_de_concordance(self):
        aujourd_hui = timezone.localdate()
        debut_brut = self.request.query_params.get("from")
        fin_brute = self.request.query_params.get("to")

        debut = parse_date(str(debut_brut).strip()) if debut_brut else aujourd_hui
        fin = parse_date(str(fin_brute).strip()) if fin_brute else debut

        if debut is None or fin is None:
            raise ValidationError({"from": "Dates illisibles. Utilisez AAAA-MM-JJ."})
        if fin < debut:
            raise ValidationError({"to": "La fin de période précède son début."})
        if (fin - debut).days + 1 > self.CONCORDANCE_MAX_JOURS:
            raise ValidationError(
                {
                    "to": f"Période trop large : {self.CONCORDANCE_MAX_JOURS} jours au maximum."
                }
            )
        return debut, fin

    def _enseignants_de_concordance(self):
        """Les enseignants a rapprocher: ceux demandes, ou tous ceux du perimetre.

        Pas seulement ceux qui ont pointe: un enseignant absent toute la
        semaine n'a aucun pointage, et c'est precisement lui qu'il faut voir.
        """
        demande = self.request.query_params.get("teacher")
        base = Teacher.objects.select_related("user", "etablissement")

        if demande not in (None, "") and str(demande).isdigit():
            base = base.filter(id=demande)

        profil = self._request_teacher_profile()
        if profil is not None:
            return base.filter(id=profil.id)

        etablissement = self._requested_etablissement()
        if etablissement is not None:
            return base.filter(etablissement=etablissement)
        if self._has_requested_scope():
            return base.none()
        if getattr(self.request.user, "role", None) == UserRole.SUPER_ADMIN:
            return base
        return base.filter(etablissement=getattr(self.request.user, "etablissement", None))

    @action(detail=False, methods=["get"], url_path="concordance")
    def concordance(self, request):
        """Ce qui devait être assuré, ce qui l'a été, et l'écart.

        L'emploi du temps et l'emargement vivaient cote a cote sans jamais
        se regarder: une seance que personne n'assurait ne remontait nulle
        part, et un pointage n'indiquait pas a quel cours il correspondait.
        """
        debut, fin = self._plage_de_concordance()

        enseignants = list(self._enseignants_de_concordance())
        if not enseignants:
            return Response(
                {
                    "from": debut.isoformat(),
                    "to": fin.isoformat(),
                    "totals": self._totaux_vides(),
                    "teachers": [],
                }
            )

        identifiants = [enseignant.id for enseignant in enseignants]

        # Trois requetes pour toute la periode, et non trois par jour et par
        # enseignant: un mois pour vingt enseignants ferait plus de mille
        # allers-retours.
        creneaux_par_enseignant = {}
        for creneau in TeacherScheduleSlot.objects.select_related(
            "assignment",
            "assignment__subject",
            "assignment__classroom",
            "assignment__classroom__academic_year",
        ).filter(
            assignment__teacher_id__in=identifiants,
            assignment__classroom__academic_year__start_date__lte=fin,
            assignment__classroom__academic_year__end_date__gte=debut,
        ):
            creneaux_par_enseignant.setdefault(
                creneau.assignment.teacher_id, []
            ).append(creneau)

        pointages_par_cle = {}
        for pointage in (
            TeacherTimeEntry.objects.filter(
                teacher_id__in=identifiants,
                entry_date__gte=debut,
                entry_date__lte=fin,
            )
            .prefetch_related("slot_coverages")
            .order_by("entry_date", "check_in_time", "id")
        ):
            pointages_par_cle.setdefault(
                (pointage.teacher_id, pointage.entry_date), []
            ).append(pointage)

        lignes = [
            self._concordance_d_un_enseignant(
                enseignant,
                debut,
                fin,
                creneaux_par_enseignant.get(enseignant.id, []),
                pointages_par_cle,
            )
            for enseignant in enseignants
        ]
        # Les ecarts les plus lourds en tete: c'est ce qu'on ouvre l'ecran
        # pour voir, et non l'ordre alphabetique.
        lignes.sort(key=lambda ligne: (-ligne["totals"]["sessions_missed"], ligne["teacher_full_name"]))

        return Response(
            {
                "from": debut.isoformat(),
                "to": fin.isoformat(),
                "totals": self._cumuler(ligne["totals"] for ligne in lignes),
                "teachers": lignes,
            }
        )

    def _totaux_vides(self):
        return {
            "planned_minutes": 0,
            "covered_minutes": 0,
            "gap_minutes": 0,
            "sessions_planned": 0,
            "sessions_assured": 0,
            "sessions_partial": 0,
            "sessions_missed": 0,
            "off_schedule_entries": 0,
        }

    def _cumuler(self, totaux):
        cumul = self._totaux_vides()
        for total in totaux:
            for cle, valeur in total.items():
                cumul[cle] += valeur
        return cumul

    def _concordance_d_un_enseignant(self, enseignant, debut, fin, creneaux, pointages_par_cle):
        creneaux_par_jour = {}
        for creneau in creneaux:
            creneaux_par_jour.setdefault(creneau.day_of_week, []).append(creneau)
        for liste in creneaux_par_jour.values():
            liste.sort(key=lambda creneau: (creneau.start_time, creneau.end_time, creneau.id))

        jours = []
        totaux = self._totaux_vides()

        jour = debut
        while jour <= fin:
            code = self.DAY_CODES[jour.weekday()]
            du_jour = creneaux_par_jour.get(code, [])
            pointages = pointages_par_cle.get((enseignant.id, jour), [])

            if du_jour or pointages:
                # L'annee scolaire de la classe doit couvrir cette date: un
                # creneau de l'an prochain n'est pas une seance manquee.
                du_jour = [
                    creneau
                    for creneau in du_jour
                    if self._annee_couvre(creneau, jour)
                ]
                ligne = self._concordance_d_un_jour(jour, du_jour, pointages)
                if ligne["sessions"] or ligne["entries"]:
                    jours.append(ligne)
                    for cle, valeur in ligne["totals"].items():
                        totaux[cle] += valeur
            jour += timedelta(days=1)

        user = enseignant.user
        nom = user.get_full_name().strip() if user else ""
        return {
            "teacher": enseignant.id,
            "teacher_full_name": nom or (user.username if user else ""),
            "teacher_employee_code": enseignant.employee_code or "",
            "totals": totaux,
            "days": jours,
        }

    @staticmethod
    def _annee_couvre(creneau, jour):
        annee = getattr(creneau.assignment.classroom, "academic_year", None)
        if annee is None:
            return True
        if annee.start_date and annee.start_date > jour:
            return False
        if annee.end_date and annee.end_date < jour:
            return False
        return True

    def _concordance_d_un_jour(self, jour, creneaux, pointages):
        couvertures_par_creneau = {}
        for pointage in pointages:
            for couverture in pointage.slot_coverages.all():
                cumul = couvertures_par_creneau.get(couverture.schedule_slot_id)
                if cumul is None:
                    couvertures_par_creneau[couverture.schedule_slot_id] = {
                        "covered_minutes": couverture.covered_minutes,
                        "planned_minutes": couverture.planned_minutes,
                        "late_minutes": couverture.late_minutes,
                    }
                    continue
                # Deux pointages sur le meme cours: l'enseignant est sorti
                # puis revenu. Les minutes s'ajoutent, mais le retard reste
                # celui de la premiere arrivee -- la plus petite valeur.
                cumul["covered_minutes"] += couverture.covered_minutes
                cumul["planned_minutes"] = couverture.planned_minutes
                cumul["late_minutes"] = min(
                    cumul["late_minutes"], couverture.late_minutes
                )

        seances = []
        totaux = self._totaux_vides()

        for creneau in creneaux:
            planifie = max(
                self._minutes_de(creneau.end_time) - self._minutes_de(creneau.start_time),
                0,
            )
            couvert = couvertures_par_creneau.get(creneau.id)
            minutes_couvertes = couvert["covered_minutes"] if couvert else 0

            if minutes_couvertes <= 0:
                statut = "missed"
                totaux["sessions_missed"] += 1
            elif minutes_couvertes >= planifie:
                statut = "assured"
                totaux["sessions_assured"] += 1
            else:
                statut = "partial"
                totaux["sessions_partial"] += 1

            totaux["sessions_planned"] += 1
            totaux["planned_minutes"] += planifie
            totaux["covered_minutes"] += min(minutes_couvertes, planifie)

            seances.append(
                {
                    "slot": creneau.id,
                    "subject_name": creneau.assignment.subject.name,
                    "classroom_name": creneau.assignment.classroom.name,
                    "room": creneau.room,
                    "start_time": creneau.start_time.strftime("%H:%M"),
                    "end_time": creneau.end_time.strftime("%H:%M"),
                    "planned_minutes": planifie,
                    "covered_minutes": minutes_couvertes,
                    "late_minutes": couvert["late_minutes"] if couvert else 0,
                    "status": statut,
                }
            )

        entrees = []
        for pointage in pointages:
            hors_planning = pointage.covered_minutes == 0
            if hors_planning:
                totaux["off_schedule_entries"] += 1
            entrees.append(
                {
                    "id": pointage.id,
                    "check_in_time": pointage.check_in_time.strftime("%H:%M"),
                    "check_out_time": pointage.check_out_time.strftime("%H:%M")
                    if pointage.check_out_time
                    else None,
                    "is_auto_closed": pointage.is_auto_closed,
                    "is_off_schedule": hors_planning,
                    "off_schedule_reason": pointage.off_schedule_reason,
                    "worked_hours": str(pointage.worked_hours),
                }
            )

        totaux["gap_minutes"] = totaux["planned_minutes"] - totaux["covered_minutes"]

        return {
            "date": jour.isoformat(),
            "weekday": self.DAY_CODES[jour.weekday()],
            "sessions": seances,
            "entries": entrees,
            "totals": totaux,
        }


class DisciplineIncidentViewSet(AnneeScolaireScopeMixin, BaseModelViewSet):
    renseigne_annee_a_la_creation = True
    access_module = "discipline"
    queryset = DisciplineIncident.objects.select_related("student", "student__user", "reported_by").all().order_by("-incident_date", "-id")
    serializer_class = DisciplineIncidentSerializer
    filterset_fields = ["student", "severity", "status", "incident_date", "parent_notified"]
    # Le fil d'incidents se lit en cherchant un eleve ou un motif, pas en
    # faisant defiler: sans ces champs le frontend ne pouvait que tronquer.
    search_fields = [
        "category",
        "description",
        "sanction",
        "student__matricule",
        "student__user__first_name",
        "student__user__last_name",
    ]
    ordering_fields = ["incident_date", "severity", "status", "created_at"]
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]





    def _resolve_target_etablissement(self):
        user = self.request.user
        requested_etablissement = self._requested_etablissement()

        if getattr(user, "role", None) == "super_admin" and requested_etablissement:
            return requested_etablissement

        return getattr(user, "etablissement", None)

    def get_queryset(self):
        queryset = super().get_queryset()
        role = getattr(self.request.user, "role", "")

        if role == UserRole.STUDENT:
            return queryset.filter(student__user_id=self.request.user.id)
        if role == UserRole.PARENT:
            return queryset.filter(student__parent__user_id=self.request.user.id)
        if role == UserRole.TEACHER:
            allowed_classroom_ids = self._teacher_allowed_classroom_ids()
            if not allowed_classroom_ids:
                return queryset.none()
            return queryset.filter(student__classroom_id__in=allowed_classroom_ids)

        requested_etablissement = self._requested_etablissement()
        if requested_etablissement is not None:
            return queryset.filter(student__etablissement=requested_etablissement)

        if self._has_requested_scope():
            return queryset.none()

        if getattr(self.request.user, "role", None) == "super_admin":
            return queryset

        return queryset.filter(student__etablissement=getattr(self.request.user, "etablissement", None))

    def _validate_scope(self, serializer, instance=None):
        student = serializer.validated_data.get("student") or (instance.student if instance else None)
        if not student:
            return
        role = getattr(self.request.user, "role", "")
        if role == UserRole.TEACHER:
            allowed_classroom_ids = self._teacher_allowed_classroom_ids()
            if student.classroom_id not in allowed_classroom_ids:
                raise ValidationError(
                    {
                        "student": (
                            "Acces refuse: vous ne pouvez declarer que les incidents "
                            "des élèves de vos classes affectées."
                        )
                    }
                )
        target_etablissement = self._resolve_target_etablissement()
        if target_etablissement and student.etablissement_id != target_etablissement.id:
            raise ValidationError({"student": "L'eleve n'appartient pas a l'etablissement actif."})

    @action(detail=False, methods=["get"])
    def categories(self, request):
        """Referentiel des motifs, servi par le serveur.

        Le frontend proposait un champ de texte libre pre-rempli: recopier
        les neuf motifs dans l'application les aurait fait diverger du
        modele des la premiere evolution.
        """
        return Response(
            [
                {"value": value, "label": label}
                for value, label in DisciplineCategory.choices
            ]
        )

    def _notifier_le_parent(self, incident):
        """Trace l'information des parents dans le module Communication.

        `parent_notified` etait purement declaratif: une case cochee dans un
        coin de la fiche, sans trace, sans destinataire, et invisible du
        module qui gere precisement les envois. La notification est creee
        non envoyee -- aucune passerelle SMS n'expedie encore quoi que ce
        soit -- mais elle apparait desormais la ou on la cherche.
        """
        student = incident.student
        parent_profile = getattr(student, "parent", None)
        destinataire = getattr(parent_profile, "user", None)
        if destinataire is None:
            return

        etablissement = getattr(student, "etablissement", None)
        # SMS des qu'une passerelle est configuree: c'est le canal qui touche
        # les familles sans smartphone, majoritaires ici.
        canal = NotificationChannel.PUSH
        if etablissement is not None and SmsProviderConfig.objects.filter(
            etablissement=etablissement, is_active=True
        ).exists():
            canal = NotificationChannel.SMS

        eleve = ""
        if student.user:
            eleve = student.user.get_full_name().strip() or student.user.username

        Notification.objects.create(
            etablissement=etablissement,
            recipient=destinataire,
            channel=canal,
            title="Incident disciplinaire",
            message=(
                f"Un incident disciplinaire du {incident.incident_date} "
                f"concernant {eleve or 'votre enfant'} vous est signale: "
                f"{incident.get_category_display()}."
                + (f" Sanction: {incident.sanction}." if incident.sanction else "")
            ),
        )

    def perform_create(self, serializer):
        self._validate_scope(serializer)
        if getattr(self.request.user, "role", "") == UserRole.TEACHER:
            serializer.save(
                reported_by=self.request.user,
                status=DisciplineStatus.OPEN,
                sanction="",
                parent_notified=False,
            )
            return
        incident = serializer.save(reported_by=self.request.user)
        if incident.parent_notified:
            self._notifier_le_parent(incident)

    def perform_update(self, serializer):
        instance = self.get_object()
        self._validate_scope(serializer, instance=instance)
        # L'etat d'avant, lu avant la sauvegarde: c'est la transition qui
        # declenche l'information, pas la case cochee. Sans cette lecture,
        # chaque modification ulterieure d'un incident deja signale aurait
        # renvoye un message aux parents.
        deja_signale = instance.parent_notified
        incident = serializer.save()
        if incident.parent_notified and not deja_signale:
            self._notifier_le_parent(incident)

    def update(self, request, *args, **kwargs):
        if getattr(request.user, "role", None) == UserRole.TEACHER:
            raise PermissionDenied(
                "Acces refuse: un enseignant peut declarer un incident mais ne peut pas le valider ni le sanctionner."
            )
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        if getattr(request.user, "role", None) == UserRole.TEACHER:
            raise PermissionDenied(
                "Acces refuse: un enseignant peut declarer un incident mais ne peut pas le valider ni le sanctionner."
            )
        return super().partial_update(request, *args, **kwargs)


class StudentFeeViewSet(AnneeScolaireScopeMixin, BaseModelViewSet):
    access_module = "finance"
    queryset = StudentFee.objects.select_related("student", "student__user", "academic_year").all().order_by("-due_date", "-id")
    serializer_class = StudentFeeSerializer
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]
    pagination_class = StandardResultsSetPagination
    filterset_fields = ["student", "academic_year", "fee_type"]





    def _resolve_target_etablissement(self):
        user = self.request.user
        requested_etablissement = self._requested_etablissement()

        if getattr(user, "role", None) == "super_admin" and requested_etablissement:
            return requested_etablissement

        return getattr(user, "etablissement", None)

    def _validate_fee_scope(self, serializer, instance=None):
        student = serializer.validated_data.get("student") or (instance.student if instance else None)
        if not student:
            return

        role = getattr(self.request.user, "role", "")
        if role == UserRole.STUDENT and student.user_id != self.request.user.id:
            raise ValidationError({"student": "Vous ne pouvez creer/modifier que vos propres frais."})
        if role == UserRole.PARENT and student.parent_id:
            if student.parent.user_id != self.request.user.id:
                raise ValidationError({"student": "Vous ne pouvez creer/modifier que les frais de vos enfants."})

        target_etablissement = self._resolve_target_etablissement()
        if target_etablissement and student.etablissement_id != target_etablissement.id:
            raise ValidationError({"student": "L'eleve n'appartient pas a l'etablissement actif."})

    @staticmethod
    def _with_financial_annotations(queryset):
        paid_amount = Coalesce(
            Sum("payments__amount", filter=Q(payments__is_cancelled=False)),
            Value(0),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        )
        return queryset.annotate(
            amount_paid_annotated=paid_amount,
            balance_annotated=ExpressionWrapper(
                F("amount_due") - paid_amount,
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
        )

    def get_queryset(self):
        queryset = self._with_financial_annotations(super().get_queryset())
        role = getattr(self.request.user, "role", "")

        if role == UserRole.STUDENT:
            return queryset.filter(student__user_id=self.request.user.id)
        if role == UserRole.PARENT:
            return queryset.filter(student__parent__user_id=self.request.user.id)

        requested_etablissement = self._requested_etablissement()
        if requested_etablissement is not None:
            return queryset.filter(student__etablissement=requested_etablissement)

        if self._has_requested_scope():
            return queryset.none()

        if getattr(self.request.user, "role", None) == "super_admin":
            return queryset

        return queryset.filter(student__etablissement=getattr(self.request.user, "etablissement", None))

    def perform_create(self, serializer):
        self._validate_fee_scope(serializer)
        serializer.save()

    def perform_update(self, serializer):
        self._validate_fee_scope(serializer, instance=self.get_object())
        serializer.save()


class PaymentViewSet(BaseModelViewSet):
    access_module = "finance"
    queryset = Payment.objects.filter(is_cancelled=False)
    serializer_class = PaymentSerializer
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]
    pagination_class = StandardResultsSetPagination
    filterset_fields = ["fee", "fee__student", "method", "received_by"]
    search_fields = [
        "reference",
        "method",
        "fee__fee_type",
        "fee__student__matricule",
        "fee__student__user__first_name",
        "fee__student__user__last_name",
    ]
    ordering_fields = ["created_at", "amount", "method"]
    # "-id" en second: sur des lignes creees dans la meme seconde,
    # "-created_at" seul laisse l_ordre indefini et une meme ligne peut
    # apparaitre sur deux pages, ou sur aucune.
    ordering = ["-created_at", "-id"]





    def _resolve_target_etablissement(self):
        user = self.request.user
        requested_etablissement = self._requested_etablissement()

        if getattr(user, "role", None) == "super_admin" and requested_etablissement:
            return requested_etablissement

        return getattr(user, "etablissement", None)

    def _validate_payment_scope(self, serializer, instance=None):
        fee = serializer.validated_data.get("fee") or (instance.fee if instance else None)
        if not fee:
            return

        student = fee.student
        target_etablissement = self._resolve_target_etablissement()
        if target_etablissement and student.etablissement_id != target_etablissement.id:
            raise ValidationError({"fee": "Le frais selectionne n'appartient pas a l'etablissement actif."})

    def _log_payment_action(self, *, action, payment, details=""):
        user = self.request.user
        ActivityLog.objects.create(
            user=user if user and user.is_authenticated else None,
            etablissement=getattr(payment, "etablissement", None),
            role=getattr(user, "role", "") or "",
            action=action,
            method=self.request.method,
            path=self.request.path,
            module="finance",
            target=f"payment:{payment.id}",
            status_code=200,
            success=True,
            details=details,
        )

    def get_queryset(self):
        user = self.request.user
        qs = Payment.objects.select_related("fee", "fee__student", "fee__student__user", "received_by").filter(
            is_cancelled=False
        ).order_by("-created_at")
        role = getattr(user, "role", "")
        if role == UserRole.STUDENT:
            return qs.filter(fee__student__user_id=user.id)
        if role == UserRole.PARENT:
            return qs.filter(fee__student__parent__user_id=user.id)
        requested_etablissement = self._requested_etablissement()
        if requested_etablissement is not None:
            return qs.filter(etablissement=requested_etablissement)
        if self._has_requested_scope():
            return qs.none()
        if hasattr(user, "role") and user.role == "super_admin":
            return qs.all()
        return qs.filter(etablissement=user.etablissement)

    def perform_create(self, serializer):
        self._validate_payment_scope(serializer)
        payment = serializer.save(
            etablissement=self._resolve_target_etablissement(),
            received_by=self.request.user,
        )
        self._log_payment_action(
            action="payment_created",
            payment=payment,
            details=f"fee={payment.fee_id};amount={payment.amount};method={payment.method}",
        )

    def perform_update(self, serializer):
        self._validate_payment_scope(serializer, instance=self.get_object())
        payment = serializer.save(etablissement=self._resolve_target_etablissement())
        self._log_payment_action(
            action="payment_updated",
            payment=payment,
            details=f"fee={payment.fee_id};amount={payment.amount};method={payment.method}",
        )

    def destroy(self, request, *args, **kwargs):
        payment = self.get_object()
        if payment.is_cancelled:
            return Response(status=status.HTTP_204_NO_CONTENT)

        reason = str(request.data.get("reason", "") if isinstance(request.data, dict) else "").strip()
        payment.cancel(user=request.user, reason=reason or "Annulation depuis ecran finance")
        self._log_payment_action(
            action="payment_cancelled",
            payment=payment,
            details=f"reason={payment.cancel_reason};amount={payment.amount};fee={payment.fee_id}",
        )
        return Response(status=status.HTTP_204_NO_CONTENT)


class ExpenseViewSet(AnneeScolaireScopeMixin, BaseModelViewSet):
    renseigne_annee_a_la_creation = True
    access_module = "finance"
    queryset = Expense.objects.select_related(
        "paid_by",
        "level_one_validated_by",
        "level_two_validated_by",
    ).all().order_by("-date", "-id")
    serializer_class = ExpenseSerializer
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]

    def get_permissions(self):
        if self.action in {"validate_level_one", "validate_level_two", "reset_validation"}:
            return [permissions.IsAuthenticated()]
        return super().get_permissions()

    @staticmethod
    def _can_validate_level_one(role):
        return role in {UserRole.CENSOR, UserRole.SUPER_ADMIN}

    @staticmethod
    def _can_validate_level_two(role):
        return role in {UserRole.ACCOUNTANT, UserRole.SUPER_ADMIN}





    def _resolve_target_etablissement(self):
        user = self.request.user
        requested_etablissement = self._requested_etablissement()

        if getattr(user, "role", None) == "super_admin" and requested_etablissement:
            return requested_etablissement

        return getattr(user, "etablissement", None)

    def get_queryset(self):
        user = self.request.user
        qs = super().get_queryset()
        requested_etablissement = self._requested_etablissement()

        if requested_etablissement is not None:
            return qs.filter(etablissement=requested_etablissement)
        if self._has_requested_scope():
            return qs.none()
        if getattr(user, "role", None) == "super_admin":
            return qs
        return qs.filter(etablissement=getattr(user, "etablissement", None))

    def perform_create(self, serializer):
        target_etablissement = self._resolve_target_etablissement()
        if target_etablissement is None:
            raise ValidationError({"detail": "Etablissement actif requis pour creer une depense."})
        serializer.save(etablissement=target_etablissement)

    def perform_update(self, serializer):
        instance = self.get_object()
        if instance.level_two_validated_at:
            raise ValidationError(
                {"detail": "La depense est validee niveau 2. Modification impossible sans reinitialisation."}
            )
        target_etablissement = self._resolve_target_etablissement()
        if target_etablissement is None:
            raise ValidationError({"detail": "Etablissement actif requis pour modifier une depense."})
        serializer.save(etablissement=target_etablissement)

    def destroy(self, request, *args, **kwargs):
        expense = self.get_object()
        if expense.level_two_validated_at:
            raise ValidationError(
                {"detail": "La depense est validee niveau 2. Suppression interdite sans reinitialisation."}
            )
        return super().destroy(request, *args, **kwargs)

    @action(
        detail=True,
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated, HasModuleAccess],
    )
    def validate_level_one(self, request, pk=None):
        role = getattr(request.user, "role", "")
        if not self._can_validate_level_one(role):
            raise ValidationError({"detail": "Acces refuse: validation niveau 1 reservee au surveillant/super admin."})

        expense = self.get_object()
        if expense.level_two_validated_at:
            raise ValidationError({"detail": "La depense est deja validee niveau 2."})

        expense.level_one_validated_by = request.user
        expense.level_one_validated_at = timezone.now()
        expense.save(update_fields=["level_one_validated_by", "level_one_validated_at", "updated_at"])

        return Response(
            {
                "detail": "Validation niveau 1 enregistree.",
                "id": expense.id,
                "validation_stage": expense.validation_stage,
            }
        )

    @action(
        detail=True,
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated, HasModuleAccess],
    )
    def validate_level_two(self, request, pk=None):
        role = getattr(request.user, "role", "")
        if not self._can_validate_level_two(role):
            raise ValidationError({"detail": "Acces refuse: validation niveau 2 reservee au comptable/super admin."})

        expense = self.get_object()
        if not expense.level_one_validated_at:
            raise ValidationError({"detail": "Validation niveau 1 requise avant la validation finale."})

        expense.level_two_validated_by = request.user
        expense.level_two_validated_at = timezone.now()
        expense.paid_by = request.user
        expense.paid_on = timezone.localdate()
        expense.save(
            update_fields=[
                "level_two_validated_by",
                "level_two_validated_at",
                "paid_by",
                "paid_on",
                "updated_at",
            ]
        )

        return Response(
            {
                "detail": "Validation niveau 2 enregistree. Depense verrouillee.",
                "id": expense.id,
                "validation_stage": expense.validation_stage,
            }
        )

    @action(
        detail=True,
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated, HasModuleAccess],
    )
    def reset_validation(self, request, pk=None):
        if getattr(request.user, "role", "") != UserRole.SUPER_ADMIN:
            raise ValidationError({"detail": "Seul le super admin peut reinitialiser la validation."})

        expense = self.get_object()
        expense.level_one_validated_by = None
        expense.level_one_validated_at = None
        expense.level_two_validated_by = None
        expense.level_two_validated_at = None
        expense.paid_by = None
        expense.paid_on = None
        expense.save(
            update_fields=[
                "level_one_validated_by",
                "level_one_validated_at",
                "level_two_validated_by",
                "level_two_validated_at",
                "paid_by",
                "paid_on",
                "updated_at",
            ]
        )

        return Response(
            {
                "detail": "Validation de la depense reinitialisee.",
                "id": expense.id,
                "validation_stage": expense.validation_stage,
            }
        )


class TeacherPayrollViewSet(AnneeScolaireScopeMixin, BaseModelViewSet):
    renseigne_annee_a_la_creation = True
    access_module = "payroll"
    queryset = TeacherPayroll.objects.select_related(
        "teacher",
        "teacher__user",
        "paid_by",
        "level_one_validated_by",
        "level_two_validated_by",
    ).all().order_by("-paid_on", "-id")
    serializer_class = TeacherPayrollSerializer
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]
    filterset_fields = ["teacher", "month"]





    def _resolve_target_etablissement(self):
        user = self.request.user
        requested_etablissement = self._requested_etablissement()

        if getattr(user, "role", None) == "super_admin" and requested_etablissement:
            return requested_etablissement

        return getattr(user, "etablissement", None)

    def _validate_payroll_scope(self, serializer, instance=None):
        teacher = serializer.validated_data.get("teacher") or (instance.teacher if instance else None)
        if not teacher:
            return
        target_etablissement = self._resolve_target_etablissement()
        if target_etablissement and teacher.etablissement_id != target_etablissement.id:
            raise ValidationError({"teacher": "L'enseignant n'appartient pas a l'etablissement actif."})

    def get_queryset(self):
        user = self.request.user
        qs = super().get_queryset()
        requested_etablissement = self._requested_etablissement()

        if requested_etablissement is not None:
            return qs.filter(
                Q(teacher__etablissement=requested_etablissement)
                | Q(teacher__etablissement__isnull=True, teacher__user__etablissement=requested_etablissement)
            )
        if self._has_requested_scope():
            return qs.none()
        if getattr(user, "role", None) == "super_admin":
            return qs
        user_etablissement = getattr(user, "etablissement", None)
        return qs.filter(
            Q(teacher__etablissement=user_etablissement)
            | Q(teacher__etablissement__isnull=True, teacher__user__etablissement=user_etablissement)
        )

    def perform_create(self, serializer):
        self._validate_payroll_scope(serializer)
        hourly_rate = serializer.validated_data.get("hourly_rate")
        hours_worked = serializer.validated_data.get("hours_worked")
        if hourly_rate is not None and hours_worked is not None:
            serializer.validated_data["amount"] = (
                Decimal(str(hourly_rate)) * Decimal(str(hours_worked))
            ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        serializer.save(paid_by=self.request.user)

    def perform_update(self, serializer):
        instance = self.get_object()
        self._validate_payroll_scope(serializer, instance=instance)

        if instance.level_two_validated_at:
            raise ValidationError(
                {"detail": "La fiche est validée niveau 2. Modification impossible sans réinitialisation."}
            )

        hourly_rate = serializer.validated_data.get("hourly_rate", instance.hourly_rate)
        hours_worked = serializer.validated_data.get("hours_worked", instance.hours_worked)
        serializer.validated_data["amount"] = (
            Decimal(str(hourly_rate)) * Decimal(str(hours_worked))
        ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        serializer.save(paid_by=self.request.user)

    def _month_range(self, month_value):
        if not month_value:
            today = timezone.now().date()
            month_start = today.replace(day=1)
        else:
            normalized = str(month_value).strip()
            if len(normalized) == 7:
                normalized = f"{normalized}-01"
            try:
                month_start = datetime.strptime(normalized, "%Y-%m-%d").date().replace(day=1)
            except ValueError:
                raise ValidationError({"month": "Format invalide. Utilisez YYYY-MM ou YYYY-MM-DD."})

        next_month = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
        month_end = next_month - timedelta(days=1)
        return month_start, month_end

    def _teacher_hours_attributed(self, teacher, month_start, month_end):
        weekday_counts = {"MON": 0, "TUE": 0, "WED": 0, "THU": 0, "FRI": 0, "SAT": 0}
        current = month_start
        day_map = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
        while current <= month_end:
            day_key = day_map[current.weekday()]
            if day_key in weekday_counts:
                weekday_counts[day_key] += 1
            current += timedelta(days=1)

        slots = TeacherScheduleSlot.objects.select_related("assignment").filter(
            assignment__teacher=teacher,
            assignment__classroom__academic_year__start_date__lte=month_end,
            assignment__classroom__academic_year__end_date__gte=month_start,
        )

        total = Decimal("0.00")
        for slot in slots:
            start_dt = datetime.combine(date.today(), slot.start_time)
            end_dt = datetime.combine(date.today(), slot.end_time)
            if end_dt <= start_dt:
                continue
            hours = Decimal(str((end_dt - start_dt).total_seconds() / 3600)).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP,
            )
            count = weekday_counts.get(slot.day_of_week, 0)
            total += (hours * Decimal(str(count))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        return total.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    def _teacher_hours_worked(self, teacher, month_start, month_end):
        total = (
            TeacherTimeEntry.objects.filter(
                teacher=teacher,
                entry_date__gte=month_start,
                entry_date__lte=month_end,
            ).aggregate(total=Sum("worked_hours"))["total"]
            or Decimal("0.00")
        )
        return Decimal(str(total)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    def _teacher_concordance_hours(self, teacher, month_start, month_end):
        """Ce que l'ecart entre attribue et travaille cache vraiment.

        Deux causes bien distinctes, qui ne se pilotent pas de la meme
        facon: des seances planifiees que personne n'a assurees, et des
        heures faites hors planning -- remplacement, reunion. Additionnees
        dans un ecart unique, elles ne disaient rien.
        """
        entrees = list(
            TeacherTimeEntry.objects.filter(
                teacher=teacher,
                entry_date__gte=month_start,
                entry_date__lte=month_end,
            ).prefetch_related("slot_coverages")
        )

        minutes_planifiees = 0
        minutes_couvertes = 0
        minutes_hors_planning = 0
        jours_pointes = set()

        for entree in entrees:
            jours_pointes.add(entree.entry_date)
            minutes_couvertes += entree.covered_minutes
            if entree.covered_minutes == 0 and entree.check_out_time:
                minutes_hors_planning += max(
                    (entree.check_out_time.hour * 60 + entree.check_out_time.minute)
                    - (entree.check_in_time.hour * 60 + entree.check_in_time.minute),
                    0,
                )

        # Les jours planifies sans le moindre pointage comptent aussi: c'est
        # meme le cas le plus grave, et aucune ligne de pointage ne le porte.
        for entree in entrees:
            minutes_planifiees += entree.planned_minutes

        minutes_planifiees += self._minutes_planifiees_sans_pointage(
            teacher, month_start, month_end, jours_pointes
        )

        manquees = max(minutes_planifiees - minutes_couvertes, 0)
        return (
            Decimal(str(manquees / 60)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
            Decimal(str(minutes_hors_planning / 60)).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            ),
        )

    def _minutes_planifiees_sans_pointage(self, teacher, month_start, month_end, jours_pointes):
        """Les cours des journees ou l'enseignant n'a pas pointe du tout."""
        day_codes = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
        creneaux_par_jour = {}
        for creneau in TeacherScheduleSlot.objects.filter(
            assignment__teacher=teacher,
            assignment__classroom__academic_year__start_date__lte=month_end,
            assignment__classroom__academic_year__end_date__gte=month_start,
        ):
            duree = max(
                (creneau.end_time.hour * 60 + creneau.end_time.minute)
                - (creneau.start_time.hour * 60 + creneau.start_time.minute),
                0,
            )
            creneaux_par_jour[creneau.day_of_week] = (
                creneaux_par_jour.get(creneau.day_of_week, 0) + duree
            )

        total = 0
        jour = month_start
        while jour <= month_end:
            if jour not in jours_pointes:
                total += creneaux_par_jour.get(day_codes[jour.weekday()], 0)
            jour += timedelta(days=1)
        return total

    def _auto_close_missing_entries(self, teacher, month_start, month_end):
        open_entries = TeacherTimeEntry.objects.filter(
            teacher=teacher,
            entry_date__gte=month_start,
            entry_date__lte=month_end,
            check_out_time__isnull=True,
        ).order_by("entry_date", "id")

        closed = 0
        for entry in open_entries:
            entry.save()
            closed += 1
        return closed

    @staticmethod
    def _can_validate_level_one(role):
        return role in {UserRole.CENSOR, UserRole.SUPER_ADMIN}

    @staticmethod
    def _can_validate_level_two(role):
        return role in {UserRole.ACCOUNTANT, UserRole.SUPER_ADMIN}

    @action(detail=False, methods=["post"], permission_classes=[permissions.IsAuthenticated, HasModuleAccess])
    def generate_monthly(self, request):
        month_start, month_end = self._month_range(request.data.get("month"))
        teacher_id = request.data.get("teacher")
        force_regenerate = bool(request.data.get("force_regenerate", False))

        if force_regenerate and getattr(request.user, "role", None) != UserRole.SUPER_ADMIN:
            raise ValidationError({"detail": "Seul le super admin peut forcer la régénération."})

        qs = Teacher.objects.select_related("etablissement", "user").all()
        requested_etablissement = self._requested_etablissement()
        if requested_etablissement is not None:
            qs = qs.filter(
                Q(etablissement=requested_etablissement)
                | Q(etablissement__isnull=True, user__etablissement=requested_etablissement)
            )
        elif self._has_requested_scope():
            qs = Teacher.objects.none()
        elif getattr(request.user, "role", None) != UserRole.SUPER_ADMIN:
            user_etablissement = getattr(request.user, "etablissement", None)
            qs = qs.filter(
                Q(etablissement=user_etablissement)
                | Q(etablissement__isnull=True, user__etablissement=user_etablissement)
            )

        if teacher_id not in (None, ""):
            qs = qs.filter(id=teacher_id)

        generated_ids = []
        skipped_final = 0
        auto_closed_entries = 0
        for teacher in qs:
            auto_closed_entries += self._auto_close_missing_entries(teacher, month_start, month_end)

            hours_attributed = self._teacher_hours_attributed(teacher, month_start, month_end)
            hours_worked = self._teacher_hours_worked(teacher, month_start, month_end)
            hours_missed, hours_off_schedule = self._teacher_concordance_hours(
                teacher, month_start, month_end
            )
            hourly_rate = Decimal(str(teacher.hourly_rate or teacher.salary_base or 0)).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP,
            )
            amount = (hours_worked * hourly_rate).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

            existing = TeacherPayroll.objects.filter(teacher=teacher, month=month_start).first()
            if existing and existing.level_two_validated_at and not force_regenerate:
                skipped_final += 1
                continue

            if existing and existing.level_two_validated_at and force_regenerate:
                existing.level_one_validated_by = None
                existing.level_one_validated_at = None
                existing.level_two_validated_by = None
                existing.level_two_validated_at = None
                existing.save(
                    update_fields=[
                        "level_one_validated_by",
                        "level_one_validated_at",
                        "level_two_validated_by",
                        "level_two_validated_at",
                        "updated_at",
                    ]
                )

            payroll, _ = TeacherPayroll.objects.update_or_create(
                teacher=teacher,
                month=month_start,
                defaults={
                    "hours_attributed": hours_attributed,
                    "hours_worked": hours_worked,
                    "hours_missed": hours_missed,
                    "hours_off_schedule": hours_off_schedule,
                    "hourly_rate": hourly_rate,
                    "amount": amount,
                    "paid_by": request.user,
                },
            )
            generated_ids.append(payroll.id)

        queryset = self.get_queryset().filter(id__in=generated_ids)
        serializer = self.get_serializer(queryset, many=True)
        return Response(
            {
                "month": month_start.strftime("%Y-%m"),
                "count": len(generated_ids),
                "auto_closed_entries": auto_closed_entries,
                "skipped_final_validated": skipped_final,
                "results": serializer.data,
            }
        )

    @action(
        detail=True,
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
    )
    def validate_level_one(self, request, pk=None):
        role = getattr(request.user, "role", "")
        if not self._can_validate_level_one(role):
            raise ValidationError({"detail": "Accès refusé: validation niveau 1 réservée au surveillant/super admin."})

        payroll = self.get_object()
        if payroll.level_two_validated_at:
            raise ValidationError({"detail": "La fiche est déjà validée niveau 2."})

        payroll.level_one_validated_by = request.user
        payroll.level_one_validated_at = timezone.now()
        payroll.save(update_fields=["level_one_validated_by", "level_one_validated_at", "updated_at"])

        return Response(
            {
                "detail": "Validation niveau 1 enregistrée.",
                "id": payroll.id,
                "validation_stage": payroll.validation_stage,
            }
        )

    @action(
        detail=True,
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
    )
    def validate_level_two(self, request, pk=None):
        role = getattr(request.user, "role", "")
        if not self._can_validate_level_two(role):
            raise ValidationError({"detail": "Accès refusé: validation niveau 2 réservée au comptable/super admin."})

        payroll = self.get_object()
        if not payroll.level_one_validated_at:
            raise ValidationError({"detail": "Validation niveau 1 requise avant la validation finale."})

        payroll.level_two_validated_by = request.user
        payroll.level_two_validated_at = timezone.now()
        payroll.paid_by = request.user
        payroll.save(
            update_fields=[
                "level_two_validated_by",
                "level_two_validated_at",
                "paid_by",
                "updated_at",
            ]
        )

        return Response(
            {
                "detail": "Validation niveau 2 enregistrée. Fiche finale verrouillée.",
                "id": payroll.id,
                "validation_stage": payroll.validation_stage,
            }
        )

    @action(
        detail=True,
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
    )
    def reset_validation(self, request, pk=None):
        if getattr(request.user, "role", "") != UserRole.SUPER_ADMIN:
            raise ValidationError({"detail": "Seul le super admin peut réinitialiser la validation."})

        payroll = self.get_object()
        payroll.level_one_validated_by = None
        payroll.level_one_validated_at = None
        payroll.level_two_validated_by = None
        payroll.level_two_validated_at = None
        payroll.save(
            update_fields=[
                "level_one_validated_by",
                "level_one_validated_at",
                "level_two_validated_by",
                "level_two_validated_at",
                "updated_at",
            ]
        )
        return Response(
            {
                "detail": "Validation réinitialisée.",
                "id": payroll.id,
                "validation_stage": payroll.validation_stage,
            }
        )


class AnnouncementViewSet(EtablissementScopedModelViewSet):
    access_module = "communication"
    queryset = Announcement.objects.select_related("author", "etablissement").all().order_by("-created_at")
    serializer_class = AnnouncementSerializer
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]

    def get_queryset(self):
        return self._filter_by_scope(super().get_queryset(), field_name="etablissement")

    def perform_create(self, serializer):
        target_etablissement = self._resolve_target_etablissement()
        if getattr(self.request.user, "role", None) == UserRole.SUPER_ADMIN and target_etablissement is None:
            raise ValidationError({"etablissement": "Selectionnez un etablissement actif."})
        serializer.save(author=self.request.user, etablissement=target_etablissement)

    def perform_update(self, serializer):
        target_etablissement = self._resolve_target_etablissement()
        if getattr(self.request.user, "role", None) == UserRole.SUPER_ADMIN and target_etablissement is None:
            raise ValidationError({"etablissement": "Selectionnez un etablissement actif."})
        serializer.save(etablissement=target_etablissement)


class NotificationViewSet(EtablissementScopedModelViewSet):
    access_module = "communication"
    queryset = Notification.objects.select_related("recipient", "etablissement").all().order_by("-created_at")
    serializer_class = NotificationSerializer
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]

    def get_queryset(self):
        return self._filter_by_scope(super().get_queryset(), field_name="etablissement")

    def _validate_scope(self, serializer):
        recipient = serializer.validated_data.get("recipient")
        target_etablissement = self._resolve_target_etablissement()
        if not recipient or not target_etablissement:
            return
        if recipient.etablissement_id not in (None, target_etablissement.id):
            raise ValidationError({"recipient": "Le destinataire n'appartient pas a l'etablissement actif."})

    def perform_create(self, serializer):
        target_etablissement = self._resolve_target_etablissement()
        if getattr(self.request.user, "role", None) == UserRole.SUPER_ADMIN and target_etablissement is None:
            raise ValidationError({"etablissement": "Selectionnez un etablissement actif."})
        self._validate_scope(serializer)
        serializer.save(etablissement=target_etablissement)

    def perform_update(self, serializer):
        target_etablissement = self._resolve_target_etablissement()
        if getattr(self.request.user, "role", None) == UserRole.SUPER_ADMIN and target_etablissement is None:
            raise ValidationError({"etablissement": "Selectionnez un etablissement actif."})
        self._validate_scope(serializer)
        serializer.save(etablissement=target_etablissement)


class SmsProviderConfigViewSet(EtablissementScopedModelViewSet):
    access_module = "sms_config"
    queryset = SmsProviderConfig.objects.select_related("etablissement").all().order_by("-id")
    serializer_class = SmsProviderConfigSerializer
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]

    def get_queryset(self):
        return self._filter_by_scope(super().get_queryset(), field_name="etablissement")

    def perform_create(self, serializer):
        target_etablissement = self._resolve_target_etablissement()
        if getattr(self.request.user, "role", None) == UserRole.SUPER_ADMIN and target_etablissement is None:
            raise ValidationError({"etablissement": "Selectionnez un etablissement actif."})
        serializer.save(etablissement=target_etablissement)

    def perform_update(self, serializer):
        target_etablissement = self._resolve_target_etablissement()
        if getattr(self.request.user, "role", None) == UserRole.SUPER_ADMIN and target_etablissement is None:
            raise ValidationError({"etablissement": "Selectionnez un etablissement actif."})
        serializer.save(etablissement=target_etablissement)


class _LibraryEtagereMixin:
    """La regle d'etagere, commune aux series, aux matieres et aux documents.

    Le fonds importe n'est pas cloisonne: ce sont les memes annales pour
    tout le monde, et les dupliquer par etablissement multiplierait des
    giga-octets identiques. Ce qu'une ecole depose, en revanche, ne regarde
    qu'elle -- son reglement interieur n'a rien a faire chez la voisine.

    D'ou la meme lecture partout: le commun, plus le sien. Et la meme
    ecriture partout: on ne touche pas au commun depuis l'application, il
    n'entre que par la commande d'import.
    """

    def _etablissement_courant(self):
        """L'ecole au nom de laquelle on lit et on ecrit.

        Pour tout le monde sauf le super-administrateur, `initial()` a deja
        force l'en-tete a l'etablissement du compte: nul ne lit l'etagere
        d'une autre ecole en changeant un parametre d'URL.
        """
        return self._resolve_effective_etablissement_for_create()

    def _filtrer_par_etagere(self, queryset, prefixe=""):
        etablissement = self._etablissement_courant()
        champ = f"{prefixe}etablissement"
        if etablissement is None:
            # Super-administrateur sans etablissement actif: la vue d'ensemble.
            if getattr(self.request.user, "role", None) == UserRole.SUPER_ADMIN:
                return queryset
            return queryset.filter(**{f"{champ}__isnull": True})
        return queryset.filter(
            Q(**{f"{champ}__isnull": True}) | Q(**{champ: etablissement})
        )

    def _refuser_le_fonds_commun(self, etablissement_id, quoi):
        """Le fonds importe ne se modifie pas depuis l'application.

        Sans cette borne, l'ecriture accordee par la matrice sur « library »
        laisserait un surveillant renommer une annale pour les neuf ecoles a
        la fois -- et la prochaine passe d'import la remettrait comme avant,
        sans que personne comprenne pourquoi.
        """
        if etablissement_id is None:
            raise ValidationError(
                {"detail": f"Le fonds commun ne peut pas être {quoi} depuis l'application."}
            )


class LibraryCollectionViewSet(_LibraryEtagereMixin, BaseModelViewSet):
    """Les series du fonds numerique, avec leurs matieres et leurs compteurs.

    Qui accede au module voit les neuf series importees, plus les etageres
    creees par son propre etablissement.
    """

    access_module = "library"
    serializer_class = LibraryCollectionSerializer
    queryset = LibraryCollection.objects.all()
    pagination_class = None

    def get_queryset(self):
        # Les compteurs viennent d'annotations: les calculer dans le
        # serializer ferait une requete par serie et par matiere.
        categories = LibraryCategory.objects.annotate(
            document_count=Count("documents")
        ).order_by("position", "name")
        return self._filtrer_par_etagere(
            LibraryCollection.objects.annotate(
                document_count=Count("categories__documents")
            )
            .prefetch_related(Prefetch("categories", queryset=categories))
            .order_by("position", "label")
        )

    def perform_create(self, serializer):
        etablissement = self._etablissement_courant()
        if etablissement is None:
            raise ValidationError(
                {"etablissement": "Sélectionnez un établissement actif."}
            )
        serializer.save(etablissement=etablissement)

    def perform_update(self, serializer):
        self._refuser_le_fonds_commun(serializer.instance.etablissement_id, "modifié")
        serializer.save()

    def perform_destroy(self, instance):
        self._refuser_le_fonds_commun(instance.etablissement_id, "supprimé")
        instance.delete()


class LibraryCategoryViewSet(_LibraryEtagereMixin, BaseModelViewSet):
    """Les matieres d'une serie, creables une a une.

    Une matiere ne vit pas sans sa serie et suit son sort: c'est la serie
    qui porte l'etablissement, la matiere n'en a pas besoin.
    """

    access_module = "library"
    serializer_class = LibraryCategoryWriteSerializer
    queryset = LibraryCategory.objects.select_related("collection").all()
    pagination_class = None

    def get_queryset(self):
        queryset = self._filtrer_par_etagere(
            super().get_queryset().annotate(document_count=Count("documents")),
            prefixe="collection__",
        )
        collection = self.request.query_params.get("collection")
        if collection not in (None, "") and str(collection).isdigit():
            queryset = queryset.filter(collection_id=collection)
        return queryset.order_by("position", "name")

    def _verifier_la_serie(self, serializer):
        collection = serializer.validated_data.get("collection")
        if collection is None:
            return
        self._refuser_le_fonds_commun(collection.etablissement_id, "modifié")
        etablissement = self._etablissement_courant()
        if etablissement and collection.etablissement_id != etablissement.id:
            raise ValidationError(
                {"collection": "Cette série n'appartient pas à l'établissement actif."}
            )

    def perform_create(self, serializer):
        self._verifier_la_serie(serializer)
        serializer.save()

    def perform_update(self, serializer):
        self._refuser_le_fonds_commun(
            serializer.instance.collection.etablissement_id, "modifié"
        )
        self._verifier_la_serie(serializer)
        serializer.save()

    def perform_destroy(self, instance):
        self._refuser_le_fonds_commun(
            instance.collection.etablissement_id, "supprimé"
        )
        instance.delete()


class LibraryDocumentViewSet(_LibraryEtagereMixin, BaseModelViewSet):
    """Les PDF du fonds: liste filtrable, lecture, et depot d'un document.

    Le fichier passe toujours par cette API, jamais par une URL exterieure
    donnee au navigateur: tant qu'un document n'est pas rapatrie, le serveur
    relaie la source. C'est ce qui rend le mode hybride invisible au client
    -- et ce qui evite un echec CORS sur un domaine tiers.

    Un document depose par une ecole ne connait pas ce mode hybride: il
    arrive avec son fichier, il est donc rapatrie des la premiere seconde.
    """

    access_module = "library"
    serializer_class = LibraryDocumentSerializer
    queryset = LibraryDocument.objects.select_related(
        "category", "category__collection", "uploaded_by"
    ).all()

    # Le relais ne sort pas de la source connue: sans cette borne, un
    # source_url modifie en base ferait du serveur un proxy ouvert.
    PROXY_ALLOWED_HOSTS = ("bkalan.ml", "www.bkalan.ml")

    def get_queryset(self):
        queryset = self._filtrer_par_etagere(super().get_queryset())

        collection = self.request.query_params.get("collection")
        if collection not in (None, ""):
            queryset = (
                queryset.filter(category__collection_id=collection)
                if str(collection).isdigit()
                else queryset.filter(category__collection__code__iexact=collection)
            )

        category = self.request.query_params.get("category")
        if category not in (None, ""):
            queryset = (
                queryset.filter(category_id=category)
                if str(category).isdigit()
                else queryset.filter(category__name__iexact=category)
            )

        search = (self.request.query_params.get("search") or "").strip()
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) | Q(description__icontains=search)
            )

        origine = (self.request.query_params.get("origin") or "").strip()
        if origine in LibraryDocumentOrigin.values:
            queryset = queryset.filter(origin=origine)

        return queryset

    # --- Depot d'un document ------------------------------------------------

    def _verifier_la_matiere(self, categorie):
        """La matiere visee doit etre une etagere de l'ecole, pas le commun."""
        if categorie is None:
            return
        self._refuser_le_fonds_commun(
            categorie.collection.etablissement_id, "alimenté"
        )
        etablissement = self._etablissement_courant()
        if etablissement and categorie.collection.etablissement_id != etablissement.id:
            raise ValidationError(
                {"category": "Cette matière n'appartient pas à l'établissement actif."}
            )

    def perform_create(self, serializer):
        etablissement = self._etablissement_courant()
        if etablissement is None:
            raise ValidationError(
                {"etablissement": "Sélectionnez un établissement actif."}
            )
        self._verifier_la_matiere(serializer.validated_data.get("category"))

        fichier = serializer.validated_data.get("file")
        serializer.save(
            etablissement=etablissement,
            origin=LibraryDocumentOrigin.UPLOAD,
            uploaded_by=self.request.user,
            # Le poids et l'etat viennent du fichier lui-meme et jamais du
            # client: un `size_bytes` annonce serait cru sur parole, et c'est
            # lui qui decide de la progression affichee a la lecture.
            size_bytes=fichier.size if fichier else 0,
            is_downloaded=bool(fichier),
        )

    def perform_update(self, serializer):
        self._refuser_le_fonds_commun(
            serializer.instance.etablissement_id, "modifié"
        )
        if "category" in serializer.validated_data:
            self._verifier_la_matiere(serializer.validated_data["category"])

        fichier = serializer.validated_data.get("file")
        if fichier is None:
            serializer.save()
            return
        # Remplacement du fichier: l'ancien ne sert plus personne et pese sur
        # le stockage. Il part apres l'enregistrement du nouveau -- l'ordre
        # inverse laisserait le document sans fichier si l'ecriture echouait.
        ancien = serializer.instance.file
        ancien_nom = ancien.name if ancien else ""
        serializer.save(size_bytes=fichier.size, is_downloaded=True)
        if ancien_nom and ancien_nom != serializer.instance.file.name:
            ancien.storage.delete(ancien_nom)

    def perform_destroy(self, instance):
        self._refuser_le_fonds_commun(instance.etablissement_id, "supprimé")
        fichier = instance.file
        nom = fichier.name if fichier else ""
        instance.delete()
        if nom:
            # Apres la suppression en base: un fichier efface alors que la
            # ligne survit laisserait un document illisible dans la liste.
            fichier.storage.delete(nom)

    # Un PDF rapatrie ne change plus jamais: son contenu est fige et son nom
    # porte l'empreinte du chemin d'origine. Une journee de cache evite de
    # retraverser le reseau pour un document qu'un eleve rouvre trois fois
    # dans la meme seance de revision.
    FILE_CACHE_SECONDS = 24 * 3600
    # Le relais, lui, ne maitrise pas ce qu'il sert: la source peut corriger
    # un fichier sans prevenir. Une heure suffit a absorber les relectures
    # sans figer une correction pour la journee.
    RELAY_CACHE_SECONDS = 3600

    # 64 Ko: le meme calibre que le rapatriement (import_bkalan). Sans decoupe
    # explicite, StreamingHttpResponse itere l'objet HTTPResponse ligne par
    # ligne -- un decoupage qui n'a aucun sens sur du binaire et qui multiplie
    # les allers-retours entre le flux amont et le client.
    RELAY_CHUNK_SIZE = 64 * 1024

    @staticmethod
    def _par_blocs(flux, taille):
        """Le flux amont, decoupe en blocs, referme a la fin.

        Le `close` est explicite: l'iterateur est consomme par le serveur bien
        apres le retour de la vue, et une connexion laissee ouverte reste
        comptee jusqu'a expiration cote source.
        """
        try:
            while True:
                bloc = flux.read(taille)
                if not bloc:
                    return
                yield bloc
        finally:
            flux.close()

    def _etag(self, document):
        """Empreinte du document tel qu'il est servi aujourd'hui.

        `is_downloaded` en fait partie: le jour ou un document passe du relais
        au stockage, le client doit cesser d'utiliser sa copie -- ce n'est
        plus la meme chaine qui le sert.
        """
        marqueur = document.updated_at.isoformat() if document.updated_at else ""
        return f'W/"doc-{document.pk}-{int(document.is_downloaded)}-{document.size_bytes}-{marqueur}"'

    @action(detail=True, methods=["get"], url_path="file")
    def file(self, request, pk=None):
        """Le PDF lui-meme: depuis le stockage, ou relaye depuis la source."""
        document = self.get_object()

        etag = self._etag(document)
        if request.headers.get("If-None-Match") == etag:
            # Rien n'a bouge: le client garde sa copie, et ni le stockage ni
            # la source ne sont sollicites.
            reponse = HttpResponseNotModified()
            reponse["ETag"] = etag
            return reponse

        if document.is_downloaded and document.file:
            redirection = self._redirection_stockage(document, etag)
            if redirection is not None:
                return redirection
            try:
                reponse = FileResponse(
                    document.file.open("rb"),
                    content_type="application/pdf",
                    filename=document.title,
                )
            except FileNotFoundError:
                # Reference en base mais absent du stockage (disque Render
                # ephemere, bucket vide): la source reste un repli valable.
                pass
            else:
                return self._avec_cache(reponse, etag, self.FILE_CACHE_SECONDS)

        return self._relay_source(document, etag)

    def _redirection_stockage(self, document, etag):
        """URL signee vers le stockage, ou None s'il faut servir le fichier.

        Rediriger evite au conteneur de relayer des dizaines de megaoctets
        qu'un stockage objet sert mieux que lui. Trois conditions, toutes
        necessaires:

        - le reglage est actif -- il ne l'est pas par defaut, car sur le web
          la redirection vers un autre domaine passe par un controle CORS que
          le bucket doit autoriser;
        - le stockage sait produire une URL absolue -- un stockage sur disque
          rend un chemin relatif, qui ne redirige nulle part;
        - cette URL est signee -- une URL nue laisserait un PDF accessible a
          qui l'a vue passer, sans expiration.
        """
        if not getattr(settings, "LIBRARY_STORAGE_REDIRECT", False):
            return None
        try:
            url = document.file.url
        except (ValueError, NotImplementedError, OSError):
            # Un stockage muet ne doit pas priver le lecteur de son document:
            # le flux local prend le relais.
            return None
        if not url.startswith(("https://", "http://")):
            return None
        if "X-Amz-Signature" not in url and "Signature=" not in url:
            return None

        reponse = HttpResponseRedirect(url)
        # Duree calee sur la signature: garder l'adresse en cache plus
        # longtemps que sa validite ferait echouer la lecture suivante sur un
        # lien expire.
        duree = min(
            self.FILE_CACHE_SECONDS,
            int(getattr(settings, "AWS_QUERYSTRING_EXPIRE", 3600)),
        )
        return self._avec_cache(reponse, etag, duree)

    def _avec_cache(self, reponse, etag, duree):
        """Ce qu'il faut pour qu'une seconde lecture ne coute rien.

        `private` et non `public`: la route est authentifiee, et un cache
        partage n'a pas a garder une reponse servie a un compte precis.
        """
        reponse["ETag"] = etag
        reponse["Cache-Control"] = f"private, max-age={duree}"
        return reponse

    def _relay_source(self, document, etag=None):
        parsed = urlparse(document.source_url or "")
        if parsed.scheme != "https" or parsed.hostname not in self.PROXY_ALLOWED_HOSTS:
            raise ValidationError(
                {"detail": "Ce document n'est pas encore disponible."}
            )

        try:
            amont = urlopen(document.source_url, timeout=30)  # noqa: S310 - hote borne
        except (HTTPError, URLError) as exc:
            raise ValidationError(
                {"detail": f"Document indisponible a la source: {exc}"}
            ) from exc

        response = StreamingHttpResponse(
            self._par_blocs(amont, self.RELAY_CHUNK_SIZE),
            content_type=amont.headers.get("Content-Type", "application/pdf"),
        )
        longueur = amont.headers.get("Content-Length")
        if longueur:
            # Transmise telle quelle: c'est elle qui permet au client
            # d'afficher une progression plutot qu'une attente sans fin.
            response["Content-Length"] = longueur
        if etag is None:
            return response
        return self._avec_cache(response, etag, self.RELAY_CACHE_SECONDS)


class BookViewSet(BaseModelViewSet):
    access_module = "library"
    queryset = Book.objects.all().order_by("title", "id")
    serializer_class = BookSerializer





    def _resolve_target_etablissement(self):
        user = self.request.user
        requested = self._requested_etablissement()
        if getattr(user, "role", None) == UserRole.SUPER_ADMIN and requested:
            return requested
        return getattr(user, "etablissement", None)

    def _filtres_du_catalogue(self, queryset):
        """Recherche et disponibilite, tels que l'ecran les demande.

        Le catalogue etait servi entier et trie par titre: chercher un
        ouvrage dans un fonds de plusieurs centaines de lignes revenait a
        faire defiler la page.
        """
        search = (self.request.query_params.get("search") or "").strip()
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search)
                | Q(author__icontains=search)
                | Q(isbn__icontains=search)
                | Q(publisher__icontains=search)
                | Q(subject__icontains=search)
                | Q(shelf_location__icontains=search)
            )

        disponibilite = (self.request.query_params.get("availability") or "").strip()
        if disponibilite == "available":
            queryset = queryset.filter(quantity_available__gt=0)
        elif disponibilite == "out":
            queryset = queryset.filter(quantity_available=0)

        return queryset

    def get_queryset(self):
        user = self.request.user
        # Le compteur des exemplaires sortis vient d'une annotation: le
        # calculer par ouvrage ferait une requete par ligne du catalogue.
        qs = Book.objects.annotate(
            exemplaires_sortis_count=Count(
                "borrows", filter=Q(borrows__returned_at__isnull=True)
            )
        ).order_by("title", "id")

        requested_etablissement = self._requested_etablissement()
        if requested_etablissement is not None:
            return self._filtres_du_catalogue(
                qs.filter(etablissement=requested_etablissement)
            )

        if self._has_requested_scope():
            return qs.none()

        if hasattr(user, "role") and user.role == "super_admin":
            return self._filtres_du_catalogue(qs)

        user_etablissement = getattr(user, "etablissement", None)
        if user_etablissement is None:
            return qs.none()
        return self._filtres_du_catalogue(qs.filter(etablissement=user_etablissement))

    def _verifier_l_isbn(self, serializer, etablissement):
        """Un ISBN ne se repete pas dans le meme fonds.

        La base le garantit, mais son refus arrive en erreur d'integrite --
        un 500 illisible la ou l'utilisateur a simplement saisi deux fois le
        meme ouvrage. Le controle ici rend le message utilisable.
        """
        isbn = serializer.validated_data.get("isbn")
        if not isbn:
            return
        doublons = Book.objects.filter(isbn=isbn, etablissement=etablissement)
        if serializer.instance is not None:
            doublons = doublons.exclude(pk=serializer.instance.pk)
        if doublons.exists():
            raise ValidationError(
                {"isbn": "Un ouvrage porte déjà cet ISBN dans cet établissement."}
            )

    def perform_create(self, serializer):
        etablissement = self._resolve_target_etablissement()
        self._verifier_l_isbn(serializer, etablissement)
        # Un ouvrage qui entre au catalogue est entierement en rayon: sa
        # disponibilite se deduit, elle ne se saisit pas.
        livre = serializer.save(
            etablissement=etablissement,
            quantity_available=serializer.validated_data.get("quantity_total", 0),
        )
        livre.recalculer_disponibilite()

    def perform_update(self, serializer):
        etablissement = self._resolve_target_etablissement()
        self._verifier_l_isbn(serializer, etablissement)
        livre = serializer.save(etablissement=etablissement)
        # Le total a pu changer: la disponibilite suit, sans jamais compter
        # comme en rayon un exemplaire qui est chez un eleve.
        livre.recalculer_disponibilite()


class BorrowViewSet(BaseModelViewSet):
    access_module = "library"
    queryset = Borrow.objects.select_related("student", "book").all().order_by("-borrowed_at", "-id")
    serializer_class = BorrowSerializer





    def _resolve_target_etablissement(self):
        user = self.request.user
        requested_etablissement = self._requested_etablissement()

        if getattr(user, "role", None) == "super_admin" and requested_etablissement:
            return requested_etablissement

        return getattr(user, "etablissement", None)

    def get_queryset(self):
        user = self.request.user
        qs = self._filtrer_par_etat(super().get_queryset())
        role = getattr(user, "role", "")

        # L'eleve et le parent ne voient que leurs propres prets: c'est
        # l'ecran « Mes emprunts », servi par la meme route.
        if role == UserRole.STUDENT:
            return qs.filter(student__user_id=user.id)
        if role == UserRole.PARENT:
            return qs.filter(student__parent__user_id=user.id)

        requested_etablissement = self._requested_etablissement()
        if requested_etablissement is not None:
            return qs.filter(student__etablissement=requested_etablissement)
        if self._has_requested_scope():
            return qs.none()
        if getattr(user, "role", None) == "super_admin":
            return qs
        return qs.filter(student__etablissement=getattr(user, "etablissement", None))

    def _validate_scope(self, serializer):
        student = serializer.validated_data.get("student")
        book = serializer.validated_data.get("book")
        if not student or not book:
            return
        target_etablissement = self._resolve_target_etablissement()
        if target_etablissement and student.etablissement_id != target_etablissement.id:
            raise ValidationError({"student": "L'eleve n'appartient pas a l'etablissement actif."})
        if target_etablissement and book.etablissement_id != target_etablissement.id:
            raise ValidationError({"book": "Le livre n'appartient pas a l'etablissement actif."})

    def _filtrer_par_etat(self, queryset):
        """En cours, rendus, en retard: les trois vues de l'ecran.

        Le suivi se faisait a l'oeil sur la liste entiere, ou les retards se
        perdaient entre les prets rendus il y a six mois.
        """
        etat = (self.request.query_params.get("status") or "").strip()
        if etat == "ongoing":
            queryset = queryset.filter(returned_at__isnull=True)
        elif etat == "returned":
            queryset = queryset.filter(returned_at__isnull=False)
        elif etat == "late":
            queryset = queryset.filter(
                returned_at__isnull=True, due_date__lt=timezone.localdate()
            )

        eleve = self.request.query_params.get("student")
        if eleve not in (None, "") and str(eleve).isdigit():
            queryset = queryset.filter(student_id=eleve)

        livre = self.request.query_params.get("book")
        if livre not in (None, "") and str(livre).isdigit():
            queryset = queryset.filter(book_id=livre)

        return queryset

    def _verifier_la_disponibilite(self, livre):
        """Pas de pret sans exemplaire en rayon.

        Le controle manquait entierement: on pretait le meme exemplaire a
        cinq eleves, et le compteur affichait toujours le fonds complet.
        """
        if livre.quantity_total <= 0:
            raise ValidationError(
                {"book": "Cet ouvrage n'a aucun exemplaire enregistré."}
            )
        if livre.exemplaires_sortis() >= livre.quantity_total:
            raise ValidationError(
                {"book": f"« {livre.title} » n'a plus d'exemplaire disponible."}
            )

    def perform_create(self, serializer):
        self._validate_scope(serializer)
        livre = serializer.validated_data.get("book")

        # Tout dans la meme transaction, le livre verrouille: deux emprunts
        # simultanes sur le dernier exemplaire passeraient tous les deux le
        # controle de disponibilite avant que l'un des deux ne l'ait pris.
        with transaction.atomic():
            if livre is not None:
                livre = Book.objects.select_for_update().get(pk=livre.pk)
                self._verifier_la_disponibilite(livre)
            emprunt = serializer.save()
            if livre is not None:
                livre.recalculer_disponibilite()
        return emprunt

    def perform_update(self, serializer):
        self._validate_scope(serializer)
        ancien_livre_id = serializer.instance.book_id
        with transaction.atomic():
            emprunt = serializer.save()
            # Les deux livres, car un emprunt peut avoir change d'ouvrage:
            # ne recalculer que le nouveau laisserait l'ancien un exemplaire
            # en moins pour toujours.
            emprunt.book.recalculer_disponibilite()
            if ancien_livre_id != emprunt.book_id:
                Book.objects.filter(pk=ancien_livre_id).first().recalculer_disponibilite()

    def perform_destroy(self, instance):
        livre = instance.book
        with transaction.atomic():
            instance.delete()
            livre.recalculer_disponibilite()

    @action(detail=True, methods=["post"], url_path="return")
    def marquer_rendu(self, request, pk=None):
        """Le retour d'un exemplaire: date, penalite, remise en rayon.

        Rien ne permettait de rendre un livre: `returned_at` n'etait ecrit
        nulle part et un pret durait indefiniment. La penalite, elle, etait
        saisie a la creation de l'emprunt -- avant tout retard -- et vaut
        desormais le tarif journalier de l'etablissement multiplie par les
        jours entames, sauf montant impose a la main.
        """
        emprunt = self.get_object()
        if emprunt.est_rendu:
            raise ValidationError(
                {"detail": "Cet emprunt a déjà été rendu le "
                           f"{emprunt.returned_at.isoformat()}."}
            )

        # Le client peut imposer la date -- un livre rendu vendredi, saisi
        # lundi -- ou la laisser au serveur.
        date_brute = request.data.get("returned_at")
        if date_brute in (None, ""):
            date_retour = timezone.localdate()
        else:
            date_retour = parse_date(str(date_brute).strip())
        if date_retour is None:
            raise ValidationError({"returned_at": "Date de retour illisible."})
        if date_retour < emprunt.borrowed_at:
            raise ValidationError(
                {"returned_at": "Le retour précède la date d'emprunt."}
            )

        with transaction.atomic():
            emprunt.returned_at = date_retour
            penalite_imposee = request.data.get("penalty_amount")
            if penalite_imposee not in (None, ""):
                try:
                    emprunt.penalty_amount = Decimal(str(penalite_imposee))
                except (InvalidOperation, ValueError) as exc:
                    raise ValidationError(
                        {"penalty_amount": "Montant de pénalité illisible."}
                    ) from exc
            else:
                emprunt.penalty_amount = emprunt.penalite_theorique()
            emprunt.save(
                update_fields=["returned_at", "penalty_amount", "updated_at"]
            )
            emprunt.book.recalculer_disponibilite()

        return Response(self.get_serializer(emprunt).data)


class CanteenMenuViewSet(BaseModelViewSet):
    access_module = "canteen"
    queryset = CanteenMenu.objects.all().order_by("-menu_date", "-id")
    serializer_class = CanteenMenuSerializer
    filterset_fields = ["menu_date", "is_active"]





    def _resolve_target_etablissement(self):
        user = self.request.user
        requested = self._requested_etablissement()
        if getattr(user, "role", None) == UserRole.SUPER_ADMIN and requested:
            return requested
        return getattr(user, "etablissement", None)

    def get_queryset(self):
        user = self.request.user
        qs = CanteenMenu.objects.all().order_by("-menu_date", "-id")

        requested_etablissement = self._requested_etablissement()
        if requested_etablissement is not None:
            return qs.filter(etablissement=requested_etablissement)

        if self._has_requested_scope():
            return qs.none()

        if hasattr(user, "role") and user.role == "super_admin":
            return qs.all()

        user_etablissement = getattr(user, "etablissement", None)
        if user_etablissement is None:
            return qs.none()
        return qs.filter(etablissement=user_etablissement)

    def perform_create(self, serializer):
        serializer.save(etablissement=self._resolve_target_etablissement())

    def perform_update(self, serializer):
        serializer.save(etablissement=self._resolve_target_etablissement())


class CanteenSubscriptionViewSet(BaseModelViewSet):
    access_module = "canteen"
    queryset = CanteenSubscription.objects.select_related("student", "student__user", "academic_year").all().order_by("-created_at")
    serializer_class = CanteenSubscriptionSerializer
    filterset_fields = ["student", "academic_year", "status"]





    def _resolve_target_etablissement(self):
        user = self.request.user
        requested_etablissement = self._requested_etablissement()

        if getattr(user, "role", None) == "super_admin" and requested_etablissement:
            return requested_etablissement

        return getattr(user, "etablissement", None)

    def get_queryset(self):
        user = self.request.user
        qs = super().get_queryset()
        role = getattr(user, "role", "")

        if role == UserRole.STUDENT:
            return qs.filter(student__user_id=user.id)
        if role == UserRole.PARENT:
            return qs.filter(student__parent__user_id=user.id)

        requested_etablissement = self._requested_etablissement()
        if requested_etablissement is not None:
            return qs.filter(student__etablissement=requested_etablissement)
        if self._has_requested_scope():
            return qs.none()
        if getattr(user, "role", None) == "super_admin":
            return qs
        return qs.filter(student__etablissement=getattr(user, "etablissement", None))

    def _validate_scope(self, serializer):
        student = serializer.validated_data.get("student")
        if not student:
            return
        target_etablissement = self._resolve_target_etablissement()
        if target_etablissement and student.etablissement_id != target_etablissement.id:
            raise ValidationError({"student": "L'eleve n'appartient pas a l'etablissement actif."})

    def perform_create(self, serializer):
        self._validate_scope(serializer)
        serializer.save()

    def perform_update(self, serializer):
        self._validate_scope(serializer)
        serializer.save()


class CanteenServiceViewSet(BaseModelViewSet):
    access_module = "canteen"
    queryset = CanteenService.objects.select_related("student", "student__user", "menu").all().order_by("-served_on", "-id")
    serializer_class = CanteenServiceSerializer
    filterset_fields = ["student", "menu", "served_on", "is_paid"]





    def _resolve_target_etablissement(self):
        user = self.request.user
        requested_etablissement = self._requested_etablissement()

        if getattr(user, "role", None) == "super_admin" and requested_etablissement:
            return requested_etablissement

        return getattr(user, "etablissement", None)

    def get_queryset(self):
        user = self.request.user
        qs = super().get_queryset()
        role = getattr(user, "role", "")

        if role == UserRole.STUDENT:
            return qs.filter(student__user_id=user.id)
        if role == UserRole.PARENT:
            return qs.filter(student__parent__user_id=user.id)

        requested_etablissement = self._requested_etablissement()
        if requested_etablissement is not None:
            return qs.filter(student__etablissement=requested_etablissement)
        if self._has_requested_scope():
            return qs.none()
        if getattr(user, "role", None) == "super_admin":
            return qs
        return qs.filter(student__etablissement=getattr(user, "etablissement", None))

    def _validate_scope(self, serializer):
        student = serializer.validated_data.get("student")
        menu = serializer.validated_data.get("menu")
        if not student or not menu:
            return
        target_etablissement = self._resolve_target_etablissement()
        if target_etablissement and student.etablissement_id != target_etablissement.id:
            raise ValidationError({"student": "L'eleve n'appartient pas a l'etablissement actif."})
        if target_etablissement and menu.etablissement_id != target_etablissement.id:
            raise ValidationError({"menu": "Le menu n'appartient pas a l'etablissement actif."})

    def perform_create(self, serializer):
        self._validate_scope(serializer)
        serializer.save()

    def perform_update(self, serializer):
        self._validate_scope(serializer)
        serializer.save()


class ExamSessionViewSet(AnneeScolaireScopeMixin, BaseModelViewSet):
    access_module = "exams"
    queryset = ExamSession.objects.select_related("academic_year").all().order_by("-id")
    serializer_class = ExamSessionSerializer
    filterset_fields = ["academic_year", "term"]
    search_fields = ["title", "academic_year__name"]
    ordering_fields = ["start_date", "end_date", "title", "term"]
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]


class ExamPlanningViewSet(BaseModelViewSet):
    access_module = "exams"
    queryset = ExamPlanning.objects.select_related("session", "classroom", "subject").all().order_by("-id")
    serializer_class = ExamPlanningSerializer
    filterset_fields = ["session", "classroom", "subject", "exam_date"]
    search_fields = ["session__title", "classroom__name", "subject__name"]
    ordering_fields = ["exam_date", "start_time", "classroom__name"]
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]





    def _resolve_target_etablissement(self):
        user = self.request.user
        requested_etablissement = self._requested_etablissement()

        if getattr(user, "role", None) == "super_admin" and requested_etablissement:
            return requested_etablissement

        return getattr(user, "etablissement", None)

    def get_queryset(self):
        user = self.request.user
        qs = super().get_queryset()
        requested_etablissement = self._requested_etablissement()

        if requested_etablissement is not None:
            return qs.filter(classroom__etablissement=requested_etablissement)
        if self._has_requested_scope():
            return qs.none()
        if getattr(user, "role", None) == "super_admin":
            return qs

        user_etablissement = getattr(user, "etablissement", None)
        if user_etablissement is None:
            return qs

        return qs.filter(classroom__etablissement=user_etablissement)

    def _validate_scope(self, serializer):
        classroom = serializer.validated_data.get("classroom")
        if not classroom:
            return
        target_etablissement = self._resolve_target_etablissement()
        if target_etablissement and classroom.etablissement_id != target_etablissement.id:
            raise ValidationError({"classroom": "La classe n'appartient pas a l'etablissement actif."})

    def perform_create(self, serializer):
        self._validate_scope(serializer)
        serializer.save()

    def perform_update(self, serializer):
        self._validate_scope(serializer)
        serializer.save()


class ExamInvigilationViewSet(BaseModelViewSet):
    access_module = "exams"
    queryset = ExamInvigilation.objects.select_related("planning", "planning__session", "planning__classroom", "planning__subject", "supervisor").all().order_by("-created_at")
    serializer_class = ExamInvigilationSerializer
    filterset_fields = ["planning", "supervisor", "planning__session"]
    search_fields = [
        "supervisor__first_name",
        "supervisor__last_name",
        "planning__classroom__name",
        "planning__subject__name",
    ]
    ordering_fields = ["planning__exam_date", "created_at"]
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]

    def get_queryset(self):
        user = self.request.user
        qs = super().get_queryset()
        requested_id = self.request.headers.get("X-Etablissement-Id") or self.request.query_params.get("etablissement")
        if requested_id not in (None, ""):
            try:
                return qs.filter(planning__classroom__etablissement_id=int(requested_id))
            except (TypeError, ValueError):
                return qs.none()
        if getattr(user, "role", None) == "super_admin":
            return qs
        return qs.filter(planning__classroom__etablissement=getattr(user, "etablissement", None))

    def _validate_scope(self, serializer):
        planning = serializer.validated_data.get("planning")
        if not planning:
            return
        user = self.request.user
        requested_id = self.request.headers.get("X-Etablissement-Id") or self.request.query_params.get("etablissement")
        target_id = None
        if requested_id not in (None, "") and getattr(user, "role", None) == "super_admin":
            try:
                target_id = int(requested_id)
            except (TypeError, ValueError):
                target_id = None
        if target_id is None:
            target_id = getattr(user, "etablissement_id", None)
        if target_id and planning.classroom.etablissement_id != target_id:
            raise ValidationError({"planning": "Le planning n'appartient pas a l'etablissement actif."})

    def perform_create(self, serializer):
        self._validate_scope(serializer)
        serializer.save()

    def perform_update(self, serializer):
        self._validate_scope(serializer)
        serializer.save()


class ExamResultViewSet(BaseModelViewSet):
    access_module = "exams"
    queryset = ExamResult.objects.select_related("session", "student", "subject").all().order_by("-id")
    serializer_class = ExamResultSerializer
    filterset_fields = ["session", "student", "subject"]
    search_fields = [
        "student__matricule",
        "student__user__first_name",
        "student__user__last_name",
        "subject__name",
    ]
    ordering_fields = ["score", "session__title", "subject__name"]
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]

    def get_queryset(self):
        user = self.request.user
        qs = super().get_queryset()
        role = getattr(user, "role", "")
        if role == UserRole.STUDENT:
            return qs.filter(student__user_id=user.id)
        if role == UserRole.PARENT:
            return qs.filter(student__parent__user_id=user.id)

        requested_id = self.request.headers.get("X-Etablissement-Id") or self.request.query_params.get("etablissement")
        if requested_id not in (None, ""):
            try:
                return qs.filter(student__etablissement_id=int(requested_id))
            except (TypeError, ValueError):
                return qs.none()
        if getattr(user, "role", None) == "super_admin":
            return qs
        return qs.filter(student__etablissement=getattr(user, "etablissement", None))

    def _validate_scope(self, serializer):
        student = serializer.validated_data.get("student")
        if not student:
            return
        user = self.request.user
        requested_id = self.request.headers.get("X-Etablissement-Id") or self.request.query_params.get("etablissement")
        target_id = None
        if requested_id not in (None, "") and getattr(user, "role", None) == "super_admin":
            try:
                target_id = int(requested_id)
            except (TypeError, ValueError):
                target_id = None
        if target_id is None:
            target_id = getattr(user, "etablissement_id", None)
        student_etablissement_id = getattr(student, "etablissement_id", None)
        classroom_etablissement_id = getattr(getattr(student, "classroom", None), "etablissement_id", None)
        if target_id and student_etablissement_id != target_id and classroom_etablissement_id != target_id:
            raise ValidationError({"student": "L'eleve n'appartient pas a l'etablissement actif."})

    def perform_create(self, serializer):
        self._validate_scope(serializer)
        serializer.save()

    def perform_update(self, serializer):
        self._validate_scope(serializer)
        serializer.save()

    @action(detail=False, methods=["post"], url_path="import-exams")
    def import_exams(self, request):
        classroom_id = request.data.get("classroom_id")
        session_id = request.data.get("session_id")
        confirm = str(request.data.get("confirm", "false")).strip().lower() in {"1", "true", "yes", "on"}

        try:
            classroom_id = int(classroom_id)
            session_id = int(session_id)
        except (TypeError, ValueError):
            raise ValidationError({"detail": "classroom_id et session_id sont requis."})

        user = request.user
        if getattr(user, "role", None) == UserRole.SUPER_ADMIN:
            requested_id = request.headers.get("X-Etablissement-Id") or request.query_params.get("etablissement")
            target_id = int(requested_id) if requested_id not in (None, "") and str(requested_id).isdigit() else getattr(user, "etablissement_id", None)
        else:
            target_id = getattr(user, "etablissement_id", None)

        classroom = get_object_or_404(ClassRoom.objects.select_related("etablissement"), id=classroom_id)
        if target_id and classroom.etablissement_id != target_id:
            raise ValidationError({"classroom_id": "La classe n'appartient pas à l'établissement actif."})

        session = get_object_or_404(ExamSession, id=session_id)
        rows = _load_import_rows(request.FILES.get("file") or request.data.get("file"))
        if not rows:
            raise ValidationError({"file": "Aucune ligne exploitable dans le fichier."})

        students_by_matricule = {
            (student.matricule or "").strip().lower(): student
            for student in Student.objects.filter(classroom=classroom, is_archived=False)
        }
        subjects = Subject.objects.filter(classroom=classroom)
        subjects_by_code = {(subject.code or "").strip().lower(): subject for subject in subjects}
        subjects_by_name = {(subject.name or "").strip().lower(): subject for subject in subjects}

        errors = []
        prepared = []
        for index, row in enumerate(rows, start=2):
            matricule = _as_text(row.get("student_matricule") or row.get("matricule")).lower()
            subject_code = _as_text(row.get("subject_code") or row.get("matiere_code")).lower()
            subject_name = _as_text(row.get("subject_name") or row.get("matiere")).lower()
            score = _as_decimal(row.get("score") or row.get("note"))

            if not matricule:
                errors.append({"row": index, "error": "student_matricule requis."})
                continue
            if score is None or score < Decimal("0") or score > Decimal("20"):
                errors.append({"row": index, "error": "score/note invalide (0..20)."})
                continue

            student = students_by_matricule.get(matricule)
            if student is None:
                errors.append({"row": index, "error": f"Élève introuvable dans la classe pour matricule '{matricule}'."})
                continue

            subject = None
            if subject_code:
                subject = subjects_by_code.get(subject_code)
            if subject is None and subject_name:
                subject = subjects_by_name.get(subject_name)
            if subject is None:
                errors.append({"row": index, "error": "Matière introuvable (subject_code/subject_name)."})
                continue

            existing = ExamResult.objects.filter(session=session, student=student, subject=subject).first()
            prepared.append(
                {
                    "row": index,
                    "student": student,
                    "subject": subject,
                    "score": score,
                    "existing": existing,
                }
            )

        to_create = sum(1 for item in prepared if item["existing"] is None)
        to_update = len(prepared) - to_create
        payload = {
            "classroom": {"id": classroom.id, "name": classroom.name},
            "session": {"id": session.id, "name": session.title, "term": session.term},
            "summary": {
                "total_rows": len(rows),
                "valid_rows": len(prepared),
                "errors": len(errors),
                "to_create": to_create,
                "to_update": to_update,
            },
            "errors": errors,
            "preview": [
                {
                    "row": item["row"],
                    "action": "create" if item["existing"] is None else "update",
                    "student": item["student"].matricule,
                    "subject": item["subject"].code,
                    "score": str(item["score"]),
                }
                for item in prepared[:150]
            ],
            "confirm_required": True,
        }

        if not confirm:
            return Response(payload)
        if errors:
            return Response({**payload, "detail": "Import bloqué: corrigez les erreurs."}, status=400)

        created = 0
        updated = 0
        with transaction.atomic():
            for item in prepared:
                data = {
                    "session": session.id,
                    "student": item["student"].id,
                    "subject": item["subject"].id,
                    "score": str(item["score"]),
                }
                if item["existing"] is None:
                    serializer = self.get_serializer(data=data)
                    serializer.is_valid(raise_exception=True)
                    self.perform_create(serializer)
                    created += 1
                else:
                    serializer = self.get_serializer(item["existing"], data=data, partial=False)
                    serializer.is_valid(raise_exception=True)
                    self.perform_update(serializer)
                    updated += 1

        return Response(
            {
                **payload,
                "result": {"created": created, "updated": updated},
                "detail": "Import notes d'examen terminé.",
            }
        )


class SupplierViewSet(EtablissementScopedModelViewSet):
    access_module = "stock"
    queryset = Supplier.objects.select_related("etablissement").all().order_by("name", "id")
    serializer_class = SupplierSerializer

    def get_queryset(self):
        return self._filter_by_scope(super().get_queryset(), field_name="etablissement")

    def perform_create(self, serializer):
        target_etablissement = self._resolve_target_etablissement()
        if getattr(self.request.user, "role", None) == UserRole.SUPER_ADMIN and target_etablissement is None:
            raise ValidationError({"etablissement": "Selectionnez un etablissement actif."})
        serializer.save(etablissement=target_etablissement)

    def perform_update(self, serializer):
        target_etablissement = self._resolve_target_etablissement()
        if getattr(self.request.user, "role", None) == UserRole.SUPER_ADMIN and target_etablissement is None:
            raise ValidationError({"etablissement": "Selectionnez un etablissement actif."})
        serializer.save(etablissement=target_etablissement)


class StockItemViewSet(EtablissementScopedModelViewSet):
    access_module = "stock"
    queryset = StockItem.objects.select_related("supplier", "etablissement").all().order_by("name", "id")
    serializer_class = StockItemSerializer

    def get_queryset(self):
        return self._filter_by_scope(super().get_queryset(), field_name="etablissement")

    def _validate_scope(self, serializer):
        supplier = serializer.validated_data.get("supplier") or getattr(serializer.instance, "supplier", None)
        target_etablissement = self._resolve_target_etablissement()
        if not supplier or not target_etablissement:
            return
        if supplier.etablissement_id != target_etablissement.id:
            raise ValidationError({"supplier": "Le fournisseur n'appartient pas a l'etablissement actif."})

    def perform_create(self, serializer):
        target_etablissement = self._resolve_target_etablissement()
        if getattr(self.request.user, "role", None) == UserRole.SUPER_ADMIN and target_etablissement is None:
            raise ValidationError({"etablissement": "Selectionnez un etablissement actif."})
        self._validate_scope(serializer)
        serializer.save(etablissement=target_etablissement)

    def perform_update(self, serializer):
        target_etablissement = self._resolve_target_etablissement()
        if getattr(self.request.user, "role", None) == UserRole.SUPER_ADMIN and target_etablissement is None:
            raise ValidationError({"etablissement": "Selectionnez un etablissement actif."})
        self._validate_scope(serializer)
        serializer.save(etablissement=target_etablissement)

    @action(detail=False, methods=["get"])
    def low_stock(self, request):
        """Les articles sous leur seuil, du plus critique au moins.

        Le tri est ce qui manquait: la liste arrivait dans l'ordre du
        catalogue, et l'article tombe a zero se lisait apres celui a qui il
        manque une unite. L'ecart au seuil dit lequel appelle un bon de
        commande aujourd'hui.
        """
        articles = sorted(
            self.get_queryset().filter(quantity__lte=F("minimum_threshold")),
            key=lambda article: (
                article.quantity - article.minimum_threshold,
                article.name,
            ),
        )
        return Response(
            {
                "count": len(articles),
                "results": self.get_serializer(articles, many=True).data,
            }
        )


class StockMovementViewSet(BaseModelViewSet):
    access_module = "stock"
    queryset = StockMovement.objects.select_related("item", "item__etablissement").all().order_by("-created_at")
    serializer_class = StockMovementSerializer





    def _resolve_target_etablissement(self):
        user = self.request.user
        requested_etablissement = self._requested_etablissement()
        if getattr(user, "role", None) == UserRole.SUPER_ADMIN and requested_etablissement:
            return requested_etablissement
        return getattr(user, "etablissement", None)

    def get_queryset(self):
        user = self.request.user
        qs = super().get_queryset()
        requested_etablissement = self._requested_etablissement()

        if requested_etablissement is not None:
            return qs.filter(item__etablissement=requested_etablissement)
        if self._has_requested_scope():
            return qs.none()
        if getattr(user, "role", None) == UserRole.SUPER_ADMIN:
            return qs

        user_etablissement = getattr(user, "etablissement", None)
        if user_etablissement is None:
            return qs.none()
        return qs.filter(item__etablissement=user_etablissement)

    def _validate_scope(self, serializer):
        item = serializer.validated_data.get("item") or getattr(serializer.instance, "item", None)
        target_etablissement = self._resolve_target_etablissement()
        if not item or not target_etablissement:
            return
        if item.etablissement_id != target_etablissement.id:
            raise ValidationError({"item": "L'article n'appartient pas a l'etablissement actif."})

    def perform_create(self, serializer):
        target_etablissement = self._resolve_target_etablissement()
        if getattr(self.request.user, "role", None) == UserRole.SUPER_ADMIN and target_etablissement is None:
            raise ValidationError({"etablissement": "Selectionnez un etablissement actif."})
        self._validate_scope(serializer)
        serializer.save()

    def perform_update(self, serializer):
        target_etablissement = self._resolve_target_etablissement()
        if getattr(self.request.user, "role", None) == UserRole.SUPER_ADMIN and target_etablissement is None:
            raise ValidationError({"etablissement": "Selectionnez un etablissement actif."})
        self._validate_scope(serializer)
        serializer.save()


class PromotionRunViewSet(EtablissementScopedModelViewSet):
    access_module = "promotion"
    queryset = (
        PromotionRun.objects.select_related(
            "etablissement",
            "source_academic_year",
            "target_academic_year",
            "executed_by",
        )
        .prefetch_related(
            "decisions",
            "decisions__student",
            "decisions__student__user",
            "decisions__source_classroom",
            "decisions__target_classroom",
        )
        .order_by("-created_at")
    )
    serializer_class = PromotionRunSerializer
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]
    filterset_fields = [
        "status",
        "source_academic_year",
        "target_academic_year",
        "etablissement",
        "created_at",
    ]
    search_fields = ["source_academic_year__name", "target_academic_year__name"]
    ordering_fields = ["created_at", "status", "total_students"]

    def get_queryset(self):
        return self._filter_by_scope(super().get_queryset())

    def _target_etablissement(self):
        target = self._resolve_target_etablissement()
        if getattr(self.request.user, "role", None) == UserRole.SUPER_ADMIN and target is None:
            raise ValidationError({"etablissement": "Selectionnez un etablissement actif."})
        return target

    def _resolve_source_year(self, payload):
        source_year_id = payload.get("source_academic_year")
        if source_year_id in (None, ""):
            active_year = AcademicYear.courante(self._resolve_target_etablissement())
            if active_year:
                return active_year
            raise ValidationError({"source_academic_year": "Aucune annee scolaire active n'est disponible."})

        source_year = AcademicYear.objects.filter(id=source_year_id).first()
        if not source_year:
            raise ValidationError({"source_academic_year": "Annee scolaire source introuvable."})
        return source_year

    def _resolve_target_year(self, payload):
        target_year_id = payload.get("target_academic_year")
        if target_year_id in (None, ""):
            return None
        target_year = AcademicYear.objects.filter(id=target_year_id).first()
        if not target_year:
            raise ValidationError({"target_academic_year": "Annee scolaire cible introuvable."})
        return target_year

    def _resolve_source_classrooms(self, payload, source_year, etablissement):
        classroom_ids = payload.get("source_classrooms")
        queryset = ClassRoom.objects.filter(academic_year=source_year)
        if etablissement is not None:
            queryset = queryset.filter(etablissement=etablissement)

        if not classroom_ids:
            return list(queryset.order_by("name", "id"))

        classrooms = list(queryset.filter(id__in=classroom_ids).order_by("name", "id"))
        if len(classrooms) != len(set(classroom_ids)):
            raise ValidationError({"source_classrooms": "Une ou plusieurs classes sources sont invalides."})
        return classrooms

    def _resolve_mapping(self, payload, source_classrooms, target_year, etablissement):
        raw_mapping = payload.get("classroom_mapping") or []
        source_ids = {classroom.id for classroom in source_classrooms}
        mapping = {}

        if isinstance(raw_mapping, dict):
            raw_mapping = [
                {"source_classroom": key, "target_classroom": value}
                for key, value in raw_mapping.items()
            ]

        if raw_mapping:
            valid_target_qs = ClassRoom.objects.all()
            if etablissement is not None:
                valid_target_qs = valid_target_qs.filter(etablissement=etablissement)
            if target_year is not None:
                valid_target_qs = valid_target_qs.filter(academic_year=target_year)

            valid_targets = {room.id: room for room in valid_target_qs}
            for item in raw_mapping:
                if not isinstance(item, dict):
                    continue
                source_id = item.get("source_classroom")
                target_id = item.get("target_classroom")
                if source_id not in source_ids:
                    continue
                if target_id in (None, ""):
                    mapping[source_id] = None
                    continue
                target_classroom = valid_targets.get(target_id)
                if target_classroom is None:
                    raise ValidationError(
                        {
                            "classroom_mapping": (
                                f"La classe cible {target_id} n'est pas valide pour la passation."
                            )
                        }
                    )
                mapping[source_id] = target_classroom

        if mapping:
            return mapping

        if target_year is None:
            return {source.id: None for source in source_classrooms}

        auto_targets = ClassRoom.objects.filter(academic_year=target_year)
        if etablissement is not None:
            auto_targets = auto_targets.filter(etablissement=etablissement)

        auto_targets = list(auto_targets.order_by("name", "id"))
        target_index = {
            room.name.strip().lower(): room for room in auto_targets
        }

        for source in source_classrooms:
            source_name = source.name.strip().lower()
            exact_target = target_index.get(source_name)
            if exact_target is not None:
                mapping[source.id] = exact_target
                continue

            mapping[source.id] = auto_targets[0] if auto_targets else None
        return mapping

    def _compute_student_average(self, student, classroom, source_year):
        grades = Grade.objects.filter(
            student=student,
            classroom=classroom,
            academic_year=source_year,
        ).select_related("subject")

        weighted_sum = Decimal("0")
        coef_sum = Decimal("0")
        for grade in grades:
            coef = Decimal(str(grade.subject.coefficient or 0))
            if coef <= 0:
                continue
            weighted_sum += Decimal(str(grade.value)) * coef
            coef_sum += coef

        if coef_sum > 0:
            return (weighted_sum / coef_sum).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        history = StudentAcademicHistory.objects.filter(
            student=student,
            academic_year=source_year,
            classroom=classroom,
        ).first()
        if history is not None:
            return Decimal(str(history.average or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        return Decimal("0.00")

    def _build_decisions(
        self,
        source_classrooms,
        source_year,
        mapping,
        min_average,
        min_conduite,
    ):
        decision_rows = []
        promoted_count = 0
        repeated_count = 0
        archived_count = 0

        for classroom in source_classrooms:
            students = list(
                Student.objects.select_related("user")
                .filter(classroom=classroom, is_archived=False)
                .order_by("user__last_name", "user__first_name", "id")
            )

            scoring_rows = []
            for student in students:
                average = self._compute_student_average(student, classroom, source_year)
                conduite = Decimal(str(student.conduite or 0)).quantize(
                    Decimal("0.01"),
                    rounding=ROUND_HALF_UP,
                )
                scoring_rows.append(
                    {
                        "student": student,
                        "average": average,
                        "conduite": conduite,
                    }
                )

            ranked_rows = sorted(scoring_rows, key=lambda item: item["average"], reverse=True)
            target_classroom = mapping.get(classroom.id)

            for index, row in enumerate(ranked_rows, start=1):
                is_eligible = row["average"] >= min_average and row["conduite"] >= min_conduite
                reason = ""
                decision = PromotionDecisionType.REPEATED
                destination = classroom
                target_is_same_class = target_classroom is not None and target_classroom.id == classroom.id

                if is_eligible and target_classroom is not None and not target_is_same_class:
                    decision = PromotionDecisionType.PROMOTED
                    destination = target_classroom
                    promoted_count += 1
                elif is_eligible and target_is_same_class:
                    repeated_count += 1
                    reason = "Classe cible identique a la classe source: promotion bloquee."
                elif is_eligible and target_classroom is None:
                    decision = PromotionDecisionType.ARCHIVED
                    destination = None
                    archived_count += 1
                    reason = "Classe terminale sans classe cible: archivage automatique."
                else:
                    repeated_count += 1
                    if row["average"] < min_average:
                        reason = "Moyenne insuffisante."
                    elif row["conduite"] < min_conduite:
                        reason = "Conduite insuffisante."

                decision_rows.append(
                    {
                        "student": row["student"],
                        "source_classroom": classroom,
                        "target_classroom": destination,
                        "decision": decision,
                        "average": row["average"],
                        "conduite": row["conduite"],
                        "rank": index,
                        "reason": reason,
                    }
                )

        summary = {
            "total_students": len(decision_rows),
            "promoted_count": promoted_count,
            "repeated_count": repeated_count,
            "archived_count": archived_count,
        }
        return decision_rows, summary

    def _create_run(self, payload, status, apply_changes):
        etablissement = self._target_etablissement()
        source_year = self._resolve_source_year(payload)
        target_year = self._resolve_target_year(payload)
        min_average = Decimal(str(payload.get("min_average", "10"))).quantize(Decimal("0.01"))
        min_conduite = Decimal(str(payload.get("min_conduite", "10"))).quantize(Decimal("0.01"))

        if target_year is None:
            raise ValidationError({"target_academic_year": "Une annee scolaire cible est requise."})
        if target_year.id == source_year.id:
            raise ValidationError(
                {"target_academic_year": "L'annee cible doit etre differente de l'annee source."}
            )

        if min_average < Decimal("0") or min_average > Decimal("20"):
            raise ValidationError({"min_average": "Le seuil de moyenne doit etre entre 0 et 20."})
        if min_conduite < Decimal("0") or min_conduite > Decimal("20"):
            raise ValidationError({"min_conduite": "Le seuil de conduite doit etre entre 0 et 20."})

        source_classrooms = self._resolve_source_classrooms(payload, source_year, etablissement)
        if not source_classrooms:
            raise ValidationError({"source_classrooms": "Aucune classe source trouvee."})

        mapping = self._resolve_mapping(payload, source_classrooms, target_year, etablissement)
        decisions_data, summary = self._build_decisions(
            source_classrooms=source_classrooms,
            source_year=source_year,
            mapping=mapping,
            min_average=min_average,
            min_conduite=min_conduite,
        )

        with transaction.atomic():
            run = PromotionRun.objects.create(
                etablissement=etablissement,
                source_academic_year=source_year,
                target_academic_year=target_year,
                status=status,
                min_average=min_average,
                min_conduite=min_conduite,
                executed_by=self.request.user,
                total_students=summary["total_students"],
                promoted_count=summary["promoted_count"],
                repeated_count=summary["repeated_count"],
                archived_count=summary["archived_count"],
                payload={
                    "source_classrooms": [room.id for room in source_classrooms],
                    "classroom_mapping": {
                        str(source.id): (mapping.get(source.id).id if mapping.get(source.id) else None)
                        for source in source_classrooms
                    },
                },
            )

            PromotionDecision.objects.bulk_create(
                [
                    PromotionDecision(
                        run=run,
                        student=row["student"],
                        source_classroom=row["source_classroom"],
                        target_classroom=row["target_classroom"],
                        decision=row["decision"],
                        average=row["average"],
                        conduite=row["conduite"],
                        rank=row["rank"],
                        reason=row["reason"],
                    )
                    for row in decisions_data
                ]
            )

            if apply_changes:
                decisions = list(
                    run.decisions.select_related("student", "target_classroom", "source_classroom")
                )
                for decision in decisions:
                    student = decision.student
                    StudentAcademicHistory.objects.update_or_create(
                        student=student,
                        academic_year=source_year,
                        classroom=decision.source_classroom,
                        defaults={
                            "average": decision.average,
                            "rank": decision.rank,
                        },
                    )

                    if decision.decision == PromotionDecisionType.PROMOTED:
                        student.classroom = decision.target_classroom
                        student.is_archived = False
                    elif decision.decision == PromotionDecisionType.ARCHIVED:
                        student.classroom = None
                        student.is_archived = True
                    else:
                        student.classroom = decision.source_classroom
                        student.is_archived = False

                    student.save(update_fields=["classroom", "is_archived", "updated_at"])

        return run

    @action(detail=False, methods=["post"], url_path="simulate")
    def simulate(self, request):
        run = self._create_run(
            payload=request.data,
            status=PromotionRunStatus.SIMULATED,
            apply_changes=False,
        )
        serializer = self.get_serializer(run)
        return Response(serializer.data)

    @action(detail=False, methods=["post"], url_path="execute")
    def execute(self, request):
        run = self._create_run(
            payload=request.data,
            status=PromotionRunStatus.EXECUTED,
            apply_changes=True,
        )
        serializer = self.get_serializer(run)
        return Response(serializer.data)


class DashboardViewSet(viewsets.ViewSet):
    access_module = "dashboard"
    permission_classes = [permissions.IsAuthenticated, HasModuleAccess]

    @staticmethod
    def _requested_etablissement_id(request):
        raw_value = (
            request.headers.get("X-Etablissement-Id")
            or request.query_params.get("etablissement")
        )
        if raw_value in (None, ""):
            return None
        try:
            parsed = int(raw_value)
        except (TypeError, ValueError):
            return None
        return parsed if parsed > 0 else None

    @staticmethod
    def _requested_etablissement_name(request):
        raw_name = (
            request.headers.get("X-Etablissement-Name")
            or request.query_params.get("etablissement_name")
        )
        if raw_name is None:
            return None
        cleaned = str(raw_name).strip()
        return cleaned or None

    def _requested_etablissement(self, request):
        requested_id = self._requested_etablissement_id(request)
        if requested_id:
            etablissement = Etablissement.objects.filter(id=requested_id).first()
            if etablissement:
                return etablissement

        requested_name = self._requested_etablissement_name(request)
        if not requested_name:
            return None

        etablissement = Etablissement.objects.filter(name__iexact=requested_name).first()
        if etablissement:
            return etablissement

        return Etablissement.objects.filter(name__icontains=requested_name).order_by("name").first()

    def _resolve_dashboard_scope(self, request):
        user = request.user
        requested = self._requested_etablissement(request)
        if getattr(user, "role", None) == "super_admin":
            return requested
        return getattr(user, "etablissement", None)

    def list(self, request):
        month_start = timezone.now().date().replace(day=1)
        active_etablissement = self._resolve_dashboard_scope(request)
        if getattr(request.user, "role", None) == UserRole.SUPER_ADMIN and active_etablissement is None:
            raise ValidationError({"etablissement": "Selectionnez un etablissement actif."})

        # Cle et duree viennent de dashboard_cache, que les signaux
        # d'invalidation utilisent aussi (voir signals.py).
        cache_key = stats_cache_key(
            active_etablissement.id if active_etablissement is not None else None,
            month_start,
        )
        cached = cache.get(cache_key)
        if cached is not None:
            return Response(cached)

        payment_qs = Payment.objects.filter(created_at__date__gte=month_start, is_cancelled=False)
        students_qs = Student.objects.filter(is_archived=False)
        attendance_qs = Attendance.objects.filter(is_absent=True, date__gte=month_start)
        classrooms_qs = ClassRoom.objects.all()
        teachers_qs = Teacher.objects.all()

        if active_etablissement is not None:
            payment_qs = payment_qs.filter(etablissement=active_etablissement)
            students_qs = students_qs.filter(etablissement=active_etablissement)
            attendance_qs = attendance_qs.filter(student__etablissement=active_etablissement)
            classrooms_qs = classrooms_qs.filter(etablissement=active_etablissement)
            teachers_qs = teachers_qs.filter(etablissement=active_etablissement)

        revenue = payment_qs.aggregate(value=Sum("amount"))["value"] or 0
        expenses_qs = Expense.objects.filter(date__gte=month_start)
        if active_etablissement is not None:
            expenses_qs = expenses_qs.filter(etablissement=active_etablissement)
        expenses = expenses_qs.aggregate(value=Sum("amount"))["value"] or 0
        students = students_qs.count()
        absences = attendance_qs.count()
        classroom_count = classrooms_qs.count()
        teacher_count = teachers_qs.count()

        etablissement_payload = None
        if active_etablissement is not None:
            etablissement_payload = {
                "id": active_etablissement.id,
                "name": active_etablissement.name,
                "address": active_etablissement.address,
                "phone": active_etablissement.phone,
                "email": active_etablissement.email,
            }

        payload = {
            "students": students,
            "monthly_revenue": revenue,
            "monthly_expenses": expenses,
            "monthly_profit": revenue - expenses,
            "monthly_absences": absences,
            "classrooms": classroom_count,
            "teachers": teacher_count,
            "active_etablissement": etablissement_payload,
        }
        cache.set(cache_key, payload, STATS_CACHE_SECONDS)
        return Response(payload)
