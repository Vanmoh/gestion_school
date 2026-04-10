from datetime import datetime
import hashlib
import json
import shutil
import tempfile
import zipfile
from pathlib import Path

from django.conf import settings
from django.apps import apps
from django.core import serializers
from django.core.files.uploadedfile import UploadedFile
from django.db import transaction
from django.db.models import Q
from django.http import FileResponse, HttpResponse
from django.utils import timezone
from rest_framework import permissions, status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.response import Response

from apps.accounts.permissions import IsAdminOrDirector
from apps.common.pagination import AuditLogPagination
from .models import ActivityLog, BackupArchive
from .serializers import ActivityLogSerializer, BackupArchiveSerializer


def _pdf_text(value) -> str:
    return str(value or "").encode("latin-1", "replace").decode("latin-1")


def _school_logo_path() -> str | None:
    raw_path = str(getattr(settings, "SCHOOL_LOGO_PATH", "") or "").strip()
    if not raw_path:
        return None

    path = Path(raw_path)
    if not path.is_absolute():
        path = Path(settings.BASE_DIR) / path

    return str(path) if path.exists() else None


class ActivityLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ActivityLog.objects.select_related("user").all()
    serializer_class = ActivityLogSerializer
    pagination_class = AuditLogPagination
    permission_classes = [permissions.IsAuthenticated, IsAdminOrDirector]
    filterset_fields = ["user", "etablissement", "role", "action", "method", "module", "success", "status_code"]
    search_fields = [
        "action",
        "path",
        "target",
        "details",
        "user__username",
        "user__first_name",
        "user__last_name",
        "etablissement__name",
    ]
    ordering_fields = ["created_at", "action", "status_code", "success", "method", "module"]
    ordering = ["-created_at"]

    def _requested_etablissement_id(self):
        raw_value = (
            self.request.headers.get("X-Etablissement-Id")
            or self.request.query_params.get("etablissement")
        )
        if raw_value in (None, ""):
            return None
        try:
            parsed = int(raw_value)
        except (TypeError, ValueError):
            return None
        return parsed if parsed > 0 else None

    def _requested_etablissement_name(self):
        raw_name = (
            self.request.headers.get("X-Etablissement-Name")
            or self.request.query_params.get("etablissement_name")
        )
        if raw_name is None:
            return None
        cleaned = str(raw_name).strip()
        return cleaned or None

    def _requested_etablissement(self):
        from apps.school.models import Etablissement

        requested_id = self._requested_etablissement_id()
        if requested_id:
            etablissement = Etablissement.objects.filter(id=requested_id).first()
            if etablissement:
                return etablissement

        requested_name = self._requested_etablissement_name()
        if not requested_name:
            return None

        etablissement = Etablissement.objects.filter(name__iexact=requested_name).first()
        if etablissement:
            return etablissement

        return Etablissement.objects.filter(name__icontains=requested_name).order_by("name").first()

    def _has_requested_scope(self):
        return self._requested_etablissement_id() is not None or self._requested_etablissement_name() is not None

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        role = getattr(user, "role", "")

        requested_etablissement = self._requested_etablissement()
        if requested_etablissement is not None:
            queryset = queryset.filter(
                Q(etablissement=requested_etablissement)
                | Q(etablissement__isnull=True, user__etablissement=requested_etablissement)
            )
        elif self._has_requested_scope():
            return queryset.none()
        elif role != "super_admin":
            user_etablissement = getattr(user, "etablissement", None)
            if user_etablissement is None:
                return queryset.none()
            queryset = queryset.filter(
                Q(etablissement=user_etablissement)
                | Q(etablissement__isnull=True, user__etablissement=user_etablissement)
            )

        date_from = self.request.query_params.get("date_from")
        date_to = self.request.query_params.get("date_to")

        if date_from:
            try:
                parsed_from = datetime.strptime(date_from, "%Y-%m-%d").date()
                queryset = queryset.filter(created_at__date__gte=parsed_from)
            except ValueError:
                pass

        if date_to:
            try:
                parsed_to = datetime.strptime(date_to, "%Y-%m-%d").date()
                queryset = queryset.filter(created_at__date__lte=parsed_to)
            except ValueError:
                pass

        return queryset

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        queryset = self.filter_queryset(self.get_queryset())[:5000]

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "JournalActivites"

        school_name = getattr(settings, "SCHOOL_NAME", "LYCEE TECHNIQUE OUMAR BAH")
        school_short = getattr(settings, "SCHOOL_SHORT", "LTOB")
        school_level = getattr(settings, "SCHOOL_LEVEL", "1er etage")
        school_phone = getattr(settings, "SCHOOL_PHONE", "")
        logo_path = _school_logo_path()

        columns = [
            "Date",
            "Utilisateur",
            "Role",
            "Action",
            "Methode",
            "Module",
            "Path",
            "Status HTTP",
            "Succes",
            "IP",
        ]
        last_col = len(columns)
        last_col_letter = get_column_letter(last_col)

        sheet.merge_cells(f"A1:{last_col_letter}1")
        sheet.merge_cells(f"A2:{last_col_letter}2")
        sheet.merge_cells(f"A3:{last_col_letter}3")

        sheet["A1"] = school_name
        sheet["A2"] = f"{school_level} | Tel: {school_phone}" if school_phone else school_level
        sheet["A3"] = f"JOURNAL DES ACTIVITES - {school_short}"

        sheet["A1"].font = Font(bold=True, size=16, color="1F3B63")
        sheet["A2"].font = Font(size=11, color="3F3F3F")
        sheet["A3"].font = Font(bold=True, size=12, color="FFFFFF")
        sheet["A3"].fill = PatternFill(fill_type="solid", fgColor="1F4E78")

        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["A2"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["A3"].alignment = Alignment(horizontal="center", vertical="center")

        sheet.row_dimensions[1].height = 28
        sheet.row_dimensions[2].height = 20
        sheet.row_dimensions[3].height = 24

        if logo_path:
            try:
                from openpyxl.drawing.image import Image as XLImage

                logo = XLImage(logo_path)
                logo.width = 50
                logo.height = 50
                logo_anchor_col = get_column_letter(max(1, last_col - 1))
                sheet.add_image(logo, f"{logo_anchor_col}1")
            except Exception:
                pass

        thin_side = Side(style="thin", color="C8CDD3")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        header_row = 5
        for col_index, title in enumerate(columns, start=1):
            cell = sheet.cell(row=header_row, column=col_index, value=title)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="3A6EA5")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        row_index = header_row + 1
        success_total = 0
        failure_total = 0

        for row in queryset:
            user_display = "Anonyme"
            if row.user:
                full_name = row.user.get_full_name().strip()
                user_display = full_name or row.user.username

            if row.success:
                success_total += 1
            else:
                failure_total += 1

            values = [
                row.created_at.strftime("%d/%m/%Y %H:%M:%S"),
                user_display,
                row.role,
                row.action,
                row.method,
                row.module,
                row.path,
                row.status_code,
                "Oui" if row.success else "Non",
                row.ip_address,
            ]

            for col_index, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_index, column=col_index, value=value)
                cell.border = thin_border
                if col_index in (8, 9):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            row_index += 1

        if row_index == header_row + 1:
            sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=last_col)
            empty_cell = sheet.cell(row=row_index, column=1, value="Aucune activite disponible.")
            empty_cell.alignment = Alignment(horizontal="center", vertical="center")
            empty_cell.font = Font(italic=True, color="6B7280")
            empty_cell.border = thin_border
            row_index += 1

        summary_row = row_index + 1
        sheet.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=5)
        summary_label = sheet.cell(row=summary_row, column=1, value="SYNTHESE")
        summary_label.font = Font(bold=True, color="1F3B63")
        summary_label.fill = PatternFill(fill_type="solid", fgColor="E8EEF7")
        summary_label.alignment = Alignment(horizontal="center", vertical="center")
        summary_label.border = thin_border

        success_cell = sheet.cell(row=summary_row, column=6, value=f"Succes: {success_total}")
        success_cell.font = Font(bold=True, color="0F5132")
        success_cell.fill = PatternFill(fill_type="solid", fgColor="D1E7DD")
        success_cell.alignment = Alignment(horizontal="center", vertical="center")
        success_cell.border = thin_border

        failure_cell = sheet.cell(row=summary_row, column=7, value=f"Echecs: {failure_total}")
        failure_cell.font = Font(bold=True, color="842029")
        failure_cell.fill = PatternFill(fill_type="solid", fgColor="F8D7DA")
        failure_cell.alignment = Alignment(horizontal="center", vertical="center")
        failure_cell.border = thin_border

        total_cell = sheet.cell(row=summary_row, column=8, value=f"Total: {success_total + failure_total}")
        total_cell.font = Font(bold=True, color="1F3B63")
        total_cell.fill = PatternFill(fill_type="solid", fgColor="E8EEF7")
        total_cell.alignment = Alignment(horizontal="center", vertical="center")
        total_cell.border = thin_border

        for col_index in range(9, last_col + 1):
            empty = sheet.cell(row=summary_row, column=col_index, value="")
            empty.fill = PatternFill(fill_type="solid", fgColor="E8EEF7")
            empty.border = thin_border

        generated_row = summary_row + 2
        generated_by = request.user.get_full_name().strip() or request.user.username
        generated_at = timezone.localtime().strftime("%d/%m/%Y %H:%M")
        sheet.merge_cells(start_row=generated_row, start_column=1, end_row=generated_row, end_column=last_col)
        generated_cell = sheet.cell(
            row=generated_row,
            column=1,
            value=f"Genere le {generated_at} par {generated_by}",
        )
        generated_cell.font = Font(italic=True, color="6B7280")
        generated_cell.alignment = Alignment(horizontal="left", vertical="center")

        widths = [22, 24, 14, 28, 12, 14, 40, 12, 10, 16]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width

        sheet.freeze_panes = "A6"

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="activity_logs.xlsx"'
        workbook.save(response)
        return response

    @action(detail=False, methods=["get"], url_path="export-pdf")
    def export_pdf(self, request):
        from fpdf import FPDF
        queryset = self.filter_queryset(self.get_queryset())[:1000]

        school_name = getattr(settings, "SCHOOL_NAME", "LYCEE TECHNIQUE OUMAR BAH")
        school_short = getattr(settings, "SCHOOL_SHORT", "LTOB")
        school_level = getattr(settings, "SCHOOL_LEVEL", "1er etage")
        school_phone = getattr(settings, "SCHOOL_PHONE", "")
        logo_path = _school_logo_path()

        pdf = FPDF(orientation="L")
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=10)

        if logo_path:
            try:
                pdf.image(logo_path, x=10, y=8, w=18)
            except Exception:
                pass

        pdf.set_xy(32 if logo_path else 10, 8)
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 6, _pdf_text(school_name), ln=True)

        pdf.set_x(32 if logo_path else 10)
        pdf.set_font("Helvetica", size=9)
        header_line = f"{school_level} | Tel: {school_phone}" if school_phone else school_level
        pdf.cell(0, 5, _pdf_text(header_line), ln=True)

        pdf.set_x(32 if logo_path else 10)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(0, 5, _pdf_text(f"Application: {school_short} - GESTION SCHOOL"), ln=True)

        top_line_y = max(pdf.get_y() + 2, 26)
        pdf.set_draw_color(60, 60, 60)
        pdf.line(10, top_line_y, 287, top_line_y)
        pdf.set_y(top_line_y + 3)

        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 8, _pdf_text("JOURNAL DES ACTIVITES"), ln=True, align="C")
        pdf.ln(1)
        pdf.set_font("Helvetica", size=8)

        headers = ["Date", "Utilisateur", "Role", "Action", "Method", "Module", "Status", "IP"]
        widths = [34, 34, 20, 70, 18, 30, 20, 30]

        pdf.set_fill_color(230, 235, 245)
        pdf.set_font("Helvetica", "B", 8)
        for idx, header in enumerate(headers):
            pdf.cell(widths[idx], 7, _pdf_text(header), border=1, fill=True)
        pdf.ln()

        pdf.set_font("Helvetica", size=8)

        for row in queryset:
            user_display = "Anonyme"
            if row.user:
                full_name = row.user.get_full_name().strip()
                user_display = full_name or row.user.username

            values = [
                row.created_at.strftime("%d/%m/%Y %H:%M"),
                user_display[:30],
                row.role[:18],
                row.action[:58],
                row.method,
                row.module[:28],
                str(row.status_code),
                row.ip_address[:20],
            ]

            for idx, value in enumerate(values):
                pdf.cell(widths[idx], 7, _pdf_text(value), border=1)
            pdf.ln()

        pdf.ln(2)
        pdf.set_font("Helvetica", size=8)
        generated_by = request.user.get_full_name().strip() or request.user.username
        generated_at = timezone.localtime().strftime("%d/%m/%Y %H:%M")
        pdf.cell(0, 5, _pdf_text(f"Genere le {generated_at} par {generated_by}"), ln=True)

        data = bytes(pdf.output())
        response = HttpResponse(data, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="activity_logs.pdf"'
        return response


class BackupArchiveViewSet(viewsets.ModelViewSet):
    queryset = BackupArchive.objects.select_related("created_by", "restored_by", "etablissement").all()
    serializer_class = BackupArchiveSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminOrDirector]
    parser_classes = [JSONParser, MultiPartParser, FormParser]
    http_method_names = ["get", "post"]
    filterset_fields = ["scope", "status", "etablissement", "created_by", "restored_by"]
    search_fields = ["filename", "notes", "etablissement__name", "created_by__username"]
    ordering_fields = ["created_at", "status", "scope", "file_size_bytes"]
    ordering = ["-created_at"]

    def _is_super_admin(self):
        return getattr(self.request.user, "role", "") == "super_admin"

    def _resolve_etablissement(self):
        from apps.school.models import Etablissement

        etablissement_id = self.request.data.get("etablissement") or self.request.data.get("etablissement_id")
        if etablissement_id in (None, ""):
            return getattr(self.request.user, "etablissement", None)
        try:
            parsed = int(etablissement_id)
        except (TypeError, ValueError):
            return None
        return Etablissement.objects.filter(id=parsed).first()

    def _backups_root(self) -> Path:
        root = Path(settings.BASE_DIR) / "backups" / "archives"
        root.mkdir(parents=True, exist_ok=True)
        return root

    def _sha256_file(self, file_path: Path) -> str:
        digest = hashlib.sha256()
        with file_path.open("rb") as f:
            for chunk in iter(lambda: f.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    def _serialize_global(self) -> str:
        serialized_payload = []
        for model in apps.get_models():
            opts = model._meta
            if opts.proxy or not opts.managed:
                continue
            if opts.app_label in {"contenttypes", "sessions", "admin"}:
                continue
            queryset = model.objects.all().order_by("pk")
            if not queryset.exists():
                continue
            serialized_payload.append(serializers.serialize("json", queryset))

        if not serialized_payload:
            return "[]"

        merged = []
        for payload in serialized_payload:
            merged.extend(json.loads(payload))
        return json.dumps(merged, ensure_ascii=False)

    def _serialize_etablissement(self, etablissement) -> str:
        serialized_payload = []
        for model in apps.get_models():
            opts = model._meta
            if opts.proxy or not opts.managed:
                continue
            if opts.app_label in {"contenttypes", "sessions", "admin"}:
                continue

            field_names = {field.name for field in opts.fields}
            queryset = None

            if model.__name__ == "Etablissement":
                queryset = model.objects.filter(pk=etablissement.pk)
            elif opts.app_label == "accounts" and model.__name__ == "User":
                queryset = model.objects.filter(etablissement=etablissement)
            elif "etablissement" in field_names:
                queryset = model.objects.filter(etablissement=etablissement)

            if queryset is None or not queryset.exists():
                continue

            serialized_payload.append(serializers.serialize("json", queryset.order_by("pk")))

        if not serialized_payload:
            return "[]"

        merged = []
        for payload in serialized_payload:
            merged.extend(json.loads(payload))
        return json.dumps(merged, ensure_ascii=False)

    def _build_archive(self, backup: BackupArchive) -> BackupArchive:
        stamp = timezone.localtime().strftime("%Y%m%d_%H%M%S")
        suffix = "global" if backup.scope == BackupArchive.Scope.GLOBAL else f"etab_{backup.etablissement_id}"
        backup_name = f"backup_{suffix}_{stamp}.zip"
        backup_dir = self._backups_root()
        archive_path = backup_dir / backup_name

        backup.status = BackupArchive.Status.RUNNING
        backup.filename = backup_name
        backup.file_path = str(archive_path)
        backup.save(update_fields=["status", "filename", "file_path", "updated_at"])

        payload_json = self._serialize_global()
        if backup.scope == BackupArchive.Scope.ETABLISSEMENT:
            payload_json = self._serialize_etablissement(backup.etablissement)

        manifest = {
            "version": 1,
            "kind": backup.kind,
            "scope": backup.scope,
            "created_at": timezone.localtime().isoformat(),
            "created_by": getattr(backup.created_by, "username", ""),
            "etablissement_id": backup.etablissement_id,
            "include_media": bool(backup.include_media),
        }

        with zipfile.ZipFile(archive_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))
            zf.writestr("data.json", payload_json)

            if backup.include_media:
                media_root = Path(settings.MEDIA_ROOT)
                if media_root.exists() and media_root.is_dir():
                    for item in media_root.rglob("*"):
                        if item.is_file():
                            relative = item.relative_to(media_root)
                            zf.write(item, arcname=str(Path("media") / relative))

        backup.file_size_bytes = archive_path.stat().st_size
        backup.sha256 = self._sha256_file(archive_path)
        backup.manifest = manifest
        backup.status = BackupArchive.Status.COMPLETED
        backup.save(
            update_fields=[
                "file_size_bytes",
                "sha256",
                "manifest",
                "status",
                "updated_at",
            ]
        )
        return backup

    def _restore_from_archive(self, backup: BackupArchive, archive_path: Path):
        restore_notes = []
        with tempfile.TemporaryDirectory(prefix="restore_backup_") as tmp:
            tmp_path = Path(tmp)
            with zipfile.ZipFile(archive_path, "r") as zf:
                zf.extractall(tmp_path)

            data_json = tmp_path / "data.json"
            if not data_json.exists():
                raise ValueError("Archive invalide: data.json manquant.")

            media_dir = tmp_path / "media"

            with transaction.atomic():
                serializers.deserialize("json", data_json.read_text(encoding="utf-8"))
                from django.core.management import call_command

                call_command("loaddata", str(data_json), verbosity=0)

            if media_dir.exists() and media_dir.is_dir():
                media_root = Path(settings.MEDIA_ROOT)
                media_root.mkdir(parents=True, exist_ok=True)
                for src in media_dir.rglob("*"):
                    if not src.is_file():
                        continue
                    dst = media_root / src.relative_to(media_dir)
                    dst.parent.mkdir(parents=True, exist_ok=True)
                    shutil.copy2(src, dst)
                restore_notes.append("Medias restaures.")

        backup.restored_by = self.request.user
        backup.restored_at = timezone.now()
        backup.restore_log = "\n".join(restore_notes) if restore_notes else "Restauration terminee."
        backup.status = BackupArchive.Status.COMPLETED
        backup.save(update_fields=["restored_by", "restored_at", "restore_log", "status", "updated_at"])

    def get_queryset(self):
        queryset = super().get_queryset()
        if self._is_super_admin():
            return queryset

        user_etablissement = getattr(self.request.user, "etablissement", None)
        if user_etablissement is None:
            return queryset.none()
        return queryset.filter(scope=BackupArchive.Scope.ETABLISSEMENT, etablissement=user_etablissement)

    def create(self, request, *args, **kwargs):
        scope = str(request.data.get("scope") or BackupArchive.Scope.ETABLISSEMENT).strip()
        include_media_raw = request.data.get("include_media", True)
        include_media = str(include_media_raw).lower() not in {"0", "false", "no"}
        notes = str(request.data.get("notes") or "").strip()

        if scope not in {BackupArchive.Scope.GLOBAL, BackupArchive.Scope.ETABLISSEMENT}:
            return Response({"detail": "Scope invalide."}, status=status.HTTP_400_BAD_REQUEST)

        if scope == BackupArchive.Scope.GLOBAL and not self._is_super_admin():
            return Response(
                {"detail": "Seul un super admin peut creer une sauvegarde globale."},
                status=status.HTTP_403_FORBIDDEN,
            )

        etablissement = None
        if scope == BackupArchive.Scope.ETABLISSEMENT:
            etablissement = self._resolve_etablissement()
            if etablissement is None:
                return Response(
                    {"detail": "Etablissement introuvable pour la sauvegarde."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if not self._is_super_admin() and getattr(request.user, "etablissement_id", None) != etablissement.id:
                return Response(
                    {"detail": "Vous ne pouvez sauvegarder que votre etablissement."},
                    status=status.HTTP_403_FORBIDDEN,
                )

        backup = BackupArchive.objects.create(
            scope=scope,
            etablissement=etablissement,
            created_by=request.user,
            include_media=include_media,
            notes=notes,
            status=BackupArchive.Status.PENDING,
        )

        try:
            backup = self._build_archive(backup)
        except Exception as exc:
            backup.status = BackupArchive.Status.FAILED
            backup.restore_log = str(exc)
            backup.save(update_fields=["status", "restore_log", "updated_at"])
            return Response({"detail": f"Echec sauvegarde: {exc}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(self.get_serializer(backup).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=["get"], url_path="download")
    def download(self, request, pk=None):
        backup = self.get_object()
        archive_path = Path(str(backup.file_path or "")).expanduser()
        if not archive_path.exists() or not archive_path.is_file():
            return Response({"detail": "Fichier backup introuvable."}, status=status.HTTP_404_NOT_FOUND)

        response = FileResponse(archive_path.open("rb"), as_attachment=True, filename=backup.filename or archive_path.name)
        return response

    @action(detail=True, methods=["post"], url_path="restore")
    def restore(self, request, pk=None):
        backup = self.get_object()
        if backup.scope == BackupArchive.Scope.GLOBAL and not self._is_super_admin():
            return Response(
                {"detail": "Seul un super admin peut restaurer une sauvegarde globale."},
                status=status.HTTP_403_FORBIDDEN,
            )

        archive_path = Path(str(backup.file_path or "")).expanduser()
        if not archive_path.exists() or not archive_path.is_file():
            return Response({"detail": "Fichier backup introuvable."}, status=status.HTTP_404_NOT_FOUND)

        try:
            self._restore_from_archive(backup, archive_path)
        except Exception as exc:
            backup.status = BackupArchive.Status.FAILED
            backup.restore_log = str(exc)
            backup.save(update_fields=["status", "restore_log", "updated_at"])
            return Response({"detail": f"Echec restauration: {exc}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(self.get_serializer(backup).data)

    @action(detail=False, methods=["post"], url_path="upload-restore")
    def upload_restore(self, request):
        uploaded = request.FILES.get("file")
        if not isinstance(uploaded, UploadedFile):
            return Response({"detail": "Fichier requis."}, status=status.HTTP_400_BAD_REQUEST)

        scope = str(request.data.get("scope") or BackupArchive.Scope.ETABLISSEMENT).strip()
        if scope not in {BackupArchive.Scope.GLOBAL, BackupArchive.Scope.ETABLISSEMENT}:
            return Response({"detail": "Scope invalide."}, status=status.HTTP_400_BAD_REQUEST)

        if scope == BackupArchive.Scope.GLOBAL and not self._is_super_admin():
            return Response(
                {"detail": "Seul un super admin peut restaurer globalement."},
                status=status.HTTP_403_FORBIDDEN,
            )

        etablissement = None
        if scope == BackupArchive.Scope.ETABLISSEMENT:
            etablissement = self._resolve_etablissement()
            if etablissement is None:
                return Response(
                    {"detail": "Etablissement introuvable pour la restauration."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if not self._is_super_admin() and getattr(request.user, "etablissement_id", None) != etablissement.id:
                return Response(
                    {"detail": "Vous ne pouvez restaurer que votre etablissement."},
                    status=status.HTTP_403_FORBIDDEN,
                )

        backup_dir = self._backups_root()
        stamp = timezone.localtime().strftime("%Y%m%d_%H%M%S")
        target = backup_dir / f"uploaded_restore_{scope}_{stamp}.zip"
        with target.open("wb") as out:
            for chunk in uploaded.chunks():
                out.write(chunk)

        backup = BackupArchive.objects.create(
            scope=scope,
            etablissement=etablissement,
            created_by=request.user,
            filename=target.name,
            file_path=str(target),
            file_size_bytes=target.stat().st_size,
            sha256=self._sha256_file(target),
            notes=str(request.data.get("notes") or "").strip(),
            status=BackupArchive.Status.RUNNING,
        )

        try:
            self._restore_from_archive(backup, target)
        except Exception as exc:
            backup.status = BackupArchive.Status.FAILED
            backup.restore_log = str(exc)
            backup.save(update_fields=["status", "restore_log", "updated_at"])
            return Response({"detail": f"Echec restauration: {exc}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(self.get_serializer(backup).data, status=status.HTTP_201_CREATED)
